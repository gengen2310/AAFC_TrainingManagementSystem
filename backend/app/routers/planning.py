"""TRGO Planning Module router.

Provides endpoints for annual training planning: planning years,
parade dates, holidays, anchor events, term planner, parade night
builder, scheduled sessions, locations, facilitators (planning view),
conflict detection, and weekly/long-range program output.
"""
from __future__ import annotations
import csv as _csv, io, json as _json, re, uuid
from datetime import date, timedelta
from typing import Optional
from fastapi import APIRouter, Depends, File, Form, HTTPException, Query, UploadFile
from fastapi.responses import StreamingResponse
from pydantic import BaseModel, Field, field_validator, model_validator
from sqlalchemy.orm import Session as DBSession

from ..database import get_db, utcnow, iso_z
from ..models import (
    Wing, Squadron, CurriculumItem, Facilitator, AuditLog, ParadeNight, TrainingArea,
    TrainingClass, SessionAudience, CurriculumPhase,
)
from ..models import Session as TrainingSession
from ..models.planning import (
    PlanningYear, HolidayPeriod, AnchorEvent,
    AnchorPrepRule, AnchorPrepPlan,
    PlanningConflict, PlanningFacilitatorLeave, PlanningNotice,
    CeaImportBatch, CeaActivity, ActivityLocalHide,
    CADET_GROUPS, IMPORTANCE_LEVELS, EVENT_TYPES,
)
# ParadeDate removed (Phase B migration a1c68e84caf5) — router endpoints that
# reference ParadeDate are handled in Task 5.
# ScheduledSession and PlanningLocation models are intentionally NOT imported here:
# both are fully superseded (TrainingSession/TrainingArea are canonical -- see
# docs/qualification/02_architecture_review.md, 03_data_integrity_review.md, and
# QUAL-001/QUAL-002 in docs/remediation/master_gap_register.csv). This router's last
# reference to either (the dead _session_out() serializer, never called) was removed
# 2026-08-08. The model classes and their tables (scheduled_sessions,
# planning_locations) still exist -- retiring them is a schema change requiring
# explicit user authorisation per .claude/rules/capability-preservation.md, not done
# here.
from ..models.training import TimingTemplate, TimingBlock, Activity, ParadeNightTimingSnapshot
from ..models.wing_calendar import WingHQEvent, SquadronEventStatus
from ..dependencies import get_principal
from ..permissions import Principal, require_role, require_can_write_squadron, require_can_view_squadron
from ..services import audit
from ..services import (visible_curriculum_item, scoped_facilitator,
                        scoped_training_area)
from ..services_year import (
    PastYearLocked, ensure_year_context, find_year_context, require_year_writable,
    selectable_years, year_display_name, year_state,
)


def _require_writable_year(db, squadron_id: str, year: int, p) -> None:
    """Translate the service-layer lock into the router's 403 contract."""
    try:
        require_year_writable(db, squadron_id, year, p)
    except PastYearLocked as exc:
        raise HTTPException(403, detail={"error": "past_year_read_only",
                                         "message": str(exc)})
from .timing import _effective_template

router = APIRouter(prefix="/api/planning", tags=["planning"])

# Maps the legacy cadet_group free-text strings to TrainingClass.stage_code values.
# Used when creating SessionAudience rows alongside TrainingSession creation (K-006).
_CADET_GROUP_STAGE_CODE: dict[str, str] = {
    "orientation": "ORI",
    "initial": "INI",
    "junior": "JNR",
    "intermediate": "INT",
    "senior": "SNR",
}
# Reverse mapping: stage_code → cadet_group (for denormalising when class IDs are provided)
_STAGE_CODE_CADET_GROUP: dict[str, str] = {v: k for k, v in _CADET_GROUP_STAGE_CODE.items()}


def _upsert_session_audience(db: DBSession, session_id: str, cadet_group: str,
                              squadron_id: str, training_year_id: str) -> None:
    """Create a SessionAudience row linking session_id to the matching TrainingClass.

    Only creates the row when exactly one active TrainingClass matches the cadet_group's
    stage_code for this squadron/year — skips silently if zero or multiple match
    (multiple classes per stage require an explicit frontend choice, not a guess).
    Idempotent: the unique constraint on (session_id, training_class_id) prevents
    duplicates if called more than once for the same pair.
    """
    stage_code = _CADET_GROUP_STAGE_CODE.get(cadet_group)
    if not stage_code:
        return
    classes = (
        db.query(TrainingClass)
        .filter(
            TrainingClass.squadron_id == squadron_id,
            TrainingClass.training_year_id == training_year_id,
            TrainingClass.stage_code == stage_code,
            TrainingClass.is_archived == False,  # noqa: E712
        )
        .all()
    )
    if len(classes) != 1:
        return
    existing = (
        db.query(SessionAudience)
        .filter(
            SessionAudience.session_id == session_id,
            SessionAudience.training_class_id == classes[0].id,
        )
        .first()
    )
    if existing:
        return
    db.add(SessionAudience(session_id=session_id, training_class_id=classes[0].id))
    db.commit()


def _resolve_scoped_classes(
    db: DBSession, class_ids: list[str] | None, squadron_id: str,
) -> list:
    """Resolve training class ids, refusing any that belong to another Squadron.

    A TrainingClass belongs to exactly one Squadron's one Training Year (see the
    model docstring). PUT /api/sessions/{id}/audience has always enforced that --
    training.py rejects `c.squadron_id != s.squadron_id` with 400
    invalid_training_class. This module's session-create, session-edit and
    assign-mission paths did not: they passed body.training_class_ids straight
    through, so the same resource had two doors and only one was locked. 703
    could post 705's class id and mint a SessionAudience row across the tenancy
    boundary (Part 82: no cross-squadron IDOR, tenancy enforced in the backend).

    Rejecting rather than skipping, and with training.py's exact status and
    error code, because two doors onto one resource answering differently is
    how the gap survived review in the first place.

    Validation happens BEFORE any write. create_session derived cadet_group from
    the first class and committed the session before touching the audience, so a
    late refusal would have left a committed session carrying a foreign class's
    stage_code.
    """
    resolved = []
    for cid in class_ids or []:
        tc = db.get(TrainingClass, cid)
        if not tc or tc.is_archived or tc.squadron_id != squadron_id:
            raise HTTPException(400, detail={
                "error": "invalid_training_class", "training_class_id": cid})
        resolved.append(tc)
    return resolved


def _create_audience_for_class_ids(
    db: DBSession, session_id: str, class_ids: list[str], squadron_id: str,
) -> None:
    """Create SessionAudience rows for explicit training_class_id values.

    Idempotent: existing rows for the same (session_id, training_class_id) pair are
    skipped (the unique constraint would catch them anyway).

    squadron_id is required, not optional. The previous signature took only the
    ids and its docstring claimed classes "that don't exist or belong to a
    different squadron/year are silently skipped" -- the code checked only
    existence, so the tenancy half of that sentence described a check that was
    never written. Callers validate up front via _resolve_scoped_classes; this
    re-checks because a helper that writes tenancy-scoped rows should not depend
    on every caller having remembered.
    """
    for tc_id in class_ids:
        tc = db.get(TrainingClass, tc_id)
        if not tc or tc.squadron_id != squadron_id:
            continue
        exists = (
            db.query(SessionAudience)
            .filter(
                SessionAudience.session_id == session_id,
                SessionAudience.training_class_id == tc_id,
            )
            .first()
        )
        if exists:
            continue
        db.add(SessionAudience(session_id=session_id, training_class_id=tc_id))
    db.commit()


_VALID_SESSION_STATUS = {
    "draft", "planned", "published", "delivered", "delivered_with_issue",
    "cancelled", "cancelled_late", "rescheduled", "not_delivered",
    "requires_review", "blocked", "closed",
}


def _check_version(obj, client_version: int | None) -> None:
    """Raise 409 if the client's version is stale (optimistic locking)."""
    if client_version is not None and obj.version != client_version:
        raise HTTPException(409, detail={
            "error": "version_conflict",
            "current_version": obj.version,
        })


def _parse_json_list(val) -> list:
    """Return val as a list, JSON-parsing it if stored as a TEXT string (Postgres TEXT vs JSON/JSONB mismatch)."""
    if not val:
        return []
    if isinstance(val, list):
        return val
    if isinstance(val, str):
        try:
            result = _json.loads(val)
            return result if isinstance(result, list) else []
        except (_json.JSONDecodeError, TypeError):
            return []
    return []


# ── Annual Program import helpers ────────────────────────────

_PROG_DATE_FMTS = ["%d/%m/%Y", "%Y-%m-%d", "%d-%m-%Y", "%m/%d/%Y", "%d %b %Y"]


def _parse_prog_date(raw: str) -> str | None:
    """Parse a date string to ISO YYYY-MM-DD."""
    if not raw:
        return None
    from datetime import datetime
    for fmt in _PROG_DATE_FMTS:
        try:
            return datetime.strptime(raw.strip(), fmt).strftime("%Y-%m-%d")
        except ValueError:
            pass
    return raw.strip()[:10] or None


def _parse_prog_time(raw: str) -> str | None:
    """Normalise a time string to HH:MM."""
    if not raw:
        return None
    raw = raw.strip()
    parts = raw.split(":")
    if len(parts) >= 2:
        try:
            return f"{int(parts[0]):02d}:{int(parts[1]):02d}"
        except ValueError:
            pass
    return raw[:5] or None


_PARADE_KWORDS = ("parade night", "training night", "parade ngt", "parade nt")


def _is_parade_row(name: str) -> bool:
    n = name.strip().lower()
    return any(kw in n for kw in _PARADE_KWORDS)


def _resolve_unit_sqn(unit_str: str, all_sqns: list) -> "Squadron | None":
    """Match a Unit column value to a Squadron by code or number prefix."""
    if not unit_str:
        return None
    u_up = unit_str.strip().upper()
    for sq in all_sqns:
        if sq.code and sq.code.strip().upper() == u_up:
            return sq
    m = re.match(r'^(\d+)', unit_str.strip())
    if m:
        num = m.group(1)
        for sq in all_sqns:
            if sq.code and re.match(r'^' + re.escape(num) + r'\b', sq.code.strip()):
                return sq
    u_lo = unit_str.strip().lower()
    for sq in all_sqns:
        if sq.name and u_lo in sq.name.lower():
            return sq
    return None


def _parse_program_file(content: bytes, is_xlsx: bool) -> list[dict]:
    """Decode CSV or XLSX bytes into a list of {header: value} dicts."""
    if is_xlsx:
        import openpyxl
        wb = openpyxl.load_workbook(io.BytesIO(content), read_only=True, data_only=True)
        ws = wb.active
        rows_iter = ws.iter_rows(values_only=True)
        header_row = next(rows_iter, None)
        if not header_row:
            return []
        headers = [str(c).strip() if c is not None else "" for c in header_row]
        result = []
        for row in rows_iter:
            if all(v is None for v in row):
                continue
            result.append({h: (str(v) if v is not None else "") for h, v in zip(headers, row)})
        return result
    try:
        text = content.decode("utf-8-sig")
    except UnicodeDecodeError:
        text = content.decode("latin-1")
    reader = _csv.DictReader(io.StringIO(text))
    return [dict(row) for row in reader]


# ─────────────────────────────────────────────────────────────
# RBAC helpers
# ─────────────────────────────────────────────────────────────

_WRITE_BLOCKED = frozenset({"sqn_general", "wing_viewer", "national_viewer", "auditor"})
_NAT_ROLES     = frozenset({"national_admin", "system_admin"})
_WING_ROLES    = frozenset({"wing_admin", *_NAT_ROLES})
_ALL_ADMIN     = frozenset({"sqn_admin", "wing_admin", "national_admin", "system_admin"})


def _require_plan_write(p: Principal) -> None:
    if p.role in _WRITE_BLOCKED:
        raise HTTPException(403, detail={"error": "forbidden"})


def _require_year_access(p: Principal, py: PlanningYear, write: bool = False,
                         db: DBSession | None = None) -> None:
    """Enforce scope: sqn_admin/sqn_general → own sqn; wing_admin → own wing; nat → all.

    Also enforces the past-year lock when `db` is supplied. It is enforced here
    rather than at each of the fifteen year-scoped write endpoints so that a new
    endpoint added later inherits the protection instead of forgetting it.
    """
    if write and p.role in _WRITE_BLOCKED:
        raise HTTPException(403, detail={"error": "forbidden"})
    if write and db is not None and py.unit_id:
        _require_writable_year(db, py.unit_id, py.year, p)
    if p.role in ("sqn_admin", "sqn_general"):
        if py.unit_id != p.squadron_id:
            raise HTTPException(403, detail={"error": "out_of_scope"})
    elif p.role == "wing_admin":
        if py.wing_id != p.wing_id:
            raise HTTPException(403, detail={"error": "out_of_scope"})
    elif p.role in ("wing_viewer", "national_viewer", "auditor"):
        if p.role == "wing_viewer" and py.wing_id != p.wing_id:
            raise HTTPException(403, detail={"error": "out_of_scope"})
    # national_admin, system_admin: unrestricted


def _get_year_or_404(year_id: str, db: DBSession) -> PlanningYear:
    py = db.get(PlanningYear, year_id)
    if not py:
        raise HTTPException(404, detail={"error": "planning_year_not_found"})
    return py



# ─────────────────────────────────────────────────────────────
# Serialisers
# ─────────────────────────────────────────────────────────────

def _year_out(py: PlanningYear, unit_code: str | None = None,
              unit_name: str | None = None, wing_code: str | None = None) -> dict:
    return {
        "planning_year_id": py.id, "unit_id": py.unit_id, "wing_id": py.wing_id,
        "year": py.year, "name": py.name, "active_status": py.active_status,
        # Derived, not stored. state is filled by the caller that knows
        # which squadron the year is being viewed for; materialised says
        # a row exists, and is False for logical-only years.
        "state": None, "materialised": True,
        "unit_code": unit_code, "unit_name": unit_name, "wing_code": wing_code,
        "created_by": py.created_by, "updated_by": py.updated_by,
        "created_at": iso_z(py.created_at) if py.created_at else None,
        "updated_at": iso_z(py.updated_at) if py.updated_at else None,
        "version": py.version,
    }


def _night_out_as_date(pn: ParadeNight) -> dict:
    """Serialise a ParadeNight in the parade-dates response shape.

    parade_date_id = pn.id — backward-compat alias used by the React PW's
    ~40 references to pd.parade_date_id. Do not rename this field.
    parade_night_id = pn.id — same value; kept for callers that used the
    old linked-night field and now read the same record.
    """
    return {
        "parade_date_id": pn.id,
        "planning_year_id": pn.planning_year_id,
        "unit_id": pn.squadron_id,
        "parade_date": pn.date,
        "parade_type": pn.parade_type,
        "is_active": pn.is_active,
        "notes": pn.notes,
        "term": pn.term,
        "week_number": pn.week_number,
        "cancellation_reason": pn.cancellation_reason,
        "parade_night_id": pn.id,
    }


def _holiday_out(h: HolidayPeriod) -> dict:
    return {
        "holiday_id": h.id, "planning_year_id": h.planning_year_id,
        "jurisdiction": h.jurisdiction, "name": h.name,
        "start_date": h.start_date, "end_date": h.end_date,
        "holiday_type": getattr(h, "holiday_type", "school_holiday"),
        "affects_parade": h.affects_parade, "notes": h.notes,
    }


def _anchor_out(a: AnchorEvent) -> dict:
    return {
        "anchor_event_id": a.id, "planning_year_id": a.planning_year_id,
        "owning_level": a.owning_level, "wing_id": a.wing_id, "unit_id": a.unit_id,
        "event_name": a.event_name, "event_type": a.event_type, "importance": a.importance,
        "importance_level": getattr(a, "importance_level", None),
        "start_date": a.start_date, "end_date": a.end_date,
        # Flat audience fields (expected by frontend TS type)
        "audience_orientation": a.audience_orientation,
        "audience_initial": a.audience_initial,
        "audience_junior": a.audience_junior,
        "audience_intermediate": a.audience_intermediate,
        "audience_senior": a.audience_senior,
        "audience_staff_only": getattr(a, "audience_staff_only", False),
        "audience_proficient": getattr(a, "audience_proficient", False),
        "audience_first_years": getattr(a, "audience_first_years", False),
        # Nested audience object (kept for backward compatibility)
        "audience": {
            "orientation": a.audience_orientation, "initial": a.audience_initial,
            "junior": a.audience_junior, "intermediate": a.audience_intermediate,
            "senior": a.audience_senior,
            "staff_only": getattr(a, "audience_staff_only", False),
            "proficient": getattr(a, "audience_proficient", False),
            "first_years": getattr(a, "audience_first_years", False),
        },
        "cea_activity_id": getattr(a, "cea_activity_id", None),
        "nomination_end_date": getattr(a, "nomination_end_date", None),
        "unit_name": getattr(a, "unit_name", None),
        "planning_impact": a.planning_impact,
        "readiness_requirements": a.readiness_requirements,
        "notes": a.notes, "is_archived": a.is_archived,
        "created_by": a.created_by,
        "created_at": iso_z(a.created_at) if a.created_at else None,
        "version": a.version,
    }


def _location_out(loc: TrainingArea) -> dict:
    # Rooms merger (master transformation plan, Phase 1): Planning Workspace's
    # Rooms tab now reads/writes the same `training_areas` table connected-
    # frontend's Resources page uses, instead of the separate `planning_locations`
    # table. This also fixes a real, live bug: create_session/update_session's
    # room-resolution (`db.get(TrainingArea, body.location_id)`) only ever looked
    # up TrainingArea rows — a location_id from the old PlanningLocation-backed
    # endpoint silently failed to resolve, so a room picked in Planning Workspace
    # would not actually attach to the session. The response shape below is kept
    # identical to the old PlanningLocation-backed JSON so no frontend changes
    # are required.
    return {
        "location_id": loc.id, "unit_id": loc.squadron_id, "name": loc.name,
        "location_type": loc.type, "capacity": loc.capacity,
        "notes": loc.notes, "active_status": loc.active_status,
        "capabilities": loc.capabilities,
    }


def _real_session_out(
    s: TrainingSession, db: DBSession,
    ci_tier: "dict[str, dict] | None" = None,
) -> dict:
    """Serialize a real training Session in the builder grid format."""
    room_name = s.training_area_name_at_time
    if not room_name and s.training_area_id:
        ra = db.get(TrainingArea, s.training_area_id)
        if ra:
            room_name = ra.name
    asst_name: str | None = None
    if s.assistant_facilitator_id:
        af = db.get(Facilitator, s.assistant_facilitator_id)
        if af:
            asst_name = " ".join(x for x in [af.current_rank, af.first_name, af.last_name] if x)
    # CLASS-21: curriculum core_status and is_optional for Foundation/Extension/Optional PW filters.
    # ci_tier pre-loaded by bulk callers (long-range endpoint); falls back to
    # identity-map PK lookup when not supplied (weekly-program, term-planner).
    core_status: str | None = None
    is_optional: bool = False
    if s.curriculum_item_id:
        if ci_tier is not None:
            t = ci_tier.get(s.curriculum_item_id)
            if t:
                core_status = t["core_status"]
                is_optional = t.get("is_optional", False)
        else:
            ci_obj = db.get(CurriculumItem, s.curriculum_item_id)
            if ci_obj:
                core_status = ci_obj.core_status
                is_optional = ci_obj.is_optional
    return {
        "session_id": s.id,
        "parade_night_id": s.parade_night_id,
        "squadron_id": s.squadron_id,
        "cadet_group": s.cadet_group,
        "session_number": s.period_number,
        "part_number": s.part_number,
        "curriculum_id": s.curriculum_item_id,
        "curriculum_code": s.curriculum_code_at_time,
        "curriculum_title": s.curriculum_title_at_time,
        "activity_title": s.curriculum_title_at_time or s.custom_title,
        "facilitator_id": s.facilitator_id,
        "facilitator_name": s.facilitator_display_name_at_time,
        "assistant_facilitator_id": s.assistant_facilitator_id,
        "assistant_facilitator_name": asst_name,
        "location_id": s.training_area_id,
        "location_name": room_name,
        "status": s.status,
        "notes": s.delivery_notes,
        "is_combined": db.query(SessionAudience).filter(SessionAudience.session_id == s.id).count() > 1,
        "override_conflict": False,
        "created_at": iso_z(s.created_at) if s.created_at else None,
        "version": s.version,
        "core_status": core_status,
        "is_optional": is_optional,
    }


def _conflict_out(c: PlanningConflict) -> dict:
    return {
        "conflict_id": c.id, "planning_year_id": c.planning_year_id,
        "parade_night_id": c.parade_night_id,
        "scheduled_session_id": c.scheduled_session_id,
        "conflict_type": c.conflict_type, "severity": c.severity,
        "message": c.message, "is_resolved": c.is_resolved,
        "override_reason": c.override_reason,
        "created_at": iso_z(c.created_at) if c.created_at else None,
    }


# ─────────────────────────────────────────────────────────────
# Planning Years
# ─────────────────────────────────────────────────────────────

class PlanningYearIn(BaseModel):
    year: int
    name: str
    unit_id: Optional[str] = None
    wing_id: Optional[str] = None
    active_status: bool = True


class PlanningYearUpdateIn(BaseModel):
    name: Optional[str] = None
    active_status: Optional[bool] = None
    version: Optional[int] = None


@router.get("/years")
def list_planning_years(
    unit_id: Optional[str] = None,
    wing_id: Optional[str] = None,
    include_unmaterialised: bool = False,
    db: DBSession = Depends(get_db),
    p: Principal = Depends(get_principal),
):
    """List planning years. `state` is derived and always present.

    Logical years -- selectable years with no row yet -- are opt-in via
    include_unmaterialised. They carry planning_year_id=None, and existing
    consumers build sub-resource URLs (/years/{id}/holidays) straight from
    this list, so returning them by default would break every one of them.
    """
    q = db.query(PlanningYear)
    if p.role in ("sqn_admin", "sqn_general"):
        q = q.filter(PlanningYear.unit_id == p.squadron_id)
    elif p.role in ("wing_admin", "wing_viewer"):
        q = q.filter(PlanningYear.wing_id == p.wing_id)
    if unit_id:
        q = q.filter(PlanningYear.unit_id == unit_id)
    if wing_id:
        q = q.filter(PlanningYear.wing_id == wing_id)
    years = q.order_by(PlanningYear.year.desc()).all()

    # Preload squadrons and wings for label enrichment
    sqn_ids = {py.unit_id for py in years if py.unit_id}
    wing_ids = {py.wing_id for py in years if py.wing_id}
    sqns = {s.id: s for s in db.query(Squadron).filter(Squadron.id.in_(sqn_ids)).all()} if sqn_ids else {}
    wings = {w.id: w for w in db.query(Wing).filter(Wing.id.in_(wing_ids)).all()} if wing_ids else {}

    out = []
    for py in years:
        sq = sqns.get(py.unit_id) if py.unit_id else None
        wg = wings.get(py.wing_id) if py.wing_id else None
        out.append(_year_out(py,
            unit_code=sq.code if sq else None,
            unit_name=sq.name if sq else None,
            wing_code=wg.code if wg else None,
        ))

    sqn_id = p.squadron_id if p.role in ("sqn_admin", "sqn_general") else unit_id

    # Years the user may select that have no row yet. Listing one does NOT
    # create it -- materialisation happens on write, in ensure_year_context.
    if include_unmaterialised and sqn_id:
        have = {row["year"] for row in out}
        for y in selectable_years(db, sqn_id):
            if y not in have:
                out.append({
                    "planning_year_id": None, "unit_id": sqn_id, "wing_id": None,
                    "year": y, "name": year_display_name(y), "active_status": True,
                    "state": year_state(db, sqn_id, y), "materialised": False,
                    "unit_code": None, "unit_name": None, "wing_code": None,
                    "created_by": None, "updated_by": None,
                    "created_at": None, "updated_at": None, "version": 0,
                })

    # Every row carries a derived state. Wing and national years have no
    # squadron to resolve a timezone against, so theirs stays None rather than
    # being guessed from the server's clock.
    for row in out:
        if row["state"] is None:
            sid = row["unit_id"] or sqn_id
            if sid:
                row["state"] = year_state(db, sid, row["year"])
    out.sort(key=lambda r: r["year"], reverse=True)
    return out


class CopySetupIn(BaseModel):
    source_year: int
    target_year: int
    copy_classes: bool = True
    copy_parade_pattern: bool = False


@router.post("/years/copy-setup")
def copy_setup(body: CopySetupIn, db: DBSession = Depends(get_db),
               p: Principal = Depends(get_principal)):
    """Copy configuration from one year into another.

    This does not create a year in the user's sense -- the year already exists
    as calendar context. It materialises that year's container and seeds
    configuration into it.

    Copies class structure and, optionally, the parade recurrence pattern.
    Never sessions, outcomes, progress, attendance, audit history or published
    status, and never date-shifted holidays -- those are re-imported, because a
    shifted public holiday is wrong far more often than it is right.
    """
    sqn_id = p.squadron_id
    if not sqn_id:
        raise HTTPException(400, detail={
            "error": "squadron_required",
            "message": "Copy setup runs for a squadron. Sign in to a squadron account."})
    require_can_write_squadron(p, sqn_id, p.wing_id)

    source = find_year_context(db, sqn_id, body.source_year)
    if source is None:
        raise HTTPException(404, detail={
            "error": "source_year_not_configured",
            "message": f"{body.source_year} has nothing set up to copy."})
    if body.target_year == body.source_year:
        raise HTTPException(400, detail={
            "error": "same_year",
            "message": "Choose a different year to copy into."})

    _require_writable_year(db, sqn_id, body.target_year, p)
    target = ensure_year_context(db, sqn_id, body.target_year, p.user_id)

    classes_copied = 0
    if body.copy_classes:
        existing = {
            c.display_name for c in db.query(TrainingClass).filter(
                TrainingClass.training_year_id == target.id,
                TrainingClass.is_archived == False).all()  # noqa: E712
        }
        for c in db.query(TrainingClass).filter(
                TrainingClass.training_year_id == source.id,
                TrainingClass.is_archived == False).all():  # noqa: E712
            if c.display_name in existing:
                continue        # re-running must not duplicate the structure
            db.add(TrainingClass(
                id=str(uuid.uuid4()), squadron_id=c.squadron_id,
                training_year_id=target.id, training_stage_id=c.training_stage_id,
                stage_code=c.stage_code, display_name=c.display_name,
                class_number=c.class_number, expected_count=c.expected_count,
                created_at=utcnow(), updated_at=utcnow()))
            classes_copied += 1

    db.commit()
    audit(db, p, object_type="planning_year", object_id=target.id,
          action="copy_setup",
          new={"source_year": body.source_year, "target_year": body.target_year,
               "classes_copied": classes_copied})
    return {"ok": True, "planning_year_id": target.id,
            "classes_copied": classes_copied, "sessions_copied": 0}


@router.get("/year-context")
def get_year_context(
    squadron_id: str, year: int,
    db: DBSession = Depends(get_db),
    p: Principal = Depends(get_principal),
):
    """A year context, whether or not a row exists for it.

    Deliberately does NOT materialise: a read must remain a read. Callers that
    need a row call a write endpoint, which uses ensure_year_context.
    """
    sqn = db.get(Squadron, squadron_id)
    if sqn is None:
        raise HTTPException(404, detail={"error": "not_found",
                                         "message": "Unknown squadron."})
    require_can_view_squadron(p, squadron_id, sqn.wing_id)
    py = find_year_context(db, squadron_id, year)
    return {
        "squadron_id": squadron_id, "year": year,
        "state": year_state(db, squadron_id, year),
        "materialised": py is not None,
        "planning_year_id": py.id if py else None,
        "name": py.name if py else year_display_name(year),
    }


@router.post("/years")
def create_planning_year(
    body: PlanningYearIn,
    db: DBSession = Depends(get_db),
    p: Principal = Depends(get_principal),
):
    _require_plan_write(p)
    unit_id = body.unit_id
    wing_id = body.wing_id
    if p.role == "sqn_admin":
        unit_id = p.squadron_id
        wing_id = p.wing_id
    elif p.role == "wing_admin":
        wing_id = p.wing_id
        if unit_id and unit_id != p.squadron_id:
            sqn = db.get(Squadron, unit_id)
            if not sqn or sqn.wing_id != p.wing_id:
                raise HTTPException(403, detail={"error": "out_of_scope"})
    # A squadron-scoped plan is a delegated write on that squadron's data --
    # require the same Proxy/Delegated Intervention state every other
    # squadron-scoped write in this app requires (require_can_write_squadron),
    # not just a bare role check. Previously wing_admin/national_admin/
    # system_admin could create a year for any unit_id with no proxy/
    # intervention at all (Stage 10, 2026-08-05).
    if unit_id:
        require_can_write_squadron(p, unit_id, wing_id)
    # REM-134: one planning year per (unit, year). There was no check here at all,
    # while POST /years/{id}/rollover has always returned 409 for exactly this --
    # the two creation paths disagreed, and repeated calls silently produced
    # duplicates that every downstream year selector then listed several times.
    # 409 matches rollover rather than inventing a second status for one condition.
    if unit_id:
        # Active rows only, matching uq_planning_years_unit_year_active. An
        # archived year of the same number is not a conflict -- archiving one and
        # creating a replacement is a supported workflow.
        dupe = db.query(PlanningYear).filter(
            PlanningYear.unit_id == unit_id,
            PlanningYear.year == body.year,
            PlanningYear.active_status == True,  # noqa: E712
        ).first()
        if dupe:
            raise HTTPException(409, detail={
                "error": "planning_year_already_exists",
                "existing_id": dupe.id,
                "message": f"Training year {body.year} already exists for this unit.",
            })
    py = PlanningYear(
        id=str(uuid.uuid4()), year=body.year, name=body.name,
        unit_id=unit_id, wing_id=wing_id, active_status=body.active_status,
        created_by=p.user_id, created_at=utcnow(), updated_at=utcnow(),
    )
    db.add(py); db.commit()
    audit(db, p, object_type="planning_year", object_id=py.id, action="create",
          new={"year": body.year, "name": body.name})
    # Auto-create the 5 standard training classes when creating a squadron-scoped year.
    # Wing/national-scoped years have no squadron_id to attach classes to, so skip them.
    if py.unit_id:
        _AUTO_STAGE_DEFAULTS = [
            ("ORI", "Orientation",   1),
            ("INI", "Initial",       2),
            ("JNR", "Junior",        3),
            ("INT", "Intermediate",  4),
            ("SNR", "Senior",        5),
        ]
        start = f"{py.year}-01-01"
        for code, name, seq in _AUTO_STAGE_DEFAULTS:
            tc = TrainingClass(
                id=str(uuid.uuid4()),
                squadron_id=py.unit_id,
                training_year_id=py.id,
                training_stage_id=None,
                stage_code=code,
                display_name=name,
                class_number=seq,
                start_date=start,
                end_date=None,
                created_at=utcnow(),
                updated_at=utcnow(),
            )
            db.add(tc)
        db.commit()
    sq = db.get(Squadron, py.unit_id) if py.unit_id else None
    wg = db.get(Wing, py.wing_id) if py.wing_id else None
    return _year_out(py,
        unit_code=sq.code if sq else None, unit_name=sq.name if sq else None,
        wing_code=wg.code if wg else None)


@router.get("/years/{year_id}")
def get_planning_year(
    year_id: str,
    db: DBSession = Depends(get_db),
    p: Principal = Depends(get_principal),
):
    py = _get_year_or_404(year_id, db)
    _require_year_access(p, py)
    sq = db.get(Squadron, py.unit_id) if py.unit_id else None
    wg = db.get(Wing, py.wing_id) if py.wing_id else None
    return _year_out(py,
        unit_code=sq.code if sq else None, unit_name=sq.name if sq else None,
        wing_code=wg.code if wg else None)


@router.patch("/years/{year_id}")
def update_planning_year(
    year_id: str,
    body: PlanningYearUpdateIn,
    db: DBSession = Depends(get_db),
    p: Principal = Depends(get_principal),
):
    py = _get_year_or_404(year_id, db)
    _require_year_access(p, py, write=True, db=db)
    # Structural edits to the year entity itself (name/active_status) require
    # Proxy Mode (wing_admin) or Delegated Intervention (national_admin/system_admin)
    # when the year is squadron-scoped. Content operations (CEA imports, parade
    # dates) carry their own require_can_write_squadron calls at their endpoints.
    if py.unit_id:
        require_can_write_squadron(p, py.unit_id, py.wing_id)
    _check_version(py, body.version)
    if body.name is not None:
        py.name = body.name
    if body.active_status is not None:
        py.active_status = body.active_status
    py.updated_by = p.user_id; py.updated_at = utcnow()
    py.version += 1
    db.commit()
    audit(db, p, object_type="planning_year", object_id=py.id, action="update")
    sq = db.get(Squadron, py.unit_id) if py.unit_id else None
    wg = db.get(Wing, py.wing_id) if py.wing_id else None
    return _year_out(py,
        unit_code=sq.code if sq else None, unit_name=sq.name if sq else None,
        wing_code=wg.code if wg else None)


@router.delete("/years/{year_id}")
def delete_planning_year(
    year_id: str,
    db: DBSession = Depends(get_db),
    p: Principal = Depends(get_principal),
):
    """Permanent delete -- only when a dependency check shows zero linked
    records of any kind. This is additive to (not a replacement for) the
    existing archive path (PATCH .../years/{id} with active_status=false):
    archive remains the default whenever any dependent exists."""
    py = _get_year_or_404(year_id, db)
    _require_year_access(p, py, write=True, db=db)
    if py.unit_id:
        require_can_write_squadron(p, py.unit_id, py.wing_id)

    dependents = {
        "parade_dates": db.query(ParadeNight).filter(ParadeNight.planning_year_id == year_id).count(),
        "holidays": db.query(HolidayPeriod).filter(HolidayPeriod.planning_year_id == year_id).count(),
        "anchor_events": db.query(AnchorEvent).filter(AnchorEvent.planning_year_id == year_id).count(),
        "notices": db.query(PlanningNotice).join(
            ParadeNight, PlanningNotice.parade_night_id == ParadeNight.id
        ).filter(ParadeNight.planning_year_id == year_id).count(),
        "cea_activities": db.query(CeaActivity).filter(CeaActivity.planning_year_id == year_id).count(),
        "cea_import_batches": db.query(CeaImportBatch).filter(CeaImportBatch.planning_year_id == year_id).count(),
        "facilitator_leave": db.query(PlanningFacilitatorLeave).filter(PlanningFacilitatorLeave.planning_year_id == year_id).count(),
        "conflicts": db.query(PlanningConflict).filter(PlanningConflict.planning_year_id == year_id).count(),
    }
    blockers = {k: v for k, v in dependents.items() if v > 0}
    if blockers:
        raise HTTPException(409, detail={
            "error": "has_dependents", "dependents": blockers,
            "message": "This Training Year has linked records and cannot be permanently deleted. Archive it instead.",
        })

    name, year_num = py.name, py.year
    db.delete(py)
    db.commit()
    audit(db, p, object_type="planning_year", object_id=year_id, action="delete",
          old={"name": name, "year": year_num})
    return {"ok": True}


# ─────────────────────────────────────────────────────────────
# Parade Dates
# ─────────────────────────────────────────────────────────────

class ParadeDateIn(BaseModel):
    parade_date: str  # ISO YYYY-MM-DD
    parade_type: str = "standard"
    is_active: bool = True
    notes: Optional[str] = None

    @field_validator("parade_date")
    @classmethod
    def _validate_parade_date(cls, v: str) -> str:
        _validate_iso_date(v, "parade_date")
        return v


class GenerateParadeDatesIn(BaseModel):
    weekday: int          # 0=Mon … 6=Sun
    start_date: str       # ISO YYYY-MM-DD
    end_date: str | None = None   # ISO YYYY-MM-DD; omit if max_repeats given
    parade_type: str = "standard"
    exclude_holidays: bool = True
    frequency: str = "weekly"          # weekly | fortnightly | monthly | yearly | daily
    excluded_dates: list[str] = []     # specific ISO dates to skip
    max_repeats: int | None = None     # alternative to end_date
    parade_start_time: str | None = None   # HH:MM override; falls back to squadron default
    parade_end_time: str | None = None     # HH:MM override; falls back to squadron default


@router.get("/years/{year_id}/parade-dates")
def list_parade_dates(
    year_id: str,
    db: DBSession = Depends(get_db),
    p: Principal = Depends(get_principal),
):
    py = _get_year_or_404(year_id, db)
    _require_year_access(p, py)
    rows = db.query(ParadeNight).filter(ParadeNight.planning_year_id == year_id)\
             .order_by(ParadeNight.date).all()
    # Annotate with holiday flags
    holidays = db.query(HolidayPeriod).filter(
        HolidayPeriod.planning_year_id == year_id,
        HolidayPeriod.affects_parade == True,  # noqa: E712
    ).all()
    def in_holiday(d: str) -> bool:
        for h in holidays:
            if h.start_date <= d <= h.end_date:
                return True
        return False
    out = []
    for pn in rows:
        r = _night_out_as_date(pn)
        r["in_holiday"] = in_holiday(pn.date)
        out.append(r)
    return out


@router.post("/years/{year_id}/parade-dates")
def add_parade_date(
    year_id: str,
    body: ParadeDateIn,
    db: DBSession = Depends(get_db),
    p: Principal = Depends(get_principal),
):
    py = _get_year_or_404(year_id, db)
    _require_year_access(p, py, write=True, db=db)
    if py.unit_id is None:
        raise HTTPException(400, detail={
            "error": "wing_national_year_not_supported",
            "message": "Parade dates cannot be added to wing or national planning years through this endpoint.",
        })
    sq = db.get(Squadron, py.unit_id)
    if sq is None:
        raise HTTPException(400, detail={"error": "squadron_not_found"})
    require_can_write_squadron(p, py.unit_id, sq.wing_id)
    # Idempotent: if a night already exists for this date in THIS year, return it.
    # Two different planning years may share the same calendar date (valid scenario).
    existing = db.query(ParadeNight).filter(
        ParadeNight.planning_year_id == year_id,
        ParadeNight.date == body.parade_date,
        ParadeNight.is_archived == False,  # noqa: E712
    ).first()
    if existing:
        return _night_out_as_date(existing)
    pn = ParadeNight(
        id=str(uuid.uuid4()),
        squadron_id=py.unit_id, wing_id=sq.wing_id,
        date=body.parade_date, planning_year_id=year_id,
        parade_type=body.parade_type or "standard",
        is_active=body.is_active if body.is_active is not None else True,
        notes=body.notes,
        created_by=p.user_id, created_at=utcnow(), updated_at=utcnow(),
    )
    db.add(pn); db.commit()
    audit(db, p, object_type="parade_night", object_id=pn.id, action="create",
          new={"date": body.parade_date, "via": "planning_add_parade_date"})
    return _night_out_as_date(pn)


def _compute_candidate_dates(body: GenerateParadeDatesIn, holidays: list) -> list[str]:
    """Return ISO date strings that should become parade dates given the body parameters.

    Supports weekly, fortnightly, monthly (first weekday in month), and daily frequencies.
    excluded_dates and max_repeats are applied here. Holiday exclusion is optional.
    """
    try:
        start = date.fromisoformat(body.start_date)
    except ValueError:
        raise HTTPException(400, detail={"error": "invalid_date_format"})

    end: date | None = None
    if body.end_date:
        try:
            end = date.fromisoformat(body.end_date)
        except ValueError:
            raise HTTPException(400, detail={"error": "invalid_date_format"})

    if end is None and body.max_repeats is None:
        raise HTTPException(400, detail={
            "error": "end_date_or_max_repeats_required",
            "message": "Provide either end_date or max_repeats.",
        })

    excluded_set = set(body.excluded_dates or [])

    def in_holiday(d: date) -> bool:
        ds = d.isoformat()
        return any(h.start_date <= ds <= h.end_date for h in holidays)

    freq = (body.frequency or "weekly").lower()
    candidates: list[str] = []
    d = start
    last_occurrence: date | None = None

    while True:
        if end and d > end:
            break
        if body.max_repeats is not None and len(candidates) >= body.max_repeats:
            break

        include = False
        if freq == "daily":
            include = True
        elif freq in ("weekly", "fortnightly"):
            if d.weekday() == body.weekday:
                if freq == "weekly":
                    include = True
                else:
                    # fortnightly: every second occurrence
                    if last_occurrence is None or (d - last_occurrence).days >= 14:
                        include = True
        elif freq == "monthly":
            # First occurrence of weekday in the calendar month
            if d.weekday() == body.weekday:
                # Is this the first occurrence of this weekday in the month?
                if d.day <= 7:
                    include = True
        elif freq == "yearly":
            # Same calendar month/day as the start date, each year.
            if d.month == start.month and d.day == start.day:
                include = True
        else:
            # Unknown frequency falls back to weekly
            if d.weekday() == body.weekday:
                include = True

        if include:
            ds = d.isoformat()
            skip = ds in excluded_set
            if body.exclude_holidays and in_holiday(d):
                skip = True
            if not skip:
                candidates.append(ds)
                last_occurrence = d

        d += timedelta(days=1)

    return candidates


def _compute_candidate_dates_classified(body: GenerateParadeDatesIn, holidays: list) -> list[dict]:
    """Like _compute_candidate_dates, but returns every date the recurrence
    pattern touches (not just the ones that would be created), each tagged
    with why it would or would not be created.

    original_instruction.md Section 9 requires the preview to classify each
    candidate rather than silently dropping holiday-conflicting and
    explicitly-skipped dates from the list with no explanation. This is a
    read-only, additive sibling of _compute_candidate_dates -- the write path
    (generate_parade_dates) still calls the original function unchanged, so
    this cannot alter what actually gets created.
    """
    try:
        start = date.fromisoformat(body.start_date)
    except ValueError:
        raise HTTPException(400, detail={"error": "invalid_date_format"})

    end: date | None = None
    if body.end_date:
        try:
            end = date.fromisoformat(body.end_date)
        except ValueError:
            raise HTTPException(400, detail={"error": "invalid_date_format"})

    if end is None and body.max_repeats is None:
        raise HTTPException(400, detail={
            "error": "end_date_or_max_repeats_required",
            "message": "Provide either end_date or max_repeats.",
        })

    excluded_set = set(body.excluded_dates or [])

    def in_holiday(d: date) -> bool:
        ds = d.isoformat()
        return any(h.start_date <= ds <= h.end_date for h in holidays)

    freq = (body.frequency or "weekly").lower()
    rows: list[dict] = []
    included_count = 0
    d = start
    last_occurrence: date | None = None

    while True:
        if end and d > end:
            break
        if body.max_repeats is not None and included_count >= body.max_repeats:
            break

        include = False
        if freq == "daily":
            include = True
        elif freq in ("weekly", "fortnightly"):
            if d.weekday() == body.weekday:
                if freq == "weekly":
                    include = True
                else:
                    if last_occurrence is None or (d - last_occurrence).days >= 14:
                        include = True
        elif freq == "monthly":
            if d.weekday() == body.weekday and d.day <= 7:
                include = True
        elif freq == "yearly":
            if d.month == start.month and d.day == start.day:
                include = True
        else:
            if d.weekday() == body.weekday:
                include = True

        if include:
            ds = d.isoformat()
            if ds in excluded_set:
                rows.append({"date": ds, "status": "explicitly_skipped"})
            elif body.exclude_holidays and in_holiday(d):
                rows.append({"date": ds, "status": "holiday_conflict"})
            else:
                rows.append({"date": ds, "status": "will_create"})
                included_count += 1
                last_occurrence = d

        d += timedelta(days=1)

    return rows


@router.post("/years/{year_id}/preview-parade-dates")
def preview_parade_dates(
    year_id: str,
    body: GenerateParadeDatesIn,
    db: DBSession = Depends(get_db),
    p: Principal = Depends(get_principal),
):
    """Return every candidate date the pattern touches, classified (not just
    which already exist) -- see _compute_candidate_dates_classified."""
    py = _get_year_or_404(year_id, db)
    _require_year_access(p, py)
    holidays = db.query(HolidayPeriod).filter(
        HolidayPeriod.planning_year_id == year_id,
        HolidayPeriod.affects_parade == True,  # noqa: E712
    ).all() if body.exclude_holidays else []
    existing = {
        pn.date for pn in
        db.query(ParadeNight).filter(ParadeNight.planning_year_id == year_id).all()
    }
    classified = _compute_candidate_dates_classified(body, holidays)
    rows = []
    for row in classified:
        status = row["status"]
        if status == "will_create" and row["date"] in existing:
            status = "already_exists"
        # "new" is kept for backward compatibility with existing callers that
        # only check r.new; it is true only for dates that will actually be
        # created by generate-parade-dates.
        rows.append({"date": row["date"], "status": status, "new": status == "will_create"})
    return {
        "dates": rows,
        "new_count": sum(1 for r in rows if r["new"]),
        "total": len(rows),
    }


@router.post("/years/{year_id}/generate-parade-dates")
def generate_parade_dates(
    year_id: str,
    body: GenerateParadeDatesIn,
    db: DBSession = Depends(get_db),
    p: Principal = Depends(get_principal),
):
    py = _get_year_or_404(year_id, db)
    _require_year_access(p, py, write=True, db=db)
    if py.unit_id is None:
        raise HTTPException(400, detail={
            "error": "wing_national_year_not_supported",
            "message": "Parade dates cannot be added to wing or national planning years through this endpoint.",
        })
    sq = db.get(Squadron, py.unit_id)
    require_can_write_squadron(p, py.unit_id, sq.wing_id if sq else py.wing_id)
    holidays = db.query(HolidayPeriod).filter(
        HolidayPeriod.planning_year_id == year_id,
        HolidayPeriod.affects_parade == True,  # noqa: E712
    ).all() if body.exclude_holidays else []

    existing_dates = {
        pn.date for pn in db.query(ParadeNight).filter(
            ParadeNight.planning_year_id == year_id
        ).all()
    }
    candidates = _compute_candidate_dates(body, holidays)
    created = []
    for ds in candidates:
        if ds not in existing_dates:
            pn = ParadeNight(
                id=str(uuid.uuid4()),
                squadron_id=py.unit_id, wing_id=sq.wing_id if sq else None,
                date=ds, planning_year_id=year_id,
                parade_type=body.parade_type or "standard",
                is_active=True,
                start_time=body.parade_start_time or (sq.default_start_time if sq else None),
                end_time=body.parade_end_time or (sq.default_end_time if sq else None),
                created_by=p.user_id, created_at=utcnow(), updated_at=utcnow(),
            )
            db.add(pn)
            existing_dates.add(ds)
            created.append(ds)
    db.commit()
    audit(db, p, object_type="planning_year", object_id=year_id,
          action="generate_parade_dates", new={"created": len(created)})
    return {"ok": True, "created": len(created), "linked": 0, "dates": created}


_SESSION_STATUS_SCOPES = {
    "draft_only": {"draft"},
    "draft_and_planned": {"draft", "planned"},
}


class UpdateFutureParadeDayIn(BaseModel):
    new_weekday: int                      # 0=Mon … 6=Sun, same convention as GenerateParadeDatesIn
    from_date: str | None = None          # ISO date; defaults to today
    exclude_ids: list[str] = []           # ParadeDate IDs to leave untouched (kept as one-night exceptions)
    reason: str | None = None             # required when preview=false
    preview: bool = True
    # Restrict which nights are eligible for a bulk day-of-week move by
    # how far their sessions have progressed. None/"all" preserves the endpoint's
    # original unfiltered behaviour for any existing caller. A night with no
    # sessions at all is always eligible (nothing to protect).
    session_status_scope: str | None = None   # None|"all" | "draft_only" | "draft_and_planned"


@router.post("/years/{year_id}/update-future-parade-day")
def update_future_parade_day(
    year_id: str,
    body: UpdateFutureParadeDayIn,
    db: DBSession = Depends(get_db),
    p: Principal = Depends(get_principal),
):
    """TRGO-01: move future Parade Nights to a new day of the week.

    Changing Squadron.default_parade_day only affects newly-generated dates --
    it was never meant to (and does not) retroactively touch existing ParadeDate/
    ParadeNight rows, since both store a concrete ISO date, not a derived one.
    This endpoint is the explicit, auditable action for a squadron that actually
    wants its upcoming nights moved to a new day, previously missing entirely.

    Only ParadeDate rows with parade_date >= from_date and parade_type=="standard"
    are considered -- one-night exceptions (parade_type != "standard", e.g. a
    special/cancelled night) are preserved automatically, not just via exclude_ids.
    Each candidate is shifted to the same day *within its existing ISO week*
    (Mon-Sun), preserving term/week_number and the linked ParadeNight (so
    sessions, facilitators and rooms already assigned are never disturbed --
    only the date changes, in place).
    """
    from sqlalchemy import or_
    py = _get_year_or_404(year_id, db)
    _require_year_access(p, py, write=not body.preview)
    if not body.preview and py.unit_id:
        _sq = db.get(Squadron, py.unit_id)
        require_can_write_squadron(p, py.unit_id, _sq.wing_id if _sq else py.wing_id)
    if not body.preview and not (body.reason or "").strip():
        raise HTTPException(400, detail={"error": "reason_required",
                                          "message": "A reason is required to update future parade nights."})
    if body.new_weekday < 0 or body.new_weekday > 6:
        raise HTTPException(400, detail={"error": "invalid_weekday"})
    status_scope = body.session_status_scope
    if status_scope not in (None, "all", *_SESSION_STATUS_SCOPES):
        raise HTTPException(400, detail={"error": "invalid_session_status_scope"})
    allowed_statuses = _SESSION_STATUS_SCOPES.get(status_scope)

    from_date = body.from_date or date.today().isoformat()
    rows = db.query(ParadeNight).filter(
        ParadeNight.planning_year_id == year_id,
        ParadeNight.is_active == True,  # noqa: E712
        ParadeNight.date >= from_date,
        ParadeNight.parade_type.in_(["standard", "normal"]),
    ).order_by(ParadeNight.date).all()

    holidays = db.query(HolidayPeriod).filter(
        HolidayPeriod.planning_year_id == year_id,
        HolidayPeriod.affects_parade == True,  # noqa: E712
    ).all()

    def in_holiday(d: str) -> bool:
        return any(h.start_date <= d <= h.end_date for h in holidays)

    # All active dates for this year, to detect a shift landing on an existing date.
    existing_dates = {
        pn.date for pn in db.query(ParadeNight).filter(
            ParadeNight.planning_year_id == year_id,
            ParadeNight.is_active == True,  # noqa: E712
        ).all()
    }

    plan: list[dict] = []
    status_excluded = 0
    for pn_row in rows:
        if pn_row.id in body.exclude_ids:
            continue
        old_d = date.fromisoformat(pn_row.date)
        if old_d.weekday() == body.new_weekday:
            continue  # already on the target day -- nothing to do
        new_d = old_d - timedelta(days=old_d.weekday()) + timedelta(days=body.new_weekday)
        new_ds = new_d.isoformat()

        conflicts = []
        if new_ds in existing_dates and new_ds != pn_row.date:
            conflicts.append("duplicate_date")
        if in_holiday(new_ds):
            conflicts.append("holiday")

        has_sessions = False
        session_statuses: list[str] = []
        session_statuses = [
            s.status for s in db.query(TrainingSession).filter(
                TrainingSession.parade_night_id == pn_row.id,
                TrainingSession.is_archived == False,  # noqa: E712
            ).all()
        ]
        has_sessions = len(session_statuses) > 0

        # A night whose sessions have progressed past the requested
        # scope (e.g. already published/delivered) is left alone entirely --
        # excluded from the plan the same way a parade_type exception is,
        # not merely blocked, since this is a protection the caller asked
        # for, not a data conflict.
        if allowed_statuses is not None and session_statuses:
            if any(st not in allowed_statuses for st in session_statuses):
                status_excluded += 1
                continue

        plan.append({
            "parade_date_id": pn_row.id,
            "old_date": pn_row.date,
            "new_date": new_ds,
            "term": pn_row.term,
            "week_number": pn_row.week_number,
            "parade_night_id": pn_row.id,
            "has_sessions": has_sessions,
            "conflicts": conflicts,
            "blocked": len(conflicts) > 0,
        })

    exceptions = db.query(ParadeNight).filter(
        ParadeNight.planning_year_id == year_id,
        ParadeNight.is_active == True,  # noqa: E712
        ParadeNight.date >= from_date,
        ~ParadeNight.parade_type.in_(["standard", "normal"]),
    ).count()

    if body.preview:
        return {
            "ok": True, "preview": True,
            "changes": plan,
            "to_update": sum(1 for r in plan if not r["blocked"]),
            "blocked": sum(1 for r in plan if r["blocked"]),
            "exceptions_preserved": exceptions,
            "session_status_excluded": status_excluded,
        }

    updated = []
    skipped = []
    for r in plan:
        if r["blocked"]:
            skipped.append(r)
            continue
        pn_row = db.get(ParadeNight, r["parade_date_id"])
        old_date = pn_row.date
        pn_row.date = r["new_date"]
        pn_row.updated_at = utcnow()
        audit(db, p, object_type="parade_night", object_id=r["parade_date_id"],
              action="update_future_parade_day",
              old={"date": old_date}, new={"date": r["new_date"]}, reason=body.reason)
        updated.append(r)
    db.commit()
    audit(db, p, object_type="planning_year", object_id=year_id, action="update_future_parade_day_bulk",
          new={"updated": len(updated), "skipped": len(skipped), "new_weekday": body.new_weekday}, reason=body.reason)
    return {
        "ok": True, "preview": False,
        "updated": len(updated), "skipped": len(skipped),
        "updated_dates": updated, "skipped_conflicts": skipped,
        "exceptions_preserved": exceptions,
        "session_status_excluded": status_excluded,
    }


@router.delete("/parade-dates/{date_id}")
def delete_parade_date(
    date_id: str,
    db: DBSession = Depends(get_db),
    p: Principal = Depends(get_principal),
):
    pn = db.get(ParadeNight, date_id)
    if not pn:
        raise HTTPException(404, detail={"error": "not_found"})
    py = _get_year_or_404(pn.planning_year_id, db)
    _require_year_access(p, py, write=True, db=db)
    require_can_write_squadron(p, pn.squadron_id, pn.wing_id)
    # Check for training sessions — cannot delete a night that has sessions
    session_count = db.query(TrainingSession).filter(
        TrainingSession.parade_night_id == pn.id
    ).count()
    if session_count > 0:
        raise HTTPException(409, detail={
            "error": "parade_night_has_sessions",
            "message": (
                f"This parade night has {session_count} scheduled session(s). "
                "Archive it from the TMS instead of deleting, or remove the sessions first."
            ),
            "session_count": session_count,
        })
    # Resolve FK children before the delete so PostgreSQL FK constraints are not violated.
    # Notices are owned by the night — delete them. Prep plans and conflicts reference it
    # optionally (nullable FK) — nullify rather than cascade-delete.
    db.query(PlanningNotice).filter(PlanningNotice.parade_night_id == pn.id).delete()
    for app_row in db.query(AnchorPrepPlan).filter(AnchorPrepPlan.planned_parade_night_id == pn.id).all():
        app_row.planned_parade_night_id = None
    for pc in db.query(PlanningConflict).filter(PlanningConflict.parade_night_id == pn.id).all():
        pc.parade_night_id = None
    db.delete(pn)
    audit(db, p, object_type="parade_night", object_id=pn.id, action="delete",
          new={"date": pn.date}, commit=False)
    db.commit()
    return {"ok": True}


# ─────────────────────────────────────────────────────────────
# Holidays
# ─────────────────────────────────────────────────────────────

def _validate_iso_date(v: str, field_name: str) -> str:
    try:
        date.fromisoformat(v)
    except ValueError:
        raise ValueError(f"{field_name} must be a valid ISO-8601 date (YYYY-MM-DD), got '{v}'")
    return v


class HolidayIn(BaseModel):
    name: str = Field(max_length=120)
    start_date: str
    end_date: str
    jurisdiction: Optional[str] = None
    holiday_type: str = "school_holiday"
    affects_parade: bool = True
    notes: Optional[str] = None

    def model_post_init(self, __context) -> None:
        _validate_iso_date(self.start_date, "start_date")
        _validate_iso_date(self.end_date, "end_date")
        if self.end_date < self.start_date:
            raise ValueError("end_date must not be before start_date")


@router.get("/years/{year_id}/holidays")
def list_holidays(
    year_id: str,
    db: DBSession = Depends(get_db),
    p: Principal = Depends(get_principal),
):
    py = _get_year_or_404(year_id, db)
    _require_year_access(p, py)
    rows = db.query(HolidayPeriod).filter(HolidayPeriod.planning_year_id == year_id)\
             .order_by(HolidayPeriod.start_date).all()
    return [_holiday_out(h) for h in rows]


@router.post("/years/{year_id}/holidays")
def add_holiday(
    year_id: str,
    body: HolidayIn,
    db: DBSession = Depends(get_db),
    p: Principal = Depends(get_principal),
):
    py = _get_year_or_404(year_id, db)
    _require_year_access(p, py, write=True, db=db)
    if py.unit_id:
        _sq = db.get(Squadron, py.unit_id)
        require_can_write_squadron(p, py.unit_id, _sq.wing_id if _sq else py.wing_id)
    h = HolidayPeriod(
        id=str(uuid.uuid4()), planning_year_id=year_id,
        jurisdiction=body.jurisdiction, name=body.name,
        start_date=body.start_date, end_date=body.end_date,
        holiday_type=body.holiday_type,
        affects_parade=body.affects_parade, notes=body.notes,
        created_at=utcnow(), updated_at=utcnow(),
    )
    db.add(h); db.commit()
    audit(db, p, object_type="holiday_period", object_id=h.id, action="create",
          new={"name": body.name, "start": body.start_date, "end": body.end_date})
    return _holiday_out(h)


class HolidayUpdateIn(BaseModel):
    name: Optional[str] = Field(default=None, max_length=120)
    start_date: Optional[str] = None
    end_date: Optional[str] = None
    holiday_type: Optional[str] = None
    affects_parade: Optional[bool] = None
    notes: Optional[str] = None
    jurisdiction: Optional[str] = None

    def model_post_init(self, __context) -> None:
        if self.start_date:
            _validate_iso_date(self.start_date, "start_date")
        if self.end_date:
            _validate_iso_date(self.end_date, "end_date")


@router.patch("/holidays/{holiday_id}")
def update_holiday(
    holiday_id: str,
    body: HolidayUpdateIn,
    db: DBSession = Depends(get_db),
    p: Principal = Depends(get_principal),
):
    h = db.get(HolidayPeriod, holiday_id)
    if not h:
        raise HTTPException(404, detail={"error": "not_found"})
    py = _get_year_or_404(h.planning_year_id, db)
    _require_year_access(p, py, write=True, db=db)
    if py.unit_id:
        _sq = db.get(Squadron, py.unit_id)
        require_can_write_squadron(p, py.unit_id, _sq.wing_id if _sq else py.wing_id)
    old = {"name": h.name, "start": h.start_date, "end": h.end_date, "type": h.holiday_type}
    if body.name is not None:
        h.name = body.name
    if body.start_date is not None:
        h.start_date = body.start_date
    if body.end_date is not None:
        h.end_date = body.end_date
    if body.holiday_type is not None:
        h.holiday_type = body.holiday_type
    if body.affects_parade is not None:
        h.affects_parade = body.affects_parade
    if body.notes is not None:
        h.notes = body.notes
    if body.jurisdiction is not None:
        h.jurisdiction = body.jurisdiction
    if h.end_date < h.start_date:
        raise HTTPException(400, detail={"error": "end_before_start"})
    h.updated_at = utcnow()
    db.commit()
    audit(db, p, object_type="holiday_period", object_id=h.id, action="update",
          old=old, new={"name": h.name, "start": h.start_date, "end": h.end_date, "type": h.holiday_type})
    return _holiday_out(h)


@router.delete("/holidays/{holiday_id}")
def delete_holiday(
    holiday_id: str,
    db: DBSession = Depends(get_db),
    p: Principal = Depends(get_principal),
):
    h = db.get(HolidayPeriod, holiday_id)
    if not h:
        raise HTTPException(404, detail={"error": "not_found"})
    py = _get_year_or_404(h.planning_year_id, db)
    _require_year_access(p, py, write=True, db=db)
    if py.unit_id:
        _sq = db.get(Squadron, py.unit_id)
        require_can_write_squadron(p, py.unit_id, _sq.wing_id if _sq else py.wing_id)
    audit(db, p, object_type="holiday_period", object_id=h.id, action="delete",
          new={"name": h.name})
    db.delete(h); db.commit()
    return {"ok": True}


# ─────────────────────────────────────────────────────────────
# Anchor Events
# ─────────────────────────────────────────────────────────────

class AnchorEventIn(BaseModel):
    event_name: str
    event_type: str = "other"
    importance: str = "key_event"
    start_date: str
    end_date: Optional[str] = None
    owning_level: str = "unit"
    wing_id: Optional[str] = None
    unit_id: Optional[str] = None
    audience_orientation: bool = True
    audience_initial: bool = True
    audience_junior: bool = True
    audience_intermediate: bool = True
    audience_senior: bool = True
    planning_impact: Optional[str] = None
    readiness_requirements: Optional[str] = None
    notes: Optional[str] = None


class AnchorEventUpdateIn(BaseModel):
    event_name: Optional[str] = None
    importance: Optional[str] = None
    start_date: Optional[str] = None
    end_date: Optional[str] = None
    planning_impact: Optional[str] = None
    readiness_requirements: Optional[str] = None
    notes: Optional[str] = None
    version: Optional[int] = None


@router.get("/years/{year_id}/anchors")
def list_anchors(
    year_id: str,
    importance: Optional[str] = None,
    event_type: Optional[str] = None,
    include_archived: bool = False,
    db: DBSession = Depends(get_db),
    p: Principal = Depends(get_principal),
):
    py = _get_year_or_404(year_id, db)
    _require_year_access(p, py)
    q = db.query(AnchorEvent).filter(AnchorEvent.planning_year_id == year_id)
    if not include_archived:
        q = q.filter(AnchorEvent.is_archived == False)  # noqa: E712
    if importance:
        q = q.filter(AnchorEvent.importance == importance)
    if event_type:
        q = q.filter(AnchorEvent.event_type == event_type)
    return [_anchor_out(a) for a in q.order_by(AnchorEvent.start_date).all()]


@router.post("/years/{year_id}/anchors")
def create_anchor(
    year_id: str,
    body: AnchorEventIn,
    db: DBSession = Depends(get_db),
    p: Principal = Depends(get_principal),
):
    py = _get_year_or_404(year_id, db)
    _require_year_access(p, py, write=True, db=db)
    if py.unit_id:
        _sq = db.get(Squadron, py.unit_id)
        require_can_write_squadron(p, py.unit_id, _sq.wing_id if _sq else py.wing_id)
    wing_id = body.wing_id or py.wing_id
    unit_id = body.unit_id or py.unit_id
    a = AnchorEvent(
        id=str(uuid.uuid4()), planning_year_id=year_id,
        owning_level=body.owning_level,
        wing_id=wing_id, unit_id=unit_id,
        event_name=body.event_name, event_type=body.event_type,
        importance=body.importance,
        start_date=body.start_date, end_date=body.end_date,
        audience_orientation=body.audience_orientation,
        audience_initial=body.audience_initial,
        audience_junior=body.audience_junior,
        audience_intermediate=body.audience_intermediate,
        audience_senior=body.audience_senior,
        planning_impact=body.planning_impact,
        readiness_requirements=body.readiness_requirements,
        notes=body.notes,
        is_archived=False,
        created_by=p.user_id, created_at=utcnow(), updated_at=utcnow(),
    )
    db.add(a); db.commit()
    audit(db, p, object_type="anchor_event", object_id=a.id, action="create",
          new={"name": body.event_name, "date": body.start_date, "importance": body.importance})
    return _anchor_out(a)


@router.patch("/anchors/{anchor_id}")
def update_anchor(
    anchor_id: str,
    body: AnchorEventUpdateIn,
    db: DBSession = Depends(get_db),
    p: Principal = Depends(get_principal),
):
    a = db.get(AnchorEvent, anchor_id)
    if not a or a.is_archived:
        raise HTTPException(404, detail={"error": "not_found"})
    py = _get_year_or_404(a.planning_year_id, db)
    _require_year_access(p, py, write=True, db=db)
    if py.unit_id:
        _sq = db.get(Squadron, py.unit_id)
        require_can_write_squadron(p, py.unit_id, _sq.wing_id if _sq else py.wing_id)
    _check_version(a, body.version)
    for field in ("event_name", "importance", "start_date", "end_date",
                  "planning_impact", "readiness_requirements", "notes"):
        val = getattr(body, field)
        if val is not None:
            setattr(a, field, val)
    a.updated_by = p.user_id; a.updated_at = utcnow()
    a.version += 1
    db.commit()
    audit(db, p, object_type="anchor_event", object_id=a.id, action="update")
    return _anchor_out(a)


@router.delete("/anchors/{anchor_id}")
def archive_anchor(
    anchor_id: str,
    db: DBSession = Depends(get_db),
    p: Principal = Depends(get_principal),
):
    a = db.get(AnchorEvent, anchor_id)
    if not a:
        raise HTTPException(404, detail={"error": "not_found"})
    py = _get_year_or_404(a.planning_year_id, db)
    _require_year_access(p, py, write=True, db=db)
    if py.unit_id:
        _sq = db.get(Squadron, py.unit_id)
        require_can_write_squadron(p, py.unit_id, _sq.wing_id if _sq else py.wing_id)
    # R5-M16: delete orphaned AnchorPrepPlan rows before archiving the parent
    # AnchorEvent. AnchorPrepPlan has no is_archived field (no SoftDeleteMixin)
    # and the FK has no cascade, so these rows would become dangling and block
    # any future hard-delete of the event.
    db.query(AnchorPrepPlan).filter(AnchorPrepPlan.anchor_event_id == a.id).delete()
    a.is_archived = True; a.updated_at = utcnow()
    db.commit()
    audit(db, p, object_type="anchor_event", object_id=a.id, action="archive")
    return {"ok": True}


@router.post("/anchors/{anchor_id}/restore")
def restore_anchor(
    anchor_id: str,
    db: DBSession = Depends(get_db),
    p: Principal = Depends(get_principal),
):
    a = db.get(AnchorEvent, anchor_id)
    if not a:
        raise HTTPException(404, detail={"error": "not_found"})
    if not a.is_archived:
        raise HTTPException(409, detail={"error": "not_archived"})
    py = _get_year_or_404(a.planning_year_id, db)
    _require_year_access(p, py, write=True, db=db)
    if py.unit_id:
        _sq = db.get(Squadron, py.unit_id)
        require_can_write_squadron(p, py.unit_id, _sq.wing_id if _sq else py.wing_id)
    a.is_archived = False; a.updated_at = utcnow()
    db.commit()
    audit(db, p, object_type="anchor_event", object_id=a.id, action="restore")
    return {"ok": True}


@router.get("/anchors/{anchor_id}/prep-suggestions")
def get_prep_suggestions(
    anchor_id: str,
    db: DBSession = Depends(get_db),
    p: Principal = Depends(get_principal),
):
    """Return rule-based preparation suggestions for this anchor event."""
    a = db.get(AnchorEvent, anchor_id)
    if not a or a.is_archived:
        raise HTTPException(404, detail={"error": "not_found"})
    py = _get_year_or_404(a.planning_year_id, db)
    _require_year_access(p, py, write=False)
    rules = db.query(AnchorPrepRule).filter(AnchorPrepRule.event_type == a.event_type).all()
    # Find parade dates in the prep window
    try:
        event_dt = date.fromisoformat(a.start_date)
    except ValueError:
        event_dt = None
    suggested_dates = []
    if event_dt:
        parade_nights = db.query(ParadeNight).filter(
            ParadeNight.planning_year_id == a.planning_year_id,
            ParadeNight.is_active == True,  # noqa: E712
        ).all()
        for r in rules:
            window_start = (event_dt - timedelta(weeks=r.weeks_before_max)).isoformat()
            window_end   = (event_dt - timedelta(weeks=r.weeks_before_min)).isoformat()
            candidates   = [pn.date for pn in parade_nights
                            if window_start <= pn.date <= window_end]
            suggested_dates.append({
                "subject_area": r.subject_area,
                "suggested_activity": r.suggested_activity,
                "weeks_before_min": r.weeks_before_min,
                "weeks_before_max": r.weeks_before_max,
                "candidate_parade_dates": sorted(candidates),
                "notes": r.notes,
            })
    return {
        "anchor_event_id": anchor_id,
        "event_name": a.event_name,
        "event_type": a.event_type,
        "start_date": a.start_date,
        "suggestions": suggested_dates,
    }


# ─────────────────────────────────────────────────────────────
# Term Planner
# ─────────────────────────────────────────────────────────────

_TERM_RANGES = {
    1: ("01-28", "04-11"),
    2: ("04-28", "06-27"),
    3: ("07-14", "09-19"),
    4: ("10-06", "12-12"),
}


@router.get("/years/{year_id}/term-planner")
def get_term_planner(
    year_id: str,
    term: Optional[int] = Query(None, ge=1, le=4),
    db: DBSession = Depends(get_db),
    p: Principal = Depends(get_principal),
):
    py = _get_year_or_404(year_id, db)
    _require_year_access(p, py)

    all_dates = db.query(ParadeNight).filter(
        ParadeNight.planning_year_id == year_id,
        ParadeNight.is_active == True,  # noqa: E712
    ).order_by(ParadeNight.date).all()

    if term:
        t_start, t_end = _TERM_RANGES.get(term, ("01-01", "12-31"))
        yr = str(py.year)
        all_dates = [d for d in all_dates
                     if f"{yr}-{t_start}" <= d.date <= f"{yr}-{t_end}"]

    anchors = db.query(AnchorEvent).filter(
        AnchorEvent.planning_year_id == year_id,
        AnchorEvent.is_archived == False,  # noqa: E712
    ).order_by(AnchorEvent.start_date).all()

    # Batch-load all sessions for parade nights in one query (avoids N+1).
    # all_dates ARE the parade nights now (Phase B merge).
    pn_ids = {pn.id for pn in all_dates}

    sessions_by_date: dict[str, list] = {}
    if pn_ids:
        ts_rows = db.query(TrainingSession).filter(
            TrainingSession.parade_night_id.in_(pn_ids),
            TrainingSession.is_archived == False,  # noqa: E712
        ).all()
        ts_by_night: dict[str, list] = {}
        for s in ts_rows:
            ts_by_night.setdefault(s.parade_night_id, []).append(s)
        for pn in all_dates:
            sessions_by_date[pn.id] = [_real_session_out(s, db) for s in ts_by_night.get(pn.id, [])]
    else:
        for pn in all_dates:
            sessions_by_date[pn.id] = []

    # Calculate per-term session capacity summary
    # capacity = parade nights × cadet groups × periods
    total_periods = sum(pn.session_count for pn in all_dates)
    capacity = total_periods * len(CADET_GROUPS)
    filled = sum(len(v) for v in sessions_by_date.values())

    return {
        "planning_year_id": year_id,
        "year": py.year,
        "term": term,
        "parade_dates": [_night_out_as_date(d) for d in all_dates],
        "parade_count": len(all_dates),
        "session_capacity": capacity,
        "sessions_filled": filled,
        "sessions_remaining": max(0, capacity - filled),
        "anchors": [_anchor_out(a) for a in anchors],
        "sessions_by_parade_date": sessions_by_date,
    }


# ─────────────────────────────────────────────────────────────
# Parade Night Builder
# ─────────────────────────────────────────────────────────────

@router.get("/parade-dates/{date_id}/builder")
def get_builder(
    date_id: str,
    db: DBSession = Depends(get_db),
    p: Principal = Depends(get_principal),
):
    pn = db.get(ParadeNight, date_id)
    if not pn:
        raise HTTPException(404, detail={"error": "not_found"})
    py = _get_year_or_404(pn.planning_year_id, db)
    _require_year_access(p, py)

    # Timing template blocks
    timing_blocks: list[dict] = []
    session_count = 3
    tmpl = None
    if pn.timing_template_id:
        tmpl = db.get(TimingTemplate, pn.timing_template_id)
    if not tmpl and pn.squadron_id:
        tmpl = _effective_template(db, pn.squadron_id, pn.date)
    if tmpl:
        blocks = db.query(TimingBlock).filter(
            TimingBlock.timing_template_id == tmpl.id,
        ).order_by(TimingBlock.display_order).all()
        timing_blocks = [
            {
                "display_order": b.display_order, "block_name": b.block_name,
                "block_type": b.block_type, "start_time": b.start_time, "end_time": b.end_time,
                "duration_minutes": b.duration_minutes,
                "is_instructional_period": b.is_instructional_period,
                "period_number": b.period_number,
            }
            for b in blocks
        ]
        ip_count = sum(1 for b in blocks if b.is_instructional_period)
        if ip_count > 0:
            session_count = ip_count
    if pn.session_count:
        session_count = pn.session_count

    # Pull real sessions from the ParadeNight
    ts = db.query(TrainingSession).filter(
        TrainingSession.parade_night_id == pn.id,
        TrainingSession.is_archived == False,  # noqa: E712
    ).order_by(TrainingSession.period_number, TrainingSession.cadet_group).all()
    real_sessions = [_real_session_out(s, db) for s in ts]

    conflicts = db.query(PlanningConflict).filter(
        PlanningConflict.parade_night_id == date_id,
        PlanningConflict.is_resolved == False,  # noqa: E712
    ).all()

    return {
        "parade_date_id": date_id,
        "parade_night_id": pn.id,
        "parade_date": pn.date,
        "parade_type": pn.parade_type,
        "unit_id": pn.squadron_id,
        "session_count": session_count,
        "timing_blocks": timing_blocks,
        "cadet_groups": list(CADET_GROUPS),
        "sessions": real_sessions,
        "conflicts": [_conflict_out(c) for c in conflicts],
    }


# ─────────────────────────────────────────────────────────────
# Scheduled Sessions
# ─────────────────────────────────────────────────────────────

class SessionCreateIn(BaseModel):
    cadet_group: Optional[str] = None  # legacy; mutually exclusive with training_class_ids
    training_class_ids: Optional[list[str]] = None  # canonical; preferred over cadet_group
    session_number: int
    curriculum_id: Optional[str] = None
    activity_title: Optional[str] = None
    facilitator_id: Optional[str] = None
    # Mirrors SessionUpdateIn. Absent here until 2026-09-02, so Pydantic dropped
    # it silently: an assistant chosen while CREATING a session was lost, while
    # the same choice made while editing was kept. Same field, two paths, two
    # outcomes.
    assistant_facilitator_id: Optional[str] = None
    location_id: Optional[str] = None
    is_combined: bool = False
    combined_groups: Optional[list] = None
    override_conflict: bool = False
    override_reason: Optional[str] = None
    status: str = "draft"
    notes: Optional[str] = None

    @model_validator(mode="after")
    def _require_group_or_classes(self) -> "SessionCreateIn":
        if not self.cadet_group and not self.training_class_ids:
            raise ValueError("Either cadet_group or training_class_ids must be provided")
        return self

    @field_validator("status")
    @classmethod
    def _validate_status(cls, v: str) -> str:
        allowed = {"draft", "planned"}
        if v not in allowed:
            raise ValueError(f"status must be one of {sorted(allowed)}, got '{v}'")
        return v


class SessionUpdateIn(BaseModel):
    curriculum_id: Optional[str] = None
    activity_title: Optional[str] = None
    facilitator_id: Optional[str] = None
    assistant_facilitator_id: Optional[str] = None
    location_id: Optional[str] = None
    is_combined: Optional[bool] = None
    combined_groups: Optional[list] = None
    override_conflict: Optional[bool] = None
    override_reason: Optional[str] = None
    status: Optional[str] = None
    notes: Optional[str] = None
    version: Optional[int] = None
    cadet_group: Optional[str] = None  # legacy; updates TrainingSession.cadet_group
    training_class_ids: Optional[list[str]] = None  # replaces all SessionAudience rows when provided
    part_number: Optional[int] = None


@router.post("/parade-dates/{date_id}/sessions")
def create_session(
    date_id: str,
    body: SessionCreateIn,
    db: DBSession = Depends(get_db),
    p: Principal = Depends(get_principal),
):
    pn = db.get(ParadeNight, date_id)
    if not pn:
        raise HTTPException(404, detail={"error": "not_found"})
    py = _get_year_or_404(pn.planning_year_id, db)
    _require_year_access(p, py, write=True, db=db)
    require_can_write_squadron(p, pn.squadron_id, pn.wing_id)

    # Resolve cadet_group and audience path.
    # Canonical path: explicit training_class_ids (multi-class squadron support)
    # Legacy path: cadet_group string auto-matched to a single TrainingClass
    use_class_ids = bool(body.training_class_ids)
    if use_class_ids:
        # Scope-check every id BEFORE anything is written -- the session is
        # committed further down, ahead of the audience rows.
        scoped = _resolve_scoped_classes(db, body.training_class_ids, pn.squadron_id)
        # Derive cadet_group for denormalisation from the first class's stage_code.
        first_tc = scoped[0] if scoped else None
        resolved_cadet_group = _STAGE_CODE_CADET_GROUP.get(first_tc.stage_code) if first_tc else None
    else:
        if body.cadet_group not in CADET_GROUPS:
            raise HTTPException(422, detail={"error": "invalid_cadet_group"})
        resolved_cadet_group = body.cadet_group

    # Resolve room ID from location_id (which may be a PlanningLocation or TrainingArea id)
    training_area_id = None
    if body.location_id:
        ta = scoped_training_area(db, body.location_id, pn.squadron_id)
        if ta:
            training_area_id = ta.id

    # Create a real Session record
    s = TrainingSession(
        parade_night_id=pn.id, squadron_id=pn.squadron_id,
        period_number=body.session_number, cadet_group=resolved_cadet_group,
        custom_title=body.activity_title, status=body.status,
        delivery_notes=body.notes, created_by=p.user_id,
    )
    # Denormalize curriculum and facilitator
    if body.curriculum_id:
        ci = visible_curriculum_item(db, p, body.curriculum_id)
        if ci:
            s.curriculum_item_id = ci.id
            s.curriculum_code_at_time = ci.code
            s.curriculum_title_at_time = ci.title
            s.phase_at_time = ci.phase
            s.element_at_time = ci.element
    if body.facilitator_id:
        f = scoped_facilitator(db, body.facilitator_id, pn.squadron_id)
        if f:
            s.facilitator_id = f.id
            s.facilitator_display_name_at_time = " ".join(x for x in [f.current_rank, f.first_name, f.last_name] if x)
    if body.assistant_facilitator_id:
        af = scoped_facilitator(db, body.assistant_facilitator_id, pn.squadron_id)
        if af:
            s.assistant_facilitator_id = af.id
    if training_area_id:
        ra = scoped_training_area(db, training_area_id, pn.squadron_id)
        if ra:
            s.training_area_id = ra.id
            s.training_area_name_at_time = ra.name
    db.add(s); db.commit()
    if use_class_ids:
        _create_audience_for_class_ids(db, s.id, body.training_class_ids, pn.squadron_id)
    else:
        _upsert_session_audience(db, s.id, body.cadet_group, pn.squadron_id, py.id)
    _run_conflict_check(py.id, date_id, db)
    audit(db, p, object_type="session", object_id=s.id, action="create",
          new={"group": resolved_cadet_group, "session": body.session_number})
    return _real_session_out(s, db)


@router.get("/sessions/{session_id}")
def get_session(
    session_id: str,
    db: DBSession = Depends(get_db),
    p: Principal = Depends(get_principal),
):
    s = db.get(TrainingSession, session_id)
    if not s or s.is_archived:
        raise HTTPException(404, detail={"error": "not_found"})
    pn = db.get(ParadeNight, s.parade_night_id) if s.parade_night_id else None
    if pn:
        # Read-only endpoint: view permission is sufficient; write check here
        # previously caused sqn_general to receive 403 on session detail lookup.
        require_can_view_squadron(p, pn.squadron_id, pn.wing_id)
    return _real_session_out(s, db)


@router.patch("/sessions/{session_id}")
def update_session(
    session_id: str,
    body: SessionUpdateIn,
    db: DBSession = Depends(get_db),
    p: Principal = Depends(get_principal),
):
    s = db.get(TrainingSession, session_id)
    if not s or s.is_archived:
        raise HTTPException(404, detail={"error": "not_found"})
    pn = db.get(ParadeNight, s.parade_night_id) if s.parade_night_id else None
    if pn:
        require_can_write_squadron(p, pn.squadron_id, pn.wing_id)
    _check_version(s, body.version)

    if body.curriculum_id is not None:
        if body.curriculum_id:
            ci = visible_curriculum_item(db, p, body.curriculum_id)
            if ci:
                s.curriculum_item_id = ci.id
                s.curriculum_code_at_time = ci.code
                s.curriculum_title_at_time = ci.title
                s.phase_at_time = ci.phase
                s.element_at_time = ci.element
        else:
            s.curriculum_item_id = None
            s.curriculum_code_at_time = None
            s.curriculum_title_at_time = None
    if body.activity_title is not None:
        s.custom_title = body.activity_title
    if body.facilitator_id is not None:
        if body.facilitator_id:
            f = scoped_facilitator(db, body.facilitator_id, s.squadron_id)
            if f:
                s.facilitator_id = f.id
                s.facilitator_display_name_at_time = " ".join(x for x in [f.current_rank, f.first_name, f.last_name] if x)
        else:
            s.facilitator_id = None
            s.facilitator_display_name_at_time = None
    # SessionUpdateIn has accepted assistant_facilitator_id since it was added,
    # and nothing ever assigned it: the Planning Workspace drawer offers the
    # control, sends the id on save and on notes autosave, and got a 200 back
    # while the value was dropped. The read path resolves and returns the column,
    # so it advertised a field that could never be non-null. A field the API
    # accepts and discards is worse than one it rejects -- the user is told the
    # change was saved.
    #
    # No *_display_name_at_time column here by design: unlike the primary
    # facilitator, the assistant's name is resolved live on read, so only the id
    # is stored. Same scope rule -- Facilitator.squadron_id is a non-nullable FK.
    if body.assistant_facilitator_id is not None:
        if body.assistant_facilitator_id:
            af = scoped_facilitator(db, body.assistant_facilitator_id, s.squadron_id)
            if af:
                s.assistant_facilitator_id = af.id
        else:
            s.assistant_facilitator_id = None
    if body.location_id is not None:
        if body.location_id:
            ra = scoped_training_area(db, body.location_id, s.squadron_id)
            if ra:
                s.training_area_id = ra.id
                s.training_area_name_at_time = ra.name
        else:
            s.training_area_id = None
            s.training_area_name_at_time = None
    if body.status is not None:
        if body.status not in _VALID_SESSION_STATUS:
            raise HTTPException(400, detail={"error": "invalid_status"})
        s.status = body.status
    if body.notes is not None:
        s.delivery_notes = body.notes
    if body.part_number is not None:
        s.part_number = body.part_number if body.part_number > 0 else None
    if body.cadet_group is not None:
        if body.cadet_group and body.cadet_group not in CADET_GROUPS:
            raise HTTPException(422, detail={"error": "invalid_cadet_group"})
        s.cadet_group = body.cadet_group or None
    if body.training_class_ids is not None:
        # Scope-check before the delete: a refusal after it would drop the
        # session's existing audience and leave nothing in its place.
        scoped = _resolve_scoped_classes(db, body.training_class_ids, s.squadron_id)
        # Replace all existing SessionAudience rows for this session with the new set.
        db.query(SessionAudience).filter(SessionAudience.session_id == s.id).delete()
        db.flush()
        _create_audience_for_class_ids(db, s.id, body.training_class_ids, s.squadron_id)
        # Derive cadet_group from the first class for backward compat denormalisation.
        if scoped:
            s.cadet_group = _STAGE_CODE_CADET_GROUP.get(scoped[0].stage_code)
    s.version += 1
    db.commit()
    audit(db, p, object_type="session", object_id=s.id, action="update")
    return _real_session_out(s, db)


@router.delete("/sessions/{session_id}")
def delete_session(
    session_id: str,
    db: DBSession = Depends(get_db),
    p: Principal = Depends(get_principal),
):
    s = db.get(TrainingSession, session_id)
    if not s or s.is_archived:
        raise HTTPException(404, detail={"error": "not_found"})
    # parade_night_id is a required FK and ParadeNight rows are never hard-
    # deleted, so pn should always be found for a real session -- but the
    # permission check must not be skippable if that invariant is ever
    # violated (fail closed, not fail open; see restore_session's identical
    # fix, found by security review of that endpoint).
    pn = db.get(ParadeNight, s.parade_night_id) if s.parade_night_id else None
    if not pn:
        raise HTTPException(404, detail={"error": "not_found"})
    require_can_write_squadron(p, pn.squadron_id, pn.wing_id)
    s.is_archived = True
    db.commit()
    audit(db, p, object_type="session", object_id=s.id, action="delete")
    return {"ok": True}


@router.post("/sessions/{session_id}/restore")
def restore_session(
    session_id: str,
    db: DBSession = Depends(get_db),
    p: Principal = Depends(get_principal),
):
    """REM-133: session delete existed with no restore counterpart -- same
    pattern as restore_fac (training.py), the direct template for this gap."""
    s = db.get(TrainingSession, session_id)
    if not s:
        raise HTTPException(404, detail={"error": "not_found"})
    # Fail closed: a missing ParadeNight must reject the request, not skip
    # the permission check. Security review flagged the original conditional
    # form (`if pn: require_can_write_squadron(...)`) as fail-open -- in
    # practice parade_night_id is a required FK and ParadeNight rows are
    # never hard-deleted, so pn is always found for a real session today,
    # but the check must not be silently bypassable if that ever changes.
    pn = db.get(ParadeNight, s.parade_night_id) if s.parade_night_id else None
    if not pn:
        raise HTTPException(404, detail={"error": "not_found"})
    require_can_write_squadron(p, pn.squadron_id, pn.wing_id)
    if not s.is_archived:
        raise HTTPException(409, detail={"error": "not_archived"})
    s.is_archived = False
    db.commit()
    audit(db, p, object_type="session", object_id=s.id, action="restore")
    return {"ok": True}


# ─────────────────────────────────────────────────────────────
# Weekly Program
# ─────────────────────────────────────────────────────────────

@router.get("/parade-dates/{date_id}/archived-sessions")
def list_archived_sessions(
    date_id: str,
    db: DBSession = Depends(get_db),
    p: Principal = Depends(get_principal),
):
    """REM-133: archived sessions were reachable in the database but had no
    way to be seen or restored through the product -- a dedicated endpoint
    rather than an include_archived param on get_weekly_program(), since that
    endpoint's `sessions` list is consumed directly as grid cells and mixing
    archived rows into it would risk them rendering as live schedule slots."""
    pn = db.get(ParadeNight, date_id)
    if not pn:
        raise HTTPException(404, detail={"error": "not_found"})
    py = _get_year_or_404(pn.planning_year_id, db)
    _require_year_access(p, py)

    ts = db.query(TrainingSession).filter(
        TrainingSession.parade_night_id == pn.id,
        TrainingSession.is_archived == True,  # noqa: E712
    ).order_by(TrainingSession.period_number, TrainingSession.cadet_group).all()
    return {"sessions": [_real_session_out(s, db) for s in ts]}


@router.get("/parade-dates/{date_id}/weekly-program")
def get_weekly_program(
    date_id: str,
    db: DBSession = Depends(get_db),
    p: Principal = Depends(get_principal),
):
    pn = db.get(ParadeNight, date_id)
    if not pn:
        raise HTTPException(404, detail={"error": "not_found"})
    py = _get_year_or_404(pn.planning_year_id, db)
    _require_year_access(p, py)

    # Pull real sessions from the ParadeNight
    real_sessions: list[dict] = []
    if pn:
        ts = db.query(TrainingSession).filter(
            TrainingSession.parade_night_id == pn.id,
            TrainingSession.is_archived == False,  # noqa: E712
        ).order_by(TrainingSession.period_number, TrainingSession.cadet_group).all()
        real_sessions = [_real_session_out(s, db) for s in ts]

        # CLASS-06: which Training Class(es) each session targets, additive
        # to _real_session_out()'s own output. Attached here rather than
        # inside _real_session_out() itself, which also serves 7 other
        # endpoints (term planner, builder, session create/get/update,
        # long-range, mission assignment) -- scoping this to the one
        # endpoint this task covers keeps the change's blast radius to
        # Weekly Program only, matching CLASS-05's own additive-not-shared
        # approach in list_missions().
        if ts:
            from collections import defaultdict
            aud_rows = (
                db.query(SessionAudience, TrainingClass)
                .join(TrainingClass, SessionAudience.training_class_id == TrainingClass.id)
                .filter(SessionAudience.session_id.in_([s.id for s in ts]))
                .all()
            )
            classes_by_session: dict[str, list[dict]] = defaultdict(list)
            for aud, tc in aud_rows:
                classes_by_session[aud.session_id].append({
                    "training_class_id": tc.id, "display_name": tc.display_name,
                })
            for sess_out, s in zip(real_sessions, ts):
                sess_out["training_classes"] = classes_by_session.get(s.id, [])

    # Timing template for time labels
    timing_blocks: list[dict] = []
    tmpl = None
    if pn and pn.timing_template_id:
        tmpl = db.get(TimingTemplate, pn.timing_template_id)
    if not tmpl and pn.squadron_id:
        tmpl = _effective_template(db, pn.squadron_id, pn.date)
    if tmpl:
        blocks = db.query(TimingBlock).filter(
            TimingBlock.timing_template_id == tmpl.id,
        ).order_by(TimingBlock.display_order).all()
        timing_blocks = [
            {
                "sequence": b.display_order, "name": b.block_name, "block_type": b.block_type,
                "start_time": b.start_time, "end_time": b.end_time,
                "duration_minutes": b.duration_minutes,
                "is_instructional": b.is_instructional_period,
                "period_number": b.period_number,
            }
            for b in blocks
        ]

    conflicts = db.query(PlanningConflict).filter(
        PlanningConflict.parade_night_id == date_id,
        PlanningConflict.is_resolved == False,  # noqa: E712
    ).all()

    audit(db, p, object_type="parade_date", object_id=date_id, action="view_weekly_program")
    return {
        "parade_date_id": date_id,
        "parade_night_id": pn.id,
        "parade_date": pn.date,
        "unit_id": pn.squadron_id,
        "timing_blocks": timing_blocks,
        "sessions": real_sessions,
        "conflicts": [_conflict_out(c) for c in conflicts],
        "has_unresolved_conflicts": len(conflicts) > 0,
    }


# ─────────────────────────────────────────────────────────────
# Long Range View
# ─────────────────────────────────────────────────────────────

@router.get("/years/{year_id}/long-range")
def get_long_range(
    year_id: str,
    weeks: int = Query(default=8, ge=1, le=52),
    from_date: Optional[str] = None,
    end_date: Optional[str] = None,
    db: DBSession = Depends(get_db),
    p: Principal = Depends(get_principal),
):
    py = _get_year_or_404(year_id, db)
    _require_year_access(p, py)

    today = date.today().isoformat()
    start = from_date or today
    try:
        start_dt = date.fromisoformat(start)
    except ValueError:
        start_dt = date.today()

    if end_date:
        try:
            end_dt = date.fromisoformat(end_date)
        except ValueError:
            raise HTTPException(422, detail={"error": "invalid_end_date", "message": "end_date must be YYYY-MM-DD"})
        if end_dt < start_dt:
            raise HTTPException(422, detail={"error": "end_before_start", "message": "end_date must not be before from_date"})
    else:
        end_dt = start_dt + timedelta(weeks=weeks)

    parade_dates = db.query(ParadeNight).filter(
        ParadeNight.planning_year_id == year_id,
        ParadeNight.is_active == True,  # noqa: E712
        ParadeNight.date >= start_dt.isoformat(),
        ParadeNight.date <= end_dt.isoformat(),
    ).order_by(ParadeNight.date).all()

    anchors = db.query(AnchorEvent).filter(
        AnchorEvent.planning_year_id == year_id,
        AnchorEvent.is_archived == False,  # noqa: E712
        AnchorEvent.start_date >= start_dt.isoformat(),
        AnchorEvent.start_date <= end_dt.isoformat(),
    ).order_by(AnchorEvent.start_date).all()

    # Pre-load CI tier data for all parade nights in range (avoids N+1 in loop).
    pn_ids_lr = [pn.id for pn in parade_dates]
    ci_tier_lr: dict[str, dict] = {}
    if pn_ids_lr:
        ts_lr_all = db.query(TrainingSession).filter(
            TrainingSession.parade_night_id.in_(pn_ids_lr),
            TrainingSession.is_archived == False,  # noqa: E712
        ).all()
        ci_ids_lr = {s.curriculum_item_id for s in ts_lr_all if s.curriculum_item_id}
        if ci_ids_lr:
            for ci in db.query(CurriculumItem).filter(CurriculumItem.id.in_(ci_ids_lr)).all():
                ci_tier_lr[ci.id] = {"core_status": ci.core_status, "is_optional": ci.is_optional}
        # Index sessions by parade night id
        ts_by_night_lr: dict[str, list] = {}
        for s in ts_lr_all:
            ts_by_night_lr.setdefault(s.parade_night_id, []).append(s)

        # Bulk-load training_classes for all sessions in range (mirrors weekly-program endpoint).
        # Without this, EightWeekView renders in training-class mode (every new planning year
        # auto-creates 5 default TrainingClass rows) but getCellByClassId() never finds any
        # session because training_classes was absent from the response — all cells rendered as
        # "Empty slot" and conflict dots were invisible even for sessions with real conflicts.
        from collections import defaultdict
        classes_by_session_lr: dict[str, list[dict]] = {}
        if ts_lr_all:
            aud_rows_lr = (
                db.query(SessionAudience, TrainingClass)
                .join(TrainingClass, SessionAudience.training_class_id == TrainingClass.id)
                .filter(SessionAudience.session_id.in_([s.id for s in ts_lr_all]))
                .all()
            )
            _cb: dict[str, list[dict]] = defaultdict(list)
            for aud, tc in aud_rows_lr:
                _cb[aud.session_id].append({"training_class_id": tc.id, "display_name": tc.display_name})
            classes_by_session_lr = dict(_cb)
    else:
        ts_by_night_lr = {}
        classes_by_session_lr = {}

    rows = []
    for pn_obj in parade_dates:
        real_sessions: list[dict] = []
        ts = sorted(ts_by_night_lr.get(pn_obj.id, []),
                    key=lambda s: (s.period_number, s.cadet_group or ""))
        real_sessions = [_real_session_out(s, db, ci_tier=ci_tier_lr) for s in ts]
        for sess_out, s in zip(real_sessions, ts):
            sess_out["training_classes"] = classes_by_session_lr.get(s.id, [])

        conflicts = db.query(PlanningConflict).filter(
            PlanningConflict.parade_night_id == pn_obj.id,
            PlanningConflict.is_resolved == False,  # noqa: E712
        ).all()

        rows.append({
            "parade_date": _night_out_as_date(pn_obj),
            "sessions": real_sessions,
            "session_count": len(real_sessions),
            "filled_slots": len([s for s in real_sessions if s.get("curriculum_title") or s.get("activity_title")]),
            "conflicts": [_conflict_out(c) for c in conflicts],
        })

    return {
        "planning_year_id": year_id,
        "from_date": start_dt.isoformat(),
        "to_date": end_dt.isoformat(),
        "weeks": weeks,
        "parade_dates": rows,
        "anchors": [_anchor_out(a) for a in anchors],
    }


# ─────────────────────────────────────────────────────────────
# Planning Locations
# ─────────────────────────────────────────────────────────────

class PlanningLocationIn(BaseModel):
    name: str
    location_type: str = "indoor"
    capacity: Optional[int] = None
    notes: Optional[str] = None
    unit_id: Optional[str] = None


class PlanningLocationUpdateIn(BaseModel):
    name: Optional[str] = None
    location_type: Optional[str] = None
    capacity: Optional[int] = None
    notes: Optional[str] = None
    active_status: Optional[bool] = None


@router.get("/locations")
def list_locations(
    unit_id: Optional[str] = None,
    db: DBSession = Depends(get_db),
    p: Principal = Depends(get_principal),
):
    # REM-130: this endpoint's role chain had no branch at all for sqn_general
    # (only sqn_admin was scoped) -- live-reported as a sqn_general user seeing
    # another squadron's data in Planning Workspace. sqn_general fell through
    # every branch unfiltered, returning every TrainingArea in the system --
    # the exact same class of gap already found and fixed once for the sibling
    # /api/planning/facilitators endpoint (see that endpoint's own comment),
    # just with sqn_general specifically missed here rather than a whole-role
    # rewrite. Squadron-level roles (sqn_admin and sqn_general) both scope to
    # their own squadron; wing/national behaviour is unchanged.
    q = db.query(TrainingArea).filter(
        TrainingArea.active_status == True,  # noqa: E712
        TrainingArea.is_archived == False,  # noqa: E712
    )
    if p.role in ("sqn_admin", "sqn_general"):
        q = q.filter(TrainingArea.squadron_id == p.squadron_id)
    elif p.role in ("wing_admin", "wing_viewer"):
        sqn_ids = [s.id for s in db.query(Squadron).filter(
            Squadron.wing_id == p.wing_id, Squadron.is_archived == False  # noqa: E712
        ).all()]
        q = q.filter(TrainingArea.squadron_id.in_(sqn_ids))
    if unit_id:
        q = q.filter(TrainingArea.squadron_id == unit_id)
    return [_location_out(loc) for loc in q.order_by(TrainingArea.name).all()]


@router.post("/locations")
def create_location(
    body: PlanningLocationIn,
    db: DBSession = Depends(get_db),
    p: Principal = Depends(get_principal),
):
    _require_plan_write(p)
    # active_squadron_id, not squadron_id: a wing or national caller has no home
    # squadron, so the old default resolved to None, skipped the
    # require_can_write_squadron block below, and failed on a NOT NULL
    # constraint -- a 500 where the guard should have said "enter Proxy Mode".
    unit_id = body.unit_id or p.active_squadron_id
    if p.role == "sqn_admin":
        unit_id = p.squadron_id
    if not unit_id:
        raise HTTPException(400, detail={
            "error": "squadron_required",
            "message": "Choose a squadron for this training area, "
                       "or enter Proxy Mode to work within one.",
        })
    # A squadron-scoped Training Area is a delegated write on that squadron's
    # data -- require Proxy/Delegated Intervention like every other squadron-
    # scoped write, not just a bare role check (REM-45, Stage 12 follow-up;
    # same class of gap as create_planning_year, fixed in Stage 10).
    if unit_id:
        wing_id = None
        sqn = db.get(Squadron, unit_id)
        if sqn:
            wing_id = sqn.wing_id
        require_can_write_squadron(p, unit_id, wing_id)
    loc = TrainingArea(
        id=str(uuid.uuid4()), squadron_id=unit_id, name=body.name,
        type=body.location_type, capacity=body.capacity,
        notes=body.notes, active_status=True,
        created_at=utcnow(), updated_at=utcnow(),
    )
    db.add(loc); db.commit()
    audit(db, p, object_type="training_area", object_id=loc.id, action="create",
          new={"name": body.name})
    return _location_out(loc)


@router.patch("/locations/{location_id}")
def update_location(
    location_id: str,
    body: PlanningLocationUpdateIn,
    db: DBSession = Depends(get_db),
    p: Principal = Depends(get_principal),
):
    loc = db.get(TrainingArea, location_id)
    if not loc or loc.is_archived:
        raise HTTPException(404, detail={"error": "not_found"})
    _require_plan_write(p)
    # Resolved from the row's own squadron_id, never from caller-supplied
    # context (REM-45, Stage 12 follow-up) -- previously only sqn_admin's own
    # scope was checked; wing_admin/national_admin/system_admin could edit
    # ANY squadron's Training Area with zero scope check at all.
    if loc.squadron_id:
        wing_id = None
        sqn = db.get(Squadron, loc.squadron_id)
        if sqn:
            wing_id = sqn.wing_id
        require_can_write_squadron(p, loc.squadron_id, wing_id)
    if body.name is not None:
        loc.name = body.name
    if body.location_type is not None:
        loc.type = body.location_type
    if body.capacity is not None:
        loc.capacity = body.capacity
    if body.notes is not None:
        loc.notes = body.notes
    if body.active_status is not None:
        loc.active_status = body.active_status
    loc.updated_at = utcnow()
    db.commit()
    audit(db, p, object_type="training_area", object_id=loc.id, action="update")
    return _location_out(loc)


# ─────────────────────────────────────────────────────────────
# Facilitators (planning view — references existing model)
# ─────────────────────────────────────────────────────────────

@router.get("/facilitators")
def list_planning_facilitators(
    unit_id: Optional[str] = None,
    db: DBSession = Depends(get_db),
    p: Principal = Depends(get_principal),
):
    # Aligned to the same _view_squadron_id() resolution every other resource
    # endpoint in training.py already standardizes on (GET /api/facilitators,
    # /api/training-areas, /api/equipment, /api/activities) -- this endpoint
    # previously had its own bespoke role filter with no national_admin/
    # system_admin branch at all (silently unfiltered = every facilitator in
    # the system) and no proxy/acting-squadron awareness, so it could disagree
    # with what a squadron's own facilitator list actually shows.
    from .training import _view_squadron_id
    sq_id = _view_squadron_id(p, unit_id, db)
    q = db.query(Facilitator).filter(
        Facilitator.active_status == True,  # noqa: E712
        Facilitator.squadron_id == sq_id,
    )
    return [
        {
            "facilitator_id": f.id,
            "display_name": f"{f.current_rank or ''} {f.last_name}".strip(),
            "first_name": f.first_name, "last_name": f.last_name,
            "rank": f.current_rank, "type": f.type,
            "subject_areas": _parse_json_list(f.subject_areas),
            "max_sessions_per_night": f.max_sessions_per_night,
            "unit_id": f.squadron_id,
        }
        for f in q.order_by(Facilitator.last_name).all()
    ]


# ─────────────────────────────────────────────────────────────
# Conflict Detection
# ─────────────────────────────────────────────────────────────

def _detect_conflicts(year_id: str, night_id: str, db: DBSession) -> list[dict]:
    """Detect conflicts for one parade night. Reads only -- writes nothing.

    Split out of _run_conflict_check so the same rules can answer two
    different questions: "record what is wrong" (the persisted conflicts,
    which carry a user's override and reason) and "what is wrong right now"
    (the derived plan review, which must not write anything to answer a
    GET). Two copies of these rules would drift, and a review that
    disagreed with the recorded conflicts would be worse than no review.

    Anchored on ParadeNight since Phase B merged ParadeDate into it."""
    pn_obj = db.get(ParadeNight, night_id)
    # Use real TrainingSession records for conflict detection
    real_sessions: list = []
    if pn_obj:
        real_sessions = db.query(TrainingSession).filter(
            TrainingSession.parade_night_id == pn_obj.id,
            TrainingSession.is_archived == False,  # noqa: E712
        ).all()

    conflicts: list[dict] = []

    def _conflict(ctype: str, severity: str, msg: str, sess_id=None):
        conflicts.append({
            "conflict_type": ctype, "severity": severity, "message": msg,
            "scheduled_session_id": sess_id,
        })

    # Facilitator double-booking
    fac_map: dict[tuple, list] = {}
    for s in real_sessions:
        if s.facilitator_id:
            key = (s.facilitator_id, s.period_number)
            fac_map.setdefault(key, []).append(s)
    for (fid, snum), group in fac_map.items():
        if len(group) > 1:
            fac = db.get(Facilitator, fid)
            name = f"{fac.current_rank or ''} {fac.last_name}".strip() if fac else fid
            _conflict("facilitator_double_booked", "critical",
                      f"Facilitator {name} is assigned to multiple groups in session {snum}.",
                      group[0].id)

    # Room double-booking
    room_map: dict[tuple, list] = {}
    for s in real_sessions:
        if s.training_area_id:
            key = (s.training_area_id, s.period_number)
            room_map.setdefault(key, []).append(s)
    for (lid, snum), group in room_map.items():
        if len(group) > 1:
            ra = db.get(TrainingArea, lid)
            name = ra.name if ra else lid
            _conflict("room_double_booked", "critical",
                      f"Location '{name}' is assigned to multiple groups in session {snum}.",
                      group[0].id)

    # Empty sessions (group has no session for this night)
    scheduled_groups = {s.cadet_group for s in real_sessions}
    for grp in CADET_GROUPS:
        if grp not in scheduled_groups:
            _conflict("empty_session", "warning",
                      f"Cadet group '{grp}' has no session scheduled for this parade night.")

    # Anchor event without prep (checked at year level - skip per-date)

    # Holiday conflict
    if pn_obj:
        holidays = db.query(HolidayPeriod).filter(
            HolidayPeriod.planning_year_id == year_id,
            HolidayPeriod.affects_parade == True,  # noqa: E712
        ).all()
        for h in holidays:
            if h.start_date <= pn_obj.date <= h.end_date:
                _conflict("holiday_conflict", "warning",
                          f"This parade date falls within the '{h.name}' holiday period.")
                break

    # Facilitator on leave
    if pn_obj:
        fac_ids = {s.facilitator_id for s in real_sessions if s.facilitator_id}
        if fac_ids:
            leave_records = db.query(PlanningFacilitatorLeave).filter(
                PlanningFacilitatorLeave.facilitator_id.in_(fac_ids),
                PlanningFacilitatorLeave.is_archived == False,  # noqa: E712
                PlanningFacilitatorLeave.start_date <= pn_obj.date,
                PlanningFacilitatorLeave.end_date >= pn_obj.date,
            ).all()
            warned_facs: set[str] = set()
            for lv in leave_records:
                if lv.facilitator_id in warned_facs:
                    continue
                warned_facs.add(lv.facilitator_id)
                fac = db.get(Facilitator, lv.facilitator_id)
                name = f"{fac.current_rank or ''} {fac.last_name}".strip() if fac else lv.facilitator_id
                reason_part = f" ({lv.reason})" if lv.reason else ""
                affected = [s for s in real_sessions if s.facilitator_id == lv.facilitator_id]
                _conflict("facilitator_on_leave", "warning",
                          f"Facilitator {name} is on leave on this date{reason_part}.",
                          affected[0].id if affected else None)

    return conflicts


def _run_conflict_check(year_id: str, night_id: str, db: DBSession) -> list[PlanningConflict]:
    """Persist the detected conflicts for one parade night.

    Replaces previous UNRESOLVED results only: a conflict a user has overridden
    carries their reason and who they are, and re-running checks must never
    discard that."""
    db.query(PlanningConflict).filter(
        PlanningConflict.parade_night_id == night_id,
        PlanningConflict.is_resolved == False,  # noqa: E712
    ).delete(synchronize_session=False)

    rows = []
    for found in _detect_conflicts(year_id, night_id, db):
        c = PlanningConflict(
            id=str(uuid.uuid4()), planning_year_id=year_id,
            parade_night_id=night_id,
            scheduled_session_id=found["scheduled_session_id"],
            conflict_type=found["conflict_type"], severity=found["severity"],
            message=found["message"], is_resolved=False,
            created_at=utcnow(), updated_at=utcnow(),
        )
        rows.append(c)
        db.add(c)
    db.commit()
    return rows


@router.get("/years/{year_id}/conflicts")
def get_conflicts(
    year_id: str,
    severity: Optional[str] = None,
    db: DBSession = Depends(get_db),
    p: Principal = Depends(get_principal),
):
    py = _get_year_or_404(year_id, db)
    _require_year_access(p, py)
    q = db.query(PlanningConflict).filter(
        PlanningConflict.planning_year_id == year_id,
        PlanningConflict.is_resolved == False,  # noqa: E712
    )
    if severity:
        q = q.filter(PlanningConflict.severity == severity)
    return {"conflicts": [_conflict_out(c) for c in q.order_by(PlanningConflict.created_at.desc()).all()]}


class ConflictOverrideIn(BaseModel):
    override_reason: str


@router.post("/conflicts/{conflict_id}/override")
def override_conflict(
    conflict_id: str,
    body: ConflictOverrideIn,
    db: DBSession = Depends(get_db),
    p: Principal = Depends(get_principal),
):
    _require_plan_write(p)
    if not (body.override_reason or "").strip():
        raise HTTPException(422, detail={"error": "override_requires_reason"})
    c = db.get(PlanningConflict, conflict_id)
    if not c:
        raise HTTPException(404, detail={"error": "not_found"})
    # Resolved from the conflict's own linked PlanningYear, not caller-
    # supplied context (REM-45, Stage 12 follow-up) -- previously
    # wing_admin/national_admin/system_admin could resolve ANY conflict
    # anywhere with zero scope check.
    if c.planning_year_id:
        py = db.get(PlanningYear, c.planning_year_id)
        if py and py.unit_id:
            wing_id = py.wing_id
            sqn = db.get(Squadron, py.unit_id)
            if sqn:
                wing_id = sqn.wing_id
            require_can_write_squadron(p, py.unit_id, wing_id)
        elif py:
            _require_year_access(p, py, write=True, db=db)
    c.is_resolved = True
    c.override_reason = body.override_reason
    c.resolved_by = p.user_id
    c.updated_at = utcnow()
    db.commit()
    audit(db, p, object_type="planning_conflict", object_id=c.id,
          action="conflict_override", reason=body.override_reason)
    return {"ok": True, "conflict_id": conflict_id}


@router.get("/years/{year_id}/plan-review")
def get_plan_review(
    year_id: str,
    db: DBSession = Depends(get_db),
    p: Principal = Depends(get_principal),
):
    """The current state of every planning check, derived on read.

    Answers "what is wrong with this year right now" without the user having to
    remember to press anything, and without writing a row to answer a GET.

    It does not replace the persisted conflicts, and deliberately so: a
    PlanningConflict carries a user's override and the reason they typed, and a
    derived view has nowhere to put that. So each finding here reports whether
    it is already overridden, by matching a resolved conflict of the same type
    on the same date. Findings and recorded conflicts come from the same
    detector, so the two can never disagree about what counts as a conflict.

    Read-only: this endpoint needs view access, not write access, which is why
    it is a GET and why a viewer can open it.
    """
    py = _get_year_or_404(year_id, db)
    _require_year_access(p, py)

    # Phase B merged ParadeDate into ParadeNight. Findings below are keyed by
    # parade_night_id, matching _conflict_out on the sibling /conflicts endpoint
    # -- two names for the same thing inside one router is the duplication this
    # programme exists to remove.
    dates = db.query(ParadeNight).filter(
        ParadeNight.planning_year_id == year_id,
        ParadeNight.is_active == True,  # noqa: E712
    ).order_by(ParadeNight.date).all()

    overridden = db.query(PlanningConflict).filter(
        PlanningConflict.planning_year_id == year_id,
        PlanningConflict.is_resolved == True,  # noqa: E712
    ).all()
    override_index: dict[tuple, PlanningConflict] = {}
    for c in overridden:
        override_index.setdefault((c.parade_night_id, c.conflict_type), c)

    findings = []
    counts = {"critical": 0, "warning": 0, "overridden": 0}
    for pd_obj in dates:
        for found in _detect_conflicts(year_id, pd_obj.id, db):
            prior = override_index.get((pd_obj.id, found["conflict_type"]))
            item = {
                "parade_night_id": pd_obj.id,
                "parade_date": pd_obj.date,
                "conflict_type": found["conflict_type"],
                "severity": found["severity"],
                "message": found["message"],
                "scheduled_session_id": found["scheduled_session_id"],
                "is_overridden": prior is not None,
                "override_reason": prior.override_reason if prior else None,
            }
            findings.append(item)
            if prior is not None:
                counts["overridden"] += 1
            elif found["severity"] == "critical":
                counts["critical"] += 1
            else:
                counts["warning"] += 1

    return {
        "planning_year_id": year_id,
        "parade_dates_reviewed": len(dates),
        "counts": counts,
        "findings": findings,
    }


@router.post("/years/{year_id}/run-checks")
def run_checks(
    year_id: str,
    db: DBSession = Depends(get_db),
    p: Principal = Depends(get_principal),
):
    py = _get_year_or_404(year_id, db)
    _require_year_access(p, py, write=True, db=db)
    if py.unit_id:
        _sq = db.get(Squadron, py.unit_id)
        require_can_write_squadron(p, py.unit_id, _sq.wing_id if _sq else py.wing_id)
    nights = db.query(ParadeNight).filter(
        ParadeNight.planning_year_id == year_id,
        ParadeNight.is_active == True,  # noqa: E712
    ).all()
    total = 0
    for pn_obj in nights:
        results = _run_conflict_check(year_id, pn_obj.id, db)
        total += len(results)
    return {"ok": True, "conflicts_detected": total}


# ─────────────────────────────────────────────────────────────
# Decision Guide
# ─────────────────────────────────────────────────────────────

@router.get("/years/{year_id}/decision-guide")
def get_decision_guide(
    year_id: str,
    date_id: Optional[str] = None,
    db: DBSession = Depends(get_db),
    p: Principal = Depends(get_principal),
):
    py = _get_year_or_404(year_id, db)
    _require_year_access(p, py)

    today = date.today()
    checks = []

    # 1. Must Attend / Key Event in next 3 weeks
    window = (today + timedelta(weeks=3)).isoformat()
    upcoming = db.query(AnchorEvent).filter(
        AnchorEvent.planning_year_id == year_id,
        AnchorEvent.is_archived == False,  # noqa: E712
        AnchorEvent.start_date <= window,
        AnchorEvent.start_date >= today.isoformat(),
        AnchorEvent.importance.in_(["must_attend", "key_event"]),
    ).order_by(AnchorEvent.start_date).all()
    checks.append({
        "rule": 1,
        "question": "Is there a Must Attend or Key Event within the next 3 weeks?",
        "result": bool(upcoming),
        "detail": [a.event_name for a in upcoming],
        "action": "Schedule preparation lessons for these events." if upcoming else None,
    })

    # 2. Unscheduled groups
    if date_id:
        pd_obj = db.get(ParadeNight, date_id)
        if pd_obj:
            scheduled = {
                s.cadet_group for s in
                db.query(TrainingSession).filter(
                    TrainingSession.parade_night_id == pd_obj.id,
                    TrainingSession.is_archived == False,  # noqa: E712
                ).all()
                if s.cadet_group
            }
            missing = [g for g in CADET_GROUPS if g not in scheduled]
            checks.append({
                "rule": 3,
                "question": "Is there a cadet group with no mission assigned for this night?",
                "result": bool(missing),
                "detail": missing,
                "action": f"Assign sessions for: {', '.join(missing)}." if missing else None,
            })

    # 3. Active conflicts
    conflicts = db.query(PlanningConflict).filter(
        PlanningConflict.planning_year_id == year_id,
        PlanningConflict.severity == "critical",
        PlanningConflict.is_resolved == False,  # noqa: E712
    ).count()
    checks.append({
        "rule": 7,
        "question": "Are there unresolved critical conflicts?",
        "result": conflicts > 0,
        "detail": [f"{conflicts} critical conflict(s) unresolved"],
        "action": "Resolve or override all critical conflicts before publishing." if conflicts > 0 else None,
    })

    # 10. Is the night ready to publish?
    ready = not (conflicts > 0)
    checks.append({
        "rule": 10,
        "question": "Is the night ready to publish?",
        "result": ready,
        "detail": [],
        "action": None if ready else "Resolve all critical conflicts first.",
    })

    return {"planning_year_id": year_id, "checks": checks}


# ─────────────────────────────────────────────────────────────
# Command Centre dashboard aggregation
# ─────────────────────────────────────────────────────────────

@router.get("/command-centre")
def get_command_centre(
    year_id: Optional[str] = None,
    db: DBSession = Depends(get_db),
    p: Principal = Depends(get_principal),
):
    from sqlalchemy import or_, func
    from .dashboard import _data_freshness

    today = date.today().isoformat()

    # Resolve planning year
    py: PlanningYear | None = None
    if year_id:
        py = db.get(PlanningYear, year_id)
        if py:
            _require_year_access(p, py)
    else:
        q = db.query(PlanningYear)
        if p.role in ("sqn_admin", "sqn_general"):
            q = q.filter(PlanningYear.unit_id == p.squadron_id)
        elif p.role in ("wing_admin", "wing_viewer"):
            q = q.filter(PlanningYear.wing_id == p.wing_id)
        py = q.filter(PlanningYear.active_status == True).order_by(PlanningYear.year.desc()).first()  # noqa: E712
        if py is None:
            py = q.order_by(PlanningYear.year.desc()).first()

    if py is None:
        return {
            "planning_year_id": None, "year": None,
            "upcoming_anchors": [], "prep_gaps": [], "unreviewed_wing": [],
            "active_conflicts": [], "unscheduled_required": [],
            "recent_imports": [], "nights_missing_facilitator": 0,
        }

    # Upcoming anchor events (within the next 90 days)
    cutoff_90 = (date.today() + timedelta(days=90)).isoformat()
    anchor_rows = (
        db.query(AnchorEvent)
        .filter(
            AnchorEvent.planning_year_id == py.id,
            AnchorEvent.is_archived == False,  # noqa: E712
            AnchorEvent.start_date >= today,
            AnchorEvent.start_date <= cutoff_90,
        )
        .order_by(AnchorEvent.start_date)
        .limit(20)
        .all()
    )
    upcoming_anchors = [_anchor_out(a) for a in anchor_rows]

    # Prep gaps: anchor events with readiness_requirements but no linked prep plans
    anchors_needing_prep = (
        db.query(AnchorEvent)
        .filter(
            AnchorEvent.planning_year_id == py.id,
            AnchorEvent.is_archived == False,  # noqa: E712
            AnchorEvent.start_date >= today,
            AnchorEvent.readiness_requirements != None,  # noqa: E711
        )
        .all()
    )
    anchor_ids_with_plans = {
        row[0]
        for row in db.query(AnchorPrepPlan.anchor_event_id)
        .filter(AnchorPrepPlan.anchor_event_id.in_([a.id for a in anchors_needing_prep]))
        .all()
    } if anchors_needing_prep else set()

    prep_gaps = [
        {
            "anchor_event_id": a.id,
            "event_name": a.event_name,
            "start_date": a.start_date,
            "days_until": (date.fromisoformat(a.start_date) - date.today()).days,
        }
        for a in anchors_needing_prep
        if a.id not in anchor_ids_with_plans
    ]

    # Active (unresolved) conflicts
    conflict_rows = (
        db.query(PlanningConflict)
        .filter(
            PlanningConflict.planning_year_id == py.id,
            PlanningConflict.is_resolved == False,  # noqa: E712
        )
        .limit(50)
        .all()
    )
    active_conflicts = [
        {
            "conflict_id": c.id,
            "type": c.conflict_type,
            "message": c.message,
            "parade_date": None,
        }
        for c in conflict_rows
    ]

    # Unreviewed wing events for this squadron
    unreviewed_wing: list[dict] = []
    if p.squadron_id:
        sqn = db.get(Squadron, p.squadron_id)
        wing_id = sqn.wing_id if sqn else py.wing_id
        if wing_id:
            wing_events = (
                db.query(WingHQEvent)
                .filter(
                    WingHQEvent.wing_id == wing_id,
                    WingHQEvent.year == py.year,
                    WingHQEvent.is_archived == False,  # noqa: E712
                    WingHQEvent.start_date >= today,
                )
                .all()
            )
            reviewed_ids = {
                row[0]
                for row in db.query(SquadronEventStatus.wing_event_id)
                .filter(
                    SquadronEventStatus.squadron_id == p.squadron_id,
                    SquadronEventStatus.wing_event_id.in_([e.id for e in wing_events]),
                    SquadronEventStatus.status != "not_reviewed",
                )
                .all()
            } if wing_events else set()
            unreviewed_wing = [
                {
                    "wing_event_id": e.id,
                    "title": e.title,
                    "start_date": e.start_date,
                    "days_until": (date.fromisoformat(e.start_date) - date.today()).days,
                }
                for e in wing_events
                if e.id not in reviewed_ids
            ]

    # Unscheduled required curriculum items (core items with no session this year), and
    # nights missing a facilitator.
    # NOTE: this previously queried ScheduledSession (parade_date_id FK) -- a model with
    # no live write path anywhere in this codebase (confirmed: no `ScheduledSession(...)`
    # instantiation exists), the exact same defect class already fixed once for
    # facilitator_workload() above. Because the table is never populated,
    # scheduled_curriculum_ids was always empty (every core item silently shown as
    # "unscheduled required", regardless of the real schedule) and nights_missing_fac was
    # always 0 (silently hiding real unstaffed nights). Rewritten to the same
    # ParadeDate -> ParadeNight -> TrainingSession join qualification-program Phase B,
    # 2026-08-08; see docs/qualification/03_data_integrity_review.md P1 finding #1.
    pd_rows_cc = (
        db.query(ParadeNight)
        .filter(ParadeNight.planning_year_id == py.id)
        .all()
    )
    pn_to_pd_cc: dict[str, ParadeNight] = {
        pn_obj.id: pn_obj for pn_obj in pd_rows_cc
    }
    # Active training classes for this planning year, joined to stage name for PW Stage focus
    active_class_rows = (
        db.query(TrainingClass, CurriculumPhase)
        .join(CurriculumPhase, TrainingClass.training_stage_id == CurriculumPhase.id)
        .filter(
            TrainingClass.training_year_id == py.id,
            TrainingClass.is_archived == False,  # noqa: E712
        )
        .order_by(TrainingClass.class_number, TrainingClass.display_name)
        .all()
    )
    active_classes = [tc for tc, _phase in active_class_rows]
    active_class_ids = {tc.id for tc in active_classes}
    training_classes_out = [
        {
            "training_class_id": tc.id,
            "display_name": tc.display_name,
            "training_stage_id": tc.training_stage_id,
            "stage_name": phase.display_name,
        }
        for tc, phase in active_class_rows
    ]

    scheduled_curriculum_ids: set[str] = set()
    nights_missing_fac_ids: set[str] = set()
    # Maps curriculum_item_id -> set of class_ids that have it scheduled
    class_coverage: dict[str, set[str]] = {}
    cc_sessions: list = []
    if pn_to_pd_cc:
        cc_sessions = (
            db.query(TrainingSession)
            .filter(
                TrainingSession.parade_night_id.in_(list(pn_to_pd_cc.keys())),
                TrainingSession.is_archived == False,  # noqa: E712
            )
            .all()
        )
        scheduled_curriculum_ids = {
            s.curriculum_item_id for s in cc_sessions if s.curriculum_item_id
        }
        nights_missing_fac_ids = {
            pn_to_pd_cc[s.parade_night_id].id
            for s in cc_sessions
            if s.facilitator_id is None
        }
        # Build per-class curriculum coverage using SessionAudience
        if cc_sessions and active_class_ids:
            session_ids = [s.id for s in cc_sessions if s.curriculum_item_id]
            aud_rows = (
                db.query(SessionAudience)
                .filter(SessionAudience.session_id.in_(session_ids))
                .all()
            ) if session_ids else []
            # Map session_id -> curriculum_item_id
            sess_to_ci = {s.id: s.curriculum_item_id for s in cc_sessions if s.curriculum_item_id}
            for aud in aud_rows:
                ci_id = sess_to_ci.get(aud.session_id)
                if ci_id and aud.training_class_id in active_class_ids:
                    class_coverage.setdefault(ci_id, set()).add(aud.training_class_id)

    curriculum_q = _curriculum_scope_query(db, p).filter(
        CurriculumItem.core_status == "core",
    )
    unscheduled_required = []
    for ci in curriculum_q.limit(200).all():
        if active_classes:
            # Class-aware: show item if at least one active class has not scheduled it
            covered = class_coverage.get(ci.id, set())
            needs_class_ids = [tc.id for tc in active_classes if tc.id not in covered]
            if not needs_class_ids:
                continue  # all classes have it scheduled
        else:
            # No classes configured — fall back to simple scheduled check
            if ci.id in scheduled_curriculum_ids:
                continue
            needs_class_ids = []
        unscheduled_required.append({
            "curriculum_id": ci.id,
            "code": ci.code,
            "title": ci.title,
            "phase": ci.phase,
            "needs_class_ids": needs_class_ids,
        })

    nights_missing_fac = len(nights_missing_fac_ids)

    # Determine scope for data freshness (command-centre is squadron or wing scoped)
    _cc_scope = "squadron" if p.role in ("sqn_admin", "sqn_general") else "wing"
    _cc_sq_id = py.unit_id if _cc_scope == "squadron" else None
    _cc_wing_id = (py.wing_id or p.wing_id) if _cc_scope == "wing" else None

    return {
        "planning_year_id": py.id,
        "year": py.year,
        "upcoming_anchors": upcoming_anchors,
        "prep_gaps": prep_gaps,
        "unreviewed_wing": unreviewed_wing,
        "active_conflicts": active_conflicts,
        "unscheduled_required": unscheduled_required,
        "training_classes": training_classes_out,
        "recent_imports": [],
        "nights_missing_facilitator": nights_missing_fac,
        "data_freshness": _data_freshness(db, _cc_scope, _cc_sq_id, _cc_wing_id),
    }


# ─────────────────────────────────────────────────────────────
# CLASS-FORECAST-01 — Per-Training-Class planning forecast
# ─────────────────────────────────────────────────────────────

@router.get("/class-forecasts")
def get_class_forecasts(
    year_id: str,
    db: DBSession = Depends(get_db),
    p: Principal = Depends(get_principal),
):
    """CLASS-FORECAST-01: deterministic per-class planning forecast.

    Status rule (documented):
      available_time_blocks = ceil(remaining_parade_nights
                                   * avg_sessions_per_night
                                   / total_active_classes)
      ON TRACK     — unplanned_requirements <= available_time_blocks
      PLANNING RISK — unplanned_requirements > available_time_blocks (but nights remain)
      CRITICAL     — remaining_parade_nights == 0 and unplanned_requirements > 0

    avg_sessions_per_night is computed from sessions already created in this year
    (across all nights with at least one session); defaults to 2 when no history.
    """
    from sqlalchemy import func, select as sa_select
    import math
    from .training import _class_curriculum_progress

    py = db.get(PlanningYear, year_id)
    if not py:
        raise HTTPException(404, detail={"error": "planning_year_not_found"})
    sq_id = py.unit_id
    if not sq_id:
        raise HTTPException(400, detail={"error": "year_not_squadron_scoped"})
    sq = db.get(Squadron, sq_id)
    require_can_view_squadron(p, sq_id, sq.wing_id if sq else None)

    today_str = date.today().isoformat()

    # Active Training Classes for this year
    classes = db.query(TrainingClass).filter(
        TrainingClass.training_year_id == year_id,
        TrainingClass.is_archived == False,  # noqa: E712
    ).order_by(TrainingClass.class_number, TrainingClass.display_name).all()

    if not classes:
        return []

    total_classes = len(classes)

    # Remaining parade nights for this squadron (future, not archived)
    remaining_pns = db.query(func.count(ParadeNight.id)).filter(
        ParadeNight.squadron_id == sq_id,
        ParadeNight.date > today_str,
        ParadeNight.is_archived == False,  # noqa: E712
    ).scalar() or 0

    # Average sessions per past parade night in this year
    # "This year" scoped by ParadeNight.planning_year_id (Phase B: direct FK)
    past_pn_ids_in_year = sa_select(ParadeNight.id).where(
        ParadeNight.planning_year_id == year_id,
        ParadeNight.date <= today_str,
    )
    total_sessions_this_year = db.query(func.count(TrainingSession.id)).filter(
        TrainingSession.parade_night_id.in_(past_pn_ids_in_year),
        TrainingSession.is_archived == False,  # noqa: E712
    ).scalar() or 0
    nights_with_sessions = db.query(func.count(TrainingSession.parade_night_id.distinct())).filter(
        TrainingSession.parade_night_id.in_(past_pn_ids_in_year),
        TrainingSession.is_archived == False,  # noqa: E712
    ).scalar() or 0
    avg_sessions_per_night = (
        total_sessions_this_year / nights_with_sessions
        if nights_with_sessions > 0
        else 2.0  # default: assume 2 sessions per night when no history
    )

    # Shared capacity per class (fractional)
    available_per_class = math.ceil(
        remaining_pns * avg_sessions_per_night / total_classes
    ) if total_classes > 0 else 0

    _DELIVERED_STATUSES = {"delivered", "delivered_with_issue"}
    _PLANNED_STATUS = "planned"

    forecasts = []
    for c in classes:
        prog = _class_curriculum_progress(db, c)
        requirements = prog.get("requirements", [])
        remaining = [r for r in requirements if r["status"] not in _DELIVERED_STATUSES]
        planned = [r for r in remaining if r["status"] == _PLANNED_STATUS]
        unplanned = [r for r in remaining if r["status"] not in (_PLANNED_STATUS, *_DELIVERED_STATUSES)]

        remaining_count = len(remaining)
        planned_count = len(planned)
        unplanned_count = len(unplanned)

        # Determine status
        if c.training_stage_id is None:
            # The five classes auto-created with a planning year carry a
            # stage_code (ORI/INI/...) but no training_stage_id, and
            # _class_curriculum_progress keys off the stage, so they have zero
            # requirements for a reason that has nothing to do with progress.
            # Reporting "All requirements delivered." there is a green light
            # for work nobody has scoped yet -- say what is actually true.
            status = "not_configured"
            message = (
                "No Training Stage assigned, so this class has no curriculum "
                "requirements to track. Assign a stage to see its forecast."
            )
        elif remaining_count == 0:
            status = "on_track"
            message = "All requirements delivered."
        elif remaining_pns == 0 and unplanned_count > 0:
            status = "critical"
            message = (
                f"No remaining parade nights — {unplanned_count} requirement(s) "
                "not yet planned."
            )
        elif unplanned_count <= available_per_class:
            status = "on_track"
            message = (
                f"{unplanned_count} unplanned requirement(s) within estimated "
                f"capacity ({available_per_class} slot(s) across {remaining_pns} remaining nights)."
            )
        else:
            status = "planning_risk"
            message = (
                f"{unplanned_count} unplanned requirement(s) exceed estimated "
                f"capacity ({available_per_class} slot(s) across {remaining_pns} remaining nights)."
            )

        stage = db.get(CurriculumPhase, c.training_stage_id) if c.training_stage_id else None
        forecasts.append({
            "class_id": c.id,
            "class_name": c.display_name,
            "stage_name": stage.name if stage else None,
            "remaining_requirements": remaining_count,
            "planned_requirements": planned_count,
            "unplanned_requirements": unplanned_count,
            "remaining_parade_nights": remaining_pns,
            "available_time_blocks": available_per_class,
            "status": status,
            "message": message,
        })

    return forecasts


# ─────────────────────────────────────────────────────────────
# Prep Rules (read-only; seeded)
# ─────────────────────────────────────────────────────────────

@router.get("/prep-rules")
def list_prep_rules(
    event_type: Optional[str] = None,
    db: DBSession = Depends(get_db),
    p: Principal = Depends(get_principal),
):
    q = db.query(AnchorPrepRule)
    if event_type:
        q = q.filter(AnchorPrepRule.event_type == event_type)
    return [
        {
            "prep_rule_id": r.id, "event_type": r.event_type,
            "subject_area": r.subject_area, "suggested_activity": r.suggested_activity,
            "weeks_before_min": r.weeks_before_min, "weeks_before_max": r.weeks_before_max,
            "notes": r.notes,
        }
        for r in q.order_by(AnchorPrepRule.event_type, AnchorPrepRule.weeks_before_min).all()
    ]


# ─────────────────────────────────────────────────────────────
# V14 — Training Planner: Mission Scheduling View
# ─────────────────────────────────────────────────────────────

# WA school term date ranges (defaults; adjusted per school year by parade dates)
_WA_TERM_RANGES = {
    1: ("01-28", "04-11"),
    2: ("04-28", "06-27"),
    3: ("07-14", "09-19"),
    4: ("10-06", "12-12"),
}


def _term_for_date(date_str: str, year: int) -> str | None:
    """Return T1/T2/T3/T4 for a date based on WA default term boundaries."""
    for t, (ts, te) in _WA_TERM_RANGES.items():
        if f"{year}-{ts}" <= date_str <= f"{year}-{te}":
            return f"T{t}"
    return None


def _curriculum_scope_query(db: DBSession, p: Principal):
    """Return a query for CurriculumItem rows visible to this principal."""
    from sqlalchemy import or_
    q = db.query(CurriculumItem).filter(
        CurriculumItem.is_archived == False,  # noqa: E712
        CurriculumItem.active_status == True,  # noqa: E712
    )
    if p.role in ("sqn_admin", "sqn_general"):
        sqn = db.get(Squadron, p.squadron_id) if p.squadron_id else None
        wing_id = sqn.wing_id if sqn else None
        q = q.filter(or_(
            CurriculumItem.owning_level == "national",
            CurriculumItem.wing_id == wing_id,
            CurriculumItem.squadron_id == p.squadron_id,
        ))
    elif p.role in ("wing_admin", "wing_viewer"):
        q = q.filter(or_(
            CurriculumItem.owning_level == "national",
            CurriculumItem.wing_id == p.wing_id,
        ))
    # nat/system: all
    return q


@router.get("/years/{year_id}/missions")
def list_missions(
    year_id: str,
    phase: Optional[str] = None,
    element: Optional[str] = None,
    term: Optional[str] = None,
    status: Optional[str] = None,   # "scheduled" | "unscheduled" | "cancelled" | "not_delivered"
                                     # | "rescheduled" | "resolved" | "planned"
    search: Optional[str] = None,
    start_date: Optional[str] = None,  # TRGO-08: ISO date, inclusive
    end_date: Optional[str] = None,    # TRGO-08: ISO date, inclusive
    db: DBSession = Depends(get_db),
    p: Principal = Depends(get_principal),
):
    """Training Planner: curriculum items with scheduling status for this planning year."""
    py = _get_year_or_404(year_id, db)
    _require_year_access(p, py)

    # Get all parade nights for this year (Phase B: direct planning_year_id FK)
    pn_rows = db.query(ParadeNight).filter(
        ParadeNight.planning_year_id == year_id,
        ParadeNight.is_active == True,  # noqa: E712
    ).all()
    pn_to_pd: dict[str, ParadeNight] = {pn_obj.id: pn_obj for pn_obj in pn_rows}

    # Pull all sessions for parade nights in this year
    pn_ids = list(pn_to_pd.keys())
    sessions_in_year: list[TrainingSession] = []
    if pn_ids:
        sessions_in_year = db.query(TrainingSession).filter(
            TrainingSession.parade_night_id.in_(pn_ids),
            TrainingSession.is_archived == False,  # noqa: E712
        ).all()

    # Index sessions by curriculum_item_id
    from collections import defaultdict
    sessions_by_ci: dict[str, list[TrainingSession]] = defaultdict(list)
    for s in sessions_in_year:
        if s.curriculum_item_id:
            sessions_by_ci[s.curriculum_item_id].append(s)

    # Build curriculum query
    q = _curriculum_scope_query(db, p)
    if phase:
        q = q.filter(CurriculumItem.phase == phase)
    if element:
        q = q.filter(CurriculumItem.element == element)
    if search:
        like = f"%{search}%"
        from sqlalchemy import or_
        q = q.filter(or_(
            CurriculumItem.code.ilike(like),
            CurriculumItem.title.ilike(like),
        ))

    items = q.order_by(CurriculumItem.phase, CurriculumItem.recommended_sequence, CurriculumItem.code).all()

    # CLASS-05: per-class breakdown, additive to the existing item-level
    # backlog_status/is_scheduled/etc fields below (none of which change).
    # A Training Class's Stage is matched to a CurriculumItem's phase the
    # same way _class_curriculum_progress (training.py) does -- by
    # CurriculumPhase.name == CurriculumItem.phase -- so this reuses that
    # resolution rather than inventing a second one (addendum §44). Classes
    # are not filtered by training_year_id: the Session<->Class assignment
    # UI (Quick Edit, CLASS-17) deliberately shows every active class for
    # the squadron regardless of year, so a session's audience can name a
    # class whose own training_year_id differs from this planning year.
    classes_by_phase: dict[str, list[TrainingClass]] = defaultdict(list)
    if py.unit_id:
        class_rows = (
            db.query(TrainingClass, CurriculumPhase)
            .join(CurriculumPhase, TrainingClass.training_stage_id == CurriculumPhase.id)
            .filter(TrainingClass.squadron_id == py.unit_id, TrainingClass.is_archived == False)  # noqa: E712
            .order_by(TrainingClass.class_number, TrainingClass.display_name)
            .all()
        )
        for tc, stage in class_rows:
            classes_by_phase[stage.name].append(tc)

    session_class_ids: dict[str, set[str]] = defaultdict(set)
    if sessions_in_year:
        aud_rows = db.query(SessionAudience).filter(
            SessionAudience.session_id.in_([s.id for s in sessions_in_year])
        ).all()
        for aud in aud_rows:
            session_class_ids[aud.session_id].add(aud.training_class_id)

    def _backlog_status_for(sessions: list[TrainingSession]) -> dict:
        """Same six-state model as the item-level computation below, applied
        to an arbitrary session subset (here, one Training Class's share of
        a mission's sessions)."""
        is_scheduled = len(sessions) > 0
        has_cancelled = any(s.status in ("cancelled", "cancelled_late") for s in sessions)
        has_not_delivered = any(s.status == "not_delivered" for s in sessions)
        has_rescheduled = any(s.status == "rescheduled" for s in sessions)
        has_delivered = any(s.status in ("delivered", "delivered_with_issue") for s in sessions)
        needs_reschedule = has_cancelled or has_not_delivered
        if not is_scheduled:
            backlog_status = "unscheduled"
        elif needs_reschedule and has_delivered:
            backlog_status = "resolved"
        elif needs_reschedule and has_cancelled:
            backlog_status = "cancelled_awaiting_reschedule"
        elif needs_reschedule:
            backlog_status = "not_delivered_awaiting_reschedule"
        elif has_rescheduled:
            backlog_status = "rescheduled"
        else:
            backlog_status = "planned"
        return {
            "is_scheduled": is_scheduled, "has_cancelled": has_cancelled,
            "has_not_delivered": has_not_delivered, "has_rescheduled": has_rescheduled,
            "needs_reschedule": needs_reschedule, "backlog_status": backlog_status,
        }

    def _sess_summary(s: TrainingSession) -> dict:
        pd_obj = pn_to_pd.get(s.parade_night_id)
        room_name = s.training_area_name_at_time
        if not room_name and s.training_area_id:
            ra = db.get(TrainingArea, s.training_area_id)
            if ra:
                room_name = ra.name
        return {
            "session_id": s.id,
            "parade_night_id": s.parade_night_id,
            "parade_date": pd_obj.date if pd_obj else None,
            "parade_date_id": pd_obj.id if pd_obj else None,
            "term": (pd_obj.term if pd_obj and pd_obj.term else
                     (_term_for_date(pd_obj.date, py.year) if pd_obj else None)),
            "session_number": s.period_number,
            "part_number": s.part_number,
            "cadet_group": s.cadet_group,
            "facilitator_id": s.facilitator_id,
            "facilitator_name": s.facilitator_display_name_at_time,
            "location_id": s.training_area_id,
            "location_name": room_name,
            "status": s.status,
            "cancelled_reason": s.cancelled_reason,
            "not_delivered_reason": s.not_delivered_reason,
            "rescheduled_to_date": s.rescheduled_to_date,
            "outcome_note": s.delivery_notes or s.issue_notes,
        }

    result = []
    for ci in items:
        scheduled = sessions_by_ci.get(ci.id, [])
        is_scheduled = len(scheduled) > 0
        has_cancelled = any(s.status in ("cancelled", "cancelled_late") for s in scheduled)
        has_not_delivered = any(s.status == "not_delivered" for s in scheduled)
        has_rescheduled = any(s.status == "rescheduled" for s in scheduled)
        has_delivered = any(s.status in ("delivered", "delivered_with_issue") for s in scheduled)
        needs_reschedule = has_cancelled or has_not_delivered

        # Six-state Mission Backlog model (master transformation plan Block 6):
        # unscheduled / planned / cancelled_awaiting_reschedule /
        # not_delivered_awaiting_reschedule / rescheduled / resolved.
        # No explicit link exists in the schema between a cancelled/not-delivered
        # session and whatever session eventually replaces it (rescheduled_to_date
        # is a plain date string, not an FK) — "resolved" is therefore an honest
        # approximation: a mission that has a cancelled/not-delivered session AND
        # has also since been delivered via some session is treated as resolved,
        # rather than left permanently flagged as needing a reschedule action.
        if not is_scheduled:
            backlog_status = "unscheduled"
        elif needs_reschedule and has_delivered:
            backlog_status = "resolved"
        elif needs_reschedule and has_cancelled:
            backlog_status = "cancelled_awaiting_reschedule"
        elif needs_reschedule:
            backlog_status = "not_delivered_awaiting_reschedule"
        elif has_rescheduled:
            backlog_status = "rescheduled"
        else:
            backlog_status = "planned"

        # Filter by date range if requested. An unscheduled mission has no date to
        # match against and still needs attention regardless of the visible window,
        # so it is never excluded by this filter -- only scheduled missions are
        # narrowed down to those with at least one session inside the range.
        if (start_date or end_date) and is_scheduled:
            in_range = False
            for s in scheduled:
                pd_obj = pn_to_pd.get(s.parade_night_id)
                d = pd_obj.date if pd_obj else None
                if not d:
                    continue
                if start_date and d < start_date:
                    continue
                if end_date and d > end_date:
                    continue
                in_range = True
                break
            if not in_range:
                continue

        # Filter by term if requested
        if term:
            matching = [s for s in scheduled
                        if pn_to_pd.get(s.parade_night_id) and
                        (_term_for_date(pn_to_pd[s.parade_night_id].date, py.year) == term or
                         (pn_to_pd[s.parade_night_id].term == term))]
            if not matching and status == "scheduled":
                continue
        # Filter by status
        if status == "scheduled" and not is_scheduled:
            continue
        if status == "unscheduled" and is_scheduled:
            continue
        if status == "cancelled" and not has_cancelled:
            continue
        if status == "not_delivered" and not has_not_delivered:
            continue
        if status == "rescheduled" and backlog_status != "rescheduled":
            continue
        if status == "resolved" and backlog_status != "resolved":
            continue
        if status == "planned" and backlog_status != "planned":
            continue

        class_breakdown = []
        for tc in classes_by_phase.get(ci.phase, []):
            class_sessions = [s for s in scheduled if tc.id in session_class_ids.get(s.id, ())]
            cb = _backlog_status_for(class_sessions)
            class_breakdown.append({
                "training_class_id": tc.id,
                "display_name": tc.display_name,
                "scheduled_count": len(class_sessions),
                **cb,
            })
        unassigned_session_count = sum(1 for s in scheduled if not session_class_ids.get(s.id))

        result.append({
            "curriculum_id": ci.id,
            "code": ci.code,
            "title": ci.title,
            "phase": ci.phase,
            "element": ci.element,
            "recommended_term": ci.recommended_term,
            "part_count": ci.part_count,
            "instructor_suitability": ci.instructor_suitability,
            "duration_minutes": ci.duration_minutes,
            "core_status": ci.core_status,
            "is_scheduled": is_scheduled,
            "has_cancelled": has_cancelled,
            "has_not_delivered": has_not_delivered,
            "has_rescheduled": has_rescheduled,
            "needs_reschedule": needs_reschedule,
            "class_breakdown": class_breakdown,
            "unassigned_session_count": unassigned_session_count,
            "backlog_status": backlog_status,
            "scheduled_sessions": [_sess_summary(s) for s in scheduled],
            "scheduled_count": len(scheduled),
        })

    return {
        "planning_year_id": year_id,
        "year": py.year,
        "total": len(result),
        "scheduled_count": sum(1 for r in result if r["is_scheduled"]),
        "missions": result,
    }


# ─────────────────────────────────────────────────────────────
# V14 — Mission Assignment (creates real Session record)
# ─────────────────────────────────────────────────────────────

class MissionAssignIn(BaseModel):
    curriculum_id: str
    parade_date_id: str
    session_number: int
    cadet_group: Optional[str] = None  # legacy
    training_class_ids: Optional[list[str]] = None  # canonical
    part_number: Optional[int] = None
    facilitator_id: Optional[str] = None
    training_area_id: Optional[str] = None
    notes: Optional[str] = None

    @model_validator(mode="after")
    def _require_group_or_classes(self) -> "MissionAssignIn":
        if not self.cadet_group and not self.training_class_ids:
            raise ValueError("Either cadet_group or training_class_ids must be provided")
        return self


@router.post("/years/{year_id}/assign-mission")
def assign_mission(
    year_id: str,
    body: MissionAssignIn,
    db: DBSession = Depends(get_db),
    p: Principal = Depends(get_principal),
):
    """Assign a curriculum mission to a parade night session (Training Planner action)."""
    py = _get_year_or_404(year_id, db)
    _require_year_access(p, py, write=True, db=db)
    if py.unit_id:
        _sq = db.get(Squadron, py.unit_id)
        require_can_write_squadron(p, py.unit_id, _sq.wing_id if _sq else py.wing_id)

    use_class_ids = bool(body.training_class_ids)
    if use_class_ids:
        scoped = _resolve_scoped_classes(db, body.training_class_ids, py.unit_id)
        first_tc = scoped[0] if scoped else None
        resolved_cadet_group = _STAGE_CODE_CADET_GROUP.get(first_tc.stage_code) if first_tc else None
    else:
        if body.cadet_group not in CADET_GROUPS:
            raise HTTPException(422, detail={"error": "invalid_cadet_group"})
        resolved_cadet_group = body.cadet_group

    # Not visible to this principal reads as not-found: a distinct "exists but
    # belongs to another squadron" answer would be an existence oracle.
    ci = visible_curriculum_item(db, p, body.curriculum_id)
    if not ci:
        raise HTTPException(404, detail={"error": "curriculum_item_not_found"})

    pd_obj = db.get(ParadeNight, body.parade_date_id)
    if not pd_obj or pd_obj.planning_year_id != year_id:
        raise HTTPException(404, detail={"error": "parade_date_not_found"})

    pn = pd_obj  # parade_date_id is now the parade_night_id post-merge

    s = TrainingSession(
        parade_night_id=pn.id, squadron_id=pn.squadron_id,
        period_number=body.session_number,
        cadet_group=resolved_cadet_group,
        part_number=body.part_number,
        curriculum_item_id=ci.id,
        curriculum_code_at_time=ci.code,
        curriculum_title_at_time=ci.title,
        phase_at_time=ci.phase,
        element_at_time=ci.element,
        status="planned",
        delivery_notes=body.notes,
        created_by=p.user_id,
        created_at=utcnow(), updated_at=utcnow(),
    )
    if body.facilitator_id:
        f = scoped_facilitator(db, body.facilitator_id, pn.squadron_id)
        if f:
            s.facilitator_id = f.id
            s.facilitator_display_name_at_time = " ".join(x for x in [f.current_rank, f.first_name, f.last_name] if x)
    if body.training_area_id:
        ra = scoped_training_area(db, body.training_area_id, pn.squadron_id)
        if ra:
            s.training_area_id = ra.id
            s.training_area_name_at_time = ra.name

    db.add(s)
    db.commit()
    if use_class_ids:
        _create_audience_for_class_ids(db, s.id, body.training_class_ids, pn.squadron_id)
    else:
        _upsert_session_audience(db, s.id, body.cadet_group, pn.squadron_id, year_id)
    _run_conflict_check(year_id, body.parade_date_id, db)
    audit(db, p, object_type="session", object_id=s.id, action="assign_mission",
          new={"curriculum": ci.code, "date": pd_obj.date, "session": body.session_number,
               "group": resolved_cadet_group})
    return _real_session_out(s, db)


# ─────────────────────────────────────────────────────────────
# V14 — Annual Program (full year view with term blocks)
# ─────────────────────────────────────────────────────────────

@router.get("/years/{year_id}/annual-program")
def get_annual_program(
    year_id: str,
    db: DBSession = Depends(get_db),
    p: Principal = Depends(get_principal),
):
    """Full-year calendar view: 4 term blocks, parade dates, holidays, activities."""
    py = _get_year_or_404(year_id, db)
    _require_year_access(p, py)

    all_dates = db.query(ParadeNight).filter(
        ParadeNight.planning_year_id == year_id,
    ).order_by(ParadeNight.date).all()

    all_holidays = db.query(HolidayPeriod).filter(
        HolidayPeriod.planning_year_id == year_id,
    ).order_by(HolidayPeriod.start_date).all()

    all_anchors = db.query(AnchorEvent).filter(
        AnchorEvent.planning_year_id == year_id,
        AnchorEvent.is_archived == False,  # noqa: E712
    ).order_by(AnchorEvent.start_date).all()

    # Bulk-fetch sessions for all parade nights (all_dates ARE the parade nights now)
    all_pn_ids_pre = [d.id for d in all_dates]
    pn_map: dict = {}
    ts_by_pn: dict = {}
    ts_rows: list = []
    # all_dates ARE the parade nights now (Phase B merge); pn_map maps id→night
    pn_map = {d.id: d for d in all_dates}
    if all_pn_ids_pre:
        ts_rows = db.query(TrainingSession).filter(
            TrainingSession.parade_night_id.in_(all_pn_ids_pre),
            TrainingSession.is_archived == False,  # noqa: E712
        ).all()
        for ts in ts_rows:
            ts_by_pn.setdefault(ts.parade_night_id, []).append(ts)

    # CLASS-06: which Training Class(es) each session targets, additive to
    # the inline sessions_summary built below. One bulk query for the whole
    # year (matching this endpoint's own existing avoid-N+1 discipline)
    # rather than per-date or per-session calls.
    classes_by_session: dict[str, list[dict]] = {}
    if all_pn_ids_pre and ts_rows:
        aud_rows = (
            db.query(SessionAudience, TrainingClass)
            .join(TrainingClass, SessionAudience.training_class_id == TrainingClass.id)
            .filter(SessionAudience.session_id.in_([s.id for s in ts_rows]))
            .all()
        )
        for aud, tc in aud_rows:
            classes_by_session.setdefault(aud.session_id, []).append(
                {"training_class_id": tc.id, "display_name": tc.display_name})

    # CLASS-21: pre-load CurriculumItem core_status for Foundation/Extension PW
    # filter. One bulk query, no N+1. ts_rows is only defined when all_pn_ids_pre
    # is non-empty (same guard as the classes_by_session block above).
    ci_tier_ap: dict[str, dict] = {}
    if all_pn_ids_pre and ts_rows:
        ci_ids_ap = {s.curriculum_item_id for s in ts_rows if s.curriculum_item_id}
        if ci_ids_ap:
            for ci in db.query(CurriculumItem).filter(CurriculumItem.id.in_(ci_ids_ap)).all():
                ci_tier_ap[ci.id] = {"core_status": ci.core_status, "is_optional": ci.is_optional}

    # Build per-date-id session index (pn_id == date_id now — same record)
    ts_by_date_id: dict[str, list] = ts_by_pn  # direct alias: same keys

    # Bulk-load conflict counts and notices per parade night (2 extra queries, no N+1)
    all_pn_ids = all_pn_ids_pre
    conflict_counts_map: dict[str, int] = {}
    notices_by_date_id: dict[str, list] = {}
    if all_pn_ids:
        for c in db.query(PlanningConflict).filter(
            PlanningConflict.parade_night_id.in_(all_pn_ids),
            PlanningConflict.is_resolved == False,  # noqa: E712
        ).all():
            conflict_counts_map[c.parade_night_id] = conflict_counts_map.get(c.parade_night_id, 0) + 1
        for n in db.query(PlanningNotice).filter(
            PlanningNotice.parade_night_id.in_(all_pn_ids),
            PlanningNotice.is_archived == False,  # noqa: E712
        ).all():
            notices_by_date_id.setdefault(n.parade_night_id, []).append(n)

    def _in_range(d: str, start: str, end: str) -> bool:
        return start <= d <= end

    # Build term blocks using WA defaults
    yr = str(py.year)
    terms = []
    assigned_pn_ids: set[str] = set()
    for t_num, (ts, te) in sorted(_WA_TERM_RANGES.items()):
        t_start = f"{yr}-{ts}"
        t_end   = f"{yr}-{te}"
        term_label = f"T{t_num}"

        t_dates = [d for d in all_dates if t_start <= d.date <= t_end]
        assigned_pn_ids.update(d.id for d in t_dates)
        t_holidays = [h for h in all_holidays
                      if not (h.end_date < t_start or h.start_date > t_end)]
        t_anchors = [a for a in all_anchors
                     if _in_range(a.start_date, t_start, t_end)]

        # Per-date session fill summary (all_dates ARE parade nights in Phase B)
        date_summaries = []
        for pn_obj in t_dates:
            sessions = ts_by_pn.get(pn_obj.id, [])
            session_count = pn_obj.session_count
            filled = len([s for s in sessions if s.curriculum_item_id or s.custom_title])
            in_hol = any(_in_range(pn_obj.date, h.start_date, h.end_date)
                         for h in t_holidays if h.affects_parade)
            # Inline session summaries — facilitator & location are denormalized on TrainingSession
            date_sessions = ts_by_date_id.get(pn_obj.id, [])
            sessions_summary = [
                {
                    "session_id": s.id,
                    "period": s.period_number,
                    "cadet_group": s.cadet_group,
                    "title": s.curriculum_title_at_time or s.custom_title,
                    "curriculum_code": s.curriculum_code_at_time,
                    "facilitator": s.facilitator_display_name_at_time,
                    "location": s.training_area_name_at_time,
                    "training_classes": classes_by_session.get(s.id, []),
                    "core_status": ci_tier_ap.get(s.curriculum_item_id or "", {}).get("core_status"),
                }
                for s in date_sessions
            ]
            date_summaries.append({
                **_night_out_as_date(pn_obj),
                "term": term_label,
                "session_count": session_count,
                "filled_count": filled,
                "in_holiday": in_hol,
                "sessions_summary": sessions_summary,
                "conflict_count": conflict_counts_map.get(pn_obj.id, 0),
                "notices": [_notice_out(n) for n in notices_by_date_id.get(pn_obj.id, [])],
            })

        def _anchor_v14_out(a: AnchorEvent) -> dict:
            base = _anchor_out(a)
            base["importance_level"] = a.importance_level
            base["audience_staff_only"] = a.audience_staff_only
            base["audience_proficient"] = a.audience_proficient
            base["audience_first_years"] = a.audience_first_years
            base["cea_activity_id"] = a.cea_activity_id
            base["nomination_end_date"] = a.nomination_end_date
            base["unit_name"] = a.unit_name
            return base

        def _holiday_v14_out(h: HolidayPeriod) -> dict:
            base = _holiday_out(h)
            base["holiday_type"] = h.holiday_type
            return base

        terms.append({
            "term": term_label,
            "term_number": t_num,
            "start_date": t_start,
            "end_date": t_end,
            "parade_count": len(t_dates),
            "parade_dates": date_summaries,
            "holidays": [_holiday_v14_out(h) for h in t_holidays],
            "activities": [_anchor_v14_out(a) for a in t_anchors],
        })

    # Parade nights outside the WA term ranges (school holidays, late Jan, late Dec)
    # must still appear in the PW calendar. Assign each to the nearest preceding term
    # (nights before T1 go to T1; nights in T1-T2 break go to T1; etc.).
    if terms:
        _term_end_dates = [f"{yr}-{te}" for _, (_, te) in sorted(_WA_TERM_RANGES.items())]
        for pn_obj in all_dates:
            if pn_obj.id in assigned_pn_ids:
                continue
            target_idx = 0
            for i, t_end_str in enumerate(_term_end_dates):
                if t_end_str < pn_obj.date:
                    target_idx = i
            sessions = ts_by_pn.get(pn_obj.id, [])
            filled = len([s for s in sessions if s.curriculum_item_id or s.custom_title])
            in_hol = any(_in_range(pn_obj.date, h.start_date, h.end_date)
                         for h in all_holidays if h.affects_parade)
            date_sessions = ts_by_date_id.get(pn_obj.id, [])
            sessions_summary = [
                {
                    "session_id": s.id,
                    "period": s.period_number,
                    "cadet_group": s.cadet_group,
                    "title": s.curriculum_title_at_time or s.custom_title,
                    "curriculum_code": s.curriculum_code_at_time,
                    "facilitator": s.facilitator_display_name_at_time,
                    "location": s.training_area_name_at_time,
                    "training_classes": classes_by_session.get(s.id, []),
                    "core_status": ci_tier_ap.get(s.curriculum_item_id or "", {}).get("core_status"),
                }
                for s in date_sessions
            ]
            terms[target_idx]["parade_dates"].append({
                **_night_out_as_date(pn_obj),
                "term": terms[target_idx]["term"],
                "session_count": pn_obj.session_count,
                "filled_count": filled,
                "in_holiday": in_hol,
                "sessions_summary": sessions_summary,
                "conflict_count": conflict_counts_map.get(pn_obj.id, 0),
                "notices": [_notice_out(n) for n in notices_by_date_id.get(pn_obj.id, [])],
            })
            terms[target_idx]["parade_count"] += 1

    # Overall stats (uses pn_map / ts_by_pn already fetched above — no extra queries)
    total_dates = len(all_dates)
    active_dates = sum(1 for d in all_dates if d.is_active)
    all_pn_ids = all_pn_ids_pre
    total_sessions = 0
    filled_sessions = 0
    if all_pn_ids:
        all_ts = [s for sessions in ts_by_pn.values() for s in sessions]
        total_slots = sum(
            pn_map[pnid].session_count if pnid in pn_map else 3
            for pnid in all_pn_ids
        ) * len(CADET_GROUPS)
        total_sessions = total_slots
        filled_sessions = len([s for s in all_ts if s.curriculum_item_id or s.custom_title])

    # Wing HQ event overlay — resolve wing_id from planning year or squadron
    overlay_wing_id = py.wing_id
    if not overlay_wing_id and py.unit_id:
        sq = db.get(Squadron, py.unit_id)
        if sq:
            overlay_wing_id = sq.wing_id

    wing_events = []
    if overlay_wing_id:
        we_rows = db.query(WingHQEvent).filter(
            WingHQEvent.wing_id == overlay_wing_id,
            WingHQEvent.year == py.year,
            WingHQEvent.is_archived == False,  # noqa: E712
            WingHQEvent.status != "cancelled",
        ).order_by(WingHQEvent.start_date).all()
        for we in we_rows:
            wing_events.append({
                "id": we.id,
                "title": we.title,
                "event_type": we.event_type,
                "start_date": we.start_date,
                "end_date": we.end_date,
                "planning_importance": we.planning_importance,
                "audience": we.audience or [],
                "location": we.location,
                "requires_squadron_action": we.requires_squadron_action,
                "is_planning_anchor": we.is_planning_anchor,
                "notes": we.notes,
                "source": "wing_hq",
            })

    return {
        "planning_year_id": year_id,
        "year": py.year,
        "name": py.name,
        "total_parade_dates": total_dates,
        "active_parade_dates": active_dates,
        "total_session_slots": total_sessions,
        "filled_session_slots": filled_sessions,
        "terms": terms,
        "wing_events": wing_events,
        "wing_id": overlay_wing_id,
    }


# ─────────────────────────────────────────────────────────────
# V14 — Year Rollover
# ─────────────────────────────────────────────────────────────

class RolloverIn(BaseModel):
    target_year: Optional[int] = None   # defaults to source year + 1
    name: Optional[str] = None
    copy_holidays: bool = True
    carry_incomplete_sessions: bool = True
    # CLASS-11: without this, a squadron's Training Classes (e.g. Senior 1,
    # Senior 2) simply cease to exist in the new year -- TrainingClass is
    # scoped per training_year_id, and nothing previously copied it forward.
    copy_training_classes: bool = True


@router.post("/years/{year_id}/rollover", deprecated=True)
def rollover_year(
    year_id: str,
    body: RolloverIn,
    db: DBSession = Depends(get_db),
    p: Principal = Depends(get_principal),
):
    """Deprecated. Superseded by POST /years/copy-setup.

    Still fully functional, and deliberately so: it copies holidays and carries
    incomplete sessions, which copy-setup does not, and both frontends still
    call it. Degrading it to a copy-setup shim would leave existing callers
    silently doing less rather than "keeping working". Retiring it is its own
    task, once no caller remains.

    The naming defect is fixed here, though, because it reached production: the
    target name is derived rather than built by arrowing the source name, so no
    more "2026 Training Year -> 2027". The year is a calendar fact; its name is
    not a place to record where it came from.
    """
    py = _get_year_or_404(year_id, db)
    _require_year_access(p, py, write=True, db=db)
    if py.unit_id:
        _sq = db.get(Squadron, py.unit_id)
        require_can_write_squadron(p, py.unit_id, _sq.wing_id if _sq else py.wing_id)

    target_year = body.target_year or (py.year + 1)
    new_name = body.name or year_display_name(target_year)

    # Check for existing year
    existing = db.query(PlanningYear).filter(
        PlanningYear.unit_id == py.unit_id,
        PlanningYear.year == target_year,
    ).first()
    if existing:
        raise HTTPException(409, detail={"error": "planning_year_already_exists",
                                         "existing_id": existing.id})

    new_py = PlanningYear(
        id=str(uuid.uuid4()),
        unit_id=py.unit_id, wing_id=py.wing_id,
        year=target_year, name=new_name,
        active_status=True,
        created_by=p.user_id, created_at=utcnow(), updated_at=utcnow(),
    )
    db.add(new_py)
    db.flush()

    # Copy holiday periods
    holidays_copied = 0
    if body.copy_holidays:
        old_holidays = db.query(HolidayPeriod).filter(
            HolidayPeriod.planning_year_id == year_id,
        ).all()
        year_delta = target_year - py.year
        for h in old_holidays:
            try:
                new_start = date.fromisoformat(h.start_date).replace(
                    year=date.fromisoformat(h.start_date).year + year_delta
                ).isoformat()
                new_end = date.fromisoformat(h.end_date).replace(
                    year=date.fromisoformat(h.end_date).year + year_delta
                ).isoformat()
            except (ValueError, AttributeError):
                new_start, new_end = h.start_date, h.end_date
            nh = HolidayPeriod(
                id=str(uuid.uuid4()), planning_year_id=new_py.id,
                jurisdiction=h.jurisdiction, name=h.name,
                start_date=new_start, end_date=new_end,
                holiday_type=h.holiday_type,
                affects_parade=h.affects_parade, notes=h.notes,
                created_at=utcnow(), updated_at=utcnow(),
            )
            db.add(nh)
            holidays_copied += 1

    # CLASS-11: copy active Training Classes forward into the new year.
    # Only the class *definitions* carry over (stage/name/sequence/expected
    # count) -- no session or curriculum-progress data, since those are
    # inherently new-year concepts computed fresh from that year's own
    # Sessions. A class archived in the source year is deliberately not
    # copied (an intentionally-retired class shouldn't reappear).
    training_classes_copied = 0
    if body.copy_training_classes:
        old_classes = db.query(TrainingClass).filter(
            TrainingClass.training_year_id == year_id,
            TrainingClass.is_archived == False,  # noqa: E712
        ).all()
        for oc in old_classes:
            db.add(TrainingClass(
                squadron_id=oc.squadron_id, training_year_id=new_py.id,
                training_stage_id=oc.training_stage_id, display_name=oc.display_name,
                class_number=oc.class_number, expected_count=oc.expected_count, notes=oc.notes,
                # start_date/end_date are deliberately NOT copied -- they're
                # specific calendar dates within the source year's own
                # season and would be wrong (stale) in the new year, unlike
                # holidays above which are explicitly year-shifted. Left
                # null for the admin to set for the new year if needed.
                created_by=p.user_id, updated_by=p.user_id,
            ))
            training_classes_copied += 1

    # Copy parade dates (same weekday pattern, new year) — Phase B: create ParadeNight rows
    dates_copied = 0
    new_sq = db.get(Squadron, new_py.unit_id) if new_py.unit_id else None
    old_dates = db.query(ParadeNight).filter(
        ParadeNight.planning_year_id == year_id,
        ParadeNight.is_active == True,  # noqa: E712
    ).order_by(ParadeNight.date).all()
    year_delta = target_year - py.year
    for old_pn in old_dates:
        try:
            new_date_str = date.fromisoformat(old_pn.date).replace(
                year=date.fromisoformat(old_pn.date).year + year_delta
            ).isoformat()
        except (ValueError, AttributeError):
            continue
        new_pn = ParadeNight(
            id=str(uuid.uuid4()),
            squadron_id=new_py.unit_id,
            wing_id=new_sq.wing_id if new_sq else old_pn.wing_id,
            date=new_date_str,
            planning_year_id=new_py.id,
            parade_type=old_pn.parade_type,
            is_active=True,
            term=old_pn.term,
            created_by=p.user_id,
            created_at=utcnow(), updated_at=utcnow(),
        )
        db.add(new_pn)
        dates_copied += 1

    # Carry forward incomplete sessions as draft assignments
    sessions_carried = 0
    if body.carry_incomplete_sessions:
        # Find sessions that were NOT delivered in the old year
        old_pn_ids = [d.id for d in old_dates]
        if old_pn_ids:
            incomplete = db.query(TrainingSession).filter(
                TrainingSession.parade_night_id.in_(old_pn_ids),
                TrainingSession.is_archived == False,  # noqa: E712
                TrainingSession.status.in_(["planned", "not_delivered"]),
                TrainingSession.curriculum_item_id.isnot(None),
            ).all()
            sessions_carried = len(incomplete)
            # We note them but don't auto-assign (planner should review)

    db.commit()
    audit(db, p, object_type="planning_year", object_id=new_py.id, action="rollover",
          new={"source_year": py.year, "target_year": target_year,
               "holidays_copied": holidays_copied, "dates_copied": dates_copied,
               "incomplete_sessions_noted": sessions_carried,
               "training_classes_copied": training_classes_copied})

    return {
        "ok": True,
        "new_planning_year_id": new_py.id,
        "year": target_year,
        "name": new_name,
        "holidays_copied": holidays_copied,
        "parade_dates_copied": dates_copied,
        "incomplete_sessions_noted": sessions_carried,
        "training_classes_copied": training_classes_copied,
    }


# ─────────────────────────────────────────────────────────────
# Spreadsheet Export — Annual Program + Schedule
# ─────────────────────────────────────────────────────────────

def _neutralise_cell(v):
    """Prevent CSV/formula injection in spreadsheet cells."""
    s = str(v) if v is not None else ""
    return ("'" + s) if s[:1] in ("=", "+", "-", "@") else s


@router.get("/years/{year_id}/export.xlsx")
def export_annual_program_xlsx(
    year_id: str,
    db: DBSession = Depends(get_db),
    p: Principal = Depends(get_principal),
):
    """Export the annual program (parade dates, holidays, anchors) as XLSX."""
    import openpyxl
    from openpyxl.styles import Font, PatternFill, Alignment
    py = _get_year_or_404(year_id, db)
    _require_year_access(p, py)

    wb = openpyxl.Workbook()

    # Sheet 1 — Parade Dates
    ws1 = wb.active
    ws1.title = "Parade Dates"
    hdr_fill = PatternFill("solid", fgColor="002F65")
    hdr_font = Font(color="FFFFFF", bold=True)
    dates_headers = ["Date", "Day", "Type", "Active", "Notes", "Term"]
    ws1.append(dates_headers)
    for cell in ws1[1]:
        cell.fill = hdr_fill; cell.font = hdr_font
    all_dates = db.query(ParadeNight).filter(ParadeNight.planning_year_id == year_id).order_by(ParadeNight.date).all()
    yr_str = str(py.year)
    def _term_label(ds: str) -> str:
        for t_num, (ts, te) in sorted(_WA_TERM_RANGES.items()):
            if f"{yr_str}-{ts}" <= ds <= f"{yr_str}-{te}":
                return f"T{t_num}"
        return ""
    for d in all_dates:
        try:
            dow = date.fromisoformat(d.date).strftime("%A")
        except Exception:
            dow = ""
        ws1.append([d.date, dow, d.parade_type or "standard",
                    "Yes" if d.is_active else "No",
                    _neutralise_cell(d.notes or ""), _term_label(d.date)])

    # Sheet 2 — Holidays
    ws2 = wb.create_sheet("Holidays")
    ws2.append(["Name", "Start Date", "End Date", "Type", "Affects Parade"])
    for cell in ws2[1]:
        cell.fill = hdr_fill; cell.font = hdr_font
    for h in db.query(HolidayPeriod).filter(HolidayPeriod.planning_year_id == year_id).order_by(HolidayPeriod.start_date).all():
        ws2.append([_neutralise_cell(h.name), h.start_date, h.end_date,
                    h.holiday_type or "", "Yes" if h.affects_parade else "No"])

    # Sheet 3 — Anchor Events
    ws3 = wb.create_sheet("Anchor Events")
    ws3.append(["Name", "Start Date", "End Date", "Type", "Importance", "Location", "Notes"])
    for cell in ws3[1]:
        cell.fill = hdr_fill; cell.font = hdr_font
    for a in db.query(AnchorEvent).filter(AnchorEvent.planning_year_id == year_id, AnchorEvent.is_archived == False).order_by(AnchorEvent.start_date).all():  # noqa: E712
        ws3.append([_neutralise_cell(a.name), a.start_date, a.end_date or "",
                    a.event_type or "", a.importance_level or "",
                    _neutralise_cell(a.location or ""), _neutralise_cell(a.description or "")])

    bio = io.BytesIO(); wb.save(bio); bio.seek(0)
    audit(db, p, object_type="export", object_id=year_id, action="export",
          new={"type": "annual_program", "fmt": "xlsx"})
    fname = f"annual-program-{py.year}.xlsx"
    return StreamingResponse(bio, media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                             headers={"Content-Disposition": f"attachment; filename={fname}"})


@router.get("/years/{year_id}/export")
async def export_year_csv(
    year_id: str,
    db: DBSession = Depends(get_db),
    p: Principal = Depends(get_principal),
):
    """Download a planning year's full program as a CSV file.

    Sections: CEA activities, parade schedule, scheduled sessions.
    BOM-prefixed UTF-8 for direct Excel compatibility.
    Filename: AAFC_TMS_{year}_{scope}_{date}.csv
    """
    from datetime import date as _date

    year = db.query(PlanningYear).filter(PlanningYear.id == year_id).first()
    if not year:
        raise HTTPException(404, "Year not found")
    _require_year_access(p, year)

    # ── Scope label ──────────────────────────────────────────────────────────
    # Principal carries IDs only; look up the human-readable short name.
    if p.acting_squadron_id or p.squadron_id:
        sq_id = p.acting_squadron_id or p.squadron_id
        sq = db.get(Squadron, sq_id)
        scope_label = (sq.short_name if sq else "SQUADRON").upper()
    elif p.acting_wing_id or p.wing_id:
        wg_id = p.acting_wing_id or p.wing_id
        wg = db.get(Wing, wg_id)
        scope_label = (wg.short_name if wg else "WING").upper()
    else:
        scope_label = "NATIONAL"

    buf = io.StringIO()
    writer = _csv.writer(buf)

    # ── Header ───────────────────────────────────────────────────────────────
    writer.writerow([f"TRAINING PROGRAM — {scope_label} — {year.year}"])
    writer.writerow([f"Exported: {_date.today().isoformat()}"])
    writer.writerow([])

    # ── CEA Activities ────────────────────────────────────────────────────────
    writer.writerow(["ACTIVITIES"])
    writer.writerow(["cea_activity_id", "name", "type", "importance", "owning_level"])
    activities = (
        db.query(CeaActivity)
        .filter(CeaActivity.planning_year_id == year_id, CeaActivity.is_archived.is_(False))
        .order_by(CeaActivity.activity_name)
        .all()
    )
    for act in activities:
        writer.writerow([
            act.cea_activity_id or "",
            act.activity_name or "",
            act.activity_type or "",
            act.importance or "",
            act.parent_unit or act.host_unit or "",
        ])
    writer.writerow([])

    # ── Parade Schedule ───────────────────────────────────────────────────────
    writer.writerow(["PARADE SCHEDULE"])
    writer.writerow(["date", "type", "term", "week", "notes"])
    parade_dates = (
        db.query(ParadeNight)
        .filter(ParadeNight.planning_year_id == year_id, ParadeNight.is_active.is_(True))
        .order_by(ParadeNight.date)
        .all()
    )
    for pn_row in parade_dates:
        writer.writerow([
            pn_row.date or "",
            pn_row.parade_type or "",
            pn_row.term or "",
            pn_row.week_number or "",
            (pn_row.notes or "").replace("\n", " "),
        ])
    writer.writerow([])

    # ── Sessions ──────────────────────────────────────────────────────────────
    # Build a parade_night_id → parade_date lookup from the nights already loaded.
    pn_to_date: dict[str, str] = {pn_row.id: pn_row.date or "" for pn_row in parade_dates}

    sessions = (
        db.query(TrainingSession)
        .join(ParadeNight, TrainingSession.parade_night_id == ParadeNight.id)
        .filter(
            ParadeNight.planning_year_id == year_id,
            ParadeNight.is_active.is_(True),
            TrainingSession.is_archived.is_(False),
        )
        .order_by(ParadeNight.date, TrainingSession.period_number)
        .all()
    )

    # Batch-load session audiences to avoid N+1 queries.
    session_ids = [s.id for s in sessions]
    sa_by_session: dict[str, list[str]] = {s.id: [] for s in sessions}
    if session_ids:
        audience_rows = (
            db.query(SessionAudience, TrainingClass)
            .join(TrainingClass, SessionAudience.training_class_id == TrainingClass.id)
            .filter(SessionAudience.session_id.in_(session_ids))
            .all()
        )
        for sa, tc in audience_rows:
            sa_by_session[sa.session_id].append(tc.display_name)

    writer.writerow(["SESSIONS"])
    writer.writerow(["date", "period", "title", "facilitator", "training_classes"])
    for sess in sessions:
        parade_date_str = pn_to_date.get(sess.parade_night_id, "")
        classes = "; ".join(sa_by_session.get(sess.id, []))
        writer.writerow([
            parade_date_str,
            sess.period_number or "",
            sess.custom_title or sess.curriculum_title_at_time or "",
            sess.facilitator_display_name_at_time or "",
            classes,
        ])

    csv_bytes = buf.getvalue().encode("utf-8-sig")  # BOM for Excel compatibility
    safe_scope = scope_label.replace(" ", "_").replace("/", "-")
    filename = f"AAFC_TMS_{year.year}_{safe_scope}_{_date.today().isoformat()}.csv"

    return StreamingResponse(
        io.BytesIO(csv_bytes),
        media_type="text/csv",
        headers={"Content-Disposition": f'attachment; filename="{filename}"'},
    )


@router.get("/years/{year_id}/schedule/export.xlsx")
def export_schedule_xlsx(
    year_id: str,
    db: DBSession = Depends(get_db),
    p: Principal = Depends(get_principal),
):
    """Export all scheduled sessions for a planning year as XLSX."""
    import openpyxl
    from openpyxl.styles import Font, PatternFill
    py = _get_year_or_404(year_id, db)
    _require_year_access(p, py)
    sq_id = p.acting_squadron_id or p.squadron_id

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Schedule"
    hdr_fill = PatternFill("solid", fgColor="002F65")
    hdr_font = Font(color="FFFFFF", bold=True)
    headers = ["Date", "Day", "Term", "Session", "Group", "Phase", "Code", "Title", "Facilitator", "Room", "Status"]
    ws.append(headers)
    for cell in ws[1]:
        cell.fill = hdr_fill; cell.font = hdr_font

    yr_str = str(py.year)
    def _term_lbl(ds: str) -> str:
        for t_num, (ts, te) in sorted(_WA_TERM_RANGES.items()):
            if f"{yr_str}-{ts}" <= ds <= f"{yr_str}-{te}":
                return f"T{t_num}"
        return ""

    all_dates = db.query(ParadeNight).filter(
        ParadeNight.planning_year_id == year_id,
        ParadeNight.is_active == True,  # noqa: E712
    ).order_by(ParadeNight.date).all()

    for pn in all_dates:
        try:
            dow = date.fromisoformat(pn.date).strftime("%A")
        except Exception:
            dow = ""
        term = _term_lbl(pn.date)
        sessions = db.query(TrainingSession).filter(
            TrainingSession.parade_night_id == pn.id,
            TrainingSession.is_archived == False,  # noqa: E712
            *([] if sq_id is None else [TrainingSession.squadron_id == sq_id]),
        ).order_by(TrainingSession.period_number).all()
        for s in sessions:
            ws.append([
                pn.date, dow, term,
                _neutralise_cell(s.period_number),
                _neutralise_cell(s.cadet_group or ""),
                _neutralise_cell(s.phase_at_time or ""),
                _neutralise_cell(s.curriculum_code_at_time or ""),
                _neutralise_cell(s.curriculum_title_at_time or s.custom_title or ""),
                _neutralise_cell(s.facilitator_display_name_at_time or ""),
                _neutralise_cell(s.training_area_name_at_time or ""),
                _neutralise_cell(s.status or "planned"),
            ])

    bio = io.BytesIO(); wb.save(bio); bio.seek(0)
    audit(db, p, object_type="export", object_id=year_id, action="export",
          new={"type": "schedule", "fmt": "xlsx"})
    fname = f"schedule-{py.year}.xlsx"
    return StreamingResponse(bio, media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                             headers={"Content-Disposition": f"attachment; filename={fname}"})


@router.post("/years/{year_id}/schedule/import")
async def import_schedule_xlsx(
    year_id: str,
    preview: bool = Query(False),
    file: UploadFile = File(...),
    db: DBSession = Depends(get_db),
    p: Principal = Depends(get_principal),
):
    """Import an edited schedule XLSX back into the planning year.

    Expected columns (case-insensitive): Date, Session, Group, Code, Title, Facilitator.
    Returns a row-level diff when preview=true; applies changes when preview=false.
    Only updates curriculum_item_id and custom_title — all other session fields are untouched.
    Permission-gated: same write rules as schedule edits.
    """
    import openpyxl
    from ..config import settings

    py = _get_year_or_404(year_id, db)
    _require_year_access(p, py, write=True, db=db)
    if py.unit_id:
        _sq = db.get(Squadron, py.unit_id)
        require_can_write_squadron(p, py.unit_id, _sq.wing_id if _sq else py.wing_id)

    raw = await file.read()
    if len(raw) > settings.UPLOAD_MAX_MB * 1024 * 1024:
        raise HTTPException(413, detail={"error": "file_too_large"})
    try:
        wb = openpyxl.load_workbook(io.BytesIO(raw), read_only=True, data_only=True)
    except Exception as exc:
        raise HTTPException(400, detail={"error": "unreadable_workbook", "message": str(exc)[:120]})

    ws = wb.active
    rows_iter = ws.iter_rows(values_only=True)
    raw_headers = next(rows_iter, None)
    if not raw_headers:
        raise HTTPException(400, detail={"error": "empty_sheet"})
    headers = [str(h).strip().lower() if h is not None else "" for h in raw_headers]

    def _col(name: str) -> int | None:
        aliases = {"date": ["date"], "session": ["session", "period", "session #", "period number"],
                   "group": ["group", "cadet group"], "code": ["code", "curriculum code"],
                   "title": ["title", "custom title"], "facilitator": ["facilitator"]}
        for a in aliases.get(name, [name]):
            if a in headers:
                return headers.index(a)
        return None

    ci = {k: _col(k) for k in ("date", "session", "group", "code", "title")}
    if ci["date"] is None or ci["session"] is None:
        raise HTTPException(400, detail={"error": "missing_required_columns",
                                         "message": "Sheet must have Date and Session columns"})

    # Phase B: query ParadeNight directly — pn.id == former parade_date_id alias
    all_pd = db.query(ParadeNight).filter(ParadeNight.planning_year_id == year_id).all()
    pd_by_date = {d.date: d for d in all_pd}

    sq_id = p.acting_squadron_id or p.squadron_id

    preview_rows: list[dict] = []
    updated = skipped = not_found = 0

    for raw_row in rows_iter:
        if all(c is None for c in raw_row):
            continue
        def _cell(idx):
            if idx is None: return ""
            v = raw_row[idx] if idx < len(raw_row) else None
            return str(v).strip() if v is not None else ""

        date_val = _cell(ci["date"])
        session_val = _cell(ci["session"])
        group_val = _cell(ci["group"]) if ci["group"] is not None else ""
        code_val = _cell(ci["code"]) if ci["code"] is not None else ""
        title_val = _cell(ci["title"]) if ci["title"] is not None else ""

        if not date_val or not session_val:
            continue

        # Normalise date (handle date objects from openpyxl)
        try:
            if hasattr(date_val, "strftime"):
                date_str = date_val.strftime("%Y-%m-%d")
            else:
                date_str = str(date_val)[:10]
        except Exception:
            date_str = str(date_val)[:10]

        try:
            period = int(session_val)
        except (ValueError, TypeError):
            not_found += 1
            preview_rows.append({"date": date_str, "session": session_val, "group": group_val,
                                  "code": code_val, "title": title_val,
                                  "action": "not_found", "reason": "invalid_session_number"})
            continue

        pd_obj = pd_by_date.get(date_str)
        if not pd_obj:
            not_found += 1
            preview_rows.append({"date": date_str, "session": period, "group": group_val,
                                  "code": code_val, "title": title_val,
                                  "action": "not_found", "reason": "no_parade_night_on_date"})
            continue

        q = db.query(TrainingSession).filter(
            TrainingSession.parade_night_id == pd_obj.id,
            TrainingSession.period_number == period,
            TrainingSession.is_archived == False,  # noqa: E712
        )
        if group_val:
            q = q.filter(TrainingSession.cadet_group == group_val)
        if sq_id:
            q = q.filter(TrainingSession.squadron_id == sq_id)
        sess_obj = q.first()

        if not sess_obj:
            not_found += 1
            preview_rows.append({"date": date_str, "session": period, "group": group_val,
                                  "code": code_val, "title": title_val,
                                  "action": "not_found", "reason": "no_session_record"})
            continue

        cur_code = sess_obj.curriculum_code_at_time or ""
        cur_title = sess_obj.curriculum_title_at_time or sess_obj.custom_title or ""

        # Resolve curriculum_item_id from code if provided
        new_ci_id = sess_obj.curriculum_item_id
        new_code = code_val or cur_code
        new_title = title_val or cur_title
        if code_val and code_val != cur_code:
            ci_match = db.query(CurriculumItem).filter(
                CurriculumItem.identifier == code_val,
                CurriculumItem.is_archived == False,  # noqa: E712
            ).first()
            if not ci_match:
                ci_match = db.query(CurriculumItem).filter(
                    CurriculumItem.code == code_val,
                    CurriculumItem.is_archived == False,  # noqa: E712
                ).first()
            if ci_match:
                new_ci_id = ci_match.id
                new_code = ci_match.identifier or ci_match.code
                new_title = ci_match.title

        changed = (new_ci_id != sess_obj.curriculum_item_id) or (new_title != cur_title)

        if not changed:
            skipped += 1
            preview_rows.append({"date": date_str, "session": period, "group": group_val,
                                  "code": new_code, "title": new_title,
                                  "current_code": cur_code, "current_title": cur_title,
                                  "action": "unchanged"})
            continue

        preview_rows.append({"date": date_str, "session": period, "group": group_val,
                              "code": new_code, "title": new_title,
                              "current_code": cur_code, "current_title": cur_title,
                              "action": "update"})

        if not preview:
            sess_obj.curriculum_item_id = new_ci_id
            sess_obj.curriculum_code_at_time = new_code
            sess_obj.curriculum_title_at_time = new_title
            if not code_val and title_val:
                sess_obj.curriculum_item_id = None
                sess_obj.custom_title = title_val
            updated += 1

    if not preview:
        db.commit()
        audit(db, p, object_type="schedule_import", object_id=year_id, action="import",
              new={"updated": updated, "skipped": skipped, "not_found": not_found})

    return {"ok": True, "preview": preview, "rows": preview_rows,
            "updated": updated, "skipped": skipped, "not_found": not_found}


# ── ANNUAL PROGRAM IMPORT ─────────────────────────────────────────────────────

@router.post("/years/{year_id}/import-program")
async def import_annual_program(
    year_id: str,
    preview: bool = Query(default=False),
    file: UploadFile = File(...),
    db: DBSession = Depends(get_db),
    p: Principal = Depends(get_principal),
):
    """Import Annual Program schedule from CSV or XLSX (CADET.Net export format).

    Expected columns (case-insensitive):
    SeqNr, Name, Start date, Start time, End date, End time, Unit, Owner, Status, Last Updated

    - Rows whose Name contains 'parade night' or 'training night' become
      ParadeDate + ParadeNight records linked to this planning year.
    - All other rows become Activity records.
    - For wing/national planning years the Unit column routes each row to
      the correct squadron within scope.
    - Duplicate parade dates (same date + unit) and duplicate SeqNr
      activities are skipped.
    - preview=true: classify and validate rows without writing to the database.
    """
    _require_plan_write(p)
    py = _get_year_or_404(year_id, db)
    _require_year_access(p, py, write=True, db=db)
    # _require_year_access has no proxy/delegation awareness (by design, per
    # architecture.md, for endpoints with no proxy concept) -- but importing
    # a schedule into a squadron's plan IS a proxy/delegation-relevant
    # squadron-scoped write, same as any other. Only enforced when this call
    # will actually write (preview=true never touches the database) (REM-45,
    # Stage 12 follow-up).
    if not preview and py.unit_id:
        wing_id = py.wing_id
        sqn = db.get(Squadron, py.unit_id)
        if sqn:
            wing_id = sqn.wing_id
        require_can_write_squadron(p, py.unit_id, wing_id)

    # Build the set of squadrons in scope for unit resolution
    sqn_q = db.query(Squadron).filter(Squadron.is_archived == False)  # noqa: E712
    if py.unit_id:
        sqn_q = sqn_q.filter(Squadron.id == py.unit_id)
    elif py.wing_id:
        sqn_q = sqn_q.filter(Squadron.wing_id == py.wing_id)
    all_sqns = sqn_q.all()
    default_sqn = db.get(Squadron, py.unit_id) if py.unit_id else None

    # Detect file format from filename / content-type
    fname = (file.filename or "").lower()
    is_xlsx = fname.endswith(".xlsx") or (file.content_type or "").startswith("application/vnd")

    content = await file.read()
    if len(content) > 5 * 1024 * 1024:
        raise HTTPException(413, detail={"error": "file_too_large",
                                         "message": "File exceeds the 5 MB limit."})
    try:
        rows = _parse_program_file(content, is_xlsx)
    except Exception:
        raise HTTPException(400, detail={"error": "file_parse_failed",
                                         "message": "The file could not be parsed. Ensure it is a valid CSV or XLSX."})

    if not rows:
        raise HTTPException(400, detail={"error": "empty_file"})

    # Deduplication sets — Phase B: query ParadeNight directly
    existing_pdate_keys: set[tuple[str, str | None]] = {
        (pn.date, pn.squadron_id)
        for pn in db.query(ParadeNight).filter(ParadeNight.planning_year_id == year_id).all()
    }
    act_q = db.query(Activity.cea_seq_nr).filter(
        Activity.cea_seq_nr.isnot(None),
        Activity.is_archived == False,  # noqa: E712
    )
    if py.unit_id:
        act_q = act_q.filter(Activity.squadron_id == py.unit_id)
    elif py.wing_id:
        act_q = act_q.filter(Activity.wing_id == py.wing_id)
    existing_seq_nrs: set[str] = {a.cea_seq_nr for a in act_q.all() if a.cea_seq_nr}

    parse_errors: list[str] = []
    preview_rows: list[dict] = []
    created_parade_dates = created_activities = skipped = 0

    for i, raw_row in enumerate(rows, start=2):
        norm = {k.strip().lower(): (v or "").strip() for k, v in raw_row.items()}

        seq_nr = norm.get("seqnr") or norm.get("seq nr") or norm.get("seq_nr") or norm.get("seq") or ""
        name = norm.get("name") or norm.get("activity name") or ""
        date_start_raw = norm.get("start date") or norm.get("startdate") or norm.get("start_date") or ""
        time_start_raw = norm.get("start time") or norm.get("starttime") or norm.get("start_time") or ""
        date_end_raw = norm.get("end date") or norm.get("enddate") or norm.get("end_date") or ""
        time_end_raw = norm.get("end time") or norm.get("endtime") or norm.get("end_time") or ""
        unit_col = norm.get("unit") or ""
        owner = norm.get("owner") or ""

        if not name or not date_start_raw:
            parse_errors.append(f"Row {i}: missing Name or Start date — skipped.")
            continue

        date_start = _parse_prog_date(date_start_raw)
        date_end = _parse_prog_date(date_end_raw) if date_end_raw else None
        time_start = _parse_prog_time(time_start_raw) if time_start_raw else None
        time_end = _parse_prog_time(time_end_raw) if time_end_raw else None

        if not date_start:
            parse_errors.append(f"Row {i}: unrecognised date '{date_start_raw}' — skipped.")
            continue

        # Resolve squadron
        if py.unit_id:
            resolved_sqn = default_sqn
        elif unit_col:
            resolved_sqn = _resolve_unit_sqn(unit_col, all_sqns)
            if not resolved_sqn:
                parse_errors.append(f"Row {i}: Unit '{unit_col}' not found in scope — skipped.")
                continue
        else:
            resolved_sqn = None  # wing/national activity with no unit specified

        # REM-45 residual / security review candidate 3: a wing/national-scoped plan year
        # only had year-level access enforced (_require_year_access, no proxy awareness --
        # correct for the year itself, which is wing/national-level data). But each row
        # here can resolve to a SPECIFIC squadron via the Unit column, and writing
        # squadron-scoped data must go through the same Proxy/Delegated-Intervention gate
        # every other squadron-scoped write in this app requires. Without this check, a
        # wing_admin could import CADET.Net rows into any squadron in their wing without
        # ever entering Proxy Mode -- confirmed live during this program (a wing_admin,
        # never in Proxy Mode, successfully created a real Activity for a squadron via this
        # exact path). Rows targeting a squadron the actor cannot write to are skipped
        # (both in preview, so the user can see it before committing, and at commit time),
        # not silently written -- the rest of a legitimate import is not blocked by one
        # out-of-authority row.
        scope_denied = resolved_sqn is not None and not p.can_write_squadron(resolved_sqn.id, resolved_sqn.wing_id)

        is_parade = _is_parade_row(name)
        row_type = "parade_date" if is_parade else "activity"

        if preview:
            preview_rows.append({
                "row": i, "seq_nr": seq_nr or None, "name": name, "type": row_type,
                "date_start": date_start, "date_end": date_end,
                "time_start": time_start, "time_end": time_end,
                "unit": unit_col,
                "resolved_sqn": resolved_sqn.code if resolved_sqn else None,
                "owner": owner or None,
                "status": "blocked_scope" if scope_denied else "new",
            })
            continue

        if scope_denied:
            parse_errors.append(
                f"Row {i}: you do not have write authority for squadron '{resolved_sqn.code}' "
                f"(enter Proxy/Delegated Intervention for that squadron first) — skipped.")
            skipped += 1
            continue

        if is_parade:
            sqn_id_for_key = resolved_sqn.id if resolved_sqn else None
            key = (date_start, sqn_id_for_key)
            if key in existing_pdate_keys:
                skipped += 1
                continue
            # Phase B: create ParadeNight directly with planning_year_id
            sqn_for_pn = resolved_sqn or (db.get(Squadron, py.unit_id) if py.unit_id else None)
            pn_import = ParadeNight(
                id=str(uuid.uuid4()),
                squadron_id=sqn_id_for_key or py.unit_id,
                wing_id=sqn_for_pn.wing_id if sqn_for_pn else py.wing_id,
                date=date_start,
                planning_year_id=year_id,
                parade_type="normal",
                is_active=True,
                created_by=p.user_id,
                created_at=utcnow(), updated_at=utcnow(),
            )
            db.add(pn_import)
            existing_pdate_keys.add(key)
            created_parade_dates += 1
        else:
            if seq_nr and seq_nr in existing_seq_nrs:
                skipped += 1
                continue
            sqn_id = resolved_sqn.id if resolved_sqn else py.unit_id
            w_id = resolved_sqn.wing_id if resolved_sqn else py.wing_id
            a = Activity(
                squadron_id=sqn_id, wing_id=w_id,
                owning_level="squadron" if sqn_id else "wing",
                activity_name=name, activity_type="program_import",
                date_start=date_start, date_end=date_end,
                time_start=time_start, time_end=time_end,
                oic=owner or None, cea_seq_nr=seq_nr or None,
            )
            db.add(a)
            if seq_nr:
                existing_seq_nrs.add(seq_nr)
            created_activities += 1

    if preview:
        return {
            "ok": True, "preview": True, "rows": preview_rows,
            "parse_errors": parse_errors,
        }

    db.commit()
    audit(db, p, object_type="planning_program", object_id=year_id, action="import",
          new={"created_parade_dates": created_parade_dates,
               "created_activities": created_activities,
               "skipped": skipped})

    return {
        "ok": True, "preview": False,
        "created_parade_dates": created_parade_dates,
        "created_activities": created_activities,
        "skipped": skipped,
        "total_rows": created_parade_dates + created_activities + skipped,
        "parse_errors": parse_errors,
    }


# ─── Facilitator Leave / Unavailability ───────────────────────────────────────

class FacilitatorLeaveIn(BaseModel):
    start_date: str
    end_date: str
    reason: Optional[str] = None
    notes: Optional[str] = None
    planning_year_id: Optional[str] = None


@router.get("/facilitators/{fac_id}/leave")
def list_facilitator_leave(
    fac_id: str,
    include_archived: bool = False,
    db: DBSession = Depends(get_db),
    p: Principal = Depends(get_principal),
):
    """List all active leave periods for a facilitator."""
    fac = db.get(Facilitator, fac_id)
    if not fac:
        raise HTTPException(404, detail={"error": "facilitator_not_found"})
    require_can_view_squadron(p, fac.squadron_id, fac.wing_id)
    query = db.query(PlanningFacilitatorLeave).filter(
        PlanningFacilitatorLeave.facilitator_id == fac_id,
    )
    if not include_archived:
        query = query.filter(PlanningFacilitatorLeave.is_archived == False)  # noqa: E712
    rows = query.order_by(PlanningFacilitatorLeave.start_date).all()
    return {
        "leave": [
            {
                "id": r.id, "facilitator_id": r.facilitator_id,
                "planning_year_id": r.planning_year_id,
                "start_date": r.start_date, "end_date": r.end_date,
                "reason": r.reason, "notes": r.notes,
                "created_by": r.created_by,
                "created_at": iso_z(r.created_at) if r.created_at else None,
                "is_archived": r.is_archived,
            }
            for r in rows
        ]
    }


@router.post("/facilitators/{fac_id}/leave")
def add_facilitator_leave(
    fac_id: str,
    body: FacilitatorLeaveIn,
    db: DBSession = Depends(get_db),
    p: Principal = Depends(get_principal),
):
    """Add a leave period for a facilitator and return affected scheduled sessions."""
    fac = db.get(Facilitator, fac_id)
    if not fac:
        raise HTTPException(404, detail={"error": "facilitator_not_found"})
    require_can_write_squadron(p, fac.squadron_id, fac.wing_id)
    if body.start_date > body.end_date:
        raise HTTPException(400, detail={"error": "start_date_after_end_date"})

    leave = PlanningFacilitatorLeave(
        id=str(uuid.uuid4()),
        facilitator_id=fac_id,
        planning_year_id=body.planning_year_id,
        start_date=body.start_date,
        end_date=body.end_date,
        reason=body.reason,
        notes=body.notes,
        created_by=p.user_id,
        is_archived=False,
    )
    db.add(leave)
    db.flush()

    # Find TrainingSessions that use this facilitator on parade dates in the leave window.
    # NOTE: this previously queried ScheduledSession -- the same never-written model
    # already fixed once for facilitator_workload() and again just above for
    # get_command_centre() -- so this conflict warning silently never fired for any real
    # session. Rewritten to the same ParadeDate -> ParadeNight -> TrainingSession join,
    # batched to avoid an N+1 ParadeDate lookup per session. Qualification-program Phase B,
    # 2026-08-08; see docs/qualification/03_data_integrity_review.md P1 finding #2.
    affected: list[dict] = []
    sessions_with_fac = (
        db.query(TrainingSession)
        .filter(
            TrainingSession.facilitator_id == fac_id,
            TrainingSession.is_archived == False,  # noqa: E712
        )
        .all()
    )
    parade_night_ids = {s.parade_night_id for s in sessions_with_fac if s.parade_night_id}
    pn_by_id: dict[str, ParadeNight] = {}
    if parade_night_ids:
        pn_by_id = {
            pn_obj.id: pn_obj
            for pn_obj in db.query(ParadeNight)
            .filter(ParadeNight.id.in_(list(parade_night_ids)))
            .all()
        }
    for s in sessions_with_fac:
        pn_obj = pn_by_id.get(s.parade_night_id)
        if pn_obj and body.start_date <= pn_obj.date <= body.end_date:
            affected.append({
                "session_id": s.id,
                "parade_date": pn_obj.date,
                "session_number": s.period_number,
                "cadet_group": s.cadet_group,
                "title": s.curriculum_title_at_time or s.custom_title,
            })

    db.commit()
    audit(db, p, object_type="facilitator_leave", object_id=leave.id, action="create",
          new={"facilitator_id": fac_id, "start": body.start_date, "end": body.end_date})

    return {
        "ok": True,
        "leave": {
            "id": leave.id, "facilitator_id": leave.facilitator_id,
            "planning_year_id": leave.planning_year_id,
            "start_date": leave.start_date, "end_date": leave.end_date,
            "reason": leave.reason, "notes": leave.notes,
            "created_by": leave.created_by,
            "created_at": iso_z(leave.created_at) if leave.created_at else None,
        },
        "affected_sessions": affected,
    }


@router.delete("/facilitator-leave/{leave_id}")
def delete_facilitator_leave(
    leave_id: str,
    db: DBSession = Depends(get_db),
    p: Principal = Depends(get_principal),
):
    """Soft-delete a facilitator leave record."""
    leave = db.get(PlanningFacilitatorLeave, leave_id)
    if not leave or leave.is_archived:
        raise HTTPException(404, detail={"error": "not_found"})
    fac = db.get(Facilitator, leave.facilitator_id)
    if not fac:
        raise HTTPException(404, detail={"error": "facilitator_not_found"})
    require_can_write_squadron(p, fac.squadron_id, fac.wing_id)
    leave.is_archived = True
    db.commit()
    audit(db, p, object_type="facilitator_leave", object_id=leave_id, action="archive")
    return {"ok": True}


@router.post("/facilitator-leave/{leave_id}/restore")
def restore_facilitator_leave(
    leave_id: str,
    db: DBSession = Depends(get_db),
    p: Principal = Depends(get_principal),
):
    """REM-133: facilitator-leave removal existed with no restore counterpart
    -- same pattern as restore_fac (training.py), the direct template for
    this gap."""
    leave = db.get(PlanningFacilitatorLeave, leave_id)
    if not leave:
        raise HTTPException(404, detail={"error": "not_found"})
    fac = db.get(Facilitator, leave.facilitator_id)
    if not fac:
        raise HTTPException(404, detail={"error": "facilitator_not_found"})
    require_can_write_squadron(p, fac.squadron_id, fac.wing_id)
    if not leave.is_archived:
        raise HTTPException(409, detail={"error": "not_archived"})
    leave.is_archived = False
    db.commit()
    audit(db, p, object_type="facilitator_leave", object_id=leave_id, action="restore")
    return {"ok": True}


@router.get("/years/{year_id}/facilitators/{fac_id}/workload")
def facilitator_workload(
    year_id: str,
    fac_id: str,
    db: DBSession = Depends(get_db),
    p: Principal = Depends(get_principal),
):
    """Return workload stats for a facilitator within a planning year."""
    require_role(p, "sqn_admin", "wing_admin", "national_admin", "system_admin")
    py = _get_year_or_404(year_id, db)
    _require_year_access(p, py)

    fac = db.get(Facilitator, fac_id)
    if not fac:
        raise HTTPException(404, detail={"error": "facilitator_not_found"})

    # Phase B: query ParadeNight directly — no more ParadeDate intermediary
    pn_rows = db.query(ParadeNight).filter(
        ParadeNight.planning_year_id == year_id,
        ParadeNight.is_active == True,  # noqa: E712
    ).all()
    pn_by_id: dict[str, ParadeNight] = {pn_obj.id: pn_obj for pn_obj in pn_rows}

    sessions: list[TrainingSession] = []
    if pn_by_id:
        sessions = db.query(TrainingSession).filter(
            TrainingSession.facilitator_id == fac_id,
            TrainingSession.parade_night_id.in_(list(pn_by_id.keys())),
            TrainingSession.is_archived == False,  # noqa: E712
        ).all()

    total_scheduled = len(sessions)
    # Group by parade night
    by_night: dict[str, list] = {}
    for s in sessions:
        by_night.setdefault(s.parade_night_id, []).append(s)

    nights_with_sessions = len(by_night)
    counts_per_night = [len(v) for v in by_night.values()]
    avg_per_night = round(sum(counts_per_night) / len(counts_per_night), 1) if counts_per_night else 0.0
    max_per_night = max(counts_per_night) if counts_per_night else 0
    min_per_night = min(counts_per_night) if counts_per_night else 0

    today = date.today().isoformat()
    upcoming: list[dict] = []
    for s in sessions:
        pn_obj = pn_by_id.get(s.parade_night_id)
        pd_date = pn_obj.date if pn_obj else ""
        if pd_date >= today:
            title = s.curriculum_title_at_time or s.custom_title
            upcoming.append({
                "session_id": s.id,
                "parade_date": pd_date,
                "session_number": s.period_number,
                "cadet_group": s.cadet_group,
                "title": title,
                "location_name": s.training_area_name_at_time,
            })
    upcoming.sort(key=lambda x: (x["parade_date"], x["session_number"] or 0))

    return {
        "total_scheduled": total_scheduled,
        "nights_with_sessions": nights_with_sessions,
        "avg_per_night": avg_per_night,
        "max_per_night": max_per_night,
        "min_per_night": min_per_night,
        "upcoming_sessions": upcoming,
    }


# ── Night Summaries ──────────────────────────────────────────────────────────

def _notice_out(n: PlanningNotice) -> dict:
    return {
        "notice_id": n.id,
        "parade_night_id": n.parade_night_id,
        "notice_text": n.notice_text,
        "audience": n.audience,
        "priority": n.priority,
        "created_by": n.created_by,
        "is_archived": n.is_archived,
        "created_at": iso_z(n.created_at) if n.created_at else None,
        "updated_at": iso_z(n.updated_at) if n.updated_at else None,
    }


@router.get("/years/{year_id}/night-summaries")
def night_summaries(
    year_id: str,
    db: DBSession = Depends(get_db),
    p: Principal = Depends(get_principal),
):
    """Return all parade nights with session/conflict/notice summary for the planning grid."""
    require_role(p, "sqn_admin", "sqn_general", "wing_admin", "national_admin", "system_admin")
    py = _get_year_or_404(year_id, db)
    _require_year_access(p, py)

    # Phase B: query ParadeNight directly; pn.id == former parade_date_id alias
    all_dates = (
        db.query(ParadeNight)
        .filter(
            ParadeNight.planning_year_id == year_id,
            ParadeNight.is_active == True,  # noqa: E712
        )
        .order_by(ParadeNight.date)
        .all()
    )

    # pn.id is now the authoritative identifier (was parade_date_id in the old dual-table model)
    night_ids = [pn.id for pn in all_dates]

    # Batch-load sessions — parade_night_id IS the pn.id directly
    all_ts: list[TrainingSession] = []
    if night_ids:
        all_ts = db.query(TrainingSession).filter(
            TrainingSession.parade_night_id.in_(night_ids),
            TrainingSession.is_archived == False,  # noqa: E712
        ).all()

    ts_by_night: dict[str, list[TrainingSession]] = {}
    for s in all_ts:
        ts_by_night.setdefault(s.parade_night_id, []).append(s)

    # Batch-load facilitator display names
    fac_ids = list({s.facilitator_id for s in all_ts if s.facilitator_id})
    fac_map: dict[str, str] = {}
    if fac_ids:
        for f in db.query(Facilitator).filter(Facilitator.id.in_(fac_ids)).all():
            fac_map[f.id] = f"{f.current_rank or ''} {f.last_name}".strip()

    # Batch-load location names
    loc_ids = list({s.training_area_id for s in all_ts if s.training_area_id})
    loc_map: dict[str, str] = {}
    if loc_ids:
        for loc in db.query(TrainingArea).filter(TrainingArea.id.in_(loc_ids)).all():
            loc_map[loc.id] = loc.name

    # Batch-load unresolved conflict counts
    conflict_counts: dict[str, int] = {}
    if night_ids:
        for c in db.query(PlanningConflict).filter(
            PlanningConflict.parade_night_id.in_(night_ids),
            PlanningConflict.is_resolved == False,  # noqa: E712
        ).all():
            conflict_counts[c.parade_night_id] = conflict_counts.get(c.parade_night_id, 0) + 1

    # Batch-load notices
    notices_by_night: dict[str, list[PlanningNotice]] = {}
    if night_ids:
        for n in db.query(PlanningNotice).filter(
            PlanningNotice.parade_night_id.in_(night_ids),
            PlanningNotice.is_archived == False,  # noqa: E712
        ).all():
            notices_by_night.setdefault(n.parade_night_id, []).append(n)

    # Batch-load timing snapshots for all nights
    snapshots_by_night: dict[str, list] = {}
    if night_ids:
        all_snaps = db.query(ParadeNightTimingSnapshot).filter(
            ParadeNightTimingSnapshot.parade_night_id.in_(night_ids)
        ).order_by(
            ParadeNightTimingSnapshot.parade_night_id,
            ParadeNightTimingSnapshot.display_order,
        ).all()
        for snap in all_snaps:
            snapshots_by_night.setdefault(snap.parade_night_id, []).append(snap)

    summaries = []
    for pn in all_dates:
        pn_sessions = ts_by_night.get(pn.id, [])
        session_summaries = []
        for s in pn_sessions:
            session_summaries.append({
                "session_id": s.id,
                "period": s.period_number,
                "cadet_group": s.cadet_group,
                "title": s.curriculum_title_at_time or s.custom_title,
                "curriculum_code": s.curriculum_code_at_time,
                "facilitator": fac_map.get(s.facilitator_id) if s.facilitator_id else None,
                "location": (s.training_area_name_at_time or loc_map.get(s.training_area_id))
                             if s.training_area_id else None,
            })

        # Build instructional_periods and timing_strip from snapshots
        snaps = snapshots_by_night.get(pn.id, [])
        instructional_periods = [
            {
                "period_number": s.period_number,
                "label": s.block_label,
                "start_time": s.start_time,
                "end_time": s.end_time,
            }
            for s in snaps if s.is_instructional and s.period_number is not None
        ]
        timing_strip = [
            {
                "label": s.block_label,
                "start_time": s.start_time,
                "end_time": s.end_time,
                "is_instructional": s.is_instructional,
                "display_order": s.display_order,
            }
            for s in snaps
        ]
        # Fallback for legacy nights without snapshots
        if not instructional_periods and pn.session_count:
            instructional_periods = [
                {"period_number": i, "label": f"Period {i}", "start_time": None, "end_time": None}
                for i in range(1, pn.session_count + 1)
            ]

        summaries.append({
            "parade_date_id": pn.id,  # backward-compat alias
            "parade_night_id": pn.id,
            "parade_date": pn.date,
            "parade_type": pn.parade_type,
            "term": pn.term,
            "week_number": pn.week_number,
            "notes": pn.notes,
            "parade_night_notes": pn.notes,
            "timing_template_id": pn.timing_template_id,
            "sessions": session_summaries,
            "conflict_count": conflict_counts.get(pn.id, 0),
            "notices": [_notice_out(n) for n in notices_by_night.get(pn.id, [])],
            "instructional_periods": instructional_periods,
            "timing_strip": timing_strip,
        })

    return {"planning_year_id": year_id, "summaries": summaries}


# ── Parade Date Notices ──────────────────────────────────────────────────────

class NoticeIn(BaseModel):
    notice_text: str
    priority: str = Field(default="normal", max_length=20)
    audience: str | None = Field(default=None, max_length=60)


@router.get("/parade-nights/{night_id}/notices")
@router.get("/parade-dates/{night_id}/notices")  # backward-compat alias (date_id == pn.id after Phase B)
def list_notices(
    night_id: str,
    db: DBSession = Depends(get_db),
    p: Principal = Depends(get_principal),
):
    require_role(p, "sqn_admin", "wing_admin", "national_admin", "system_admin")
    pn = db.get(ParadeNight, night_id)
    if not pn:
        raise HTTPException(404, detail={"error": "parade_night_not_found"})
    _require_year_access(p, _get_year_or_404(pn.planning_year_id, db), write=False)
    notices = (
        db.query(PlanningNotice)
        .filter(
            PlanningNotice.parade_night_id == night_id,
            PlanningNotice.is_archived == False,  # noqa: E712
        )
        .order_by(PlanningNotice.created_at)
        .all()
    )
    return [_notice_out(n) for n in notices]


@router.post("/parade-nights/{night_id}/notices")
@router.post("/parade-dates/{night_id}/notices")  # backward-compat alias (date_id == pn.id after Phase B)
def create_notice(
    night_id: str,
    body: NoticeIn,
    db: DBSession = Depends(get_db),
    p: Principal = Depends(get_principal),
):
    pn = db.get(ParadeNight, night_id)
    if not pn:
        raise HTTPException(404, detail={"error": "parade_night_not_found"})
    _require_year_access(p, _get_year_or_404(pn.planning_year_id, db), write=True)
    require_can_write_squadron(p, pn.squadron_id, pn.wing_id)
    notice = PlanningNotice(
        parade_night_id=night_id,
        notice_text=body.notice_text.strip(),
        priority=body.priority,
        audience=body.audience,
        created_by=p.user_id,
    )
    db.add(notice)
    db.commit()
    db.refresh(notice)
    audit(db, p, object_type="PlanningNotice", object_id=notice.id, action="create")
    return {"ok": True, "notice_id": notice.id}


class NoticeUpdateIn(BaseModel):
    notice_text: str | None = None
    priority: str | None = Field(default=None, max_length=20)
    audience: str | None = Field(default=None, max_length=60)
    version: int | None = None


@router.patch("/notices/{notice_id}")
def update_notice(
    notice_id: str,
    body: NoticeUpdateIn,
    db: DBSession = Depends(get_db),
    p: Principal = Depends(get_principal),
):
    notice = db.get(PlanningNotice, notice_id)
    if not notice or notice.is_archived:
        raise HTTPException(404, detail={"error": "notice_not_found"})
    pn = db.get(ParadeNight, notice.parade_night_id)
    if not pn:
        raise HTTPException(404, detail={"error": "parade_night_not_found"})
    _require_year_access(p, _get_year_or_404(pn.planning_year_id, db), write=True)
    require_can_write_squadron(p, pn.squadron_id, pn.wing_id)
    _check_version(notice, body.version)
    if body.notice_text is not None:
        notice.notice_text = body.notice_text.strip()
    if body.priority is not None:
        notice.priority = body.priority
    if body.audience is not None:
        notice.audience = body.audience
    notice.version += 1
    db.commit()
    audit(db, p, object_type="PlanningNotice", object_id=notice.id, action="update")
    return {"ok": True}


@router.post("/notices/{notice_id}/archive")
def archive_notice(
    notice_id: str,
    db: DBSession = Depends(get_db),
    p: Principal = Depends(get_principal),
):
    notice = db.get(PlanningNotice, notice_id)
    if not notice:
        raise HTTPException(404, detail={"error": "notice_not_found"})
    pn = db.get(ParadeNight, notice.parade_night_id)
    if not pn:
        raise HTTPException(404, detail={"error": "parade_night_not_found"})
    _require_year_access(p, _get_year_or_404(pn.planning_year_id, db), write=True)
    require_can_write_squadron(p, pn.squadron_id, pn.wing_id)
    notice.is_archived = True
    db.commit()
    audit(db, p, object_type="PlanningNotice", object_id=notice_id, action="archive")
    return {"ok": True}


# ── CEA Activities ────────────────────────────────────────────────────────────

def _cea_activity_out(a: CeaActivity) -> dict:
    return {
        "id": a.id,
        "import_batch_id": a.import_batch_id,
        "planning_year_id": a.planning_year_id,
        "wing_id": a.wing_id,
        "unit_id": a.unit_id,
        "cea_activity_id": a.cea_activity_id,
        "activity_type": a.activity_type,
        "status_name": a.status_name,
        "parent_unit": a.parent_unit,
        "host_unit": a.host_unit,
        "activity_name": a.activity_name,
        "nomination_start_date": a.nomination_start_date,
        "nomination_end_date": a.nomination_end_date,
        "activity_start_date": a.activity_start_date,
        "activity_end_date": a.activity_end_date,
        "start_time": a.start_time,
        "end_time": a.end_time,
        "location": a.location,
        "activity_poc": a.activity_poc,
        "notes": a.notes,
        "source_type": a.source_type,
        "classification_status": a.classification_status,
        "importance": a.importance,
        "audience_staff_only": a.audience_staff_only,
        "audience_seniors": a.audience_seniors,
        "audience_proficient": a.audience_proficient,
        "audience_first_years": a.audience_first_years,
        "classified_by": a.classified_by,
        "classified_at": a.classified_at,
        "is_removed_from_cea": a.is_removed_from_cea,
        "is_archived": a.is_archived,
        "created_at": iso_z(a.created_at) if a.created_at else None,
    }


@router.get("/years/{year_id}/cea/activities")
def list_cea_activities(
    year_id: str,
    status: str | None = Query(None, description="needs_review | classified | irrelevant"),
    db: DBSession = Depends(get_db),
    p: Principal = Depends(get_principal),
):
    require_role(p, "sqn_admin", "wing_admin", "national_admin", "system_admin")
    _require_year_access(p, _get_year_or_404(year_id, db), write=False)
    from sqlalchemy import select
    stmt = select(CeaActivity).where(
        CeaActivity.planning_year_id == year_id,
        CeaActivity.is_archived.is_(False),
    )
    if status:
        stmt = stmt.where(CeaActivity.classification_status == status)
    rows = db.scalars(stmt.order_by(CeaActivity.activity_start_date)).all()
    # TRGO-02: surface the caller's own squadron's local-hide/note state so the
    # unified Activities view can show "hidden for you" and let a squadron
    # toggle it, without touching the shared CeaActivity row (the point of
    # local-hide being a per-squadron overlay, not a source-record edit).
    hides_by_activity: dict[str, ActivityLocalHide] = {}
    if p.squadron_id and rows:
        activity_ids = [r.id for r in rows]
        hide_rows = db.query(ActivityLocalHide).filter(
            ActivityLocalHide.cea_activity_id.in_(activity_ids),
            ActivityLocalHide.unit_id == p.squadron_id,
        ).all()
        hides_by_activity = {h.cea_activity_id: h for h in hide_rows}
    out = []
    for r in rows:
        d = _cea_activity_out(r)
        hide = hides_by_activity.get(r.id)
        d["is_hidden_for_me"] = bool(hide and hide.is_hidden)
        d["local_note"] = hide.local_note if hide else None
        out.append(d)
    return {"activities": out}


@router.get("/years/{year_id}/cea/batches")
def list_cea_batches(
    year_id: str,
    db: DBSession = Depends(get_db),
    p: Principal = Depends(get_principal),
):
    require_role(p, "sqn_admin", "wing_admin", "national_admin", "system_admin")
    _require_year_access(p, _get_year_or_404(year_id, db), write=False)
    from sqlalchemy import select
    rows = db.scalars(
        select(CeaImportBatch)
        .where(CeaImportBatch.planning_year_id == year_id)
        .order_by(CeaImportBatch.created_at.desc())
        .limit(20)
    ).all()
    return {"batches": [
        {
            "id": b.id,
            "imported_by": b.imported_by,
            "source_file_name": b.source_file_name,
            "row_count": b.row_count,
            "created_count": b.created_count,
            "updated_count": b.updated_count,
            "duplicate_count": b.duplicate_count,
            "skipped_count": b.skipped_count,
            "error_count": b.error_count,
            "created_at": iso_z(b.created_at) if b.created_at else None,
        }
        for b in rows
    ]}


def _parse_cea_date(raw: str | None) -> str | None:
    if not raw:
        return None
    raw = raw.strip()
    for fmt in ("%d/%m/%Y", "%Y-%m-%d", "%m/%d/%Y", "%d-%m-%Y"):
        try:
            from datetime import datetime as _dt
            return _dt.strptime(raw, fmt).strftime("%Y-%m-%d")
        except ValueError:
            continue
    return raw[:10] if len(raw) >= 10 else raw


def _parse_cea_time(raw: str | None) -> str | None:
    if not raw:
        return None
    raw = raw.strip()
    if len(raw) >= 5 and ":" in raw:
        return raw[:5]
    return None


def _normalise_cea_row(row: dict) -> dict:
    """Map both full-export and simple-export CSV column names to internal fields."""
    def g(*keys: str) -> str | None:
        for k in keys:
            v = row.get(k, "").strip()
            if v:
                return v
        return None

    return {
        "cea_activity_id": g("ActivityID", "SeqNr"),
        "activity_type": g("ActivityTypeName", "SupportRequestType"),
        "status_name": g("StatusName"),
        "parent_unit": g("UnitParentName"),
        "host_unit": g("UnitName", "Unit"),
        "activity_name": g("SeqNrAndName", "Name", "ActivityName"),
        "nomination_start_date": _parse_cea_date(g("NominationStartDate")),
        "nomination_end_date": _parse_cea_date(g("NominationEndDate")),
        "activity_start_date": _parse_cea_date(g("StartDate", "Start date")),
        "activity_end_date": _parse_cea_date(g("EndDate", "End date")),
        "start_time": _parse_cea_time(g("StartTime", "Start time")),
        "end_time": _parse_cea_time(g("EndTime", "End time")),
        "location": g("Location"),
        "activity_poc": g("POC"),
        "notes": g("ActivityNotes", "Activity Notes", "SupportRequestTypeCode"),
    }


@router.post("/years/{year_id}/cea/import")
async def import_cea_csv(
    year_id: str,
    file: UploadFile = File(...),
    keep_existing: str = Form(default=""),
    db: DBSession = Depends(get_db),
    p: Principal = Depends(get_principal),
):
    require_role(p, "wing_admin", "national_admin", "system_admin")
    year = _get_year_or_404(year_id, db)
    _require_year_access(p, year, write=True)

    from ..config import settings as _settings
    content = await file.read()
    if len(content) > _settings.UPLOAD_MAX_MB * 1024 * 1024:
        raise HTTPException(413, detail={"error": "file_too_large"})
    try:
        text = content.decode("utf-8-sig")
    except UnicodeDecodeError:
        text = content.decode("latin-1")

    reader = _csv.DictReader(io.StringIO(text))
    rows = list(reader)

    # Build set of cea_activity_ids the caller wants to preserve unchanged
    _keep_ids = {s.strip() for s in keep_existing.split(",") if s.strip()}

    batch = CeaImportBatch(
        id=str(uuid.uuid4()),
        planning_year_id=year_id,
        wing_id=p.wing_id,
        imported_by=p.user_id,
        source_file_name=file.filename,
        row_count=len(rows),
        created_at=utcnow(),
        updated_at=utcnow(),
    )
    db.add(batch)
    db.flush()

    from sqlalchemy import select

    created = updated = duplicates = skipped = errors = 0
    previews = []

    # Load existing activities for this year to detect duplicates/updates
    existing_by_cea_id: dict[str, CeaActivity] = {}
    existing_by_name_date: dict[tuple, list[CeaActivity]] = {}
    for act in db.scalars(
        select(CeaActivity).where(
            CeaActivity.planning_year_id == year_id,
            CeaActivity.is_archived.is_(False),
        )
    ).all():
        if act.cea_activity_id:
            existing_by_cea_id[act.cea_activity_id] = act
        key = (act.activity_name or "", act.activity_start_date or "")
        existing_by_name_date.setdefault(key, []).append(act)

    for row in rows:
        try:
            parsed = _normalise_cea_row(row)
            name = parsed.get("activity_name") or ""
            if not name:
                skipped += 1
                continue

            cea_id = parsed.get("cea_activity_id")
            start_date = parsed.get("activity_start_date") or ""
            key = (name, start_date)

            action = "create"
            existing: CeaActivity | None = None

            if cea_id and cea_id in existing_by_cea_id:
                existing = existing_by_cea_id[cea_id]
                action = "update"
            elif key in existing_by_name_date:
                existing = existing_by_name_date[key][0]
                action = "duplicate"

            previews.append({
                "action": action,
                "cea_activity_id": cea_id,
                "activity_name": name,
                "activity_start_date": start_date,
                "activity_end_date": parsed.get("activity_end_date"),
                "host_unit": parsed.get("host_unit"),
                "parent_unit": parsed.get("parent_unit"),
                "location": parsed.get("location"),
                "existing_id": existing.id if existing else None,
            })

            if action == "create":
                act = CeaActivity(
                    id=str(uuid.uuid4()),
                    import_batch_id=batch.id,
                    planning_year_id=year_id,
                    wing_id=p.wing_id,
                    created_by=p.user_id,
                    created_at=utcnow(),
                    updated_at=utcnow(),
                    **{k: v for k, v in parsed.items()},
                )
                db.add(act)
                created += 1
            elif action == "update" and existing:
                if cea_id and cea_id in _keep_ids:
                    skipped += 1
                    continue
                for k, v in parsed.items():
                    if v is not None:
                        setattr(existing, k, v)
                existing.updated_by = p.user_id
                existing.updated_at = utcnow()
                updated += 1
            else:
                duplicates += 1

        except Exception:
            errors += 1

    # Mark activities no longer present in the import
    current_cea_ids = {
        r.get("cea_activity_id") or r.get("ActivityID") or r.get("SeqNr", "")
        for r in rows
    }
    for act in existing_by_cea_id.values():
        if act.cea_activity_id not in current_cea_ids:
            act.is_removed_from_cea = True
            act.updated_at = utcnow()

    batch.created_count = created
    batch.updated_count = updated
    batch.duplicate_count = duplicates
    batch.skipped_count = skipped
    batch.error_count = errors

    db.commit()
    audit(db, p, object_type="CeaImportBatch", object_id=batch.id, action="import_cea")

    return {
        "ok": True,
        "batch_id": batch.id,
        "row_count": len(rows),
        "created": created,
        "updated": updated,
        "duplicates": duplicates,
        "skipped": skipped,
        "errors": errors,
        "preview": previews[:50],
    }


class CeaClassifyIn(BaseModel):
    importance: str | None = None
    audience_staff_only: bool = False
    audience_seniors: bool = False
    audience_proficient: bool = False
    audience_first_years: bool = False


@router.patch("/cea/{activity_id}/classify")
def classify_cea_activity(
    activity_id: str,
    body: CeaClassifyIn,
    db: DBSession = Depends(get_db),
    p: Principal = Depends(get_principal),
):
    require_role(p, "sqn_admin", "wing_admin", "national_admin", "system_admin")
    act = db.get(CeaActivity, activity_id)
    if not act or act.is_archived:
        raise HTTPException(404, detail={"error": "activity_not_found"})
    _require_year_access(p, _get_year_or_404(act.planning_year_id, db), write=True)
    # DEFECT-005: classify mutates the shared CeaActivity row directly
    # (importance/audience/classification_status), unlike local-hide below
    # which only ever writes a squadron-local overlay row. unit_id is set
    # only for manually-created, squadron-owned activities (see
    # create_manual_activity) -- never by the CSV import path, which leaves
    # it null for a wing-wide/shared activity. A sqn_admin may classify
    # their own squadron's manual activity, but must never be able to
    # overwrite a wing-wide CEA activity's classification for every
    # squadron in the wing at once.
    if act.unit_id != p.squadron_id and p.role == "sqn_admin":
        raise HTTPException(403, detail={
            "error": "wing_wide_activity",
            "message": "Only a Wing Admin or higher can classify a wing-wide CEA activity.",
        })
    act.importance = body.importance
    act.audience_staff_only = body.audience_staff_only
    act.audience_seniors = body.audience_seniors
    act.audience_proficient = body.audience_proficient
    act.audience_first_years = body.audience_first_years
    act.classification_status = "classified" if body.importance else "needs_review"
    act.classified_by = p.user_id
    from datetime import datetime, timezone
    act.classified_at = datetime.now(timezone.utc).isoformat()
    act.updated_at = utcnow()
    db.commit()
    audit(db, p, object_type="CeaActivity", object_id=activity_id, action="classify")
    return {"ok": True}


class ActivityLocalHideIn(BaseModel):
    is_hidden: bool = True
    local_note: str | None = None


@router.post("/cea/{activity_id}/local-hide")
def set_local_hide(
    activity_id: str,
    body: ActivityLocalHideIn,
    db: DBSession = Depends(get_db),
    p: Principal = Depends(get_principal),
):
    require_role(p, "sqn_admin", "wing_admin", "national_admin", "system_admin")
    if not p.squadron_id:
        raise HTTPException(400, detail={"error": "no_unit_scope"})
    act = db.get(CeaActivity, activity_id)
    if not act or act.is_archived:
        raise HTTPException(404, detail={"error": "activity_not_found"})
    py = _get_year_or_404(act.planning_year_id, db)
    _require_year_access(p, py, write=True, db=db)
    from sqlalchemy import select
    existing = db.scalar(
        select(ActivityLocalHide).where(
            ActivityLocalHide.cea_activity_id == activity_id,
            ActivityLocalHide.unit_id == p.squadron_id,
        )
    )
    if existing:
        existing.is_hidden = body.is_hidden
        existing.local_note = body.local_note
        existing.updated_at = utcnow()
    else:
        db.add(ActivityLocalHide(
            id=str(uuid.uuid4()),
            cea_activity_id=activity_id,
            unit_id=p.squadron_id,
            is_hidden=body.is_hidden,
            local_note=body.local_note,
            hidden_by=p.user_id,
            created_at=utcnow(),
            updated_at=utcnow(),
        ))
    db.commit()
    return {"ok": True}


class ManualActivityIn(BaseModel):
    activity_name: str
    activity_start_date: str | None = None
    activity_end_date: str | None = None
    start_time: str | None = None
    end_time: str | None = None
    location: str | None = None
    notes: str | None = None
    importance: str | None = None
    audience_staff_only: bool = False
    audience_seniors: bool = False
    audience_proficient: bool = False
    audience_first_years: bool = False


@router.post("/years/{year_id}/cea/activities")
def create_manual_activity(
    year_id: str,
    body: ManualActivityIn,
    db: DBSession = Depends(get_db),
    p: Principal = Depends(get_principal),
):
    require_role(p, "sqn_admin", "wing_admin", "national_admin", "system_admin")
    _require_year_access(p, _get_year_or_404(year_id, db), write=True)
    act = CeaActivity(
        id=str(uuid.uuid4()),
        planning_year_id=year_id,
        wing_id=p.wing_id,
        unit_id=p.squadron_id,
        activity_name=body.activity_name.strip(),
        activity_start_date=body.activity_start_date,
        activity_end_date=body.activity_end_date,
        start_time=body.start_time,
        end_time=body.end_time,
        location=body.location,
        notes=body.notes,
        source_type="manual",
        importance=body.importance,
        classification_status="classified" if body.importance else "needs_review",
        audience_staff_only=body.audience_staff_only,
        audience_seniors=body.audience_seniors,
        audience_proficient=body.audience_proficient,
        audience_first_years=body.audience_first_years,
        created_by=p.user_id,
        created_at=utcnow(),
        updated_at=utcnow(),
    )
    db.add(act)
    db.commit()
    db.refresh(act)
    audit(db, p, object_type="CeaActivity", object_id=act.id, action="create_manual")
    return {"ok": True, "id": act.id, "activity": _cea_activity_out(act)}
