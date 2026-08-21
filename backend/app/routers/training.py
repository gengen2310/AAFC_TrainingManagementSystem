from __future__ import annotations

import json as _json

from fastapi import APIRouter, Depends, Request, HTTPException, UploadFile, File, Header
from pydantic import BaseModel, Field, field_validator
from typing import List, Literal, Optional
from sqlalchemy.orm import Session as DBSession

from ..database import get_db, utcnow
from ..models import (CurriculumItem, CurriculumElement, CurriculumPhase, ParadeNight, Session, SessionStatusHistory,
                      Facilitator, FacilitatorRankHistory, SubjectAreaTag, FacilitatorTypeTag, SessionStatusReasonTag, TrainingArea, Equipment,
                      Activity, Cadet, Squadron, TimingTemplate, TimingBlock, SystemSetting, TrainingClass, PlanningYear,
                      SessionAudience, CadetClassMembership,
                      ParadeNightTemplate, ParadeNightTemplateSession)
from ..models.planning import ActivityLocalOverride
from ..models.training import ELEMENT_SCOPE_LEVELS, PHASE_SCOPE_LEVELS, STAGE_CODES
from .timing import _effective_template
from ..dependencies import get_principal, client_meta
from ..permissions import (Principal, require_can_view_squadron, require_can_write_squadron,
                          require_can_view_wing, require_can_write_activity, require_role, require_system_admin,
                          NATIONAL_LEVEL)
from ..services import (audit, score_parade, publish_blockers, close_blockers)
from ..services_readiness import parade_night_readiness

router = APIRouter(prefix="/api", tags=["training"])


def _check_version(obj, client_version: int | None) -> None:
    """Raise 409 if the client's version is stale (optimistic locking)."""
    if client_version is not None and obj.version != client_version:
        raise HTTPException(409, detail={
            "error": "version_conflict",
            "current_version": obj.version,
        })


VALID_STATUS = {"draft", "planned", "published", "delivered", "delivered_with_issue",
                "cancelled", "cancelled_late", "rescheduled", "not_delivered",
                "requires_review", "blocked", "closed"}


def _check_http_url(v: str | None) -> str | None:
    """Raise ValueError if v is a non-http/https URL (blocks javascript: URIs)."""
    if v is None:
        return v
    from urllib.parse import urlsplit
    scheme = urlsplit(str(v)).scheme.lower()
    if scheme not in ("http", "https"):
        raise ValueError("URL must use http or https scheme")
    return v


def _parse_json_list(val) -> list:
    """Return val as a list, JSON-parsing it if stored as a TEXT string (Postgres TEXT vs JSON/JSONB mismatch)."""
    if not val:
        return []
    if isinstance(val, list):
        return val
    if isinstance(val, str):
        try:
            result = _json.loads(val)
            return result if isinstance(result, list) else []
        except (_json.JSONDecodeError, TypeError):
            return []
    return []


def _active_squadron(p: Principal) -> str:
    """The squadron a write should target: proxy/intervention target, else home."""
    if p.acting_squadron_id:
        return p.acting_squadron_id
    return p.squadron_id


def _view_squadron_id(p: Principal, squadron_id: str | None, db: DBSession) -> str | None:
    """Resolve which squadron's data a READ should return. An explicit squadron_id
    (validated via require_can_view_squadron, the same check dashboard charts and
    /api/parade-nights already use) lets a wing/national viewer see a specific
    squadron's operational pages WITHOUT needing to enter Proxy/Delegated
    Intervention Mode — viewing is broad by design (permissions.py's
    Principal.can_view_squadron), only writing is proxy-gated (see
    require_can_write_squadron / Block 7's canWriteSquadron fix). Falls back to
    _active_squadron() when no squadron_id is given, so existing squadron-scoped
    callers are unaffected.

    Master transformation plan Block 8: this closes the inconsistency where
    /api/parade-nights and /api/dashboard/charts already supported this pattern
    but /api/facilitators, /api/training-areas, /api/equipment, and /api/activities
    did not — a wing/national viewer's squadron selector must behave the same way
    on every squadron page, not degrade differently depending which one they're on."""
    if squadron_id:
        s = db.get(Squadron, squadron_id)
        if not s:
            # A bogus squadron_id must 404, never silently fall back to the
            # caller's own scope — that would mask a broken link/typo as "no
            # results" instead of a clear error.
            raise HTTPException(404, detail={"error": "squadron_not_found"})
        require_can_view_squadron(p, s.id, s.wing_id)
        return s.id
    return _active_squadron(p)


def _sess_dict(s: Session) -> dict:
    d = {c.name: getattr(s, c.name) for c in s.__table__.columns}
    d["session_id"] = s.id  # alias for Night Builder compatibility
    return d


# ── CURRICULUM ──
@router.get("/curriculum")
def list_curriculum(squadron_id: str | None = None, include_archived: bool = False,
                    db: DBSession = Depends(get_db), p: Principal = Depends(get_principal)):
    sq_id = _view_squadron_id(p, squadron_id, db)
    # Resolve the wing for the acting scope
    wing_id: str | None = p.acting_wing_id or p.wing_id
    if sq_id:
        s = db.get(Squadron, sq_id)
        if s:
            wing_id = s.wing_id

    from sqlalchemy import or_
    conditions = [CurriculumItem.owning_level == "national"]
    if wing_id:
        conditions.append(
            (CurriculumItem.owning_level == "wing") & (CurriculumItem.wing_id == wing_id))
    elif p.role in _NAT_ADMIN_ROLES:
        # National admin with no proxy/wing scope sees all wing curriculum across all wings
        conditions.append(CurriculumItem.owning_level == "wing")
    if sq_id:
        conditions.append(CurriculumItem.squadron_id == sq_id)

    query = db.query(CurriculumItem).filter(or_(*conditions))
    if not include_archived:
        query = query.filter(CurriculumItem.is_archived == False)  # noqa: E712
    items = query.order_by(CurriculumItem.recommended_sequence).all()

    # Preload all sessions for this squadron in one query, then group by curriculum_item_id.
    item_ids = [i.id for i in items]
    all_sess = db.query(Session).filter(
        Session.curriculum_item_id.in_(item_ids),
        Session.squadron_id == sq_id,
        Session.is_archived == False,  # noqa: E712
    ).all() if item_ids else []
    from collections import defaultdict
    sess_by_item: dict[str, list] = defaultdict(list)
    for s in all_sess:
        sess_by_item[s.curriculum_item_id].append(s)

    out = []
    for i in items:
        statuses = [s.status for s in sess_by_item[i.id]]
        out.append({"curriculum_id": i.id, "code": i.code, "identifier": i.identifier,
                    "part_number": i.part_number, "title": i.title, "phase": i.phase,
                    "element": i.element, "duration_minutes": i.duration_minutes,
                    "part_count": i.part_count, "instructor_suitability": i.instructor_suitability,
                    "core_status": i.core_status, "learning_hub_url": i.learning_hub_url,
                    "recommended_term": i.recommended_term, "owning_level": i.owning_level,
                    "wing_id": i.wing_id, "squadron_id": i.squadron_id,
                    "session_count": len(statuses), "progress": _progress(statuses),
                    "is_archived": i.is_archived})
    return {"items": out}


def _progress(statuses: list[str]) -> str:
    if not statuses:
        return "unscheduled"
    if "delivered" in statuses or "delivered_with_issue" in statuses:
        return "delivered"
    if all(s in ("cancelled", "cancelled_late") for s in statuses):
        return "cancelled"
    if "rescheduled" in statuses:
        return "rescheduled"
    if "not_delivered" in statuses:
        return "not_delivered"
    return "planned"


def _lh_last_segment(url: str) -> str:
    """The final non-empty path segment of a URL, ignoring query string and
    trailing slash -- used only to compare against a curriculum item's
    identifier, never to construct/guess a URL."""
    from urllib.parse import urlsplit
    path = urlsplit(url).path.rstrip("/")
    return path.rsplit("/", 1)[-1] if path else ""


@router.get("/curriculum/learning-hub/validation")
def learning_hub_validation_report(db: DBSession = Depends(get_db), p: Principal = Depends(get_principal)):
    """REM-19 (original_instruction.md Section 20): a read-only report flagging
    curriculum items whose stored learning_hub_url doesn't match their own
    identifier, and items missing a link entirely -- surfaced for admin
    review via the existing single-item PATCH /curriculum/{cid} or the bulk
    CSV/XLSX re-import (import_curriculum). This never invents or corrects a
    URL itself -- point 4 of Section 20 explicitly prohibits guessing -- it
    only flags for a human to check against the real Learning Hub site.
    """
    require_role(p, "wing_admin", "national_admin", "system_admin")
    items = db.query(CurriculumItem).filter(CurriculumItem.is_archived == False).all()  # noqa: E712
    flagged = []
    for i in items:
        if not i.learning_hub_url:
            flagged.append({
                "curriculum_id": i.id, "code": i.code, "identifier": i.identifier,
                "title": i.title, "owning_level": i.owning_level,
                "learning_hub_url": None, "issue": "missing_link",
            })
            continue
        if not i.identifier:
            continue
        last_segment = _lh_last_segment(i.learning_hub_url)
        if last_segment.lower() != i.identifier.lower():
            flagged.append({
                "curriculum_id": i.id, "code": i.code, "identifier": i.identifier,
                "title": i.title, "owning_level": i.owning_level,
                "learning_hub_url": i.learning_hub_url, "issue": "identifier_mismatch",
            })
    return {"checked": len(items), "flagged_count": len(flagged), "flagged": flagged}


@router.get("/curriculum/export.xlsx")
def export_curriculum_xlsx(db: DBSession = Depends(get_db), p: Principal = Depends(get_principal)):
    """Export the visible curriculum catalogue as XLSX."""
    import io, openpyxl
    from fastapi.responses import StreamingResponse
    from openpyxl.styles import Font, PatternFill
    from ..services import audit as _audit

    sq_id = _active_squadron(p)
    wing_id: str | None = p.acting_wing_id or p.wing_id
    if sq_id:
        s = db.get(Squadron, sq_id)
        if s:
            wing_id = s.wing_id
    from sqlalchemy import or_
    conditions = [CurriculumItem.owning_level == "national"]
    if wing_id:
        conditions.append((CurriculumItem.owning_level == "wing") & (CurriculumItem.wing_id == wing_id))
    elif p.role in _NAT_ADMIN_ROLES:
        conditions.append(CurriculumItem.owning_level == "wing")
    if sq_id:
        conditions.append(CurriculumItem.squadron_id == sq_id)
    items = db.query(CurriculumItem).filter(
        CurriculumItem.is_archived == False,  # noqa: E712
        or_(*conditions)
    ).order_by(CurriculumItem.recommended_sequence).all()

    wb = openpyxl.Workbook()
    ws = wb.active; ws.title = "Curriculum"
    hdr_fill = PatternFill("solid", fgColor="002F65")
    hdr_font = Font(color="FFFFFF", bold=True)
    headers = ["Code", "Identifier", "Title", "Phase", "Element", "Duration (min)",
               "Core Status", "Instructor Suitability", "Recommended Term",
               "Owning Level", "Learning Hub URL"]
    ws.append(headers)
    for cell in ws[1]:
        cell.fill = hdr_fill; cell.font = hdr_font

    def _n(v):
        s = str(v) if v is not None else ""
        return ("'" + s) if s[:1] in ("=", "+", "-", "@") else s

    for i in items:
        ws.append([_n(i.code), _n(i.identifier), _n(i.title), _n(i.phase),
                   _n(i.element), i.duration_minutes or "",
                   _n(i.core_status or ""), _n(i.instructor_suitability or ""),
                   _n(i.recommended_term or ""), _n(i.owning_level),
                   _n(i.learning_hub_url or "")])

    bio = io.BytesIO(); wb.save(bio); bio.seek(0)
    _audit(db, p, object_type="export", object_id=None, action="export",
           new={"type": "curriculum", "fmt": "xlsx", "rows": len(items)})
    return StreamingResponse(bio,
                             media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                             headers={"Content-Disposition": "attachment; filename=curriculum.xlsx"})


@router.get("/curriculum/{cid}/sessions")
def curriculum_sessions(cid: str, db: DBSession = Depends(get_db), p: Principal = Depends(get_principal)):
    sq_id = _active_squadron(p)
    sess = db.query(Session).filter(Session.curriculum_item_id == cid,
                                    Session.squadron_id == sq_id).all()
    return [_sess_dict(s) for s in sess]


# ── PARADE NIGHTS ──
_VALID_PARADE_TYPE = Literal["normal", "activity", "ceremonial", "admin", "stand_down", "cancelled"]

class ParadeIn(BaseModel):
    date: str
    term: str | None = Field(default="T1", max_length=10)
    session_count: int | None = None  # None = use effective timing template or default
    parade_type: _VALID_PARADE_TYPE | None = "normal"

    @field_validator("date")
    @classmethod
    def _validate_date(cls, v: str) -> str:
        from datetime import date as _date
        try:
            _date.fromisoformat(v)
        except ValueError:
            raise ValueError(f"date must be a valid ISO-8601 date (YYYY-MM-DD), got '{v}'")
        return v


@router.get("/parade-nights")
def list_parades(squadron_id: str | None = None, db: DBSession = Depends(get_db),
                 p: Principal = Depends(get_principal)):
    sq_id = squadron_id or _active_squadron(p)
    # R5-L09: always resolve the squadron; raise 404 for a supplied but non-existent
    # sq_id rather than skipping the permission check and returning an empty list.
    s = db.get(Squadron, sq_id) if sq_id else None
    if sq_id and not s:
        raise HTTPException(404, detail={"error": "squadron_not_found"})
    if s:
        require_can_view_squadron(p, s.id, s.wing_id)
    pns = db.query(ParadeNight).filter(ParadeNight.squadron_id == sq_id,
                                       ParadeNight.is_archived == False).order_by(ParadeNight.date).all()  # noqa: E712

    from collections import defaultdict
    pn_ids = [pn.id for pn in pns]
    all_sess = db.query(Session).filter(
        Session.parade_night_id.in_(pn_ids), Session.is_archived == False,  # noqa: E712
    ).all() if pn_ids else []
    sess_by_pn: dict[str, list[Session]] = defaultdict(list)
    for x in all_sess:
        sess_by_pn[x.parade_night_id].append(x)

    # CLASS-06: attach each session's Training Class audience, additive to
    # _sess_dict()'s output. _sess_dict() itself is a raw column reflection
    # with no DB access of its own and 8 call sites across this file -- this
    # is scoped to just this endpoint (the one connected-frontend's real,
    # live Weekly Program page (renderWP(), populated from S.pns) actually
    # reads sessions from) rather than adding a db param to that shared
    # helper for every caller.
    classes_by_session: dict[str, list[dict]] = defaultdict(list)
    if all_sess:
        aud_rows = (
            db.query(SessionAudience, TrainingClass)
            .join(TrainingClass, SessionAudience.training_class_id == TrainingClass.id)
            .filter(SessionAudience.session_id.in_([x.id for x in all_sess]))
            .all()
        )
        for aud, tc in aud_rows:
            classes_by_session[aud.session_id].append(
                {"training_class_id": tc.id, "display_name": tc.display_name})

    out = []
    for pn in pns:
        sess_dicts = []
        for x in sess_by_pn.get(pn.id, []):
            d = _sess_dict(x)
            d["training_classes"] = classes_by_session.get(x.id, [])
            sess_dicts.append(d)
        out.append({**_pn_dict(pn), "sessions": sess_dicts})
    return out


@router.get("/parade-nights/{pnid}")
def get_parade(pnid: str, db: DBSession = Depends(get_db), p: Principal = Depends(get_principal)):
    pn = db.get(ParadeNight, pnid)
    if not pn or pn.is_archived:
        raise HTTPException(404, detail={"error": "not_found"})
    require_can_view_squadron(p, pn.squadron_id, pn.wing_id)
    sess = [_sess_dict(x) for x in db.query(Session).filter(Session.parade_night_id == pn.id,
                                                            Session.is_archived == False).all()]  # noqa: E712
    # Computed live from the authoritative services_readiness module (not the stored
    # ParadeNight.readiness_score/planning_status columns) so this response can never
    # drift from the list endpoint or the Dashboard even if _recompute() wasn't
    # triggered by some earlier edit path — same source of truth, always fresh.
    readiness = parade_night_readiness(sess)
    r = score_parade(sess)  # legacy shape (score/band/deductions) kept for existing callers of this field
    return {**_pn_dict(pn), "sessions": sess, "readiness": r, "readiness_v2": readiness,
            "publish_blockers": publish_blockers(sess)}


@router.post("/parade-nights")
def create_parade(body: ParadeIn, request: Request, db: DBSession = Depends(get_db),
                  p: Principal = Depends(get_principal)):
    sq_id = _active_squadron(p)
    # Roles that can never write squadron data get a clean 403 first.
    if p.role in ("sqn_general", "wing_viewer", "national_viewer", "auditor"):
        raise HTTPException(403, detail={"error": "forbidden"})
    if not sq_id:
        # Wing/National admins must enter Proxy / Delegated Intervention to gain a squadron scope.
        require_can_write_squadron(p, "none", None)
        raise HTTPException(400, detail={"error": "no_squadron_scope"})
    s = db.get(Squadron, sq_id)
    if not s:
        raise HTTPException(400, detail={"error": "no_squadron_scope"})
    require_can_write_squadron(p, s.id, s.wing_id)
    existing = db.query(ParadeNight).filter(
        ParadeNight.squadron_id == s.id,
        ParadeNight.date == body.date,
        ParadeNight.is_archived == False,  # noqa: E712
    ).first()
    if existing:
        raise HTTPException(409, detail={"error": "duplicate_date", "existing_id": existing.id})
    # Determine session count from effective timing template if one exists.
    # session_count in the body can still override (e.g. user explicitly changes it).
    effective_tmpl = _effective_template(db, s.id, body.date)
    if effective_tmpl and body.session_count is None:
        ip_count = sum(1 for b in effective_tmpl.blocks if b.is_instructional_period)
        session_count = ip_count if ip_count > 0 else (s.default_session_count or 3)
    else:
        session_count = body.session_count or s.default_session_count or 3

    pn = ParadeNight(squadron_id=s.id, wing_id=s.wing_id, date=body.date, term=body.term,
                     start_time=s.default_start_time, end_time=s.default_end_time,
                     session_count=session_count, parade_type=body.parade_type or "normal",
                     timing_template_id=effective_tmpl.id if effective_tmpl else None,
                     created_by=p.user_id)
    db.add(pn); db.flush()

    # R5-M10: capture whether a planning-year link was created. None means the
    # squadron has no active PlanningYear; the parade night is created but won't
    # appear in Planning Workspace's calendar (which keys on ParadeDate rows).
    pd_linked = _find_or_create_parade_date_for_night(db, pn, body.term)

    db.commit()
    meta = client_meta(request)
    audit(db, p, object_type="parade_night", object_id=pn.id, action="create",
          new={"date": body.date}, ip=meta["ip"], ua=meta["ua"])
    return {"ok": True, "parade_night_id": pn.id, "linked_to_planning_year": pd_linked is not None}


def _find_or_create_parade_date_for_night(db: DBSession, pn: "ParadeNight", term: str | None = None):
    """REM-129/REM-34: Planning Workspace's main canvas/command-centre view
    (and PlanningNotice, which FKs to ParadeDate, not ParadeNight) is built
    entirely around ParadeDate (joined via ParadeDate.planning_year_id) -- a
    ParadeNight created through connected-frontend's plain "add one parade
    night" flow had no matching ParadeDate row at all, so it was fully
    visible via GET /api/parade-nights (and therefore inside TMS itself) but
    invisible to anything keyed off ParadeDate. generate_parade_dates() and
    the CSV import path already create this link correctly; this was the one
    gap, now shared by both create_parade (REM-129) and the parade-night-
    scoped Notices endpoints below (REM-34).

    Returns the linked/created ParadeDate, or None if the squadron has no
    active Planning Year to attach one to (a real, surfaceable state -- see
    callers for how each handles it; this helper never invents a Planning
    Year).

    Only auto-links to a squadron-scoped active Planning Year (unit_id ==
    this squadron) -- the same scope Planning Workspace's own
    `GET /api/planning/years?unit_id=...` resolves for a squadron-role
    session -- so this never silently reaches into a wing/national-scoped
    year.
    """
    from ..models.planning import ParadeDate, PlanningYear

    existing_link = db.query(ParadeDate).filter(ParadeDate.parade_night_id == pn.id).first()
    if existing_link:
        return existing_link

    active_year = (
        db.query(PlanningYear)
        .filter(PlanningYear.unit_id == pn.squadron_id, PlanningYear.active_status == True)  # noqa: E712
        .order_by(PlanningYear.year.desc())
        .first()
    )
    if not active_year:
        return None

    existing_pd = db.query(ParadeDate).filter(
        ParadeDate.planning_year_id == active_year.id,
        ParadeDate.parade_date == pn.date,
    ).first()
    if existing_pd:
        if not existing_pd.parade_night_id:
            existing_pd.parade_night_id = pn.id
        return existing_pd

    pd = ParadeDate(
        planning_year_id=active_year.id, unit_id=pn.squadron_id, parade_date=pn.date,
        parade_type="standard", term=term, parade_night_id=pn.id,
    )
    db.add(pd)
    db.flush()
    return pd


def _pn_dict(pn: ParadeNight) -> dict:
    return {"parade_night_id": pn.id, "squadron_id": pn.squadron_id, "date": pn.date, "term": pn.term,
            "start_time": pn.start_time, "end_time": pn.end_time, "session_count": pn.session_count,
            "parade_type": pn.parade_type, "notes": pn.notes, "published_status": pn.published_status,
            "readiness_score": pn.readiness_score, "closeout_status": pn.closeout_status,
            "timing_template_id": pn.timing_template_id, "version": pn.version}


class ParadeNightUpdateIn(BaseModel):
    date: str | None = None
    term: str | None = Field(default=None, max_length=10)
    start_time: str | None = None
    end_time: str | None = None
    session_count: int | None = None
    parade_type: _VALID_PARADE_TYPE | None = None
    notes: str | None = None
    version: int | None = None

    @field_validator("date")
    @classmethod
    def _validate_date(cls, v: str | None) -> str | None:
        if v is None:
            return v
        from datetime import date as _date
        try:
            _date.fromisoformat(v)
        except ValueError:
            raise ValueError(f"date must be a valid ISO-8601 date (YYYY-MM-DD), got '{v}'")
        return v


@router.patch("/parade-nights/{pnid}")
def update_parade_night(pnid: str, body: ParadeNightUpdateIn, db: DBSession = Depends(get_db),
                        p: Principal = Depends(get_principal)):
    """Edit a Parade Night's own core fields after creation -- previously no
    endpoint existed for this in either frontend (only publish/close/archive,
    remediation program Section 9, Stage 5). Closed nights are historical
    records and protected; open (including published-but-not-yet-delivered)
    nights remain editable, matching this codebase's existing distinction
    between 'published' (a workflow milestone) and 'closed' (the actual
    protected/historical state)."""
    pn = db.get(ParadeNight, pnid)
    if not pn or pn.is_archived:
        raise HTTPException(404, detail={"error": "not_found"})
    require_can_write_squadron(p, pn.squadron_id, pn.wing_id)
    if pn.closeout_status == "closed":
        raise HTTPException(409, detail={"error": "parade_night_closed",
                                          "message": "This Parade Night is closed and its core details can no longer be edited."})
    _check_version(pn, body.version)
    if body.date is not None and body.date != pn.date:
        existing = db.query(ParadeNight).filter(
            ParadeNight.squadron_id == pn.squadron_id,
            ParadeNight.date == body.date,
            ParadeNight.is_archived == False,  # noqa: E712
            ParadeNight.id != pn.id,
        ).first()
        if existing:
            raise HTTPException(409, detail={"error": "duplicate_date", "existing_id": existing.id})
        pn.date = body.date
    if body.term is not None:
        pn.term = body.term
    if body.start_time is not None:
        pn.start_time = body.start_time
    if body.end_time is not None:
        pn.end_time = body.end_time
    if body.session_count is not None:
        pn.session_count = body.session_count
    if body.parade_type is not None:
        pn.parade_type = body.parade_type
    if body.notes is not None:
        pn.notes = body.notes
    pn.version += 1
    db.commit()
    audit(db, p, object_type="parade_night", object_id=pn.id, action="update", new={"date": pn.date})
    return _pn_dict(pn)


# ── PARADE NIGHT NOTICES ──
# REM-34: connected-frontend has always worked in terms of ParadeNight, never
# ParadeDate (the Planning-module concept PlanningNotice actually FKs to) --
# these endpoints translate transparently via _find_or_create_parade_date_for_night
# so connected-frontend never needs to know ParadeDate exists at all. Once a
# notice is created, editing/archiving it goes through the existing notice_id-
# based /api/planning/notices/{notice_id} endpoints unchanged (no new duplicate
# logic needed there -- a notice's identity doesn't depend on which frontend
# created it).

class NightNoticeIn(BaseModel):
    notice_text: str
    priority: str = Field(default="normal", max_length=20)
    audience: str | None = Field(default=None, max_length=60)


@router.get("/parade-nights/{pnid}/notices")
def list_night_notices(pnid: str, db: DBSession = Depends(get_db), p: Principal = Depends(get_principal)):
    from ..models.planning import ParadeDate, PlanningNotice
    pn = db.get(ParadeNight, pnid)
    if not pn or pn.is_archived:
        raise HTTPException(404, detail={"error": "not_found"})
    require_can_view_squadron(p, pn.squadron_id, pn.wing_id)
    pd = db.query(ParadeDate).filter(ParadeDate.parade_night_id == pnid).first()
    if not pd:
        # No ParadeDate yet means no notice could ever have been created for
        # this night (create_notice always resolves/creates one first) -- an
        # empty list is correct, not an error.
        return []
    notices = (
        db.query(PlanningNotice)
        .filter(PlanningNotice.parade_date_id == pd.id, PlanningNotice.is_archived == False)  # noqa: E712
        .order_by(PlanningNotice.created_at)
        .all()
    )
    return [_notice_out(n) for n in notices]


@router.post("/parade-nights/{pnid}/notices")
def create_night_notice(pnid: str, body: NightNoticeIn, db: DBSession = Depends(get_db),
                        p: Principal = Depends(get_principal)):
    from ..models.planning import PlanningNotice
    pn = db.get(ParadeNight, pnid)
    if not pn or pn.is_archived:
        raise HTTPException(404, detail={"error": "not_found"})
    require_can_write_squadron(p, pn.squadron_id, pn.wing_id)
    pd = _find_or_create_parade_date_for_night(db, pn, pn.term)
    if not pd:
        raise HTTPException(400, detail={
            "error": "no_active_planning_year",
            "message": "This squadron has no active Training Year, so a notice cannot be attached to this Parade Night yet. Set up a Training Year first.",
        })
    notice = PlanningNotice(
        planning_year_id=pd.planning_year_id,
        parade_date_id=pd.id,
        notice_text=body.notice_text.strip(),
        priority=body.priority,
        audience=body.audience,
        created_by=p.user_id,
    )
    db.add(notice)
    db.commit()
    db.refresh(notice)
    audit(db, p, object_type="PlanningNotice", object_id=notice.id, action="create")
    return {"ok": True, "notice_id": notice.id}


def _notice_out(n) -> dict:
    return {
        "notice_id": n.id,
        "planning_year_id": n.planning_year_id,
        "parade_date_id": n.parade_date_id,
        "notice_text": n.notice_text,
        "audience": n.audience,
        "priority": n.priority,
        "created_by": n.created_by,
        "is_archived": n.is_archived,
        "version": n.version,
        "created_at": n.created_at.isoformat() if n.created_at else None,
        "updated_at": n.updated_at.isoformat() if n.updated_at else None,
    }


# ── SESSIONS ──
class SessionIn(BaseModel):
    parade_night_id: str
    period_number: int = 1
    cadet_group: str | None = None  # orientation/initial/junior/intermediate/senior
    phase_at_time: str | None = "B. Initial"
    curriculum_item_id: str | None = None
    custom_title: str | None = None
    facilitator_id: str | None = None
    training_area_id: str | None = None
    expected_attendance: int | None = None
    version: int | None = None
    # Optional outcome/status transition, applied atomically with the field edits
    # above in one transaction (see edit_session) — a client never ends up with the
    # fields saved but the status change lost (or vice versa) from a network blip or
    # validation failure between two separate calls.
    status: str | None = None
    reason: str | None = None
    rescheduled_to_date: str | None = None
    actual_attendance: int | None = None
    # Skip the synchronous resource-conflict check below (Phase 3 scheduling UX) --
    # the caller has already shown the conflict to the user and confirmed the move.
    override_conflict: bool = False


class StatusIn(BaseModel):
    status: str
    reason: str | None = None
    rescheduled_to_date: str | None = None
    actual_attendance: int | None = None
    version: int | None = None  # optimistic-lock guard (optional for backward compat)


# Statuses where a bare state change with no explanation would leave an untrustworthy
# record — matches Session.{not_delivered_reason,cancelled_reason,issue_notes}, the
# only three status-specific free-text reason fields the model already has.
REASON_REQUIRED_STATUSES = {"not_delivered", "cancelled", "cancelled_late", "delivered_with_issue"}


def _apply_status_transition(s: Session, new_status: str, reason: str | None,
                              rescheduled_to_date: str | None, actual_attendance: int | None) -> str:
    """Validate and apply a status transition to `s` in memory (no commit). Returns the
    old status. Raises HTTPException on invalid input — call this before any other
    mutation/commit so a rejected transition never partially applies."""
    if new_status not in VALID_STATUS:
        raise HTTPException(400, detail={"error": "invalid_status"})
    if new_status in REASON_REQUIRED_STATUSES and not (reason or "").strip():
        raise HTTPException(400, detail={"error": f"reason_required_{new_status}"})
    old_status = s.status
    s.status = new_status
    if new_status == "not_delivered":
        s.not_delivered_reason = reason
    if new_status in ("cancelled", "cancelled_late"):
        s.cancelled_reason = reason
    if new_status == "delivered_with_issue":
        s.issue_notes = reason
    if new_status == "rescheduled":
        s.rescheduled_to_date = rescheduled_to_date
    if actual_attendance is not None:
        s.actual_attendance = actual_attendance
    return old_status


def _recompute(db: DBSession, pn: ParadeNight):
    """The one place ParadeNight's readiness fields are written. Every session
    create/edit/status-change calls this, so it's the natural point to keep
    planning_status/data_quality (the authoritative fields) and readiness_score
    (the legacy numeric projection, derived from the same computation — never
    computed independently) in sync with each other."""
    sess = [_sess_dict(x) for x in db.query(Session).filter(Session.parade_night_id == pn.id,
                                                            Session.is_archived == False).all()]  # noqa: E712
    result = parade_night_readiness(sess)
    pn.planning_status = result["planning_status"]
    pn.data_quality = result["data_quality"]
    pn.readiness_score = result["legacy_score"]
    db.commit()


@router.post("/sessions")
def create_session(body: SessionIn, db: DBSession = Depends(get_db), p: Principal = Depends(get_principal)):
    pn = db.get(ParadeNight, body.parade_night_id)
    if not pn or pn.is_archived:
        raise HTTPException(404, detail={"error": "parade_night_not_found"})
    require_can_write_squadron(p, pn.squadron_id, pn.wing_id)

    # ── Same synchronous resource-conflict check as edit_session below (Stage 8) --
    # previously only PUT /sessions/{sid} was checked, so a facilitator/room could
    # still be double-booked with zero warning by creating a brand new session
    # rather than editing an existing one into a clash. ──
    if not body.override_conflict:
        siblings = db.query(Session).filter(
            Session.parade_night_id == pn.id,
            Session.period_number == body.period_number, Session.is_archived == False,  # noqa: E712
        ).all()
        conflicts = []
        for sib in siblings:
            if body.facilitator_id and sib.facilitator_id == body.facilitator_id:
                conflicts.append({"type": "facilitator_clash", "session_id": sib.id,
                                  "resource_id": sib.facilitator_id,
                                  "resource_name": sib.facilitator_display_name_at_time})
            if body.training_area_id and sib.training_area_id == body.training_area_id:
                conflicts.append({"type": "room_clash", "session_id": sib.id,
                                  "resource_id": sib.training_area_id,
                                  "resource_name": sib.training_area_name_at_time})
        if conflicts:
            raise HTTPException(409, detail={"error": "resource_conflict", "conflicts": conflicts})

    initial_status = body.status if body.status else "planned"
    if initial_status not in ("draft", "planned"):
        raise HTTPException(400, detail={"error": "invalid_initial_status",
                                         "message": "New sessions must be created with status 'draft' or 'planned'."})
    s = Session(parade_night_id=pn.id, squadron_id=pn.squadron_id, period_number=body.period_number,
                cadet_group=body.cadet_group, phase_at_time=body.phase_at_time, custom_title=body.custom_title,
                expected_attendance=body.expected_attendance, status=initial_status, created_by=p.user_id)
    _denormalise(db, s, body.curriculum_item_id, body.facilitator_id, body.training_area_id)
    db.add(s); db.commit()
    _recompute(db, pn)
    audit(db, p, object_type="session", object_id=s.id, action="create")
    return {"ok": True, "session_id": s.id}


@router.put("/sessions/{sid}")
def edit_session(sid: str, body: SessionIn, db: DBSession = Depends(get_db), p: Principal = Depends(get_principal)):
    """Edit session fields, optionally with an atomic status transition in the same
    request (body.status). All validation happens before any mutation, and the field
    edit + status transition + status-history row + audit record share ONE commit —
    if any check fails, nothing is written; if the commit itself fails, nothing is
    written either, since no earlier partial commit exists to leave stale."""
    s = db.get(Session, sid)
    if not s or s.is_archived:
        raise HTTPException(404, detail={"error": "not_found"})
    pn = db.get(ParadeNight, s.parade_night_id)
    require_can_write_squadron(p, s.squadron_id, pn.wing_id if pn else None)
    _check_version(s, body.version)

    # ── Validate everything first — no mutation happens above this line ──
    if body.curriculum_item_id and not db.get(CurriculumItem, body.curriculum_item_id):
        raise HTTPException(400, detail={"error": "invalid_curriculum_item"})
    if body.facilitator_id and not db.get(Facilitator, body.facilitator_id):
        raise HTTPException(400, detail={"error": "invalid_facilitator"})
    if body.training_area_id and not db.get(TrainingArea, body.training_area_id):
        raise HTTPException(400, detail={"error": "invalid_training_area"})
    status_changing = body.status is not None and body.status != s.status
    if status_changing and body.status not in VALID_STATUS:
        raise HTTPException(400, detail={"error": "invalid_status"})
    if status_changing and body.status in REASON_REQUIRED_STATUSES and not (body.reason or "").strip():
        raise HTTPException(400, detail={"error": f"reason_required_{body.status}"})

    # ── Moving to a different Parade Night (accessible "Move to session" control,
    # Phase 3) — body.parade_night_id was previously accepted but never read here,
    # so a session could never actually move. Same squadron only; cross-squadron
    # session transfer is out of scope for this control. ──
    target_pn = pn
    if body.parade_night_id != s.parade_night_id:
        target_pn = db.get(ParadeNight, body.parade_night_id)
        if not target_pn or target_pn.is_archived:
            raise HTTPException(404, detail={"error": "target_parade_night_not_found"})
        if target_pn.squadron_id != s.squadron_id:
            raise HTTPException(403, detail={"error": "cross_squadron_move_forbidden"})

    # ── Synchronous resource-conflict check (Phase 3) — adapts GET /resources/clashes's
    # same-period/same-resource logic, which was read-only and never wired into any
    # write path, so a facilitator/room double-booking was only ever visible after
    # the fact. Runs against the TARGET parade night + period, so it correctly covers
    # both a plain reorder and a move to a different Parade Night. A pure swap between
    # two sessions that didn't already clash can never trigger this against each other
    # (relabelling two non-overlapping period numbers doesn't create new overlap) --
    # only a genuine new collision against a third session's resource does. ──
    if not body.override_conflict:
        siblings = db.query(Session).filter(
            Session.parade_night_id == target_pn.id, Session.id != s.id,
            Session.period_number == body.period_number, Session.is_archived == False,  # noqa: E712
        ).all()
        conflicts = []
        for sib in siblings:
            if body.facilitator_id and sib.facilitator_id == body.facilitator_id:
                conflicts.append({"type": "facilitator_clash", "session_id": sib.id,
                                  "resource_id": sib.facilitator_id,
                                  "resource_name": sib.facilitator_display_name_at_time})
            if body.training_area_id and sib.training_area_id == body.training_area_id:
                conflicts.append({"type": "room_clash", "session_id": sib.id,
                                  "resource_id": sib.training_area_id,
                                  "resource_name": sib.training_area_name_at_time})
        if conflicts:
            raise HTTPException(409, detail={"error": "resource_conflict", "conflicts": conflicts})

    # ── All validation passed — apply every change, then a single commit ──
    old = {"facilitator_id": s.facilitator_id, "training_area_id": s.training_area_id}
    s.parade_night_id = target_pn.id
    s.period_number = body.period_number
    s.cadet_group = body.cadet_group
    s.phase_at_time = body.phase_at_time
    s.custom_title = body.custom_title
    s.expected_attendance = body.expected_attendance
    _denormalise(db, s, body.curriculum_item_id, body.facilitator_id, body.training_area_id)
    s.version += 1

    old_status = s.status
    if status_changing:
        old_status = _apply_status_transition(s, body.status, body.reason,
                                               body.rescheduled_to_date, body.actual_attendance)
        db.add(SessionStatusHistory(session_id=s.id, old_status=old_status, new_status=body.status,
                                    changed_by=p.user_id, reason=body.reason))

    audit(db, p, object_type="session", object_id=s.id, action="edit", old=old,
          new={"facilitator_id": s.facilitator_id, "training_area_id": s.training_area_id}, commit=False)
    if status_changing:
        audit(db, p, object_type="session", object_id=s.id, action="status_change",
              old={"status": old_status}, new={"status": body.status}, reason=body.reason, commit=False)

    db.commit()  # single commit: field edits + status transition + history + audit rows

    if pn:
        _recompute(db, pn)  # old Parade Night, if this was a move -- its readiness changed too
    if target_pn and target_pn.id != (pn.id if pn else None):
        _recompute(db, target_pn)
    return {"ok": True, "version": s.version}


@router.post("/sessions/{sid}/status")
def set_status(sid: str, body: StatusIn, db: DBSession = Depends(get_db), p: Principal = Depends(get_principal)):
    """Status-only transition (no field edits). Kept for callers that only ever
    change status (e.g. the React app's status-only form) — internally atomic in the
    same way edit_session is: validate first, one commit covering the status change,
    history row, and audit record together."""
    s = db.get(Session, sid)
    if not s or s.is_archived:
        raise HTTPException(404, detail={"error": "not_found"})
    pn = db.get(ParadeNight, s.parade_night_id)
    require_can_write_squadron(p, s.squadron_id, pn.wing_id if pn else None)
    _check_version(s, body.version)

    old = _apply_status_transition(s, body.status, body.reason, body.rescheduled_to_date, body.actual_attendance)

    db.add(SessionStatusHistory(session_id=s.id, old_status=old, new_status=body.status,
                                changed_by=p.user_id, reason=body.reason))
    audit(db, p, object_type="session", object_id=s.id, action="status_change",
          old={"status": old}, new={"status": body.status}, reason=body.reason, commit=False)
    db.commit()  # single commit: status transition + history + audit row

    if pn:
        _recompute(db, pn)
    return {"ok": True}


@router.get("/sessions/{sid}/status-history")
def get_status_history(sid: str, db: DBSession = Depends(get_db), p: Principal = Depends(get_principal)):
    """Read-only timeline of a session's status transitions. SessionStatusHistory
    rows were already written on every transition (set_status/edit_session above)
    but nothing ever read them back -- squadron-level users had no way to see a
    session's history at all (only national/auditor/system_admin via the generic
    /api/audit log). Stage 7."""
    s = db.get(Session, sid)
    if not s or s.is_archived:
        raise HTTPException(404, detail={"error": "not_found"})
    pn = db.get(ParadeNight, s.parade_night_id)
    require_can_view_squadron(p, s.squadron_id, pn.wing_id if pn else None)
    rows = db.query(SessionStatusHistory).filter(
        SessionStatusHistory.session_id == sid).order_by(SessionStatusHistory.timestamp).all()
    return [{"old_status": r.old_status, "new_status": r.new_status, "changed_by": r.changed_by,
             "reason": r.reason, "timestamp": r.timestamp.isoformat()} for r in rows]


def _denormalise(db, s: Session, cid, fid, rid):
    if cid is not None:
        s.curriculum_item_id = cid
        c = db.get(CurriculumItem, cid) if cid else None
        if c:
            s.curriculum_code_at_time = c.code
            s.curriculum_title_at_time = c.title
            s.phase_at_time = c.phase
            s.element_at_time = c.element
    if fid is not None:
        s.facilitator_id = fid
        f = db.get(Facilitator, fid) if fid else None
        if f:
            s.facilitator_rank_at_time = f.current_rank
            s.facilitator_display_name_at_time = " ".join(x for x in [f.current_rank, f.first_name, f.last_name] if x)
        elif fid is None:
            s.facilitator_display_name_at_time = None
    if rid is not None:
        s.training_area_id = rid
        r = db.get(TrainingArea, rid) if rid else None
        s.training_area_name_at_time = r.name if r else None


@router.post("/parade-nights/{pnid}/publish")
def publish(pnid: str, db: DBSession = Depends(get_db), p: Principal = Depends(get_principal)):
    pn = db.get(ParadeNight, pnid)
    if not pn or pn.is_archived:
        raise HTTPException(404, detail={"error": "not_found"})
    require_can_write_squadron(p, pn.squadron_id, pn.wing_id)
    sess = [_sess_dict(x) for x in db.query(Session).filter(Session.parade_night_id == pn.id,
                                                            Session.is_archived == False).all()]  # noqa: E712
    blockers = publish_blockers(sess)
    if blockers:
        raise HTTPException(409, detail={"error": "publish_blocked", "blockers": blockers})
    pn.published_status = True; pn.published_by = p.user_id; pn.published_at = utcnow()
    db.query(Session).filter(Session.parade_night_id == pn.id,
                             Session.status.in_(("draft", "planned"))).update({"status": "published"})
    db.commit()
    audit(db, p, object_type="parade_night", object_id=pn.id, action="publish")
    return {"ok": True}


@router.post("/parade-nights/{pnid}/mark-remaining-delivered")
def mark_remaining_delivered(pnid: str, db: DBSession = Depends(get_db), p: Principal = Depends(get_principal)):
    """"Mark all delivered, then flag exceptions" bulk action (data-entry UX
    ask from the risk-register submission): transitions every session still
    sitting in draft/planned/published to delivered, leaving untouched any
    session the user has already individually flagged as cancelled/
    not_delivered/delivered_with_issue/rescheduled -- the intended workflow
    is to flag the few known exceptions first, then bulk-clear the rest in
    one action rather than clicking "Delivered" on every session
    individually. Each session still gets its own SessionStatusHistory row
    and audit entry (per-session commit, like batch_archive_accounts), not a
    silent bulk UPDATE that skips the normal record-keeping."""
    import uuid as _uuid
    pn = db.get(ParadeNight, pnid)
    if not pn or pn.is_archived:
        raise HTTPException(404, detail={"error": "not_found"})
    require_can_write_squadron(p, pn.squadron_id, pn.wing_id)
    if pn.closeout_status == "closed":
        raise HTTPException(409, detail={"error": "parade_night_closed",
                                          "message": "This Parade Night is closed and sessions can no longer be bulk-updated."})

    batch_id = str(_uuid.uuid4())
    sessions = db.query(Session).filter(
        Session.parade_night_id == pn.id,
        Session.is_archived == False,  # noqa: E712
        Session.status.in_(("draft", "planned", "published")),
    ).all()
    updated_ids = []
    for s in sessions:
        old_status = _apply_status_transition(s, "delivered", None, None, None)
        db.add(SessionStatusHistory(session_id=s.id, old_status=old_status, new_status="delivered",
                                    changed_by=p.user_id, reason="Bulk: mark remaining delivered"))
        audit(db, p, object_type="session", object_id=s.id, action="status_change",
              old={"status": old_status}, new={"status": "delivered"},
              reason="bulk_mark_remaining_delivered", batch_id=batch_id, commit=False)
        updated_ids.append(s.id)

    if updated_ids:
        _recompute(db, pn)
    audit(db, p, object_type="parade_night_bulk", object_id=batch_id, action="bulk_mark_remaining_delivered",
          new={"parade_night_id": pn.id, "sessions_updated": len(updated_ids)}, batch_id=batch_id)
    return {"ok": True, "batch_id": batch_id, "sessions_updated": len(updated_ids), "session_ids": updated_ids}


class CancelAllIn(BaseModel):
    reason: str
    notes: str | None = None


@router.post("/parade-nights/{pnid}/cancel-all")
def cancel_all_sessions(pnid: str, body: CancelAllIn, db: DBSession = Depends(get_db),
                        p: Principal = Depends(get_principal)):
    """Bulk-cancel all planned/published sessions in a parade night.

    Marks every session in draft/planned/published state as 'cancelled' with the
    supplied reason, leaving already-finalised sessions (delivered, not_delivered,
    cancelled, rescheduled) untouched.  Mirrors mark_remaining_delivered in all
    audit and history bookkeeping."""
    import uuid as _uuid
    pn = db.get(ParadeNight, pnid)
    if not pn or pn.is_archived:
        raise HTTPException(404, detail={"error": "not_found"})
    require_can_write_squadron(p, pn.squadron_id, pn.wing_id)
    if pn.closeout_status == "closed":
        raise HTTPException(409, detail={"error": "parade_night_closed",
                                          "message": "This Parade Night is closed and sessions can no longer be bulk-cancelled."})
    if not body.reason or not body.reason.strip():
        raise HTTPException(422, detail={"error": "reason_required"})

    batch_id = str(_uuid.uuid4())
    sessions = db.query(Session).filter(
        Session.parade_night_id == pn.id,
        Session.is_archived == False,  # noqa: E712
        Session.status.in_(("draft", "planned", "published")),
    ).all()
    updated_ids = []
    for s in sessions:
        old_status = _apply_status_transition(s, "cancelled", body.reason, None, None)
        db.add(SessionStatusHistory(session_id=s.id, old_status=old_status, new_status="cancelled",
                                    changed_by=p.user_id, reason=body.reason))
        audit(db, p, object_type="session", object_id=s.id, action="status_change",
              old={"status": old_status}, new={"status": "cancelled", "reason": body.reason},
              reason="bulk_cancel_all", batch_id=batch_id, commit=False)
        updated_ids.append(s.id)

    if updated_ids:
        _recompute(db, pn)
    audit(db, p, object_type="parade_night_bulk", object_id=batch_id, action="bulk_cancel_all",
          new={"parade_night_id": pn.id, "sessions_updated": len(updated_ids)})
    return {"ok": True, "batch_id": batch_id, "sessions_updated": len(updated_ids), "session_ids": updated_ids}


@router.post("/parade-nights/{pnid}/close")
def close(pnid: str, db: DBSession = Depends(get_db), p: Principal = Depends(get_principal)):
    pn = db.get(ParadeNight, pnid)
    if not pn or pn.is_archived:
        raise HTTPException(404, detail={"error": "not_found"})
    require_can_write_squadron(p, pn.squadron_id, pn.wing_id)
    sess = [_sess_dict(x) for x in db.query(Session).filter(Session.parade_night_id == pn.id,
                                                            Session.is_archived == False).all()]  # noqa: E712
    blockers = close_blockers(sess)
    if blockers:
        raise HTTPException(409, detail={"error": "close_blocked", "blockers": blockers})
    pn.closeout_status = "closed"; pn.closed_by = p.user_id; pn.closed_at = utcnow()
    db.commit()
    audit(db, p, object_type="parade_night", object_id=pn.id, action="close")
    return {"ok": True}


# ── FACILITATORS ──
_MAX_TAGS = 20
_MAX_TAG_LEN = 80

# Canonical AAFC rank list (DEF-04) — exposed via GET /api/facilitators/ranks
# so the frontend datalist is populated from a single authoritative source.
# This is advisory: the API accepts non-catalogue values so existing/edge-case
# data (e.g. "External Guest", "Civilian Instructor") is never rejected.
_AAFC_RANKS: list[dict] = [
    # Officer ranks (AAFC)
    {"code": "PLTOFF(AAFC)", "label": "Pilot Officer (AAFC)", "category": "Officer"},
    {"code": "FLGOFF(AAFC)", "label": "Flying Officer (AAFC)", "category": "Officer"},
    {"code": "FLTLT(AAFC)",  "label": "Flight Lieutenant (AAFC)", "category": "Officer"},
    {"code": "SQNLDR(AAFC)", "label": "Squadron Leader (AAFC)", "category": "Officer"},
    {"code": "WGCDR(AAFC)",  "label": "Wing Commander (AAFC)", "category": "Officer"},
    {"code": "GPCAPT(AAFC)", "label": "Group Captain (AAFC)", "category": "Officer"},
    {"code": "AIRCDRE(AAFC)","label": "Air Commodore (AAFC)", "category": "Officer"},
    {"code": "AVM(AAFC)",    "label": "Air Vice-Marshal (AAFC)", "category": "Officer"},
    # Airmen/SNCO ranks (AAFC)
    {"code": "LAC(AAFC)",    "label": "Leading Aircraftman (AAFC)", "category": "Airmen"},
    {"code": "CPL(AAFC)",    "label": "Corporal (AAFC)", "category": "Airmen"},
    {"code": "SGT(AAFC)",    "label": "Sergeant (AAFC)", "category": "Airmen"},
    {"code": "FSGT(AAFC)",   "label": "Flight Sergeant (AAFC)", "category": "Airmen"},
    {"code": "WOFF(AAFC)",   "label": "Warrant Officer (AAFC)", "category": "Airmen"},
    # Senior cadet ranks
    {"code": "CCPL",  "label": "Cadet Corporal", "category": "Senior Cadet"},
    {"code": "CSGT",  "label": "Cadet Sergeant", "category": "Senior Cadet"},
    {"code": "CFSGT", "label": "Cadet Flight Sergeant", "category": "Senior Cadet"},
    {"code": "CWO",   "label": "Cadet Warrant Officer", "category": "Senior Cadet"},
    {"code": "CUO",   "label": "Cadet Under Officer", "category": "Senior Cadet"},
    # Civilian / other
    {"code": "CIV",   "label": "Civilian", "category": "Civilian"},
]
_AAFC_RANK_CODES: frozenset[str] = frozenset(r["code"] for r in _AAFC_RANKS)


def _normalise_rank(rank: str | None) -> str | None:
    """Strip whitespace; truncate to DB column limit. Does NOT reject non-catalogue values."""
    if rank is None:
        return None
    rank = rank.strip()
    if not rank:
        return None
    return rank[:40]  # matches Facilitator.current_rank String(40)


def _validate_subject_areas(raw: list[str] | None) -> list[str]:
    """Trim, deduplicate (case-insensitive), and enforce limits on tag list."""
    if not raw:
        return []
    seen: dict[str, str] = {}
    for item in raw:
        if not isinstance(item, str):
            raise HTTPException(422, detail={"error": "subject_areas_invalid",
                                             "message": "Each subject area must be a string."})
        tag = item.strip()
        if not tag:
            continue
        if len(tag) > _MAX_TAG_LEN:
            raise HTTPException(422, detail={"error": "subject_areas_invalid",
                                             "message": f"Tags must be {_MAX_TAG_LEN} characters or fewer."})
        seen.setdefault(tag.lower(), tag)
    result = list(seen.values())
    if len(result) > _MAX_TAGS:
        raise HTTPException(422, detail={"error": "subject_areas_invalid",
                                         "message": f"A maximum of {_MAX_TAGS} subject areas is allowed."})
    return result


class FacIn(BaseModel):
    first_name: str | None = Field(default=None, max_length=80)
    last_name: str = Field(max_length=80)
    current_rank: str | None = None
    type: str | None = Field(default="Staff", max_length=30)
    subject_areas: list[str] | None = None
    confirm_duplicate: bool = False


@router.get("/facilitators/ranks")
def list_ranks(p: Principal = Depends(get_principal)):
    """Return the canonical AAFC rank catalogue (DEF-04).
    Authenticated; used to populate rank pickers in both frontends."""
    return {"ranks": _AAFC_RANKS}


@router.get("/facilitators")
def list_facs(squadron_id: str | None = None, include_archived: bool = False,
              db: DBSession = Depends(get_db), p: Principal = Depends(get_principal)):
    from datetime import date, timedelta
    from ..models.planning import PlanningFacilitatorLeave
    sq_id = _view_squadron_id(p, squadron_id, db)
    q = db.query(Facilitator).filter(Facilitator.squadron_id == sq_id)
    if not include_archived:
        q = q.filter(Facilitator.is_archived == False)  # noqa: E712
    facs = q.all()
    today = date.today().isoformat()
    horizon = (date.today() + timedelta(days=90)).isoformat()
    out = []
    for f in facs:
        leave = (db.query(PlanningFacilitatorLeave)
                 .filter(PlanningFacilitatorLeave.facilitator_id == f.id,
                         PlanningFacilitatorLeave.is_archived == False,  # noqa: E712
                         PlanningFacilitatorLeave.end_date >= today,
                         PlanningFacilitatorLeave.start_date <= horizon)
                 .order_by(PlanningFacilitatorLeave.start_date).all())
        out.append({"facilitator_id": f.id, "first_name": f.first_name, "last_name": f.last_name,
                    "current_rank": f.current_rank, "type": f.type,
                    "subject_areas": _parse_json_list(f.subject_areas), "is_archived": f.is_archived,
                    "upcoming_leave": [{"id": lv.id, "start_date": lv.start_date, "end_date": lv.end_date,
                                        "reason": lv.reason} for lv in leave]})
    return out


@router.post("/facilitators")
def add_fac(body: FacIn, db: DBSession = Depends(get_db), p: Principal = Depends(get_principal),
            idempotency_key: str | None = Header(None, alias="Idempotency-Key")):
    from sqlalchemy import func
    from ..security import idempotency_get, idempotency_set
    if p.role in _WRITE_BLOCKED:
        raise HTTPException(403, detail={"error": "forbidden"})
    # A retried POST after a client-perceived timeout must not create a second
    # facilitator, most importantly when confirm_duplicate=true (the "Add
    # anyway" resubmit) -- that path has no other duplicate protection at all.
    idem_cache_key = f"{p.user_id}:facilitators:{idempotency_key}" if idempotency_key else None
    if idem_cache_key:
        cached = idempotency_get(idem_cache_key)
        if cached is not None:
            return cached[1]
    sq_id = _active_squadron(p)
    if not sq_id:
        require_can_write_squadron(p, "none", None)
    s = db.get(Squadron, sq_id)
    if not s:
        raise HTTPException(400, detail={"error": "no_squadron_scope"})
    require_can_write_squadron(p, s.id, s.wing_id)
    # TRGO-07: warn (don't silently allow) when a facilitator with the same
    # name already exists in this squadron -- the same person re-added by
    # accident is common; a same-name-different-person case is legitimate
    # and rare, so this blocks once and lets the caller explicitly confirm
    # rather than either always blocking or never warning at all.
    if not body.confirm_duplicate:
        existing = db.query(Facilitator).filter(
            Facilitator.squadron_id == s.id,
            Facilitator.is_archived == False,  # noqa: E712
            func.lower(Facilitator.last_name) == body.last_name.strip().lower(),
            func.lower(func.coalesce(Facilitator.first_name, "")) == (body.first_name or "").strip().lower(),
        ).first()
        if existing:
            # REM-96/REM-97: the caller needs enough of the existing facilitator's
            # profile to actually tell "same person, re-added by accident" from
            # "different person, same name" without a second round trip --
            # everything here is already loaded on `existing` from the query
            # above, so it's returned directly rather than requiring the
            # frontend to make a separate GET to fetch a profile card's worth
            # of detail. Rank is surfaced in the message text itself (REM-96's
            # specific ask) since that's the single most likely distinguishing
            # detail between two same-named facilitators.
            msg = (f"A facilitator named {(body.first_name or '').strip()} "
                   f"{body.last_name.strip()} already exists in this squadron").strip()
            if existing.current_rank:
                msg += f" (rank: {existing.current_rank})"
            msg += "."
            raise HTTPException(409, detail={
                "error": "possible_duplicate",
                "message": msg,
                "existing_facilitator_id": existing.id,
                "existing_rank": existing.current_rank,
                "existing_type": existing.type,
                "existing_subject_areas": _parse_json_list(existing.subject_areas),
                "existing_active_status": existing.active_status,
                "existing_updated_at": existing.updated_at.isoformat() if existing.updated_at else None,
            })
    rank = _normalise_rank(body.current_rank)
    f = Facilitator(squadron_id=s.id, wing_id=s.wing_id, first_name=body.first_name,
                    last_name=body.last_name, current_rank=rank, type=body.type or "Staff",
                    subject_areas=_validate_subject_areas(body.subject_areas))
    db.add(f); db.commit()
    db.add(FacilitatorRankHistory(facilitator_id=f.id, rank=rank, effective_from=str(utcnow().date())))
    db.commit()
    audit(db, p, object_type="facilitator", object_id=f.id, action="create")
    result = {"ok": True, "facilitator_id": f.id}
    if idem_cache_key:
        idempotency_set(idem_cache_key, 200, result)
    return result


@router.get("/facilitators/import/template.csv")
def facilitator_import_template(p: Principal = Depends(get_principal)):
    """TRGO-05: downloadable CSV template matching the columns import_facilitators_csv accepts."""
    import io
    from fastapi.responses import StreamingResponse
    if p.role in _WRITE_BLOCKED:
        raise HTTPException(403, detail={"error": "forbidden"})
    bio = io.BytesIO(
        b"rank,first_name,last_name,type,subject_areas,active_status\r\n"
        b"FLTLT,Jordan,Smith,Staff,Drill;Air_Space,true\r\n"
    )
    return StreamingResponse(bio, media_type="text/csv",
                             headers={"Content-Disposition": "attachment; filename=facilitator_import_template.csv"})


_FAC_CSV_FIELDS = {
    "rank": ("rank", "current_rank"),
    "first_name": ("first_name", "first name", "firstname", "given name"),
    "last_name": ("last_name", "last name", "lastname", "surname", "family name"),
    "type": ("type", "facilitator type"),
    "subject_areas": ("subject_areas", "subject areas", "subjects"),
    "active_status": ("active_status", "active", "active status"),
}


def _neutralise_csv_cell(v: str) -> str:
    """Strip a leading formula-trigger character (=, +, -, @, tab, CR) so a
    re-exported/opened-in-Excel copy of this data can never execute a
    formula injected via an uploaded CSV cell (CSV/Excel formula injection)."""
    v = (v or "").strip()
    while v and v[0] in "=+-@\t\r":
        v = v[1:].strip()
    return v


def _fac_csv_get(row: dict, key: str) -> str:
    norm = {k.strip().lower(): (v or "").strip() for k, v in row.items()}
    for alias in _FAC_CSV_FIELDS[key]:
        if alias in norm and norm[alias]:
            return _neutralise_csv_cell(norm[alias])
    return ""


class FacilitatorImportIn(BaseModel):
    preview: bool = True
    # Row indices (0-based, matching the preview's row order) the caller has
    # explicitly reviewed and wants created despite a name match -- the CSV
    # equivalent of confirm_duplicate on the single-facilitator endpoint.
    confirm_duplicate_rows: list[int] = []


@router.post("/facilitators/import")
async def import_facilitators_csv(
    file: UploadFile = File(...),
    preview: bool = True,
    confirm_duplicate_rows: str = "",
    db: DBSession = Depends(get_db),
    p: Principal = Depends(get_principal),
):
    """TRGO-05: governed bulk facilitator import.

    Reuses the CEA import pipeline's preview-then-commit shape and the
    TRGO-07 duplicate-detection already shipped for single-facilitator
    create (case-insensitive first+last name match within the squadron) --
    the same "same person added twice by accident is common, but a genuine
    same-name-different-person case must still be possible" reasoning
    applies identically to a bulk import.

    No stable per-row identifier exists on Facilitator today (no email/
    service-number field), so matching is name-based, same as the single-
    create endpoint -- this is a real, named limitation, not a silent gap:
    a CSV with two genuinely different people who happen to share a name
    needs `confirm_duplicate_rows` to include both.

    CSV formula injection: any cell starting with =, +, -, @ (a formula
    trigger in Excel/Sheets) is neutralised before it ever reaches storage
    or a later export.
    """
    if p.role in _WRITE_BLOCKED:
        raise HTTPException(403, detail={"error": "forbidden"})
    sq_id = _active_squadron(p)
    if not sq_id:
        require_can_write_squadron(p, "none", None)
    s = db.get(Squadron, sq_id)
    if not s:
        raise HTTPException(400, detail={"error": "no_squadron_scope"})
    if not preview:
        require_can_write_squadron(p, s.id, s.wing_id)

    from ..config import settings as _settings
    content = await file.read()
    if len(content) > _settings.UPLOAD_MAX_MB * 1024 * 1024:
        raise HTTPException(413, detail={"error": "file_too_large"})
    try:
        text = content.decode("utf-8-sig")
    except UnicodeDecodeError:
        text = content.decode("latin-1")

    import csv, io
    reader = csv.DictReader(io.StringIO(text))
    if reader.fieldnames is None:
        raise HTTPException(400, detail={"error": "invalid_header", "message": "CSV has no header row."})

    existing = {
        ((e.first_name or "").strip().lower(), e.last_name.strip().lower())
        for e in db.query(Facilitator).filter(
            Facilitator.squadron_id == s.id, Facilitator.is_archived == False,  # noqa: E712
        ).all()
    }
    confirm_idx = set()
    if confirm_duplicate_rows:
        try:
            confirm_idx = {int(x) for x in confirm_duplicate_rows.split(",") if x.strip()}
        except ValueError:
            raise HTTPException(400, detail={"error": "invalid_confirm_rows"})

    rows_out = []
    seen_in_file: set[tuple[str, str]] = set()
    for i, row in enumerate(reader):
        last_name = _fac_csv_get(row, "last_name")
        first_name = _fac_csv_get(row, "first_name")
        if not last_name:
            rows_out.append({"row": i, "action": "error", "message": "Missing last name.", "last_name": last_name, "first_name": first_name})
            continue
        key = (first_name.lower(), last_name.lower())
        action = "create"
        if key in existing:
            action = "duplicate"
        elif key in seen_in_file:
            action = "duplicate_in_file"
        seen_in_file.add(key)

        rec = {
            "row": i, "action": action,
            "first_name": first_name or None, "last_name": last_name,
            "current_rank": _fac_csv_get(row, "rank") or None,
            "type": _fac_csv_get(row, "type") or "Staff",
            "subject_areas": [x.strip() for x in _fac_csv_get(row, "subject_areas").split(";") if x.strip()],
            "active_status": _fac_csv_get(row, "active_status").lower() not in ("false", "0", "no", "inactive"),
        }
        rows_out.append(rec)

    if preview:
        return {
            "ok": True, "preview": True, "rows": rows_out,
            "to_create": sum(1 for r in rows_out if r["action"] == "create" or r["row"] in confirm_idx),
            "duplicates": sum(1 for r in rows_out if r["action"] in ("duplicate", "duplicate_in_file")),
            "errors": sum(1 for r in rows_out if r["action"] == "error"),
        }

    created = skipped = errors = 0
    created_ids = []
    for r in rows_out:
        if r["action"] == "error":
            errors += 1
            continue
        if r["action"] in ("duplicate", "duplicate_in_file") and r["row"] not in confirm_idx:
            skipped += 1
            continue
        csv_rank = _normalise_rank(r["current_rank"])
        f = Facilitator(
            squadron_id=s.id, wing_id=s.wing_id,
            first_name=r["first_name"], last_name=r["last_name"],
            current_rank=csv_rank, type=r["type"] or "Staff",
            subject_areas=_validate_subject_areas(r["subject_areas"]),
            active_status=r["active_status"],
        )
        db.add(f); db.flush()
        db.add(FacilitatorRankHistory(facilitator_id=f.id, rank=csv_rank, effective_from=str(utcnow().date())))
        created += 1
        created_ids.append(f.id)
    db.commit()
    audit(db, p, object_type="facilitator", object_id=None, action="csv_import",
          new={"created": created, "skipped": skipped, "errors": errors, "source_file_name": file.filename})
    return {
        "ok": True, "preview": False,
        "created": created, "skipped": skipped, "errors": errors,
        "created_ids": created_ids, "total": len(rows_out),
    }


class FacUpdateIn(BaseModel):
    subject_areas: list[str] | None = None
    type: str | None = Field(default=None, max_length=30)
    current_rank: str | None = None
    first_name: str | None = Field(default=None, max_length=80)
    last_name: str | None = Field(default=None, max_length=80)


@router.patch("/facilitators/{fid}")
def update_fac(fid: str, body: FacUpdateIn, db: DBSession = Depends(get_db),
               p: Principal = Depends(get_principal)):
    f = db.get(Facilitator, fid)
    if not f:
        raise HTTPException(404, detail={"error": "not_found"})
    require_can_write_squadron(p, f.squadron_id, f.wing_id)
    if body.subject_areas is not None:
        f.subject_areas = _validate_subject_areas(body.subject_areas)
    if body.type is not None:
        f.type = body.type
    new_rank = _normalise_rank(body.current_rank) if body.current_rank is not None else None
    rank_changed = new_rank is not None and new_rank != f.current_rank
    if new_rank is not None:
        f.current_rank = new_rank
    if body.first_name is not None:
        f.first_name = body.first_name
    if body.last_name is not None:
        f.last_name = body.last_name
    db.commit()
    if rank_changed:
        db.add(FacilitatorRankHistory(facilitator_id=f.id, rank=new_rank,
                                      effective_from=str(utcnow().date())))
        db.commit()
    audit(db, p, object_type="facilitator", object_id=f.id, action="update",
          reason="subject-area tags" if body.subject_areas is not None else "update")
    return {"ok": True, "facilitator_id": f.id,
            "subject_areas": _parse_json_list(f.subject_areas)}


@router.get("/facilitators/{fid}/stats")
def fac_stats(fid: str, db: DBSession = Depends(get_db), p: Principal = Depends(get_principal)):
    f = db.get(Facilitator, fid)
    if not f:
        raise HTTPException(404, detail={"error": "not_found"})
    require_can_view_squadron(p, f.squadron_id, f.wing_id)
    sess = db.query(Session).filter(Session.facilitator_id == fid).all()
    counts, by_phase = {}, {}
    for s in sess:
        counts[s.status] = counts.get(s.status, 0) + 1
        by_phase[s.phase_at_time] = by_phase.get(s.phase_at_time, 0) + 1
    return {"facilitator": {"facilitator_id": f.id, "name": " ".join(x for x in [f.current_rank, f.first_name, f.last_name] if x)},
            "counts": counts, "by_phase": by_phase, "load_score": len(sess) * 2,
            "sessions": [_sess_dict(s) for s in sess]}


# ── RESOURCES ──
@router.get("/training-areas")
def list_rooms(squadron_id: str | None = None, include_archived: bool = False,
               db: DBSession = Depends(get_db), p: Principal = Depends(get_principal)):
    sq_id = _view_squadron_id(p, squadron_id, db)
    q = db.query(TrainingArea).filter(TrainingArea.squadron_id == sq_id)
    if not include_archived:
        q = q.filter(TrainingArea.is_archived == False)  # noqa: E712
    rows = q.all()
    return [{"training_area_id": r.id, "name": r.name, "type": r.type, "capacity": r.capacity,
             "indoor_outdoor": r.indoor_outdoor, "is_archived": r.is_archived} for r in rows]


@router.get("/equipment")
def list_equipment(squadron_id: str | None = None, include_archived: bool = False,
                   db: DBSession = Depends(get_db), p: Principal = Depends(get_principal)):
    sq_id = _view_squadron_id(p, squadron_id, db)
    q = db.query(Equipment).filter(Equipment.squadron_id == sq_id)
    if not include_archived:
        q = q.filter(Equipment.is_archived == False)  # noqa: E712
    rows = q.all()
    return [{"equipment_id": e.id, "name": e.name, "type": e.type, "quantity": e.quantity,
             "available_quantity": e.available_quantity, "is_archived": e.is_archived} for e in rows]


@router.get("/resources/clashes")
def resource_clashes(date: str, db: DBSession = Depends(get_db), p: Principal = Depends(get_principal)):
    """Detect room/facilitator double-booking on a given parade date."""
    sq_id = _active_squadron(p)
    pn = db.query(ParadeNight).filter(ParadeNight.squadron_id == sq_id, ParadeNight.date == date).first()
    clashes = []
    if pn:
        sess = db.query(Session).filter(Session.parade_night_id == pn.id).all()
        rooms, facs = {}, {}
        for s in sess:
            if s.training_area_id:
                rooms.setdefault(s.training_area_id, []).append(s.period_number)
            if s.facilitator_id:
                facs.setdefault(s.facilitator_id, []).append(s.period_number)
        for rid, periods in rooms.items():
            if len(periods) != len(set(periods)):
                clashes.append({"type": "room_clash", "resource_id": rid})
        for fid, periods in facs.items():
            if len(periods) != len(set(periods)):
                clashes.append({"type": "facilitator_clash", "resource_id": fid})
    return {"date": date, "clashes": clashes}


# ── CADETS (sensitive; sqn_general blocked) ──
@router.get("/cadets")
def list_cadets(db: DBSession = Depends(get_db), p: Principal = Depends(get_principal)):
    if p.role == "sqn_general":
        raise HTTPException(403, detail={"error": "forbidden"})
    sq_id = _active_squadron(p)
    rows = db.query(Cadet).filter(Cadet.squadron_id == sq_id, Cadet.is_archived == False).all()  # noqa: E712
    # support_notes (sensitive) only for admins
    can_sensitive = p.role in ("sqn_admin", "wing_admin", "national_admin", "system_admin")
    out = []
    for c in rows:
        d = {"cadet_id": c.id, "service_number": c.service_number, "rank": c.rank,
             "first_name": c.first_name, "last_name": c.last_name, "phase": c.phase,
             "attendance_percentage": c.attendance_percentage,
             "sitrep_part_1_status": c.sitrep_part_1_status, "support_flag": c.support_flag}
        if can_sensitive:
            d["support_notes"] = c.support_notes
        out.append(d)
    if any(c.support_notes for c in rows) and can_sensitive:
        audit(db, p, object_type="cadet", object_id=sq_id, action="view_sensitive")
    return out


@router.get("/cadets/risk")
def cadet_risk(db: DBSession = Depends(get_db), p: Principal = Depends(get_principal)):
    if p.role == "sqn_general":
        raise HTTPException(403, detail={"error": "forbidden"})
    sq_id = _active_squadron(p)
    rows = db.query(Cadet).filter(Cadet.squadron_id == sq_id, Cadet.is_archived == False).all()  # noqa: E712
    flags = []
    for c in rows:
        reasons = []
        if (c.attendance_percentage or 100) < 75:
            reasons.append("Attendance below 75%")
        if c.recent_attendance_trend == "declining":
            reasons.append("Attendance declining")
        if c.sitrep_part_1_status == "pending":
            reasons.append("SITREP Part 1 missing")
        if reasons:
            flags.append({"cadet": f"{c.rank} {c.first_name} {c.last_name}", "reasons": reasons})
    return flags


# ── CADET CLASS MEMBERSHIP (CLASS-09) ───────────────────────────────────────
# Individual Cadet<->Training Class tracking, explicitly product-scope-
# confirmed before being built (addendum §38/§39 -- see the model's own
# docstring in models/training.py). Enables a Cadet to hold concurrent
# Foundation (e.g. Senior 1) and Extension (e.g. Bronze CLP) membership,
# which Cadet.phase's single free-text column cannot express. Cadet.phase
# is untouched.

class CadetClassMembershipIn(BaseModel):
    training_class_id: str
    start_date: Optional[str] = None
    source: Optional[str] = "manual"


class CadetClassMembershipUpdateIn(BaseModel):
    end_date: Optional[str] = None
    active_status: Optional[bool] = None
    client_version: Optional[int] = None


def _membership_dict(m: CadetClassMembership, tc: TrainingClass | None = None) -> dict:
    return {
        "membership_id": m.id, "cadet_id": m.cadet_id, "training_class_id": m.training_class_id,
        "training_class_name": tc.display_name if tc else None,
        "training_stage_id": tc.training_stage_id if tc else None,
        "start_date": m.start_date, "end_date": m.end_date, "active_status": m.active_status,
        "source": m.source, "version": m.version,
        "created_at": m.created_at.isoformat() if m.created_at else None,
    }


def _require_cadet_and_squadron(db: DBSession, p: Principal, cadet_id: str, write: bool) -> Cadet:
    cadet = db.get(Cadet, cadet_id)
    if not cadet or cadet.is_archived:
        raise HTTPException(404, detail={"error": "cadet_not_found"})
    s = db.get(Squadron, cadet.squadron_id)
    if write:
        require_can_write_squadron(p, cadet.squadron_id, s.wing_id if s else None)
    else:
        require_can_view_squadron(p, cadet.squadron_id, s.wing_id if s else None)
    return cadet


@router.get("/cadets/{cadet_id}/class-memberships")
def list_cadet_class_memberships(
    cadet_id: str, include_archived: bool = False,
    db: DBSession = Depends(get_db), p: Principal = Depends(get_principal),
):
    if p.role == "sqn_general":
        raise HTTPException(403, detail={"error": "forbidden"})
    _require_cadet_and_squadron(db, p, cadet_id, write=False)
    q = db.query(CadetClassMembership).filter(CadetClassMembership.cadet_id == cadet_id)
    if not include_archived:
        q = q.filter(CadetClassMembership.is_archived == False)  # noqa: E712
    rows = q.order_by(CadetClassMembership.created_at.desc()).all()
    tc_ids = [m.training_class_id for m in rows]
    tc_map = {tc.id: tc for tc in db.query(TrainingClass).filter(TrainingClass.id.in_(tc_ids)).all()} if tc_ids else {}
    return [_membership_dict(m, tc_map.get(m.training_class_id)) for m in rows]


@router.post("/cadets/{cadet_id}/class-memberships")
def add_cadet_class_membership(
    cadet_id: str, body: CadetClassMembershipIn,
    db: DBSession = Depends(get_db), p: Principal = Depends(get_principal),
):
    if p.role in _WRITE_BLOCKED:
        raise HTTPException(403, detail={"error": "forbidden"})
    cadet = _require_cadet_and_squadron(db, p, cadet_id, write=True)

    tc = db.get(TrainingClass, body.training_class_id)
    if not tc or tc.is_archived:
        raise HTTPException(400, detail={"error": "invalid_training_class"})
    if tc.squadron_id != cadet.squadron_id:
        raise HTTPException(400, detail={"error": "training_class_wrong_squadron"})

    existing = db.query(CadetClassMembership).filter(
        CadetClassMembership.cadet_id == cadet_id,
        CadetClassMembership.training_class_id == body.training_class_id,
        CadetClassMembership.is_archived == False,  # noqa: E712
        CadetClassMembership.active_status == True,  # noqa: E712
    ).first()
    if existing:
        raise HTTPException(409, detail={"error": "already_a_member", "membership_id": existing.id})

    m = CadetClassMembership(
        cadet_id=cadet_id, training_class_id=body.training_class_id,
        start_date=body.start_date, source=body.source or "manual",
        created_by=p.user_id, updated_by=p.user_id,
    )
    db.add(m)
    db.commit()
    audit(db, p, object_type="cadet_class_membership", object_id=m.id, action="create",
          new={"cadet_id": cadet_id, "training_class_id": body.training_class_id})
    return _membership_dict(m, tc)


@router.patch("/cadet-class-memberships/{membership_id}")
def update_cadet_class_membership(
    membership_id: str, body: CadetClassMembershipUpdateIn,
    db: DBSession = Depends(get_db), p: Principal = Depends(get_principal),
):
    m = db.get(CadetClassMembership, membership_id)
    if not m or m.is_archived:
        raise HTTPException(404, detail={"error": "not_found"})
    _require_cadet_and_squadron(db, p, m.cadet_id, write=True)
    _check_version(m, body.client_version)

    if body.end_date is not None:
        m.end_date = body.end_date
    if body.active_status is not None:
        m.active_status = body.active_status
    m.version += 1
    m.updated_by = p.user_id
    db.commit()
    audit(db, p, object_type="cadet_class_membership", object_id=m.id, action="update")
    tc = db.get(TrainingClass, m.training_class_id)
    return _membership_dict(m, tc)


@router.delete("/cadet-class-memberships/{membership_id}")
def archive_cadet_class_membership(
    membership_id: str, db: DBSession = Depends(get_db), p: Principal = Depends(get_principal),
):
    m = db.get(CadetClassMembership, membership_id)
    if not m or m.is_archived:
        raise HTTPException(404, detail={"error": "not_found"})
    _require_cadet_and_squadron(db, p, m.cadet_id, write=True)
    m.is_archived = True
    m.archived_at = utcnow()
    db.commit()
    audit(db, p, object_type="cadet_class_membership", object_id=m.id, action="archive")
    return {"ok": True}


@router.get("/training-classes/{cid}/members")
def list_training_class_members(
    cid: str, db: DBSession = Depends(get_db), p: Principal = Depends(get_principal),
):
    if p.role == "sqn_general":
        raise HTTPException(403, detail={"error": "forbidden"})
    tc = db.get(TrainingClass, cid)
    if not tc:
        raise HTTPException(404, detail={"error": "not_found"})
    s = db.get(Squadron, tc.squadron_id)
    require_can_view_squadron(p, tc.squadron_id, s.wing_id if s else None)

    rows = db.query(CadetClassMembership, Cadet).join(Cadet, CadetClassMembership.cadet_id == Cadet.id).filter(
        CadetClassMembership.training_class_id == cid,
        CadetClassMembership.is_archived == False,  # noqa: E712
        CadetClassMembership.active_status == True,  # noqa: E712
        Cadet.is_archived == False,  # noqa: E712
    ).all()
    return [
        {
            "membership_id": m.id, "cadet_id": c.id,
            "rank": c.rank, "first_name": c.first_name, "last_name": c.last_name,
            "start_date": m.start_date,
        }
        for m, c in rows
    ]


# ── PARADE NIGHT BUILDER ────────────────────────────────────────────────────
@router.get("/parade-nights/{pnid}/builder")
def parade_night_builder(pnid: str, db: DBSession = Depends(get_db), p: Principal = Depends(get_principal)):
    """Return timing template blocks + sessions grid (cadet_group × period_number) for Night Builder."""
    pn = db.get(ParadeNight, pnid)
    if not pn:
        raise HTTPException(404, detail={"error": "not_found"})
    require_can_view_squadron(p, pn.squadron_id, pn.wing_id)

    # Timing template blocks
    timing_blocks: list[dict] = []
    session_count = pn.session_count or 3
    tmpl = None
    if pn.timing_template_id:
        tmpl = db.get(TimingTemplate, pn.timing_template_id)
    if not tmpl:
        tmpl = _effective_template(db, pn.squadron_id, pn.date)
    if tmpl:
        blocks = db.query(TimingBlock).filter(
            TimingBlock.timing_template_id == tmpl.id,
        ).order_by(TimingBlock.display_order).all()
        timing_blocks = [
            {
                "display_order": b.display_order, "block_name": b.block_name,
                "block_type": b.block_type, "start_time": b.start_time, "end_time": b.end_time,
                "duration_minutes": b.duration_minutes,
                "is_instructional_period": b.is_instructional_period,
                "period_number": b.period_number,
            }
            for b in blocks
        ]
        ip_count = sum(1 for b in blocks if b.is_instructional_period)
        if ip_count > 0:
            session_count = ip_count

    sessions = db.query(Session).filter(
        Session.parade_night_id == pnid,
        Session.is_archived == False,  # noqa: E712
    ).order_by(Session.period_number, Session.cadet_group).all()

    return {
        "parade_night_id": pnid,
        "parade_date": pn.date,
        "parade_type": pn.parade_type,
        "squadron_id": pn.squadron_id,
        "session_count": session_count,
        "timing_template_id": tmpl.id if tmpl else None,
        "timing_blocks": timing_blocks,
        "cadet_groups": ["orientation", "initial", "junior", "intermediate", "senior"],
        "sessions": [_sess_dict(s) for s in sessions],
    }


# ── PARADE NIGHT DELETE ──────────────────────────────────────────────────────
@router.delete("/parade-nights/{pnid}")
def delete_parade(pnid: str, db: DBSession = Depends(get_db), p: Principal = Depends(get_principal)):
    pn = db.get(ParadeNight, pnid)
    if not pn:
        raise HTTPException(404, detail={"error": "not_found"})
    require_can_write_squadron(p, pn.squadron_id, pn.wing_id)
    pn.is_archived = True
    pn.archived_at = utcnow()
    db.query(Session).filter(Session.parade_night_id == pn.id).update(
        {"is_archived": True}, synchronize_session=False)
    db.commit()
    audit(db, p, object_type="parade_night", object_id=pn.id, action="archive")
    return {"ok": True}


@router.post("/parade-nights/{pnid}/copy-sessions-to/{target_pnid}")
def copy_sessions_to_parade_night(
    pnid: str, target_pnid: str,
    db: DBSession = Depends(get_db), p: Principal = Depends(get_principal)
):
    """Copy planning fields from all sessions in source PN to target PN.

    Outcomes, publication state, and attendance data are not copied — only
    the planning intent (curriculum item, facilitator, room, class audience,
    period, custom title, cadet group). Sessions that already exist at the
    same period in the target PN are skipped, not overwritten.
    """
    import uuid as _uuid
    source = db.get(ParadeNight, pnid)
    if not source or source.is_archived:
        raise HTTPException(404, detail={"error": "source_not_found"})
    require_can_write_squadron(p, source.squadron_id, source.wing_id)
    target = db.get(ParadeNight, target_pnid)
    if not target or target.is_archived:
        raise HTTPException(404, detail={"error": "target_not_found"})
    require_can_write_squadron(p, target.squadron_id, target.wing_id)
    if source.squadron_id != target.squadron_id:
        raise HTTPException(400, detail={"error": "cross_squadron_copy_not_permitted"})

    source_sessions = (
        db.query(Session)
        .filter(Session.parade_night_id == pnid, Session.is_archived.is_(False))
        .all()
    )
    existing_periods = {
        s.period_number
        for s in db.query(Session).filter(
            Session.parade_night_id == target_pnid, Session.is_archived.is_(False)
        )
    }

    copied, skipped = 0, 0
    for src in source_sessions:
        if src.period_number in existing_periods:
            skipped += 1
            continue
        new_sess = Session(
            id=str(_uuid.uuid4()),
            parade_night_id=target_pnid,
            squadron_id=target.squadron_id,
            period_number=src.period_number,
            curriculum_item_id=src.curriculum_item_id,
            curriculum_code_at_time=src.curriculum_code_at_time,
            curriculum_title_at_time=src.curriculum_title_at_time,
            custom_title=src.custom_title,
            phase_at_time=src.phase_at_time,
            element_at_time=src.element_at_time,
            facilitator_id=src.facilitator_id,
            facilitator_rank_at_time=src.facilitator_rank_at_time,
            facilitator_display_name_at_time=src.facilitator_display_name_at_time,
            training_area_id=src.training_area_id,
            training_area_name_at_time=src.training_area_name_at_time,
            cadet_group=src.cadet_group,
            status="planned",
        )
        db.add(new_sess)
        db.flush()
        # Copy class audience (SessionAudience rows)
        src_audience = db.query(SessionAudience).filter(
            SessionAudience.session_id == src.id
        ).all()
        for sa in src_audience:
            db.add(SessionAudience(
                id=str(_uuid.uuid4()),
                session_id=new_sess.id,
                training_class_id=sa.training_class_id,
            ))
        existing_periods.add(src.period_number)
        copied += 1

    db.commit()
    audit(db, p, object_type="parade_night", object_id=target_pnid,
          action="copy_sessions_from", new={"source_parade_night_id": pnid, "copied": copied})
    return {"ok": True, "copied": copied, "skipped": skipped}


# ── FACILITATOR DELETE ───────────────────────────────────────────────────────
@router.delete("/facilitators/{fid}")
def delete_fac(fid: str, db: DBSession = Depends(get_db), p: Principal = Depends(get_principal)):
    f = db.get(Facilitator, fid)
    if not f:
        raise HTTPException(404, detail={"error": "not_found"})
    require_can_write_squadron(p, f.squadron_id, f.wing_id)
    f.is_archived = True
    f.archived_at = utcnow()
    db.commit()
    audit(db, p, object_type="facilitator", object_id=f.id, action="archive")
    return {"ok": True}


@router.post("/facilitators/{fid}/restore")
def restore_fac(fid: str, db: DBSession = Depends(get_db), p: Principal = Depends(get_principal)):
    """REM-133: facilitator archive existed with no restore counterpart --
    same pattern as restore_flight (accounts.py), the direct template for
    this exact gap."""
    f = db.get(Facilitator, fid)
    if not f:
        raise HTTPException(404, detail={"error": "not_found"})
    require_can_write_squadron(p, f.squadron_id, f.wing_id)
    if not f.is_archived:
        raise HTTPException(409, detail={"error": "not_archived"})
    f.is_archived = False
    f.archived_at = None
    db.commit()
    audit(db, p, object_type="facilitator", object_id=f.id, action="restore")
    return {"ok": True}


# ── FACILITATOR MERGE ────────────────────────────────────────────────────────
class FacAbsorbIn(BaseModel):
    source_id: str


@router.post("/facilitators/{fid}/absorb")
def absorb_fac(fid: str, body: FacAbsorbIn, db: DBSession = Depends(get_db),
               p: Principal = Depends(get_principal)):
    """Merge source facilitator into target (fid): reassign all session references
    and leave records, then archive the source.  Both must belong to the same
    squadron and neither may already be archived."""
    from ..models.planning import PlanningFacilitatorLeave
    if p.role in _WRITE_BLOCKED:
        raise HTTPException(403, detail={"error": "forbidden"})
    target = db.get(Facilitator, fid)
    if not target or target.is_archived:
        raise HTTPException(404, detail={"error": "not_found"})
    source = db.get(Facilitator, body.source_id)
    if not source or source.is_archived:
        raise HTTPException(404, detail={"error": "source_not_found",
                                         "message": "Source facilitator not found or already archived."})
    if target.id == source.id:
        raise HTTPException(400, detail={"error": "same_facilitator",
                                         "message": "Source and target must be different facilitators."})
    require_can_write_squadron(p, target.squadron_id, target.wing_id)
    if source.squadron_id != target.squadron_id:
        raise HTTPException(400, detail={"error": "cross_squadron",
                                         "message": "Both facilitators must belong to the same squadron."})
    moved = (db.query(Session)
             .filter(Session.facilitator_id == source.id)
             .update({"facilitator_id": target.id}, synchronize_session=False))
    db.query(Session).filter(Session.assistant_facilitator_id == source.id).update(
        {"assistant_facilitator_id": target.id}, synchronize_session=False)
    db.query(Session).filter(Session.backup_facilitator_id == source.id).update(
        {"backup_facilitator_id": target.id}, synchronize_session=False)
    db.query(PlanningFacilitatorLeave).filter(
        PlanningFacilitatorLeave.facilitator_id == source.id
    ).update({"facilitator_id": target.id}, synchronize_session=False)
    source.is_archived = True
    source.archived_at = utcnow()
    db.commit()
    audit(db, p, object_type="facilitator", object_id=target.id, action="absorb",
          reason=f"absorbed source={source.id} sessions_moved={moved}")
    audit(db, p, object_type="facilitator", object_id=source.id, action="archive",
          reason=f"merged into target={target.id}")
    return {"ok": True, "target_id": target.id, "source_id": source.id, "sessions_moved": moved}


# ── TRAINING AREA WRITE ──────────────────────────────────────────────────────
class TrainingAreaIn(BaseModel):
    name: str
    type: str | None = None
    capacity: int | None = None
    indoor_outdoor: str | None = None
    notes: str | None = None


class TrainingAreaUpdateIn(BaseModel):
    name: str | None = None
    type: str | None = None
    capacity: int | None = None
    indoor_outdoor: str | None = None
    notes: str | None = None


_WRITE_BLOCKED = ("sqn_general", "wing_viewer", "national_viewer", "auditor")


@router.post("/training-areas")
def create_room(body: TrainingAreaIn, db: DBSession = Depends(get_db), p: Principal = Depends(get_principal)):
    if p.role in _WRITE_BLOCKED:
        raise HTTPException(403, detail={"error": "forbidden"})
    sq_id = _active_squadron(p)
    if not sq_id:
        require_can_write_squadron(p, "none", None)
    s = db.get(Squadron, sq_id)
    if not s:
        raise HTTPException(400, detail={"error": "no_squadron_scope"})
    require_can_write_squadron(p, s.id, s.wing_id)
    r = TrainingArea(squadron_id=s.id, name=body.name, type=body.type,
                     capacity=body.capacity, indoor_outdoor=body.indoor_outdoor, notes=body.notes)
    db.add(r)
    db.commit()
    audit(db, p, object_type="training_area", object_id=r.id, action="create")
    return {"ok": True, "training_area_id": r.id}


@router.patch("/training-areas/{rid}")
def update_room(rid: str, body: TrainingAreaUpdateIn, db: DBSession = Depends(get_db),
                p: Principal = Depends(get_principal)):
    r = db.get(TrainingArea, rid)
    if not r:
        raise HTTPException(404, detail={"error": "not_found"})
    s = db.get(Squadron, r.squadron_id)
    require_can_write_squadron(p, r.squadron_id, s.wing_id if s else None)
    if body.name is not None:
        r.name = body.name
    if body.type is not None:
        r.type = body.type
    if body.capacity is not None:
        r.capacity = body.capacity
    if body.indoor_outdoor is not None:
        r.indoor_outdoor = body.indoor_outdoor
    if body.notes is not None:
        r.notes = body.notes
    db.commit()
    audit(db, p, object_type="training_area", object_id=r.id, action="update")
    return {"ok": True, "training_area_id": r.id}


@router.delete("/training-areas/{rid}")
def delete_room(rid: str, db: DBSession = Depends(get_db), p: Principal = Depends(get_principal)):
    r = db.get(TrainingArea, rid)
    if not r:
        raise HTTPException(404, detail={"error": "not_found"})
    s = db.get(Squadron, r.squadron_id)
    require_can_write_squadron(p, r.squadron_id, s.wing_id if s else None)
    r.is_archived = True
    r.archived_at = utcnow()
    db.commit()
    audit(db, p, object_type="training_area", object_id=r.id, action="archive")
    return {"ok": True}


@router.post("/training-areas/{rid}/restore")
def restore_room(rid: str, db: DBSession = Depends(get_db), p: Principal = Depends(get_principal)):
    """REM-133: training area archive existed with no restore counterpart."""
    r = db.get(TrainingArea, rid)
    if not r:
        raise HTTPException(404, detail={"error": "not_found"})
    s = db.get(Squadron, r.squadron_id)
    require_can_write_squadron(p, r.squadron_id, s.wing_id if s else None)
    if not r.is_archived:
        raise HTTPException(409, detail={"error": "not_archived"})
    r.is_archived = False
    r.archived_at = None
    db.commit()
    audit(db, p, object_type="training_area", object_id=r.id, action="restore")
    return {"ok": True}


# ── TRAINING CLASSES ─────────────────────────────────────────────────────────
# CLASS-01 (docs/remediation/master_gap_register.csv): a Squadron-specific
# local group undertaking a Training Stage (CurriculumPhase) during a
# Training Year. See docs/product-review/parallel-class-impact-analysis.md
# for the full design rationale. This is the base CRUD surface only --
# Session<->TrainingClass audience linkage (CLASS-03), Mission Backlog/
# dashboard class-awareness (CLASS-04..07), and bulk setup (addendum §36)
# are separate, dependent follow-up work, not built in this pass.

def _training_class_dict(c: TrainingClass) -> dict:
    return {
        "training_class_id": c.id,
        "squadron_id": c.squadron_id,
        "training_year_id": c.training_year_id,
        "training_stage_id": c.training_stage_id,
        "stage_code": c.stage_code,
        "display_name": c.display_name,
        "sequence": c.sequence,
        "start_date": c.start_date,
        "end_date": c.end_date,
        "expected_count": c.expected_count,
        "notes": c.notes,
        "is_archived": c.is_archived,
        "version": c.version,
    }


def _require_visible_stage(db: DBSession, p: Principal, training_stage_id: str) -> CurriculumPhase:
    """A Training Class's stage must be a CurriculumPhase actually visible to
    this principal (national/own-wing/own-squadron), not any row in the
    catalogue -- reuses the exact same visibility rule _visible_phases()
    already applies to /api/curriculum/phases, so a class can never be
    silently linked to a Stage the caller cannot see."""
    visible_ids = {ph.id for ph in _visible_phases(db, p)}
    if training_stage_id not in visible_ids:
        raise HTTPException(400, detail={"error": "training_stage_not_visible"})
    stage = db.get(CurriculumPhase, training_stage_id)
    if not stage:
        raise HTTPException(404, detail={"error": "training_stage_not_found"})
    return stage


def _require_own_year(db: DBSession, sq_id: str, training_year_id: str) -> PlanningYear:
    year = db.get(PlanningYear, training_year_id)
    if not year or year.unit_id != sq_id:
        raise HTTPException(400, detail={"error": "training_year_not_found_for_squadron"})
    return year


class TrainingClassIn(BaseModel):
    training_year_id: str
    training_stage_id: str
    display_name: str = Field(max_length=80)
    stage_code: str | None = None  # ORI|INI|JNR|INT|SNR
    sequence: int = 0
    start_date: str | None = None
    end_date: str | None = None
    expected_count: int | None = None
    notes: str | None = None


class TrainingClassUpdateIn(BaseModel):
    display_name: str | None = Field(default=None, max_length=80)
    training_stage_id: str | None = None
    stage_code: str | None = None  # ORI|INI|JNR|INT|SNR
    sequence: int | None = None
    start_date: str | None = None
    end_date: str | None = None
    expected_count: int | None = None
    notes: str | None = None
    client_version: int | None = None


@router.get("/training-classes")
def list_training_classes(squadron_id: str | None = None, training_year_id: str | None = None,
                           include_archived: bool = False,
                           db: DBSession = Depends(get_db), p: Principal = Depends(get_principal)):
    sq_id = _view_squadron_id(p, squadron_id, db)
    if not sq_id:
        return []
    q = db.query(TrainingClass).filter(TrainingClass.squadron_id == sq_id)
    if training_year_id:
        q = q.filter(TrainingClass.training_year_id == training_year_id)
    if not include_archived:
        q = q.filter(TrainingClass.is_archived == False)  # noqa: E712
    rows = q.order_by(TrainingClass.sequence, TrainingClass.display_name).all()
    return [_training_class_dict(c) for c in rows]


@router.post("/training-classes")
def create_training_class(body: TrainingClassIn, db: DBSession = Depends(get_db),
                           p: Principal = Depends(get_principal)):
    if p.role in _WRITE_BLOCKED:
        raise HTTPException(403, detail={"error": "forbidden"})
    sq_id = _active_squadron(p)
    if not sq_id:
        require_can_write_squadron(p, "none", None)
    s = db.get(Squadron, sq_id)
    if not s:
        raise HTTPException(400, detail={"error": "no_squadron_scope"})
    require_can_write_squadron(p, s.id, s.wing_id)
    _require_own_year(db, s.id, body.training_year_id)
    _require_visible_stage(db, p, body.training_stage_id)
    if body.stage_code and body.stage_code not in STAGE_CODES:
        raise HTTPException(400, detail={"error": "invalid_stage_code",
                                         "valid": sorted(STAGE_CODES)})
    c = TrainingClass(
        squadron_id=s.id, training_year_id=body.training_year_id,
        training_stage_id=body.training_stage_id, display_name=body.display_name,
        stage_code=body.stage_code,
        sequence=body.sequence, start_date=body.start_date, end_date=body.end_date,
        expected_count=body.expected_count, notes=body.notes,
        created_by=p.user_id, updated_by=p.user_id,
    )
    db.add(c)
    db.commit()
    audit(db, p, object_type="training_class", object_id=c.id, action="create")
    return {"ok": True, "training_class_id": c.id}


@router.patch("/training-classes/{cid}")
def update_training_class(cid: str, body: TrainingClassUpdateIn, db: DBSession = Depends(get_db),
                           p: Principal = Depends(get_principal)):
    c = db.get(TrainingClass, cid)
    if not c:
        raise HTTPException(404, detail={"error": "not_found"})
    s = db.get(Squadron, c.squadron_id)
    require_can_write_squadron(p, c.squadron_id, s.wing_id if s else None)
    _check_version(c, body.client_version)
    if body.training_stage_id is not None:
        _require_visible_stage(db, p, body.training_stage_id)
        c.training_stage_id = body.training_stage_id
    if body.stage_code is not None:
        if body.stage_code not in STAGE_CODES:
            raise HTTPException(400, detail={"error": "invalid_stage_code",
                                             "valid": sorted(STAGE_CODES)})
        c.stage_code = body.stage_code
    if body.display_name is not None:
        c.display_name = body.display_name
    if body.sequence is not None:
        c.sequence = body.sequence
    if body.start_date is not None:
        c.start_date = body.start_date
    if body.end_date is not None:
        c.end_date = body.end_date
    if body.expected_count is not None:
        c.expected_count = body.expected_count
    if body.notes is not None:
        c.notes = body.notes
    c.version += 1
    c.updated_by = p.user_id
    db.commit()
    audit(db, p, object_type="training_class", object_id=c.id, action="update")
    return {"ok": True, "training_class_id": c.id, "version": c.version}


@router.delete("/training-classes/{cid}")
def archive_training_class(cid: str, db: DBSession = Depends(get_db), p: Principal = Depends(get_principal)):
    c = db.get(TrainingClass, cid)
    if not c:
        raise HTTPException(404, detail={"error": "not_found"})
    s = db.get(Squadron, c.squadron_id)
    require_can_write_squadron(p, c.squadron_id, s.wing_id if s else None)
    c.is_archived = True
    c.archived_at = utcnow()
    db.commit()
    audit(db, p, object_type="training_class", object_id=c.id, action="archive")
    return {"ok": True}


@router.post("/training-classes/{cid}/restore")
def restore_training_class(cid: str, db: DBSession = Depends(get_db), p: Principal = Depends(get_principal)):
    """CLASS-10: direct counterpart to archive_training_class, following the
    same archive/restore template as restore_flight (accounts.py) and
    restore_squadron (organisations.py) -- needed because merge-into (below)
    makes archiving a routine, expected outcome of a normal operation, not a
    rare edge case, so an accidental merge needs an undo path."""
    c = db.get(TrainingClass, cid)
    if not c:
        raise HTTPException(404, detail={"error": "not_found"})
    s = db.get(Squadron, c.squadron_id)
    require_can_write_squadron(p, c.squadron_id, s.wing_id if s else None)
    if not c.is_archived:
        raise HTTPException(409, detail={"error": "not_archived"})
    c.is_archived = False
    c.archived_at = None
    c.updated_by = p.user_id
    db.commit()
    audit(db, p, object_type="training_class", object_id=c.id, action="restore")
    return {"ok": True}


# ── TRAINING CLASS SPLIT/MERGE (CLASS-10) ───────────────────────────────────
# addendum §62/§63: split and merge must preserve historical delivered-session
# data. reassign_class_members is the one primitive both split (compose with
# the existing POST /training-classes create + this endpoint) and merge
# (merge_training_class, below) build on. It only ever mutates
# CadetClassMembership -- a lifecycle join (start_date/end_date/active_status)
# designed for exactly this kind of move, see the model's own docstring --
# and never touches SessionAudience (which class(es) a Session was actually
# delivered to: a point-in-time historical fact with no "at-time" snapshot
# field, unlike Session's facilitator snapshot fields). This is deliberately
# NOT the Facilitator absorb() pattern (blind FK reassignment of every child
# row onto the survivor) -- copying that here would silently rewrite which
# class a past, already-delivered session belonged to.

class ReassignClassMembersIn(BaseModel):
    cadet_ids: list[str]
    to_training_class_id: str
    effective_date: Optional[str] = None
    reason: Optional[str] = None


def _require_same_squadron_stage_year(db: DBSession, p: Principal, from_c: TrainingClass,
                                       to_c: TrainingClass) -> Squadron:
    if from_c.squadron_id != to_c.squadron_id:
        raise HTTPException(400, detail={"error": "cross_squadron",
                                         "message": "Both classes must belong to the same squadron."})
    if from_c.training_stage_id != to_c.training_stage_id:
        raise HTTPException(400, detail={"error": "cross_stage",
                                         "message": "Both classes must be undertaking the same Training Stage."})
    if from_c.training_year_id != to_c.training_year_id:
        raise HTTPException(400, detail={"error": "cross_year",
                                         "message": "Both classes must belong to the same Training Year."})
    s = db.get(Squadron, from_c.squadron_id)
    require_can_write_squadron(p, from_c.squadron_id, s.wing_id if s else None)
    return s


def _reassign_members(db: DBSession, p: Principal, from_c: TrainingClass, to_c: TrainingClass,
                       cadet_ids: list[str], effective_date: str | None,
                       source_label: str) -> tuple[list[str], list[str]]:
    moved, skipped = [], []
    for cadet_id in cadet_ids:
        m = db.query(CadetClassMembership).filter(
            CadetClassMembership.cadet_id == cadet_id,
            CadetClassMembership.training_class_id == from_c.id,
            CadetClassMembership.is_archived == False,  # noqa: E712
            CadetClassMembership.active_status == True,  # noqa: E712
        ).first()
        if not m:
            skipped.append(cadet_id)
            continue
        m.end_date = effective_date
        m.active_status = False
        m.version += 1
        m.updated_by = p.user_id
        new_m = CadetClassMembership(
            cadet_id=cadet_id, training_class_id=to_c.id,
            start_date=effective_date, source=source_label,
            created_by=p.user_id, updated_by=p.user_id,
        )
        db.add(new_m)
        moved.append(cadet_id)
    return moved, skipped


@router.post("/training-classes/{from_id}/reassign-members")
def reassign_class_members(from_id: str, body: ReassignClassMembersIn,
                            db: DBSession = Depends(get_db), p: Principal = Depends(get_principal)):
    """Core split/move primitive: ends each named cadet's active membership in
    from_id and starts a new one in to_training_class_id, effective
    body.effective_date. Never touches SessionAudience."""
    if p.role in _WRITE_BLOCKED:
        raise HTTPException(403, detail={"error": "forbidden"})
    from_c = db.get(TrainingClass, from_id)
    if not from_c:
        raise HTTPException(404, detail={"error": "not_found"})
    to_c = db.get(TrainingClass, body.to_training_class_id)
    if not to_c or to_c.is_archived:
        raise HTTPException(400, detail={"error": "invalid_target_class"})
    if from_c.id == to_c.id:
        raise HTTPException(400, detail={"error": "same_class",
                                         "message": "Source and target must be different classes."})
    _require_same_squadron_stage_year(db, p, from_c, to_c)
    if not body.cadet_ids:
        raise HTTPException(400, detail={"error": "no_cadets"})

    moved, skipped = _reassign_members(db, p, from_c, to_c, body.cadet_ids, body.effective_date, "reassigned")
    db.commit()
    audit(db, p, object_type="training_class", object_id=to_c.id, action="reassign_members",
          reason=body.reason,
          new={"from_training_class_id": from_c.id, "to_training_class_id": to_c.id,
               "moved": moved, "skipped": skipped})
    return {"ok": True, "moved": moved, "skipped": skipped}


class MergeTrainingClassIn(BaseModel):
    target_training_class_id: str


@router.post("/training-classes/{source_id}/merge-into")
def merge_training_class(source_id: str, body: MergeTrainingClassIn,
                          db: DBSession = Depends(get_db), p: Principal = Depends(get_principal)):
    """Move every active member of source into target via the reassign-members
    primitive, then archive source (restorable via restore_training_class
    above). SessionAudience rows referencing source are left completely
    untouched -- a past session's delivered-to class list is historical fact,
    not something a later merge may rewrite (addendum §62/§63)."""
    if p.role in _WRITE_BLOCKED:
        raise HTTPException(403, detail={"error": "forbidden"})
    source = db.get(TrainingClass, source_id)
    if not source or source.is_archived:
        raise HTTPException(404, detail={"error": "not_found"})
    target = db.get(TrainingClass, body.target_training_class_id)
    if not target or target.is_archived:
        raise HTTPException(400, detail={"error": "invalid_target_class"})
    if source.id == target.id:
        raise HTTPException(400, detail={"error": "same_class",
                                         "message": "Source and target must be different classes."})
    _require_same_squadron_stage_year(db, p, source, target)

    cadet_ids = [m.cadet_id for m in db.query(CadetClassMembership).filter(
        CadetClassMembership.training_class_id == source.id,
        CadetClassMembership.is_archived == False,  # noqa: E712
        CadetClassMembership.active_status == True,  # noqa: E712
    ).all()]
    moved, skipped = _reassign_members(db, p, source, target, cadet_ids, None, "merged")

    source.is_archived = True
    source.archived_at = utcnow()
    db.commit()
    audit(db, p, object_type="training_class", object_id=target.id, action="merge_members_in",
          reason=f"merged source={source.id} cadets_moved={len(moved)}",
          new={"source_training_class_id": source.id, "moved": moved, "skipped": skipped})
    audit(db, p, object_type="training_class", object_id=source.id, action="archive",
          reason=f"merged into target={target.id}")
    return {"ok": True, "target_id": target.id, "source_id": source.id, "moved": moved, "skipped": skipped}


# ── CLASS-SPECIFIC CURRICULUM PROGRESS ──────────────────────────────────────
# CLASS-04: curriculum progress derived PER Training Class, not blended
# across every class sharing a Training Stage. Derived entirely from
# existing operational data (CurriculumItem, Session, SessionAudience) --
# no second, manually-maintained progress database (addendum §44). One
# recorded fact (a Session's status, or its per-class outcome_override)
# drives this read model; nothing here is itself written to directly.
_ITEM_STATUS_PRIORITY = ["delivered", "delivered_with_issue", "not_delivered",
                         "cancelled", "planned", "rescheduled"]


def _class_curriculum_progress(db: DBSession, c: TrainingClass) -> dict:
    stage = db.get(CurriculumPhase, c.training_stage_id)
    stage_name = stage.name if stage else None
    s = db.get(Squadron, c.squadron_id)
    wing_id = s.wing_id if s else None

    from sqlalchemy import or_
    conditions = [CurriculumItem.owning_level == "national"]
    if wing_id:
        conditions.append((CurriculumItem.owning_level == "wing") & (CurriculumItem.wing_id == wing_id))
    conditions.append(CurriculumItem.squadron_id == c.squadron_id)
    items = db.query(CurriculumItem).filter(
        CurriculumItem.is_archived == False,  # noqa: E712
        CurriculumItem.phase == stage_name,
        or_(*conditions),
    ).order_by(CurriculumItem.recommended_sequence).all()

    # One query for every Session linked to this class via SessionAudience,
    # joined back to the Session row for its curriculum_item_id/status.
    linked = (
        db.query(SessionAudience, Session)
        .join(Session, SessionAudience.session_id == Session.id)
        .filter(SessionAudience.training_class_id == c.id, Session.is_archived == False)  # noqa: E712
        .all()
    )
    from collections import defaultdict
    by_item: dict[str, list[dict]] = defaultdict(list)
    for aud, sess in linked:
        if not sess.curriculum_item_id:
            continue
        effective = aud.outcome_override or sess.status or "planned"
        by_item[sess.curriculum_item_id].append({"session_id": sess.id, "status": effective})

    requirements = []
    summary = {"total": 0, "delivered": 0, "planned": 0, "not_delivered": 0, "cancelled": 0, "not_started": 0}
    for item in items:
        sessions = by_item.get(item.id, [])
        if not sessions:
            status = "not_started"
        else:
            present = {row["status"] for row in sessions}
            status = next((cand for cand in _ITEM_STATUS_PRIORITY if cand in present), "planned")
        bucket = "delivered" if status in ("delivered", "delivered_with_issue") else status
        summary["total"] += 1
        summary[bucket] = summary.get(bucket, 0) + 1
        requirements.append({
            "curriculum_id": item.id, "code": item.code, "title": item.title,
            "status": status, "sessions": sessions,
        })

    return {
        "training_class_id": c.id,
        "training_stage_id": c.training_stage_id,
        "stage_name": stage_name,
        "requirements": requirements,
        "summary": summary,
    }


@router.get("/training-classes/{cid}/curriculum-progress")
def get_class_curriculum_progress(cid: str, db: DBSession = Depends(get_db), p: Principal = Depends(get_principal)):
    c = db.get(TrainingClass, cid)
    if not c:
        raise HTTPException(404, detail={"error": "not_found"})
    s = db.get(Squadron, c.squadron_id)
    require_can_view_squadron(p, c.squadron_id, s.wing_id if s else None)
    return _class_curriculum_progress(db, c)


@router.get("/curriculum/class-matrix")
def get_class_matrix(
    year_id: str,
    db: DBSession = Depends(get_db),
    p: Principal = Depends(get_principal),
):
    """CLASS-MATRIX-01: curriculum item × Training Class progress matrix.

    Returns stages with their curriculum items, each item having a per-class cell
    showing delivery status and contributing session IDs for the given planning year.
    """
    from sqlalchemy import or_
    from collections import defaultdict

    py = db.get(PlanningYear, year_id)
    if not py:
        raise HTTPException(404, detail={"error": "planning_year_not_found"})
    sq_id = py.unit_id
    if not sq_id:
        raise HTTPException(400, detail={"error": "year_not_squadron_scoped"})
    sq = db.get(Squadron, sq_id)
    require_can_view_squadron(p, sq_id, sq.wing_id if sq else None)

    # Load all active Training Classes for this year
    classes = db.query(TrainingClass).filter(
        TrainingClass.training_year_id == year_id,
        TrainingClass.is_archived == False,  # noqa: E712
    ).order_by(TrainingClass.sequence, TrainingClass.display_name).all()

    if not classes:
        return {"year_id": year_id, "year": py.year, "classes": [], "stages": []}

    class_ids = [c.id for c in classes]

    # Load stage info for each unique stage used by these classes
    stage_ids = list({c.training_stage_id for c in classes})
    stages_by_id = {
        s.id: s
        for s in db.query(CurriculumPhase).filter(CurriculumPhase.id.in_(stage_ids)).all()
    }

    # Load curriculum items for each relevant stage (same scope logic as _class_curriculum_progress)
    wing_id = sq.wing_id if sq else None
    conditions = [CurriculumItem.owning_level == "national"]
    if wing_id:
        conditions.append((CurriculumItem.owning_level == "wing") & (CurriculumItem.wing_id == wing_id))
    conditions.append(CurriculumItem.squadron_id == sq_id)
    stage_names = [s.name for s in stages_by_id.values() if s.name]
    all_items = db.query(CurriculumItem).filter(
        CurriculumItem.is_archived == False,  # noqa: E712
        CurriculumItem.phase.in_(stage_names),
        or_(*conditions),
    ).order_by(CurriculumItem.recommended_sequence).all()

    # Load all SessionAudience + Session rows for these classes in one joined query
    linked = (
        db.query(SessionAudience, Session)
        .join(Session, SessionAudience.session_id == Session.id)
        .filter(
            SessionAudience.training_class_id.in_(class_ids),
            Session.is_archived == False,  # noqa: E712
        )
        .all()
    )

    # Build lookup: (curriculum_item_id, class_id) -> list of {session_id, status}
    cell_data: dict[tuple[str, str], list[dict]] = defaultdict(list)
    for aud, sess in linked:
        if not sess.curriculum_item_id:
            continue
        effective_status = aud.outcome_override or sess.status or "planned"
        cell_data[(sess.curriculum_item_id, aud.training_class_id)].append({
            "session_id": sess.id,
            "status": effective_status,
        })

    # Map stage_name -> stage_id for grouping items
    stage_name_to_id = {s.name: s.id for s in stages_by_id.values() if s.name}

    # Build per-stage item lists
    stages_out = []
    for stage in sorted(stages_by_id.values(), key=lambda s: (s.name or "")):
        stage_items = [item for item in all_items if item.phase == stage.name]
        # Only include classes that use this stage
        stage_classes = [c for c in classes if c.training_stage_id == stage.id]
        stage_class_ids = {c.id for c in stage_classes}

        items_out = []
        for item in stage_items:
            cells: dict[str, dict] = {}
            for tc in stage_classes:
                sessions = cell_data.get((item.id, tc.id), [])
                if not sessions:
                    status = "not_started"
                else:
                    present = {row["status"] for row in sessions}
                    status = next(
                        (cand for cand in _ITEM_STATUS_PRIORITY if cand in present),
                        "planned",
                    )
                cells[tc.id] = {
                    "status": status,
                    "session_ids": [r["session_id"] for r in sessions],
                    "elements_delivered": sum(
                        1 for r in sessions
                        if r["status"] in ("delivered", "delivered_with_issue")
                    ),
                    "elements_total": len(sessions),
                }
            items_out.append({
                "curriculum_id": item.id,
                "code": item.code or "",
                "title": item.title,
                "cells": cells,
            })

        stages_out.append({
            "stage_id": stage.id,
            "stage_name": stage.name or "",
            "class_ids": [c.id for c in stage_classes],
            "items": items_out,
        })

    classes_out = [
        {
            "class_id": c.id,
            "display_name": c.display_name,
            "stage_id": c.training_stage_id,
            "stage_name": stages_by_id.get(c.training_stage_id, None) and stages_by_id[c.training_stage_id].name or "",
        }
        for c in classes
    ]

    return {
        "year_id": year_id,
        "year": py.year,
        "classes": classes_out,
        "stages": stages_out,
    }


@router.get("/curriculum/phases/{stage_id}/class-progress")
def get_stage_class_progress(stage_id: str, squadron_id: str, db: DBSession = Depends(get_db),
                              p: Principal = Depends(get_principal)):
    """Stage-level aggregate across every active class undertaking it, using
    weighted requirement completion (sum of delivered / sum of applicable
    across classes) -- NOT an average of each class's own percentage, which
    addendum §74 explicitly warns can conceal a struggling class behind a
    healthy-looking blended number. Also returns the per-class breakdown so
    a Squadron dashboard can show both the aggregate and which class, if
    any, needs attention (addendum §75)."""
    sq_id = _view_squadron_id(p, squadron_id, db)
    stage = db.get(CurriculumPhase, stage_id)
    if not stage:
        raise HTTPException(404, detail={"error": "training_stage_not_found"})
    classes = db.query(TrainingClass).filter(
        TrainingClass.squadron_id == sq_id, TrainingClass.training_stage_id == stage_id,
        TrainingClass.is_archived == False,  # noqa: E712
    ).order_by(TrainingClass.sequence, TrainingClass.display_name).all()

    per_class = []
    total_delivered = 0
    total_applicable = 0
    for c in classes:
        prog = _class_curriculum_progress(db, c)
        delivered = prog["summary"]["delivered"]
        total = prog["summary"]["total"]
        total_delivered += delivered
        total_applicable += total
        per_class.append({
            "training_class_id": c.id, "display_name": c.display_name,
            "delivered": delivered, "total": total,
            "coverage_pct": round(100 * delivered / total) if total else None,
        })

    return {
        "training_stage_id": stage_id,
        "stage_name": stage.name,
        "squadron_id": sq_id,
        "class_count": len(classes),
        "total_delivered": total_delivered,
        "total_applicable": total_applicable,
        "coverage_pct": round(100 * total_delivered / total_applicable) if total_applicable else None,
        "classes": per_class,
    }


# ── SESSION AUDIENCE ─────────────────────────────────────────────────────────
# CLASS-03: which Training Class(es) a Session was planned for/delivered to.
# Session.cadet_group stays untouched as a free-text compatibility field
# (addendum §86-87) -- this is the new, structured many-to-many alongside it.
# Unblocks CLASS-04 (class-specific curriculum progress), CLASS-05 (Mission
# Backlog), CLASS-06 (Weekly Program), CLASS-13 (combined-Session exceptions)
# -- none of those consumers are built yet; this pass is the linkage itself.

def _require_session_write(db: DBSession, p: Principal, sid: str) -> Session:
    s = db.get(Session, sid)
    if not s or s.is_archived:
        raise HTTPException(404, detail={"error": "not_found"})
    pn = db.get(ParadeNight, s.parade_night_id)
    require_can_write_squadron(p, s.squadron_id, pn.wing_id if pn else None)
    return s


def _session_audience_dict(row: SessionAudience) -> dict:
    return {
        "session_audience_id": row.id,
        "session_id": row.session_id,
        "training_class_id": row.training_class_id,
        "outcome_override": row.outcome_override,
        "outcome_override_reason": row.outcome_override_reason,
    }


class SessionAudienceSetIn(BaseModel):
    training_class_ids: List[str]
    override_conflict: bool = False


class SessionAudienceOutcomeIn(BaseModel):
    outcome_override: str | None = None
    outcome_override_reason: str | None = None


@router.get("/sessions/{sid}/facilitator-suggestions")
def get_facilitator_suggestions(
    sid: str,
    parade_date: str | None = None,
    period: int | None = None,
    db: DBSession = Depends(get_db),
    p: Principal = Depends(get_principal),
):
    """FAC-SUG-01: ranked, explainable facilitator suggestions for a session.

    Scoring (higher = better match):
      +5  subject_area match with curriculum item
      -1000  on leave for the parade date (shown as UNAVAILABLE)
      -500   already assigned to the same night/period (CONFLICT)
      -workload  sessions already allocated this term (prefer less-loaded)

    Returns [] with a "no_recommendation" flag when no usable data is available.
    """
    from ..models.planning import PlanningFacilitatorLeave, ParadeDate
    from sqlalchemy import func, and_

    s = db.get(Session, sid)
    if not s or s.is_archived:
        raise HTTPException(404, detail={"error": "not_found"})
    pn = db.get(ParadeNight, s.parade_night_id)
    require_can_view_squadron(p, s.squadron_id, pn.wing_id if pn else None)

    sq_id = s.squadron_id
    check_date = parade_date or (pn.date if pn else None)
    check_period = period or s.period_number

    # Curriculum item subject areas
    ci = db.get(CurriculumItem, s.curriculum_item_id) if s.curriculum_item_id else None
    ci_phase = ci.phase if ci else None  # phase name is the subject area category

    # Active facilitators for the squadron
    facs = (
        db.query(Facilitator)
        .filter(
            Facilitator.squadron_id == sq_id,
            Facilitator.active_status.is_(True),
            Facilitator.is_archived.is_(False),
        )
        .all()
    )
    if not facs:
        return {"suggestions": [], "no_recommendation": True, "reason": "No active facilitators found"}

    # On-leave set for the check_date
    on_leave_ids: set[str] = set()
    if check_date:
        leave_rows = (
            db.query(PlanningFacilitatorLeave.facilitator_id)
            .filter(
                PlanningFacilitatorLeave.start_date <= check_date,
                PlanningFacilitatorLeave.end_date >= check_date,
            )
            .all()
        )
        on_leave_ids = {r[0] for r in leave_rows}

    # Conflict set: facilitators already assigned to this period on this night
    conflict_ids: set[str] = set()
    if pn and check_period is not None:
        conflict_rows = (
            db.query(Session.facilitator_id)
            .filter(
                Session.parade_night_id == pn.id,
                Session.period_number == check_period,
                Session.is_archived.is_(False),
                Session.id != sid,
                Session.facilitator_id.isnot(None),
            )
            .all()
        )
        conflict_ids = {r[0] for r in conflict_rows}

    # Workload: sessions allocated this term per facilitator
    term = pn.term if pn else None
    workload: dict[str, int] = {}
    if term and pn:
        # All parade nights in same squadron + term
        term_pn_ids = [
            r[0]
            for r in db.query(ParadeNight.id)
            .filter(
                ParadeNight.squadron_id == sq_id,
                ParadeNight.term == term,
                ParadeNight.is_archived.is_(False),
            )
            .all()
        ]
        if term_pn_ids:
            wl_rows = (
                db.query(Session.facilitator_id, func.count(Session.id))
                .filter(
                    Session.parade_night_id.in_(term_pn_ids),
                    Session.facilitator_id.isnot(None),
                    Session.is_archived.is_(False),
                )
                .group_by(Session.facilitator_id)
                .all()
            )
            workload = {r[0]: r[1] for r in wl_rows}

    suggestions = []
    for f in facs:
        score = 0
        reasons: list[str] = []
        conflicts: list[str] = []

        # Subject area match
        fac_areas = f.subject_areas or []
        subject_match = bool(ci_phase and any(
            ci_phase.lower() in (a or "").lower() or (a or "").lower() in ci_phase.lower()
            for a in fac_areas
        ))
        if subject_match:
            score += 5
            matched = [a for a in fac_areas if ci_phase.lower() in a.lower() or a.lower() in ci_phase.lower()]
            reasons.append(f"Suitable for {matched[0] if matched else ci_phase}")
        elif fac_areas:
            reasons.append(f"Subject areas: {', '.join(fac_areas)}")

        # Leave check
        on_leave = f.id in on_leave_ids
        if on_leave:
            score -= 1000
            conflicts.append("On leave this date")

        # Conflict check
        has_conflict = f.id in conflict_ids
        if has_conflict:
            score -= 500
            conflicts.append(f"Already assigned to period {check_period} on this night")

        # Workload
        wl = workload.get(f.id, 0)
        score -= wl
        reasons.append(f"{wl} session{'s' if wl != 1 else ''} allocated this term")

        # Rank label
        if on_leave:
            rank_label = "UNAVAILABLE"
        elif has_conflict:
            rank_label = "CONFLICT"
        elif subject_match and not on_leave and not has_conflict:
            rank_label = "SUGGESTED"
        else:
            rank_label = "AVAILABLE"

        display = (
            f"{f.current_rank or ''} {f.last_name or ''}".strip()
            if (f.last_name or f.current_rank)
            else f.id
        )

        suggestions.append({
            "facilitator_id": f.id,
            "display_name": display,
            "rank_label": rank_label,
            "subject_areas": fac_areas,
            "reasons": reasons,
            "conflicts": conflicts,
            "workload_sessions_this_term": wl,
            "score": score,
        })

    suggestions.sort(key=lambda x: -x["score"])
    # Remove internal score from output
    for s_item in suggestions:
        del s_item["score"]

    return {"suggestions": suggestions, "no_recommendation": len(suggestions) == 0}


@router.get("/sessions/{sid}/audience")
def list_session_audience(sid: str, db: DBSession = Depends(get_db), p: Principal = Depends(get_principal)):
    s = db.get(Session, sid)
    if not s or s.is_archived:
        raise HTTPException(404, detail={"error": "not_found"})
    pn = db.get(ParadeNight, s.parade_night_id)
    require_can_view_squadron(p, s.squadron_id, pn.wing_id if pn else None)
    rows = db.query(SessionAudience).filter(SessionAudience.session_id == sid).all()
    return [_session_audience_dict(r) for r in rows]


@router.put("/sessions/{sid}/audience")
def set_session_audience(sid: str, body: SessionAudienceSetIn, db: DBSession = Depends(get_db),
                          p: Principal = Depends(get_principal)):
    """Replace the full set of Training Classes this Session targets --
    idempotent set, not append. Selecting Senior 1 + Senior 2 for one
    Session (addendum §40 Example 2) is a single PUT with both IDs; a
    "split Session" (addendum §59.2) is a second Session created via the
    normal POST /sessions plus its own PUT .../audience with the remaining
    class(es), not a special endpoint -- the audience relationship itself
    doesn't need split-specific logic."""
    s = _require_session_write(db, p, sid)
    classes = []
    for cid in body.training_class_ids:
        c = db.get(TrainingClass, cid)
        if not c or c.is_archived or c.squadron_id != s.squadron_id:
            raise HTTPException(400, detail={"error": "invalid_training_class", "training_class_id": cid})
        classes.append(c)

    # MBACK-05: detect class-schedule clashes — a Training Class cannot be in
    # two sessions at the same parade night and period simultaneously.  Follow
    # the same override pattern as facilitator/room clash in create/edit_session.
    if not body.override_conflict and classes and s.period_number is not None:
        sibling_ids = [
            sib.id for sib in db.query(Session).filter(
                Session.parade_night_id == s.parade_night_id,
                Session.period_number == s.period_number,
                Session.id != sid,
                Session.is_archived == False,  # noqa: E712
            ).all()
        ]
        if sibling_ids:
            sib_audiences = db.query(SessionAudience).filter(
                SessionAudience.session_id.in_(sibling_ids)
            ).all()
            sib_class_sessions: dict[str, list[str]] = {}
            for aud in sib_audiences:
                sib_class_sessions.setdefault(aud.training_class_id, []).append(aud.session_id)
            class_conflicts = [
                {"type": "class_clash",
                 "training_class_id": c.id,
                 "training_class_name": c.display_name,
                 "session_ids": sib_class_sessions[c.id]}
                for c in classes if c.id in sib_class_sessions
            ]
            if class_conflicts:
                raise HTTPException(409, detail={"error": "class_conflict", "conflicts": class_conflicts})

    existing = {r.training_class_id: r for r in
                db.query(SessionAudience).filter(SessionAudience.session_id == sid).all()}
    keep_ids = {c.id for c in classes}
    for cid, row in existing.items():
        if cid not in keep_ids:
            db.delete(row)
    for c in classes:
        if c.id not in existing:
            db.add(SessionAudience(session_id=sid, training_class_id=c.id,
                                   created_by=p.user_id, updated_by=p.user_id))
    db.commit()
    audit(db, p, object_type="session_audience", object_id=sid, action="set",
          new={"training_class_ids": sorted(keep_ids)})
    rows = db.query(SessionAudience).filter(SessionAudience.session_id == sid).all()
    return [_session_audience_dict(r) for r in rows]


@router.patch("/sessions/{sid}/audience/{cid}")
def set_session_audience_outcome(sid: str, cid: str, body: SessionAudienceOutcomeIn,
                                  db: DBSession = Depends(get_db), p: Principal = Depends(get_principal)):
    """Record a per-class outcome exception (addendum §48) -- e.g. a
    Session delivered jointly to three classes where one class didn't
    actually receive it. outcome_override=None means "use the parent
    Session's own status", which is also how a caller clears a previously
    recorded exception."""
    _require_session_write(db, p, sid)
    row = db.query(SessionAudience).filter(
        SessionAudience.session_id == sid, SessionAudience.training_class_id == cid,
    ).first()
    if not row:
        raise HTTPException(404, detail={"error": "not_found"})
    if body.outcome_override is not None and body.outcome_override not in VALID_STATUS:
        raise HTTPException(400, detail={"error": "invalid_status"})
    if (body.outcome_override in REASON_REQUIRED_STATUSES
            and not (body.outcome_override_reason or "").strip()):
        raise HTTPException(400, detail={"error": f"reason_required_{body.outcome_override}"})
    row.outcome_override = body.outcome_override
    row.outcome_override_reason = body.outcome_override_reason
    row.updated_by = p.user_id
    db.commit()
    audit(db, p, object_type="session_audience", object_id=row.id, action="outcome_override",
          new={"outcome_override": row.outcome_override})
    return _session_audience_dict(row)


@router.delete("/sessions/{sid}/audience/{cid}")
def remove_session_audience_class(sid: str, cid: str, db: DBSession = Depends(get_db),
                                   p: Principal = Depends(get_principal)):
    _require_session_write(db, p, sid)
    row = db.query(SessionAudience).filter(
        SessionAudience.session_id == sid, SessionAudience.training_class_id == cid,
    ).first()
    if not row:
        raise HTTPException(404, detail={"error": "not_found"})
    db.delete(row)
    db.commit()
    audit(db, p, object_type="session_audience", object_id=row.id, action="remove")
    return {"ok": True}


# ── EQUIPMENT WRITE ──────────────────────────────────────────────────────────
class EquipIn(BaseModel):
    name: str
    type: str | None = None
    quantity: int = 1
    notes: str | None = None


class EquipUpdateIn(BaseModel):
    name: str | None = None
    type: str | None = None
    quantity: int | None = None
    notes: str | None = None


@router.post("/equipment")
def create_equip(body: EquipIn, db: DBSession = Depends(get_db), p: Principal = Depends(get_principal)):
    if p.role in _WRITE_BLOCKED:
        raise HTTPException(403, detail={"error": "forbidden"})
    sq_id = _active_squadron(p)
    if not sq_id:
        require_can_write_squadron(p, "none", None)
    s = db.get(Squadron, sq_id)
    if not s:
        raise HTTPException(400, detail={"error": "no_squadron_scope"})
    require_can_write_squadron(p, s.id, s.wing_id)
    e = Equipment(squadron_id=s.id, name=body.name, type=body.type,
                  quantity=body.quantity, available_quantity=body.quantity, notes=body.notes)
    db.add(e)
    db.commit()
    audit(db, p, object_type="equipment", object_id=e.id, action="create")
    return {"ok": True, "equipment_id": e.id}


@router.patch("/equipment/{eid}")
def update_equip(eid: str, body: EquipUpdateIn, db: DBSession = Depends(get_db),
                 p: Principal = Depends(get_principal)):
    e = db.get(Equipment, eid)
    if not e:
        raise HTTPException(404, detail={"error": "not_found"})
    s = db.get(Squadron, e.squadron_id)
    require_can_write_squadron(p, e.squadron_id, s.wing_id if s else None)
    if body.name is not None:
        e.name = body.name
    if body.type is not None:
        e.type = body.type
    if body.quantity is not None:
        e.quantity = body.quantity
        e.available_quantity = body.quantity
    if body.notes is not None:
        e.notes = body.notes
    db.commit()
    audit(db, p, object_type="equipment", object_id=e.id, action="update")
    return {"ok": True, "equipment_id": e.id}


@router.delete("/equipment/{eid}")
def delete_equip(eid: str, db: DBSession = Depends(get_db), p: Principal = Depends(get_principal)):
    e = db.get(Equipment, eid)
    if not e:
        raise HTTPException(404, detail={"error": "not_found"})
    s = db.get(Squadron, e.squadron_id)
    require_can_write_squadron(p, e.squadron_id, s.wing_id if s else None)
    e.is_archived = True
    e.archived_at = utcnow()
    db.commit()
    audit(db, p, object_type="equipment", object_id=e.id, action="archive")
    return {"ok": True}


@router.post("/equipment/{eid}/restore")
def restore_equip(eid: str, db: DBSession = Depends(get_db), p: Principal = Depends(get_principal)):
    """REM-133: equipment archive existed with no restore counterpart."""
    e = db.get(Equipment, eid)
    if not e:
        raise HTTPException(404, detail={"error": "not_found"})
    s = db.get(Squadron, e.squadron_id)
    require_can_write_squadron(p, e.squadron_id, s.wing_id if s else None)
    if not e.is_archived:
        raise HTTPException(409, detail={"error": "not_archived"})
    e.is_archived = False
    e.archived_at = None
    db.commit()
    audit(db, p, object_type="equipment", object_id=e.id, action="restore")
    return {"ok": True}


# ── ACTIVITIES CRUD ──────────────────────────────────────────────────────────
class ActivityIn(BaseModel):
    activity_name: str
    activity_type: str | None = None
    date_start: str
    date_end: str | None = None
    time_start: str | None = None
    time_end: str | None = None
    location: str | None = None
    audience: list[str] | None = None
    notes: str | None = None


class ActivityUpdateIn(BaseModel):
    activity_name: str | None = None
    activity_type: str | None = None
    date_start: str | None = None
    date_end: str | None = None
    time_start: str | None = None
    time_end: str | None = None
    location: str | None = None
    audience: list[str] | None = None
    notes: str | None = None
    workflow_status: str | None = None
    cea_seq_nr: str | None = None


class ActivityWingIn(ActivityIn):
    wing_id: str | None = None


def _local_override_out(o: ActivityLocalOverride) -> dict:
    return {
        "override_id": o.id, "local_date_start": o.local_date_start, "local_date_end": o.local_date_end,
        "local_time_start": o.local_time_start, "local_time_end": o.local_time_end,
        "local_notes": o.local_notes, "is_hidden": o.is_hidden,
        "updated_by": o.updated_by, "version": o.version,
    }


def _activity_out(a: Activity, p: Principal, view_scope_type: str | None,
                  org_names: dict[str, str] | None = None,
                  local_override: ActivityLocalOverride | None = None) -> dict:
    org_names = org_names or {}
    owning_scope_id = (a.national_id if a.owning_level == "national"
                       else a.wing_id if a.owning_level == "wing" else a.squadron_id)
    is_inherited = (view_scope_type == "squadron" and a.owning_level != "squadron") or \
                   (view_scope_type == "wing" and a.owning_level == "national")
    # REM-103: a local override is only ever meaningful for an inherited row
    # (see the local-override endpoints below, which enforce this at write
    # time too) -- source values are always returned unchanged alongside it,
    # per the addendum's own "display source vs local value side-by-side" ask.
    return {
        "activity_id": a.id, "source": a.owning_level, "owning_level": a.owning_level,
        "owning_scope_id": owning_scope_id,
        "owning_org_label": org_names.get(owning_scope_id, "National") if a.owning_level == "national" else org_names.get(owning_scope_id),
        "wing_id": a.wing_id, "squadron_id": a.squadron_id,
        "activity_name": a.activity_name,
        "activity_type": a.activity_type, "date_start": a.date_start,
        "date_end": a.date_end, "time_start": a.time_start, "time_end": a.time_end,
        "location": a.location, "audience": a.audience or [],
        "notes": a.notes, "workflow_status": a.workflow_status,
        "cea_seq_nr": a.cea_seq_nr, "is_archived": a.is_archived,
        "is_inherited": is_inherited,
        "read_only": not p.can_write_activity(a.owning_level, a.wing_id, a.squadron_id),
        "local_override": _local_override_out(local_override) if local_override else None,
    }


def _activities_visibility_filter(scope_type: str, scope_id: str | None, wing_id_of_scope: str | None):
    """OR-branched inheritance filter -- mirrors CurriculumItem's already-proven
    national/wing/squadron pattern (list_curriculum above) rather than a UNION.
    National activities are always visible; Wing activities are visible at Wing
    scope (that wing) and Squadron scope (that squadron's wing); Squadron
    activities are visible only at that Squadron's own scope -- never
    "inherited upward" and never copied into a sibling squadron's view."""
    from sqlalchemy import or_
    conditions = [Activity.owning_level == "national"]
    if scope_type in ("wing", "squadron") and wing_id_of_scope:
        conditions.append((Activity.owning_level == "wing") & (Activity.wing_id == wing_id_of_scope))
    if scope_type == "squadron" and scope_id:
        conditions.append((Activity.owning_level == "squadron") & (Activity.squadron_id == scope_id))
    return or_(*conditions)


def _cea_source_items(db: DBSession, wing_id: str | None, squadron_id: str | None,
                      date_from: str | None, date_end: str | None, viewer_squadron_id: str | None) -> list[dict]:
    """CEA activities are read-time merged, never copied into `Activity` --
    only classification_status=='classified' rows are shown (needs_review is
    an internal wing-admin review state), excluded if removed from the source
    import or locally hidden for the viewing squadron."""
    from ..models.planning import CeaActivity, ActivityLocalHide
    q = db.query(CeaActivity).filter(
        CeaActivity.classification_status == "classified",
        CeaActivity.is_removed_from_cea == False,  # noqa: E712
        CeaActivity.is_archived == False,  # noqa: E712
    )
    if squadron_id:
        q = q.filter(CeaActivity.unit_id == squadron_id)
    elif wing_id:
        q = q.filter(CeaActivity.wing_id == wing_id)
    else:
        return []  # CEA import is wing-scoped only -- no meaningful "all-national" CEA view
    if date_from:
        q = q.filter((CeaActivity.activity_end_date == None) | (CeaActivity.activity_end_date >= date_from))  # noqa: E711
    if date_end:
        q = q.filter((CeaActivity.activity_start_date == None) | (CeaActivity.activity_start_date <= date_end))  # noqa: E711
    rows = q.order_by(CeaActivity.activity_start_date).limit(2000).all()
    hidden_ids: set[str] = set()
    if viewer_squadron_id:
        hides = db.query(ActivityLocalHide.cea_activity_id).filter(
            ActivityLocalHide.unit_id == viewer_squadron_id, ActivityLocalHide.is_hidden == True).all()  # noqa: E712
        hidden_ids = {h[0] for h in hides}
    out = []
    for c in rows:
        if c.id in hidden_ids:
            continue
        out.append({
            "activity_id": c.id, "source": "cea", "owning_level": None, "owning_scope_id": None,
            "owning_org_label": "CEA", "wing_id": c.wing_id, "squadron_id": c.unit_id,
            "activity_name": c.activity_name, "activity_type": c.activity_type,
            "date_start": c.activity_start_date, "date_end": c.activity_end_date,
            "time_start": c.start_time, "time_end": c.end_time,
            "location": c.location, "audience": [], "notes": c.notes,
            "workflow_status": c.status_name, "cea_seq_nr": c.cea_activity_id,
            "is_archived": False, "is_inherited": True, "read_only": True,
        })
    return out


def _holiday_source_items(db: DBSession, scope_type: str, wing_id: str | None, squadron_id: str | None,
                          date_from: str | None, date_end: str | None) -> list[dict]:
    """Holidays are year-scoped (via PlanningYear), not scope-tagged directly,
    so relevance is resolved by joining through this scope's own planning
    years -- Wing-level years for a Wing view, that Squadron's (and its
    Wing's) years for a Squadron view. National view shows none (no single
    "national planning year" concept exists) -- a deliberate, documented
    simplification, not an oversight."""
    from ..models.planning import HolidayPeriod, PlanningYear
    year_ids: list[str] = []
    if scope_type == "wing" and wing_id:
        year_ids = [y.id for y in db.query(PlanningYear.id).filter(PlanningYear.wing_id == wing_id).all()]
    elif scope_type == "squadron" and squadron_id:
        sq = db.get(Squadron, squadron_id)
        yq = db.query(PlanningYear.id).filter(PlanningYear.unit_id == squadron_id)
        year_ids = [y.id for y in yq.all()]
        if sq and sq.wing_id:
            year_ids += [y.id for y in db.query(PlanningYear.id).filter(PlanningYear.wing_id == sq.wing_id).all()]
    if not year_ids:
        return []
    q = db.query(HolidayPeriod).filter(HolidayPeriod.planning_year_id.in_(year_ids))
    if date_from:
        q = q.filter(HolidayPeriod.end_date >= date_from)
    if date_end:
        q = q.filter(HolidayPeriod.start_date <= date_end)
    rows = q.order_by(HolidayPeriod.start_date).limit(500).all()
    return [{
        "activity_id": h.id, "source": "holiday", "owning_level": None, "owning_scope_id": None,
        "owning_org_label": "Holiday", "wing_id": wing_id, "squadron_id": squadron_id,
        "activity_name": h.name, "activity_type": h.holiday_type,
        "date_start": h.start_date, "date_end": h.end_date,
        "time_start": None, "time_end": None, "location": None, "audience": [],
        "notes": h.notes, "workflow_status": None, "cea_seq_nr": None,
        "is_archived": False, "is_inherited": True, "read_only": True,
    } for h in rows]


_GETTING_HELP_KEY = "activities_getting_help_content"


@router.get("/activities/getting-help")
def get_activities_getting_help(db: DBSession = Depends(get_db), p: Principal = Depends(get_principal)):
    row = db.get(SystemSetting, _GETTING_HELP_KEY)
    return {"content": row.value if row else "", "updated_at": row.updated_at if row else None}


class GettingHelpIn(BaseModel):
    content: str = ""


@router.put("/activities/getting-help")
def update_activities_getting_help(body: GettingHelpIn, db: DBSession = Depends(get_db), p: Principal = Depends(get_principal)):
    require_system_admin(p)
    row = db.get(SystemSetting, _GETTING_HELP_KEY)
    old_value = row.value if row else None
    if row is None:
        row = SystemSetting(key=_GETTING_HELP_KEY)
        db.add(row)
    row.value = body.content
    row.updated_at = utcnow()
    row.updated_by = p.user_id
    db.commit()
    audit(db, p, object_type="system_setting", object_id=_GETTING_HELP_KEY,
          action="getting_help_content_updated", old={"content": old_value}, new={"content": body.content})
    return {"content": row.value, "updated_at": row.updated_at}


@router.get("/activities")
def list_activities(
    squadron_id: str | None = None,
    scope_type: str | None = None,       # national|wing|squadron -- opt-in to the new inheritance-aware shape
    scope_id: str | None = None,
    view: str = "list",                  # list|calendar|upcoming|historical|archived
    sources: str | None = None,          # comma-separated: activity,cea,holiday (default all)
    date_from: str | None = None,
    date_end: str | None = None,
    activity_type: str | None = None,
    status: str | None = None,
    location: str | None = None,
    cea_seq_nr: str | None = None,
    q: str | None = None,
    include_subordinates: bool = False,  # Wing view convenience rollup of descendant squadron-local activities
    db: DBSession = Depends(get_db), p: Principal = Depends(get_principal),
):
    if not scope_type:
        # Backward-compatible path: existing single-squadron behaviour, unchanged.
        sq_id = _view_squadron_id(p, squadron_id, db)
        rows = db.query(Activity).filter(Activity.squadron_id == sq_id,
                                         Activity.is_archived == False).order_by(Activity.date_start).all()  # noqa: E712
        return [_activity_out(a, p, "squadron") for a in rows]

    if scope_type not in ("national", "wing", "squadron"):
        raise HTTPException(400, detail={"error": "invalid_scope_type"})

    wing_id_of_scope: str | None = None
    resolved_squadron_id: str | None = None
    if scope_type == "national":
        require_role(p, *NATIONAL_LEVEL)
    elif scope_type == "wing":
        if not scope_id:
            raise HTTPException(400, detail={"error": "scope_id_required"})
        require_can_view_wing(p, scope_id)
        wing_id_of_scope = scope_id
    else:  # squadron
        if not scope_id:
            raise HTTPException(400, detail={"error": "scope_id_required"})
        s = db.get(Squadron, scope_id)
        if not s:
            raise HTTPException(404, detail={"error": "squadron_not_found"})
        require_can_view_squadron(p, s.id, s.wing_id)
        wing_id_of_scope = s.wing_id
        resolved_squadron_id = s.id

    show_archived = view == "archived"
    a_conditions = _activities_visibility_filter(scope_type, resolved_squadron_id, wing_id_of_scope)
    aq = db.query(Activity).filter(Activity.is_archived == show_archived, a_conditions)
    if scope_type == "wing" and include_subordinates:
        from sqlalchemy import or_
        sqn_ids = [x.id for x in db.query(Squadron.id).filter(Squadron.wing_id == wing_id_of_scope,
                                                               Squadron.is_archived == False).all()]  # noqa: E712
        if sqn_ids:
            aq = db.query(Activity).filter(
                Activity.is_archived == show_archived,
                or_(a_conditions, (Activity.owning_level == "squadron") & (Activity.squadron_id.in_(sqn_ids))))
    if activity_type:
        aq = aq.filter(Activity.activity_type == activity_type)
    if status:
        aq = aq.filter(Activity.workflow_status == status)
    if location:
        aq = aq.filter(Activity.location.ilike(f"%{location}%"))
    if cea_seq_nr:
        aq = aq.filter(Activity.cea_seq_nr == cea_seq_nr)
    if q:
        like = f"%{q}%"
        from sqlalchemy import or_ as _or
        aq = aq.filter(_or(Activity.activity_name.ilike(like), Activity.notes.ilike(like)))
    if date_from:
        aq = aq.filter((Activity.date_end == None) | (Activity.date_end >= date_from))  # noqa: E711
    if date_end:
        aq = aq.filter(Activity.date_start <= date_end)
    activity_rows = aq.order_by(Activity.date_start).limit(2000).all()

    src_set = set((sources or "activity,cea,holiday").split(","))
    truncated = len(activity_rows) >= 2000

    # Batch-resolve owning org labels (Wing/Squadron names) for the response.
    wing_ids = {a.wing_id for a in activity_rows if a.owning_level == "wing" and a.wing_id}
    sqn_ids = {a.squadron_id for a in activity_rows if a.owning_level == "squadron" and a.squadron_id}
    org_names: dict[str, str] = {}
    if wing_ids:
        from ..models import Wing as WingModel
        for w in db.query(WingModel).filter(WingModel.id.in_(wing_ids)).all():
            org_names[w.id] = w.name
    if sqn_ids:
        for s in db.query(Squadron).filter(Squadron.id.in_(sqn_ids)).all():
            org_names[s.id] = s.name

    # REM-103: batch-fetch this viewing squadron's own local overrides for
    # the activities in this page -- only meaningful at squadron scope (a
    # local override always belongs to exactly one squadron's view).
    overrides_by_activity: dict[str, ActivityLocalOverride] = {}
    if scope_type == "squadron" and resolved_squadron_id and activity_rows:
        for o in db.query(ActivityLocalOverride).filter(
            ActivityLocalOverride.squadron_id == resolved_squadron_id,
            ActivityLocalOverride.activity_id.in_([a.id for a in activity_rows]),
        ).all():
            overrides_by_activity[o.activity_id] = o

    # Deliberately NOT hard-filtering is_hidden rows out of the list here,
    # unlike _cea_source_items' own is_hidden handling above -- that CEA hide
    # mechanism has no "show hidden / unhide" surface anywhere in either
    # frontend today (confirmed: zero references to ActivityLocalHide outside
    # the backend), so a hidden CEA item is permanently invisible with no way
    # back. Returning is_hidden as a flag here instead lets the frontend
    # de-emphasise a locally-irrelevant activity while keeping it
    # discoverable and reversible.
    items = [_activity_out(a, p, scope_type, org_names, overrides_by_activity.get(a.id))
             for a in activity_rows] if "activity" in src_set else []
    if "cea" in src_set and not show_archived:
        items += _cea_source_items(db, wing_id_of_scope, resolved_squadron_id, date_from, date_end,
                                   resolved_squadron_id or _active_squadron(p))
        truncated = truncated or len(items) >= 2000
    if "holiday" in src_set and not show_archived:
        items += _holiday_source_items(db, scope_type, wing_id_of_scope, resolved_squadron_id, date_from, date_end)

    now = utcnow().date().isoformat()
    if view == "upcoming":
        items = [i for i in items if (i["date_end"] or i["date_start"] or "") >= now]
    elif view == "historical":
        items = [i for i in items if (i["date_end"] or i["date_start"] or "") < now]

    items.sort(key=lambda i: (i["date_start"] or "", i["time_start"] or "", i["activity_name"] or ""))
    return {"items": items, "total": len(items), "truncated": truncated}


@router.get("/activities/{aid}")
def get_activity(aid: str, db: DBSession = Depends(get_db), p: Principal = Depends(get_principal)):
    a = db.get(Activity, aid)
    if not a:
        raise HTTPException(404, detail={"error": "not_found"})
    if a.owning_level == "national":
        # REM-103 fix: national activities are always visible to every
        # authenticated principal, matching _activities_visibility_filter's
        # own "National activities are always visible" rule (list_activities,
        # above) -- previously gated to NATIONAL_LEVEL roles only here, so a
        # wing/squadron user could see a national activity in their list but
        # get a 403 viewing that exact same activity by ID. This blocked the
        # inherited-view case a squadron admin needs before they can create a
        # local override on it.
        view_scope = "national"
    elif a.owning_level == "wing":
        # REM-103 fix: visible to that wing AND every squadron under it
        # (inherited) -- previously required require_can_view_wing, which
        # every squadron role always fails (can_view_wing has no squadron
        # branch at all), for the same reason as the national case above.
        viewer_wing_id = p.wing_id if p.role in ("sqn_admin", "sqn_general") else None
        if not (p.can_view_wing(a.wing_id) or (viewer_wing_id and viewer_wing_id == a.wing_id)):
            require_can_view_wing(p, a.wing_id)  # always raises here; reuses its exact error message
        view_scope = "wing"
    else:
        s = db.get(Squadron, a.squadron_id) if a.squadron_id else None
        require_can_view_squadron(p, a.squadron_id, s.wing_id if s else None)
        view_scope = "squadron"
    org_names: dict[str, str] = {}
    if a.owning_level == "wing" and a.wing_id:
        from ..models import Wing as WingModel
        w = db.get(WingModel, a.wing_id)
        if w:
            org_names[w.id] = w.name
    elif a.owning_level == "squadron" and a.squadron_id:
        s = db.get(Squadron, a.squadron_id)
        if s:
            org_names[s.id] = s.name
    # REM-103: a squadron-scope caller viewing an inherited (wing/national)
    # activity may have their own local override for it -- view_scope above
    # reflects the ACTIVITY's own owning_level, not the viewer's squadron, so
    # this is resolved separately via the caller's active squadron.
    local_override = None
    viewer_sq_id = _active_squadron(p)
    if viewer_sq_id and a.owning_level != "squadron":
        local_override = db.query(ActivityLocalOverride).filter(
            ActivityLocalOverride.activity_id == a.id,
            ActivityLocalOverride.squadron_id == viewer_sq_id,
        ).first()
    return _activity_out(a, p, view_scope, org_names, local_override)


class ActivityLocalOverrideIn(BaseModel):
    local_date_start: str | None = None
    local_date_end: str | None = None
    local_time_start: str | None = None
    local_time_end: str | None = None
    local_notes: str | None = None
    is_hidden: bool = False
    version: int | None = None  # required on update, ignored on first create


@router.put("/activities/{aid}/local-override")
def upsert_activity_local_override(aid: str, body: ActivityLocalOverrideIn,
                                   db: DBSession = Depends(get_db), p: Principal = Depends(get_principal)):
    """REM-103: create or update the caller's own squadron's local
    date/time/notes/relevance adjustment to a Wing/National-owned inherited
    Activity. The source Activity row is never modified -- this is purely a
    per-squadron annotation, upserted (one row per activity+squadron, see
    the model's unique constraint)."""
    a = db.get(Activity, aid)
    if not a:
        raise HTTPException(404, detail={"error": "not_found"})
    sq_id = _active_squadron(p)
    if not sq_id:
        require_can_write_squadron(p, "none", None)
    s = db.get(Squadron, sq_id)
    if not s:
        raise HTTPException(400, detail={"error": "no_squadron_scope"})
    require_can_write_squadron(p, s.id, s.wing_id)
    if a.owning_level == "squadron" and a.squadron_id == sq_id:
        raise HTTPException(400, detail={
            "error": "cannot_override_own_activity",
            "message": "A squadron cannot create a local override on its own activity -- edit it directly instead.",
        })

    existing = db.query(ActivityLocalOverride).filter(
        ActivityLocalOverride.activity_id == aid, ActivityLocalOverride.squadron_id == sq_id,
    ).first()
    if existing:
        _check_version(existing, body.version)
        existing.local_date_start = body.local_date_start
        existing.local_date_end = body.local_date_end
        existing.local_time_start = body.local_time_start
        existing.local_time_end = body.local_time_end
        existing.local_notes = body.local_notes
        existing.is_hidden = body.is_hidden
        existing.updated_by = p.user_id
        existing.version += 1
        override = existing
        action = "update"
    else:
        override = ActivityLocalOverride(
            activity_id=aid, squadron_id=sq_id,
            local_date_start=body.local_date_start, local_date_end=body.local_date_end,
            local_time_start=body.local_time_start, local_time_end=body.local_time_end,
            local_notes=body.local_notes, is_hidden=body.is_hidden,
            created_by=p.user_id, updated_by=p.user_id,
        )
        db.add(override)
        action = "create"
    db.commit()
    audit(db, p, object_type="activity_local_override", object_id=override.id, action=action,
         new={"activity_id": aid, "squadron_id": sq_id, "is_hidden": body.is_hidden})
    return {"ok": True, **_local_override_out(override)}


@router.delete("/activities/{aid}/local-override")
def delete_activity_local_override(aid: str, db: DBSession = Depends(get_db), p: Principal = Depends(get_principal)):
    """REM-103: revert to the source Activity's own values -- removes this
    squadron's override row entirely rather than soft-clearing fields, so a
    later GET/list simply shows the unmodified source data again."""
    sq_id = _active_squadron(p)
    if not sq_id:
        require_can_write_squadron(p, "none", None)
    s = db.get(Squadron, sq_id)
    if not s:
        raise HTTPException(400, detail={"error": "no_squadron_scope"})
    require_can_write_squadron(p, s.id, s.wing_id)

    existing = db.query(ActivityLocalOverride).filter(
        ActivityLocalOverride.activity_id == aid, ActivityLocalOverride.squadron_id == sq_id,
    ).first()
    if not existing:
        raise HTTPException(404, detail={"error": "not_found"})
    override_id = existing.id
    db.delete(existing)
    db.commit()
    audit(db, p, object_type="activity_local_override", object_id=override_id, action="delete",
         old={"activity_id": aid, "squadron_id": sq_id})
    return {"ok": True}


@router.post("/activities")
def create_activity(body: ActivityIn, db: DBSession = Depends(get_db), p: Principal = Depends(get_principal)):
    if p.role in _WRITE_BLOCKED:
        raise HTTPException(403, detail={"error": "forbidden"})
    sq_id = _active_squadron(p)
    if not sq_id:
        require_can_write_squadron(p, "none", None)
    s = db.get(Squadron, sq_id)
    if not s:
        raise HTTPException(400, detail={"error": "no_squadron_scope"})
    require_can_write_squadron(p, s.id, s.wing_id)
    a = Activity(squadron_id=s.id, wing_id=s.wing_id, owning_level="squadron",
                 activity_name=body.activity_name,
                 activity_type=body.activity_type, date_start=body.date_start,
                 date_end=body.date_end, time_start=body.time_start, time_end=body.time_end,
                 location=body.location, audience=body.audience, notes=body.notes)
    db.add(a)
    db.commit()
    audit(db, p, object_type="activity", object_id=a.id, action="create", new={"owning_level": "squadron"})
    return {"ok": True, "activity_id": a.id}


@router.post("/activities/wing")
def create_wing_activity(body: ActivityWingIn, db: DBSession = Depends(get_db), p: Principal = Depends(get_principal)):
    """Create a Wing-owned activity, automatically visible in every subordinate
    Squadron's Activities list/calendar -- no republish or per-squadron import step."""
    wing_id = body.wing_id or p.acting_wing_id or p.wing_id
    if not wing_id:
        raise HTTPException(400, detail={"error": "no_wing_scope",
                                          "message": "Provide wing_id in the request body or enter Delegated Intervention first."})
    require_can_write_activity(p, "wing", wing_id, None)
    from ..models import Wing as WingModel
    w = db.get(WingModel, wing_id)
    if not w or w.is_archived:
        raise HTTPException(404, detail={"error": "wing_not_found"})
    a = Activity(wing_id=wing_id, squadron_id=None, owning_level="wing",
                 activity_name=body.activity_name,
                 activity_type=body.activity_type, date_start=body.date_start,
                 date_end=body.date_end, time_start=body.time_start, time_end=body.time_end,
                 location=body.location, audience=body.audience, notes=body.notes)
    db.add(a)
    db.commit()
    audit(db, p, object_type="activity", object_id=a.id, action="create", new={"owning_level": "wing", "wing_id": wing_id})
    return {"ok": True, "activity_id": a.id}


@router.post("/activities/national")
def create_national_activity(body: ActivityIn, db: DBSession = Depends(get_db), p: Principal = Depends(get_principal)):
    """Create a National activity, automatically visible in every Wing and
    Squadron's Activities list/calendar -- no republish or per-unit import step."""
    require_can_write_activity(p, "national", None, None)
    a = Activity(wing_id=None, squadron_id=None, owning_level="national",
                 activity_name=body.activity_name,
                 activity_type=body.activity_type, date_start=body.date_start,
                 date_end=body.date_end, time_start=body.time_start, time_end=body.time_end,
                 location=body.location, audience=body.audience, notes=body.notes)
    db.add(a)
    db.commit()
    audit(db, p, object_type="activity", object_id=a.id, action="create", new={"owning_level": "national"})
    return {"ok": True, "activity_id": a.id}


@router.patch("/activities/{aid}")
def update_activity(aid: str, body: ActivityUpdateIn, db: DBSession = Depends(get_db),
                    p: Principal = Depends(get_principal)):
    a = db.get(Activity, aid)
    if not a:
        raise HTTPException(404, detail={"error": "not_found"})
    require_can_write_activity(p, a.owning_level, a.wing_id, a.squadron_id)
    if body.activity_name is not None:
        a.activity_name = body.activity_name
    if body.activity_type is not None:
        a.activity_type = body.activity_type
    if body.date_start is not None:
        a.date_start = body.date_start
    if body.date_end is not None:
        a.date_end = body.date_end
    if body.time_start is not None:
        a.time_start = body.time_start
    if body.time_end is not None:
        a.time_end = body.time_end
    if body.location is not None:
        a.location = body.location
    if body.audience is not None:
        a.audience = body.audience
    if body.notes is not None:
        a.notes = body.notes
    if body.workflow_status is not None:
        a.workflow_status = body.workflow_status
    if body.cea_seq_nr is not None:
        a.cea_seq_nr = body.cea_seq_nr
    db.commit()
    audit(db, p, object_type="activity", object_id=a.id, action="update")
    return {"ok": True, "activity_id": a.id}


@router.delete("/activities/{aid}")
def delete_activity(aid: str, db: DBSession = Depends(get_db), p: Principal = Depends(get_principal)):
    a = db.get(Activity, aid)
    if not a:
        raise HTTPException(404, detail={"error": "not_found"})
    require_can_write_activity(p, a.owning_level, a.wing_id, a.squadron_id)
    a.is_archived = True
    a.archived_at = utcnow()
    db.commit()
    audit(db, p, object_type="activity", object_id=a.id, action="archive")
    return {"ok": True}


@router.post("/activities/{aid}/restore")
def restore_activity(aid: str, db: DBSession = Depends(get_db), p: Principal = Depends(get_principal)):
    a = db.get(Activity, aid)
    if not a or not a.is_archived:
        raise HTTPException(404, detail={"error": "not_found"})
    require_can_write_activity(p, a.owning_level, a.wing_id, a.squadron_id)
    a.is_archived = False
    a.archived_at = None
    db.commit()
    audit(db, p, object_type="activity", object_id=a.id, action="restore")
    return {"ok": True}


# ── ACTIVITY GENERATION ──────────────────────────────────────────────────────

RECURRENCE_OPTS = frozenset({"daily", "weekly", "fortnightly", "monthly", "yearly"})


class ActivityGenerateIn(BaseModel):
    activity_name: str
    activity_type: str = "Optional"
    start_date: str           # YYYY-MM-DD — first possible date
    end_date: str             # YYYY-MM-DD — last possible date (inclusive)
    time_start: str | None = None
    time_end: str | None = None
    recurrence: str = "weekly"      # daily|weekly|fortnightly|monthly|yearly
    weekday: int | None = None      # 0=Mon…6=Sun — required for weekly/fortnightly
    excluded_dates: list[str] = []  # YYYY-MM-DD dates to skip
    repeat_count: int | None = None # cap total occurrences if set
    location: str | None = None
    audience: list[str] = []
    notes: str | None = None
    planning_year_id: str | None = None   # optional; used for holiday lookups
    preview_only: bool = True


def _generate_dates(start: str, end: str, recurrence: str, weekday: int | None,
                    excluded: set[str], repeat_count: int | None):
    """Yield ISO date strings matching the recurrence rule within [start, end]."""
    from datetime import date, timedelta
    import calendar as _cal

    sd = date.fromisoformat(start)
    ed = date.fromisoformat(end)

    if recurrence == "daily":
        cur = sd
        count = 0
        while cur <= ed:
            ds = cur.isoformat()
            if ds not in excluded:
                yield ds
                count += 1
                if repeat_count and count >= repeat_count:
                    return
            cur += timedelta(days=1)

    elif recurrence in ("weekly", "fortnightly"):
        if weekday is None:
            weekday = sd.weekday()
        step = timedelta(days=7 if recurrence == "weekly" else 14)
        # Find first occurrence on or after start
        days_ahead = (weekday - sd.weekday()) % 7
        cur = sd + timedelta(days=days_ahead)
        count = 0
        while cur <= ed:
            ds = cur.isoformat()
            if ds not in excluded:
                yield ds
                count += 1
                if repeat_count and count >= repeat_count:
                    return
            cur += step

    elif recurrence == "monthly":
        from datetime import date as _date
        cur = sd.replace(day=1)
        target_day = sd.day
        count = 0
        while cur <= ed:
            last = _cal.monthrange(cur.year, cur.month)[1]
            day = min(target_day, last)
            candidate = _date(cur.year, cur.month, day)
            if sd <= candidate <= ed:
                ds = candidate.isoformat()
                if ds not in excluded:
                    yield ds
                    count += 1
                    if repeat_count and count >= repeat_count:
                        return
            if cur.month == 12:
                cur = cur.replace(year=cur.year + 1, month=1)
            else:
                cur = cur.replace(month=cur.month + 1)

    elif recurrence == "yearly":
        from datetime import date as _date
        import calendar as _cal
        cur_year = sd.year
        count = 0
        while True:
            try:
                candidate = _date(cur_year, sd.month, sd.day)
            except ValueError:
                last = _cal.monthrange(cur_year, sd.month)[1]
                candidate = _date(cur_year, sd.month, last)
            if candidate > ed:
                break
            if candidate >= sd:
                ds = candidate.isoformat()
                if ds not in excluded:
                    yield ds
                    count += 1
                    if repeat_count and count >= repeat_count:
                        return
            cur_year += 1


@router.post("/activities/generate")
def generate_activities(
    body: ActivityGenerateIn,
    db: DBSession = Depends(get_db),
    p: Principal = Depends(get_principal),
):
    """Preview or batch-create recurring activities.

    When preview_only=True (default), returns a preview list without writing.
    When preview_only=False, creates all 'include' records and returns a summary.
    """
    if body.recurrence not in RECURRENCE_OPTS:
        raise HTTPException(400, detail={"error": "invalid_recurrence",
                                         "allowed": sorted(RECURRENCE_OPTS)})
    if p.role in _WRITE_BLOCKED and not body.preview_only:
        raise HTTPException(403, detail={"error": "forbidden"})

    sq_id = _active_squadron(p)
    s = db.get(Squadron, sq_id) if sq_id else None
    if not s:
        raise HTTPException(400, detail={"error": "no_squadron_scope"})
    if not body.preview_only:
        require_can_write_squadron(p, s.id, s.wing_id)

    # Build holiday date set for this squadron (from planning year if given)
    holiday_dates: dict[str, str] = {}  # date → holiday name
    if body.planning_year_id:
        try:
            from ..models.planning import HolidayPeriod, PlanningYear
            from datetime import date, timedelta
            py = db.get(PlanningYear, body.planning_year_id)
            if py and (py.unit_id == s.id or py.wing_id == s.wing_id):
                hols = db.query(HolidayPeriod).filter(
                    HolidayPeriod.planning_year_id == body.planning_year_id
                ).all()
                for h in hols:
                    hd = date.fromisoformat(h.start_date)
                    end_hd = date.fromisoformat(h.end_date)
                    while hd <= end_hd:
                        holiday_dates[hd.isoformat()] = h.name
                        hd += timedelta(days=1)
        except Exception:
            pass  # holiday lookup is best-effort; do not block generation

    # Existing active activities for duplicate detection
    existing = db.query(Activity).filter(
        Activity.squadron_id == s.id,
        Activity.is_archived == False,  # noqa: E712
        Activity.activity_name == body.activity_name,
    ).all()
    existing_map: dict[str, str] = {a.date_start: a.id for a in existing}

    excluded_set = set(body.excluded_dates or [])
    preview_rows = []
    for ds in _generate_dates(body.start_date, body.end_date, body.recurrence,
                               body.weekday, excluded_set, body.repeat_count):
        holiday_name = holiday_dates.get(ds)
        dup_id = existing_map.get(ds)
        if dup_id:
            status, reason = "skip", "duplicate"
        elif holiday_name:
            status, reason = "skip", "holiday"
        else:
            status, reason = "include", None

        preview_rows.append({
            "date": ds,
            "start_time": body.time_start,
            "finish_time": body.time_end,
            "status": status,
            "reason": reason,
            "existing_activity_id": dup_id,
            "holiday_name": holiday_name,
            "scope": "squadron",
        })

    if body.preview_only:
        return {
            "preview": preview_rows,
            "would_create": sum(1 for r in preview_rows if r["status"] == "include"),
            "would_skip": sum(1 for r in preview_rows if r["status"] == "skip"),
        }

    # Actual creation
    created_ids = []
    skipped = []
    batch_note = f"batch_generate:{body.recurrence}:{body.start_date}:{body.end_date}"
    for row in preview_rows:
        if row["status"] != "include":
            skipped.append({"date": row["date"], "reason": row["reason"]})
            continue
        a = Activity(
            squadron_id=s.id,
            wing_id=s.wing_id,
            activity_name=body.activity_name,
            activity_type=body.activity_type,
            date_start=row["date"],
            time_start=body.time_start,
            time_end=body.time_end,
            location=body.location,
            audience=body.audience or [],
            notes=body.notes,
            cea_seq_nr=batch_note,
        )
        db.add(a)
        db.flush()
        created_ids.append(a.id)

    db.commit()
    audit(db, p, object_type="activity", object_id="batch",
          action="generate_batch",
          new={"count": len(created_ids), "name": body.activity_name,
               "recurrence": body.recurrence, "ids": created_ids})
    return {
        "ok": True,
        "created_count": len(created_ids),
        "skipped_count": len(skipped),
        "created_ids": created_ids,
        "skipped": skipped,
    }


# ── CURRICULUM ELEMENTS ──────────────────────────────────────────────────────

class ElementIn(BaseModel):
    name: str           # short code, e.g. "Air_Space"
    display_name: str   # human label, e.g. "Air & Space"
    scope_level: str = "national"
    wing_id: str | None = None
    squadron_id: str | None = None


def _can_create_element(p: Principal, scope_level: str,
                        wing_id: str | None = None, squadron_id: str | None = None) -> None:
    """Raise 403 if the actor cannot create an element at the requested scope."""
    if p.role in _WRITE_BLOCKED:
        raise HTTPException(403, detail={"error": "forbidden",
                                          "message": "Viewers and auditors cannot create elements."})
    if scope_level not in ELEMENT_SCOPE_LEVELS:
        raise HTTPException(400, detail={"error": "invalid_scope",
                                          "message": f"scope_level must be one of: {sorted(ELEMENT_SCOPE_LEVELS)}"})
    if scope_level == "system":
        if p.role != "system_admin":
            raise HTTPException(403, detail={"error": "forbidden",
                                              "message": "Only system_admin can create system-scope elements."})
    elif scope_level == "national":
        if p.role not in _NAT_ADMIN_ROLES:
            raise HTTPException(403, detail={"error": "forbidden",
                                              "message": "Only national_admin or system_admin can create national elements."})
    elif scope_level == "wing":
        if p.role not in _WING_WRITE_ROLES:
            raise HTTPException(403, detail={"error": "forbidden",
                                              "message": "Only wing_admin or above can create wing elements."})
        effective_wing = wing_id or p.wing_id
        if p.role == "wing_admin" and effective_wing != p.wing_id:
            raise HTTPException(403, detail={"error": "out_of_scope",
                                              "message": "Wing admin can only create elements for their own wing."})
    elif scope_level == "squadron":
        if p.role not in {*_WING_WRITE_ROLES, "sqn_admin"}:
            raise HTTPException(403, detail={"error": "forbidden"})
        if p.role == "sqn_admin" and squadron_id and squadron_id != p.squadron_id:
            raise HTTPException(403, detail={"error": "out_of_scope",
                                              "message": "Squadron admin can only create elements for their own squadron."})


def _visible_elements(db: DBSession, p: Principal) -> list[CurriculumElement]:
    """Return elements visible to this principal (scoped by role)."""
    from sqlalchemy import or_
    conditions = [
        CurriculumElement.scope_level == "national",
        CurriculumElement.scope_level == "system",
    ]
    wing_id = p.acting_wing_id or p.wing_id
    sq_id = p.acting_squadron_id or p.squadron_id
    if wing_id:
        conditions.append(
            (CurriculumElement.scope_level == "wing") & (CurriculumElement.wing_id == wing_id))
    elif p.role in _NAT_ADMIN_ROLES:
        conditions.append(CurriculumElement.scope_level == "wing")
    if sq_id:
        conditions.append(
            (CurriculumElement.scope_level == "squadron") & (CurriculumElement.squadron_id == sq_id))
    elif p.role == "wing_admin":
        pass  # wing admin: no sqn-scope elements unless proxied
    elif p.role in _NAT_ADMIN_ROLES:
        conditions.append(CurriculumElement.scope_level == "squadron")
    return db.query(CurriculumElement).filter(
        CurriculumElement.is_archived == False,  # noqa: E712
        CurriculumElement.active_status == True,  # noqa: E712
        or_(*conditions),
    ).order_by(CurriculumElement.scope_level, CurriculumElement.display_name).all()


@router.get("/curriculum/elements")
def list_elements(db: DBSession = Depends(get_db), p: Principal = Depends(get_principal)):
    """Return all curriculum elements visible to the caller's scope."""
    return [{"element_id": e.id, "name": e.name, "display_name": e.display_name,
             "scope_level": e.scope_level, "wing_id": e.wing_id, "squadron_id": e.squadron_id}
            for e in _visible_elements(db, p)]


@router.post("/curriculum/elements")
def create_element(body: ElementIn, db: DBSession = Depends(get_db),
                   p: Principal = Depends(get_principal)):
    """Create a curriculum element at the requested scope."""
    scope = body.scope_level
    _can_create_element(p, scope, body.wing_id, body.squadron_id)
    name = (body.name or "").strip().replace(" ", "_")
    if not name:
        raise HTTPException(400, detail={"error": "name_required"})
    wing_id = body.wing_id or (p.wing_id if scope in ("wing", "squadron") else None)
    sq_id = body.squadron_id or (p.squadron_id if scope == "squadron" else None)
    # Idempotent: return existing element if same name + scope
    existing = db.query(CurriculumElement).filter(
        CurriculumElement.name == name,
        CurriculumElement.scope_level == scope,
        CurriculumElement.wing_id == wing_id,
        CurriculumElement.squadron_id == sq_id,
        CurriculumElement.is_archived == False,  # noqa: E712
    ).first()
    if existing:
        return {"ok": True, "element_id": existing.id, "name": existing.name,
                "display_name": existing.display_name, "scope_level": existing.scope_level, "existed": True}
    el = CurriculumElement(
        name=name, display_name=body.display_name or name,
        scope_level=scope, wing_id=wing_id, squadron_id=sq_id,
        active_status=True, created_by=p.user_id)
    db.add(el)
    db.commit()
    audit(db, p, object_type="curriculum_element", object_id=el.id, action="create",
          new={"name": name, "scope": scope})
    return {"ok": True, "element_id": el.id, "name": el.name,
            "display_name": el.display_name, "scope_level": el.scope_level, "existed": False}


@router.post("/curriculum/elements/{eid}/archive")
def archive_element(eid: str, db: DBSession = Depends(get_db),
                    p: Principal = Depends(get_principal)):
    el = db.get(CurriculumElement, eid)
    if not el:
        raise HTTPException(404, detail={"error": "not_found"})
    # Only the owning scope or above can archive
    _can_create_element(p, el.scope_level, el.wing_id, el.squadron_id)
    el.is_archived = True
    el.archived_at = utcnow()
    db.commit()
    audit(db, p, object_type="curriculum_element", object_id=el.id, action="archive")
    return {"ok": True}


def _upsert_element(db: DBSession, name: str, scope: str = "national",
                    wing_id: str | None = None, sq_id: str | None = None) -> str:
    """Find or create an element by name+scope; return its id. Used by workbook import."""
    name = (name or "").strip().replace(" ", "_")
    if not name:
        return None
    existing = db.query(CurriculumElement).filter(
        CurriculumElement.name == name,
        CurriculumElement.scope_level == scope,
        CurriculumElement.wing_id == wing_id,
        CurriculumElement.squadron_id == sq_id,
        CurriculumElement.is_archived == False,  # noqa: E712
    ).first()
    if existing:
        return existing.id
    el = CurriculumElement(name=name, display_name=name.replace("_", " "),
                            scope_level=scope, wing_id=wing_id, squadron_id=sq_id,
                            active_status=True)
    db.add(el)
    db.flush()  # get id without full commit
    return el.id


# ── CURRICULUM PHASES ─────────────────────────────────────────────────────────
# Master transformation plan Block 10 / addendum section II.2: a governed
# training-phase catalogue, deliberately built as CurriculumElement's sibling
# (same scope model, same idempotent-create/archive shape) rather than a
# discriminator on curriculum_elements itself — elements are subject-matter
# categories, phases are program-progression stages, and a discriminator
# risked confusing list_elements' existing callers. CustomPhase (the
# pre-existing, fully-unwired dead table) is deliberately left alone, not
# built on — it lacks the scope columns this table needs, so extending it
# would have needed the same work as starting fresh here.

class PhaseIn(BaseModel):
    name: str            # e.g. "A. Orientation" — matches CurriculumItem.phase/Session.phase_at_time
    display_name: str
    scope_level: str = "national"
    wing_id: str | None = None
    squadron_id: str | None = None
    sort_order: int = 0


def _can_create_phase(p: Principal, scope_level: str,
                      wing_id: str | None = None, squadron_id: str | None = None) -> None:
    """Raise 403 if the actor cannot create a phase at the requested scope.

    DEFECT-001 (general-release qualification): the squadron-scope branch
    used to allow wing_admin/national_admin/system_admin to create a
    squadron-scope phase for ANY squadron with no Proxy/Delegated
    Intervention check at all — inconsistent with require_can_write_squadron,
    the sanctioned mechanism used everywhere else in this app for a higher
    role writing into a lower organisational scope. Now enforced the same
    way: sqn_admin writes their own squadron directly; wing_admin needs an
    active Proxy session targeting that squadron; national_admin/system_admin
    need active Delegated Intervention targeting that squadron.
    """
    if p.role in _WRITE_BLOCKED:
        raise HTTPException(403, detail={"error": "forbidden",
                                          "message": "Viewers and auditors cannot create phases."})
    if scope_level not in PHASE_SCOPE_LEVELS:
        raise HTTPException(400, detail={"error": "invalid_scope",
                                          "message": f"scope_level must be one of: {sorted(PHASE_SCOPE_LEVELS)}"})
    if scope_level == "system":
        if p.role != "system_admin":
            raise HTTPException(403, detail={"error": "forbidden",
                                              "message": "Only system_admin can create system-scope phases."})
    elif scope_level == "national":
        if p.role not in _NAT_ADMIN_ROLES:
            raise HTTPException(403, detail={"error": "forbidden",
                                              "message": "Only national_admin or system_admin can create national phases."})
    elif scope_level == "wing":
        if p.role not in _WING_WRITE_ROLES:
            raise HTTPException(403, detail={"error": "forbidden",
                                              "message": "Only wing_admin or above can create wing phases."})
        effective_wing = wing_id or p.wing_id
        if p.role == "wing_admin" and effective_wing != p.wing_id:
            raise HTTPException(403, detail={"error": "out_of_scope",
                                              "message": "Wing admin can only create phases for their own wing."})
    elif scope_level == "squadron":
        if p.role == "sqn_admin":
            if squadron_id and squadron_id != p.squadron_id:
                raise HTTPException(403, detail={"error": "out_of_scope",
                                                  "message": "Squadron admin can only create phases for their own squadron."})
        elif p.role == "wing_admin":
            if not p.acting_squadron_id:
                raise HTTPException(403, detail={"error": "proxy_required",
                                                  "message": "Wing Admin must enter Proxy Mode to create a squadron-scope phase."})
            if squadron_id and squadron_id != p.acting_squadron_id:
                raise HTTPException(403, detail={"error": "out_of_scope",
                                                  "message": "Can only create a squadron-scope phase for the squadron currently in Proxy Mode."})
        elif p.role in ("national_admin", "system_admin"):
            if not p.acting_squadron_id:
                raise HTTPException(403, detail={"error": "intervention_required",
                                                  "message": "National Admin must enter Delegated Intervention Mode to create a squadron-scope phase."})
            if squadron_id and squadron_id != p.acting_squadron_id:
                raise HTTPException(403, detail={"error": "out_of_scope",
                                                  "message": "Can only create a squadron-scope phase for the squadron currently in Delegated Intervention Mode."})
        else:
            raise HTTPException(403, detail={"error": "forbidden"})


def _visible_phases(db: DBSession, p: Principal) -> list[CurriculumPhase]:
    """Return phases visible to this principal (scoped by role) — identical
    visibility rule to _visible_elements. This is the "phase is available to
    this unit" concept from addendum section II.3, distinct from "phase
    exists in catalogue" (any row) and "phase is scheduled tonight" (a
    Session's phase_at_time)."""
    from sqlalchemy import or_
    conditions = [
        CurriculumPhase.scope_level == "national",
        CurriculumPhase.scope_level == "system",
    ]
    wing_id = p.acting_wing_id or p.wing_id
    sq_id = p.acting_squadron_id or p.squadron_id
    if wing_id:
        conditions.append(
            (CurriculumPhase.scope_level == "wing") & (CurriculumPhase.wing_id == wing_id))
    elif p.role in _NAT_ADMIN_ROLES:
        conditions.append(CurriculumPhase.scope_level == "wing")
    if sq_id:
        conditions.append(
            (CurriculumPhase.scope_level == "squadron") & (CurriculumPhase.squadron_id == sq_id))
    elif p.role == "wing_admin":
        pass  # wing admin: no sqn-scope phases unless proxied
    elif p.role in _NAT_ADMIN_ROLES:
        conditions.append(CurriculumPhase.scope_level == "squadron")
    return db.query(CurriculumPhase).filter(
        CurriculumPhase.is_archived == False,  # noqa: E712
        CurriculumPhase.active_status == True,  # noqa: E712
        or_(*conditions),
    ).order_by(CurriculumPhase.sort_order, CurriculumPhase.scope_level, CurriculumPhase.display_name).all()


@router.get("/curriculum/phases")
def list_phases(db: DBSession = Depends(get_db), p: Principal = Depends(get_principal)):
    """Return all training phases visible to the caller's scope, in
    training-progression order — the phase picker's data source."""
    return [{"phase_id": ph.id, "name": ph.name, "display_name": ph.display_name,
             "scope_level": ph.scope_level, "wing_id": ph.wing_id, "squadron_id": ph.squadron_id,
             "sort_order": ph.sort_order}
            for ph in _visible_phases(db, p)]


@router.post("/curriculum/phases")
def create_phase(body: PhaseIn, db: DBSession = Depends(get_db),
                 p: Principal = Depends(get_principal)):
    """Create a training phase at the requested scope."""
    scope = body.scope_level
    _can_create_phase(p, scope, body.wing_id, body.squadron_id)
    name = (body.name or "").strip()
    if not name:
        raise HTTPException(400, detail={"error": "name_required"})
    # DEFECT-001: fall back to the caller's ACTING wing/squadron (set while a
    # Proxy/Delegated Intervention session is active), not just their own —
    # otherwise a wing_admin/national_admin who omits squadron_id here would
    # have it resolve to their own (nonexistent) squadron_id and silently
    # create an orphaned scope_level="squadron", squadron_id=None row.
    wing_id = body.wing_id or ((p.acting_wing_id or p.wing_id) if scope in ("wing", "squadron") else None)
    sq_id = body.squadron_id or ((p.acting_squadron_id or p.squadron_id) if scope == "squadron" else None)
    # Idempotent: return existing phase if same name + scope
    existing = db.query(CurriculumPhase).filter(
        CurriculumPhase.name == name,
        CurriculumPhase.scope_level == scope,
        CurriculumPhase.wing_id == wing_id,
        CurriculumPhase.squadron_id == sq_id,
        CurriculumPhase.is_archived == False,  # noqa: E712
    ).first()
    if existing:
        return {"ok": True, "phase_id": existing.id, "name": existing.name,
                "display_name": existing.display_name, "scope_level": existing.scope_level, "existed": True}
    ph = CurriculumPhase(
        name=name, display_name=body.display_name or name,
        scope_level=scope, wing_id=wing_id, squadron_id=sq_id, sort_order=body.sort_order,
        active_status=True, created_by=p.user_id)
    db.add(ph)
    db.commit()
    audit(db, p, object_type="curriculum_phase", object_id=ph.id, action="create",
          new={"name": name, "scope": scope})
    return {"ok": True, "phase_id": ph.id, "name": ph.name,
            "display_name": ph.display_name, "scope_level": ph.scope_level, "existed": False}


@router.post("/curriculum/phases/{phid}/archive")
def archive_phase(phid: str, db: DBSession = Depends(get_db),
                  p: Principal = Depends(get_principal)):
    ph = db.get(CurriculumPhase, phid)
    if not ph:
        raise HTTPException(404, detail={"error": "not_found"})
    # Only the owning scope or above can archive
    _can_create_phase(p, ph.scope_level, ph.wing_id, ph.squadron_id)
    ph.is_archived = True
    ph.archived_at = utcnow()
    db.commit()
    audit(db, p, object_type="curriculum_phase", object_id=ph.id, action="archive")
    return {"ok": True}


# ── CURRICULUM WRITE (SQN-owned items only) ──────────────────────────────────
class CurriculumIn(BaseModel):
    code: str = Field(max_length=40)
    title: str = Field(max_length=200)
    identifier: str | None = None  # globally unique mission key, e.g. "ORI-M01-01(2)"
    part_number: int = 1
    phase: str = "B. Initial"
    element: str | None = None
    recommended_term: str | None = Field(default=None, max_length=10)
    duration_minutes: int = 60
    part_count: int = 1
    instructor_suitability: str | None = None
    learning_hub_url: str | None = None
    wing_id: str | None = None  # NAT admins pass this for wing-owned curriculum

    @field_validator("learning_hub_url")
    @classmethod
    def validate_url(cls, v): return _check_http_url(v)


class CurriculumUpdateIn(BaseModel):
    title: str | None = Field(default=None, max_length=200)
    identifier: str | None = None
    part_number: int | None = None
    phase: str | None = None
    element: str | None = None
    recommended_term: str | None = Field(default=None, max_length=10)
    duration_minutes: int | None = None
    part_count: int | None = None
    instructor_suitability: str | None = None
    learning_hub_url: str | None = None

    @field_validator("learning_hub_url")
    @classmethod
    def validate_url(cls, v): return _check_http_url(v)


class CurriculumImportItem(BaseModel):
    identifier: str | None = None    # primary unique key
    code: str = Field(max_length=40)  # Module_Code (may repeat across parts)
    part_number: int = 1
    title: str = Field(max_length=200)
    phase: str = "B. Initial"
    element: str | None = None
    duration_minutes: int = 60
    part_count: int = 1
    instructor_suitability: str | None = None
    learning_hub_url: str | None = None
    recommended_term: str | None = Field(default=None, max_length=10)

    @field_validator("learning_hub_url")
    @classmethod
    def validate_url(cls, v): return _check_http_url(v)
    core_status: str | None = None    # "core" | "additional" -- None means "not specified by caller"
    # Scheduling fields — used to link to existing parade night sessions
    scheduled_date: str | None = None  # ISO date string e.g. "2026-02-13"
    session_number: int | None = None  # period number (1, 2, 3)
    facilitator_name: str | None = None
    location: str | None = None
    room: str | None = None


class CurriculumImportIn(BaseModel):
    items: List[CurriculumImportItem]
    squadron_id: str | None = None  # if provided, link scheduled items to this sqn
    owning_level: str = "national"  # national | wing | squadron
    # Phase 3.4: compute and return the create/update/skip/failed breakdown
    # without writing anything -- default False preserves this endpoint's
    # original immediate-commit behaviour for every existing caller.
    preview: bool = False


_NAT_ADMIN_ROLES = frozenset({"national_admin", "system_admin"})
_WING_WRITE_ROLES = frozenset({"wing_admin", "national_admin", "system_admin"})


def _find_existing_curriculum(db: DBSession, body: CurriculumIn,
                              owning_level: str,
                              wing_id: str | None = None,
                              squadron_id: str | None = None) -> CurriculumItem | None:
    """Find an existing (non-archived) curriculum item by identifier or (code, part_number).

    Multiple items can share the same code (Module_Code) for different parts, so
    code alone is never a uniqueness criterion. The check hierarchy is:
    1. If identifier is set: match by identifier + owning scope.
    2. Else: match by (code, part_number) + owning scope.
    """
    q = db.query(CurriculumItem).filter(
        CurriculumItem.owning_level == owning_level,
        CurriculumItem.is_archived == False)  # noqa: E712
    if owning_level == "wing":
        q = q.filter(CurriculumItem.wing_id == wing_id)
    elif owning_level == "squadron":
        q = q.filter(CurriculumItem.squadron_id == squadron_id)

    if body.identifier:
        return q.filter(CurriculumItem.identifier == body.identifier).first()
    # Fallback: (code, part_number) — allows same code with different parts
    return q.filter(
        CurriculumItem.code == body.code,
        CurriculumItem.part_number == body.part_number,
    ).first()


@router.post("/curriculum")
def create_curriculum(body: CurriculumIn, db: DBSession = Depends(get_db),
                      p: Principal = Depends(get_principal)):
    """Create a squadron-owned curriculum item (owning_level=squadron)."""
    if p.role in _WRITE_BLOCKED:
        raise HTTPException(403, detail={"error": "forbidden"})
    sq_id = _active_squadron(p)
    if not sq_id:
        raise HTTPException(400, detail={"error": "no_squadron_scope"})
    s = db.get(Squadron, sq_id)
    if not s:
        raise HTTPException(400, detail={"error": "no_squadron_scope"})
    require_can_write_squadron(p, s.id, s.wing_id)
    exists = _find_existing_curriculum(db, body, "squadron", squadron_id=sq_id)
    if exists:
        raise HTTPException(409, detail={
            "error": "already_exists",
            "message": f"Curriculum item '{body.identifier or body.code}' already exists.",
            "curriculum_id": exists.id,
        })
    ci = CurriculumItem(
        squadron_id=s.id, wing_id=s.wing_id, owning_level="squadron",
        identifier=body.identifier, code=body.code, part_number=body.part_number,
        title=body.title, phase=body.phase, element=body.element,
        recommended_term=body.recommended_term, duration_minutes=body.duration_minutes,
        part_count=body.part_count, instructor_suitability=body.instructor_suitability,
        learning_hub_url=body.learning_hub_url, core_status="additional")
    db.add(ci)
    db.commit()
    audit(db, p, object_type="curriculum_item", object_id=ci.id, action="create",
          new={"owning_level": "squadron"})
    return {"ok": True, "curriculum_id": ci.id}


@router.post("/curriculum/wing")
def create_wing_curriculum(body: CurriculumIn, db: DBSession = Depends(get_db),
                           p: Principal = Depends(get_principal)):
    """Create a Wing-owned curriculum item visible to all squadrons under that Wing."""
    if p.role not in _WING_WRITE_ROLES:
        raise HTTPException(403, detail={"error": "forbidden",
                                          "message": "Only Wing or NAT HQ admin can create Wing curriculum."})
    wing_id = body.wing_id or p.acting_wing_id or p.wing_id
    if not wing_id:
        raise HTTPException(400, detail={"error": "no_wing_scope",
                                          "message": "Provide wing_id in the request body or enter proxy mode first."})
    from ..models import Wing as WingModel
    w = db.get(WingModel, wing_id)
    if not w or w.is_archived:
        raise HTTPException(404, detail={"error": "wing_not_found"})
    exists = _find_existing_curriculum(db, body, "wing", wing_id=wing_id)
    if exists:
        raise HTTPException(409, detail={
            "error": "already_exists",
            "message": f"Curriculum item '{body.identifier or body.code}' already exists in this Wing.",
            "curriculum_id": exists.id,
        })
    ci = CurriculumItem(
        wing_id=wing_id, squadron_id=None, owning_level="wing",
        identifier=body.identifier, code=body.code, part_number=body.part_number,
        title=body.title, phase=body.phase, element=body.element,
        recommended_term=body.recommended_term, duration_minutes=body.duration_minutes,
        part_count=body.part_count, instructor_suitability=body.instructor_suitability,
        learning_hub_url=body.learning_hub_url, core_status="additional")
    db.add(ci)
    db.commit()
    audit(db, p, object_type="curriculum_item", object_id=ci.id, action="create",
          new={"owning_level": "wing", "wing_id": wing_id})
    return {"ok": True, "curriculum_id": ci.id}


@router.post("/curriculum/national")
def create_national_curriculum(body: CurriculumIn, db: DBSession = Depends(get_db),
                               p: Principal = Depends(get_principal)):
    """Create a National curriculum item visible to all Wings and Squadrons.

    Uniqueness is by identifier (if provided) or (code, part_number).
    Multiple parts of the same module share the same code but have distinct
    identifiers / part_numbers — they are NOT duplicates.
    """
    if p.role not in _NAT_ADMIN_ROLES:
        raise HTTPException(403, detail={"error": "forbidden",
                                          "message": "Only NAT HQ admin can create National curriculum."})
    exists = _find_existing_curriculum(db, body, "national")
    if exists:
        raise HTTPException(409, detail={
            "error": "already_exists",
            "message": f"Curriculum item '{body.identifier or body.code}' (part {body.part_number}) already exists.",
            "curriculum_id": exists.id,
        })
    ci = CurriculumItem(
        wing_id=None, squadron_id=None, owning_level="national",
        identifier=body.identifier, code=body.code, part_number=body.part_number,
        title=body.title, phase=body.phase, element=body.element,
        recommended_term=body.recommended_term, duration_minutes=body.duration_minutes,
        part_count=body.part_count, instructor_suitability=body.instructor_suitability,
        learning_hub_url=body.learning_hub_url, core_status="core")
    db.add(ci)
    db.commit()
    audit(db, p, object_type="curriculum_item", object_id=ci.id, action="create",
          new={"owning_level": "national"})
    return {"ok": True, "curriculum_id": ci.id}


@router.patch("/curriculum/{cid}")
def update_curriculum(cid: str, body: CurriculumUpdateIn, db: DBSession = Depends(get_db),
                      p: Principal = Depends(get_principal)):
    ci = db.get(CurriculumItem, cid)
    if not ci:
        raise HTTPException(404, detail={"error": "not_found"})
    # Ownership check by level
    if ci.owning_level == "national":
        if p.role not in _NAT_ADMIN_ROLES:
            raise HTTPException(403, detail={"error": "cannot_edit_national_curriculum"})
    elif ci.owning_level == "wing":
        if p.role not in _WING_WRITE_ROLES:
            raise HTTPException(403, detail={"error": "cannot_edit_wing_curriculum"})
        # NAT admins may edit any wing's curriculum; wing admins are scoped to their own wing
        if p.role not in _NAT_ADMIN_ROLES:
            actor_wing = p.acting_wing_id or p.wing_id
            if ci.wing_id != actor_wing:
                raise HTTPException(403, detail={"error": "out_of_scope"})
    else:
        sq_id = _active_squadron(p)
        if ci.squadron_id != sq_id:
            raise HTTPException(403, detail={"error": "forbidden"})
    if body.title is not None:
        ci.title = body.title
    if body.phase is not None:
        ci.phase = body.phase
    if body.element is not None:
        ci.element = body.element
    if body.recommended_term is not None:
        ci.recommended_term = body.recommended_term
    if body.duration_minutes is not None:
        ci.duration_minutes = body.duration_minutes
    if body.learning_hub_url is not None:
        ci.learning_hub_url = body.learning_hub_url
    db.commit()
    audit(db, p, object_type="curriculum_item", object_id=ci.id, action="update")
    return {"ok": True}


@router.delete("/curriculum/{cid}")
def delete_curriculum(cid: str, db: DBSession = Depends(get_db), p: Principal = Depends(get_principal)):
    ci = db.get(CurriculumItem, cid)
    if not ci:
        raise HTTPException(404, detail={"error": "not_found"})
    if ci.owning_level == "national":
        if p.role not in _NAT_ADMIN_ROLES:
            raise HTTPException(403, detail={"error": "cannot_delete_national_curriculum"})
    elif ci.owning_level == "wing":
        if p.role not in _WING_WRITE_ROLES:
            raise HTTPException(403, detail={"error": "cannot_delete_wing_curriculum"})
        if p.role not in _NAT_ADMIN_ROLES:
            actor_wing = p.acting_wing_id or p.wing_id
            if ci.wing_id != actor_wing:
                raise HTTPException(403, detail={"error": "out_of_scope"})
    else:
        sq_id = _active_squadron(p)
        if ci.squadron_id != sq_id:
            raise HTTPException(403, detail={"error": "forbidden"})
    ci.is_archived = True
    ci.archived_at = utcnow()
    db.commit()
    audit(db, p, object_type="curriculum_item", object_id=ci.id, action="archive")
    return {"ok": True}


@router.post("/curriculum/{cid}/restore")
def restore_curriculum(cid: str, db: DBSession = Depends(get_db), p: Principal = Depends(get_principal)):
    """REM-133: curriculum-item archive existed with no restore counterpart --
    same pattern as restore_fac, same ownership-level permission checks as
    this item's own delete_curriculum."""
    ci = db.get(CurriculumItem, cid)
    if not ci:
        raise HTTPException(404, detail={"error": "not_found"})
    if ci.owning_level == "national":
        if p.role not in _NAT_ADMIN_ROLES:
            raise HTTPException(403, detail={"error": "cannot_edit_national_curriculum"})
    elif ci.owning_level == "wing":
        if p.role not in _WING_WRITE_ROLES:
            raise HTTPException(403, detail={"error": "cannot_edit_wing_curriculum"})
        if p.role not in _NAT_ADMIN_ROLES:
            actor_wing = p.acting_wing_id or p.wing_id
            if ci.wing_id != actor_wing:
                raise HTTPException(403, detail={"error": "out_of_scope"})
    else:
        sq_id = _active_squadron(p)
        if ci.squadron_id != sq_id:
            raise HTTPException(403, detail={"error": "forbidden"})
    if not ci.is_archived:
        raise HTTPException(409, detail={"error": "not_archived"})
    ci.is_archived = False
    ci.archived_at = None
    db.commit()
    audit(db, p, object_type="curriculum_item", object_id=ci.id, action="restore")
    return {"ok": True}


# ── CURRICULUM BULK IMPORT ────────────────────────────────────────────────────

@router.post("/curriculum/import")
def import_curriculum(body: CurriculumImportIn, db: DBSession = Depends(get_db),
                      p: Principal = Depends(get_principal)):
    """Bulk upsert curriculum items.

    Uniqueness key: identifier (if set), else (code, part_number).
    Multiple rows may share the same code (Module_Code) for different parts —
    this is NOT a duplicate. A duplicate is a row whose identifier (or
    code+part_number combo) already exists in the database.

    For each item:
    - If not found: create.
    - If found and any field changed: update.
    - If found and identical: skip.

    Never raises 409 — returns per-item status in the response.

    If squadron_id is supplied and a row has a scheduled_date + session_number,
    the endpoint attempts to link the curriculum item to the corresponding
    parade-night session for that squadron.
    """
    if p.role not in _NAT_ADMIN_ROLES:
        raise HTTPException(403, detail={
            "error": "forbidden",
            "message": "Only national_admin or system_admin can bulk-import curriculum.",
        })

    owning_level = body.owning_level if body.owning_level in {"national", "wing", "squadron"} else "national"
    sqn_id = body.squadron_id

    created = updated = skipped = failed = 0
    results = []

    # Pre-fetch facilitators + training areas for schedule linking
    fac_by_name: dict[str, str] = {}  # display_name -> id
    room_by_name: dict[str, str] = {}  # name -> id
    if sqn_id:
        for f in db.query(Facilitator).filter(Facilitator.squadron_id == sqn_id).all():
            key = (f.display_name or "").strip().lower()
            if key:
                fac_by_name[key] = f.id
        for r in db.query(TrainingArea).filter(TrainingArea.squadron_id == sqn_id).all():
            key = (r.name or "").strip().lower()
            if key:
                room_by_name[key] = r.id

    for item in body.items:
        try:
            # Locate existing item
            q = db.query(CurriculumItem).filter(
                CurriculumItem.owning_level == owning_level,
                CurriculumItem.is_archived == False)  # noqa: E712
            if owning_level == "squadron":
                q = q.filter(CurriculumItem.squadron_id == sqn_id)

            existing: CurriculumItem | None = None
            if item.identifier:
                existing = q.filter(CurriculumItem.identifier == item.identifier).first()
            if existing is None:
                existing = q.filter(
                    CurriculumItem.code == item.code,
                    CurriculumItem.part_number == item.part_number,
                ).first()

            if existing is None:
                # Ensure the element exists in the managed table (idempotent)
                if item.element:
                    _upsert_element(db, item.element, scope="national")
                # CREATE
                ci = CurriculumItem(
                    owning_level=owning_level,
                    squadron_id=sqn_id if owning_level == "squadron" else None,
                    identifier=item.identifier,
                    code=item.code,
                    part_number=item.part_number,
                    title=item.title,
                    phase=item.phase or "B. Initial",
                    element=item.element,
                    duration_minutes=item.duration_minutes or 60,
                    part_count=item.part_count or 1,
                    instructor_suitability=item.instructor_suitability,
                    learning_hub_url=item.learning_hub_url,
                    recommended_term=item.recommended_term,
                    core_status=item.core_status or ("core" if owning_level == "national" else "additional"),
                    active_status=True,
                )
                db.add(ci)
                db.flush()
                created += 1
                results.append({"identifier": item.identifier or item.code, "status": "created",
                                "code": item.code, "title": item.title, "phase": item.phase})
            else:
                ci = existing
                # Check if any field changed
                changed = False
                for field, val in [
                    ("title", item.title),
                    ("phase", item.phase),
                    ("element", item.element),
                    ("duration_minutes", item.duration_minutes),
                    ("part_count", item.part_count),
                    ("instructor_suitability", item.instructor_suitability),
                    ("learning_hub_url", item.learning_hub_url),
                    ("recommended_term", item.recommended_term),
                    ("identifier", item.identifier),
                    ("part_number", item.part_number),
                    ("core_status", item.core_status),
                ]:
                    if val is not None and getattr(ci, field) != val:
                        setattr(ci, field, val)
                        changed = True
                if changed:
                    updated += 1
                    results.append({"identifier": item.identifier or item.code, "status": "updated",
                                    "code": item.code, "title": item.title, "phase": item.phase})
                else:
                    skipped += 1
                    results.append({"identifier": item.identifier or item.code, "status": "skipped",
                                    "code": item.code, "title": item.title, "phase": item.phase})

            # Schedule linking: connect to parade night session if date provided
            if sqn_id and item.scheduled_date and item.session_number:
                _link_session(db, ci, sqn_id, item, fac_by_name, room_by_name)

        except Exception as exc:
            failed += 1
            results.append({
                "identifier": item.identifier or item.code,
                "status": "failed",
                "error": str(exc),
                "code": item.code, "title": item.title, "phase": item.phase,
            })

    summary = {"created": created, "updated": updated, "skipped": skipped, "failed": failed,
               "total": len(body.items)}
    # preview=True: discard every pending add/flush/setattr from the loop above
    # (SQLAlchemy rollback undoes flushed-but-uncommitted INSERTs too) so the
    # caller sees exactly what WOULD happen with nothing actually written --
    # same preview-then-commit shape as the Facilitator CSV import.
    if body.preview:
        db.rollback()
    else:
        db.commit()
        audit(db, p, object_type="curriculum_item", object_id="import", action="bulk_import",
              new=summary)

    return {"ok": True, "preview": body.preview, **summary, "results": results}


def _link_session(db: DBSession, ci: CurriculumItem, sqn_id: str,
                  item: CurriculumImportItem,
                  fac_by_name: dict, room_by_name: dict) -> None:
    """Try to find the parade-night session for this date/period and assign the curriculum item."""
    pn = db.query(ParadeNight).filter(
        ParadeNight.squadron_id == sqn_id,
        ParadeNight.date == item.scheduled_date,
        ParadeNight.is_archived == False,  # noqa: E712
    ).first()
    if not pn:
        return

    sess = db.query(Session).filter(
        Session.parade_night_id == pn.id,
        Session.period_number == item.session_number,
        Session.is_archived == False,  # noqa: E712
    ).first()
    if not sess:
        return

    # Only update if not already assigned to a different curriculum item
    if sess.curriculum_item_id and sess.curriculum_item_id != ci.id:
        return

    fac_id = None
    if item.facilitator_name:
        fac_id = fac_by_name.get(item.facilitator_name.strip().lower())

    room_id = None
    if item.room:
        room_id = room_by_name.get(item.room.strip().lower())

    sess.curriculum_item_id = ci.id
    sess.curriculum_code_at_time = ci.code
    sess.curriculum_title_at_time = ci.title
    sess.phase_at_time = ci.phase
    sess.element_at_time = ci.element
    if fac_id:
        sess.facilitator_id = fac_id
        sess.facilitator_display_name_at_time = item.facilitator_name
    elif item.facilitator_name:
        sess.facilitator_display_name_at_time = item.facilitator_name
    if room_id:
        sess.training_area_id = room_id
        sess.training_area_name_at_time = item.room
    elif item.room:
        sess.training_area_name_at_time = item.room


@router.post("/curriculum/import-xlsm")
async def import_curriculum_xlsm(
    file: UploadFile = File(...),
    squadron_id: str | None = None,
    db: DBSession = Depends(get_db),
    p: Principal = Depends(get_principal),
):
    """Accept an .xlsm workbook upload and import curriculum from 'zz - Program backend' sheet.

    Header row: 4. Unique key: Identifier (col 4), fallback (Module_Code, Part).
    Non-curriculum rows (missing Module_Code or Title) are silently skipped.
    """
    if p.role not in _NAT_ADMIN_ROLES:
        raise HTTPException(403, detail={"error": "forbidden",
                                         "message": "Only national_admin or system_admin can import curriculum."})

    from ..config import settings as _settings
    content = await file.read()
    if len(content) > _settings.UPLOAD_MAX_MB * 1024 * 1024:
        raise HTTPException(413, detail={"error": "file_too_large"})
    try:
        import io
        import openpyxl
        wb = openpyxl.load_workbook(io.BytesIO(content), read_only=True, data_only=True)
    except Exception as exc:
        raise HTTPException(400, detail={"error": "invalid_file",
                                          "message": f"Could not open workbook: {exc}"})

    sheet_name = "zz - Program backend"
    if sheet_name not in wb.sheetnames:
        raise HTTPException(400, detail={
            "error": "sheet_not_found",
            "message": f"Sheet '{sheet_name}' not found. Available: {wb.sheetnames}",
        })

    ws = wb[sheet_name]
    items = _parse_program_backend_sheet(ws)

    body = CurriculumImportIn(items=items, squadron_id=squadron_id, owning_level="national")
    return import_curriculum(body, db=db, p=p)


_VALID_PROGRAMS = frozenset({
    "A. Orientation", "B. Initial", "C. Junior", "D. Intermediate",
    "E. Senior", "I. Bronze", "J. Silver", "K. Gold", "M. CDT Skills",
})


def _parse_program_backend_sheet(ws) -> list[CurriculumImportItem]:
    """Parse 'zz - Program backend' worksheet rows into CurriculumImportItem objects.

    Header row: 4. Skips rows where Module_Code or Module_Title is blank,
    and rows where Program is not a recognised curriculum program.
    """
    from datetime import datetime as dt

    hdr_row = list(ws[4])
    HDR = {cell.value: idx for idx, cell in enumerate(hdr_row) if cell.value}

    def col(row, name: str, default=None):
        idx = HDR.get(name)
        return row[idx] if idx is not None and idx < len(row) else default

    items: list[CurriculumImportItem] = []
    for row in ws.iter_rows(min_row=5, values_only=True):
        if not any(v is not None for v in row[:10]):
            continue  # blank row

        program = col(row, 'Program')
        module_code = col(row, 'Module_Code')
        title = col(row, 'Module_Title')

        # Skip non-curriculum rows (header/name rows, non-programme programs)
        if not module_code or not title:
            continue
        if program and str(program).strip() not in _VALID_PROGRAMS:
            continue

        identifier = col(row, 'Identifier')
        part = col(row, 'Part')
        duration = col(row, 'Duration_Min')
        url = col(row, 'URL')
        suggested_parts = col(row, 'Suggested_#_Parts')
        elements = col(row, 'Elements')
        instr = col(row, 'Instructor_Suitability')
        scheduled = col(row, 'Scheduled')
        session_num = col(row, 'Session')

        # Convert scheduled datetime to ISO date string
        sched_date = None
        if scheduled:
            if isinstance(scheduled, dt):
                sched_date = scheduled.strftime('%Y-%m-%d')
            else:
                sched_date = str(scheduled)[:10] if scheduled else None

        items.append(CurriculumImportItem(
            identifier=str(identifier).strip() if identifier else None,
            code=str(module_code).strip().upper(),
            part_number=int(part) if part else 1,
            title=str(title).strip(),
            phase=str(program).strip() if program else "B. Initial",
            element=str(elements).strip() if elements else None,
            duration_minutes=int(duration) if duration else 60,
            part_count=int(suggested_parts) if suggested_parts else 1,
            instructor_suitability=str(instr).strip() if instr else None,
            learning_hub_url=str(url).strip() if url else None,
            scheduled_date=sched_date,
            session_number=int(session_num) if session_num else None,
            facilitator_name=str(col(row, 'Facilitator') or '').strip() or None,
            location=str(col(row, 'Location') or '').strip() or None,
            room=str(col(row, 'Room') or '').strip() or None,
        ))

    return items


# ── CSV CURRICULUM IMPORT ─────────────────────────────────────────────────────

_CSV_CURR_COL_MAP = {
    # CSV header (lower-stripped) → CurriculumImportItem field
    "training phase": "phase",
    "phase": "phase",
    "experiential code": "code",
    "code": "code",
    "module code": "code",
    "module_code": "code",
    "title": "title",
    "module title": "title",
    "module_title": "title",
    "elements": "element",
    "element": "element",
    "foundation or extension": "core_status",
    "type": "core_status",
    "instructor suitability": "instructor_suitability",
    "instructor_suitability": "instructor_suitability",
    "timing": "duration_minutes",
    "duration": "duration_minutes",
    "duration_min": "duration_minutes",
    "location": "location",
    "learning hub link": "learning_hub_url",
    "learning hub url": "learning_hub_url",
    "url": "learning_hub_url",
}


def _parse_duration(raw: str) -> int:
    """Parse a duration string like '60', '60 min', '1 hr', '1.5 hrs' → minutes."""
    if not raw:
        return 60
    raw = raw.strip().lower()
    import re
    m = re.match(r'(\d+(?:\.\d+)?)\s*(hr|hour|h)', raw)
    if m:
        return int(float(m.group(1)) * 60)
    m = re.match(r'(\d+)', raw)
    if m:
        return int(m.group(1))
    return 60


@router.post("/curriculum/import-csv")
async def import_curriculum_csv(
    file: UploadFile = File(...),
    owning_level: str = "national",
    preview: bool = False,
    db: DBSession = Depends(get_db),
    p: Principal = Depends(get_principal),
):
    """Import curriculum from a CSV file.

    Expected columns (case-insensitive, order-independent):
    Training Phase, Experiential Code, Title, Elements,
    Foundation or Extension, Instructor Suitability, Timing,
    Location, Learning Hub Link

    Returns the same summary as /curriculum/import. preview=true (Phase 3.4)
    parses and evaluates every row exactly as a real import would -- same
    create/update/skip/failed classification -- but writes nothing, so the
    caller can review before resubmitting with preview=false to commit.
    """
    import csv, io
    if p.role not in _NAT_ADMIN_ROLES:
        raise HTTPException(403, detail={
            "error": "forbidden",
            "message": "Only national_admin or system_admin can import curriculum.",
        })

    from ..config import settings as _settings
    content = await file.read()
    if len(content) > _settings.UPLOAD_MAX_MB * 1024 * 1024:
        raise HTTPException(413, detail={"error": "file_too_large"})
    try:
        text = content.decode("utf-8-sig")
    except UnicodeDecodeError:
        text = content.decode("latin-1")

    reader = csv.DictReader(io.StringIO(text))
    items: list[CurriculumImportItem] = []
    parse_errors: list[str] = []

    for i, row in enumerate(reader, start=2):
        # Normalise headers to lower-stripped
        norm = {k.strip().lower(): (v or "").strip() for k, v in row.items()}
        mapped: dict = {}
        for hdr, field in _CSV_CURR_COL_MAP.items():
            if hdr in norm and norm[hdr]:
                mapped[field] = norm[hdr]

        code = mapped.get("code", "").upper()
        title = mapped.get("title", "")
        if not code or not title:
            parse_errors.append(f"Row {i}: missing code or title — skipped.")
            continue

        # duration: parse "60 min", "1 hr", bare integer
        dur_raw = mapped.get("duration_minutes", "60")
        duration = _parse_duration(dur_raw)

        # core_status: "Foundation" → "core", "Extension" → "additional"
        cs_raw = mapped.get("core_status", "").lower()
        core_status = "core" if "foundation" in cs_raw or cs_raw in ("core", "f") else "additional"

        items.append(CurriculumImportItem(
            code=code,
            title=title,
            phase=mapped.get("phase", "B. Initial"),
            element=mapped.get("element") or None,
            duration_minutes=duration,
            instructor_suitability=mapped.get("instructor_suitability") or None,
            learning_hub_url=mapped.get("learning_hub_url") or None,
            location=mapped.get("location") or None,
            core_status=core_status,
        ))

    if not items:
        msg = "No valid rows found. " + "; ".join(parse_errors[:5]) if parse_errors else "File is empty or contains no data rows."
        raise HTTPException(400, detail={"error": "csv_parse_failed", "message": msg})

    import_body = CurriculumImportIn(items=items, owning_level=owning_level, preview=preview)
    result = import_curriculum(import_body, db, p)
    result["parse_errors"] = parse_errors
    return result


# ── CEA ACTIVITY IMPORT (retired -- see docstring below) ─────────────────────

@router.post("/activities/import-cea")
async def import_activities_cea(
    preview: bool = False,
    p: Principal = Depends(get_principal),
):
    """RETIRED (DEFECT-005, general-release qualification).

    Connected-frontend's legacy CEA import wrote squadron-scoped Activity
    rows with no persisted review/classification step and per-squadron-only
    dedup (by cea_seq_nr). Planning Workspace's CEA import
    (POST /api/planning/years/{year_id}/cea/import) has since become the
    single sanctioned pipeline: wing_admin+ only, full needs-review/classify
    workflow, import-batch history, and a squadron-local hide/note overlay
    (POST /api/planning/cea/{id}/local-hide) instead of writing squadron
    data directly. Consolidating onto one reviewed pipeline, as required.

    Still requires authentication (not publicly reachable) so this responds
    with a clear, actionable 410 rather than a bare 404 for any caller that
    still has this URL bookmarked or scripted -- see
    docs/beta/15_known_limitations.md.
    """
    raise HTTPException(410, detail={
        "error": "retired",
        "message": ("CEA import has moved to the Planning Workspace (Import CEA, "
                    "wing_admin or higher). This endpoint no longer accepts imports."),
    })


# ═══════════════════════════════════════════════════
#  SUBJECT AREA TAGS
# ═══════════════════════════════════════════════════

def _normalise_tag(name: str) -> str:
    """Lower-case, collapse internal whitespace, strip edges."""
    import re
    return re.sub(r"\s+", " ", name.strip().lower())


def _can_create_tag(p: Principal, scope: str, wing_id: str | None = None, squadron_id: str | None = None) -> None:
    """Raise 403/400 if the actor cannot create/archive a subject-area or
    facilitator-type tag at the requested scope. Mirrors _can_create_phase's
    proven role/proxy pattern (global ~ CurriculumPhase's national+system
    combined; tags only have 3 scope levels, not 4).

    Fixes a real defect found while building reference-data management UI
    (risk-register ask: every admin role must be able to create reference
    data "respective to their scope"): the prior create endpoints
    unconditionally required p.squadron_id (db.get(Squadron, p.squadron_id)),
    which wing_admin/national_admin/system_admin never have -- even while
    Proxy/Delegated Intervention is active, only p.acting_squadron_id
    changes -- so those roles could never create a tag at ANY scope,
    contradicting the ask outright.
    """
    if p.role in _WRITE_BLOCKED:
        raise HTTPException(403, detail={"error": "forbidden",
                                          "message": "Viewers and auditors cannot create tags."})
    if scope not in ("global", "wing", "squadron"):
        raise HTTPException(400, detail={"error": "invalid_scope",
                                          "message": "scope must be one of: global, wing, squadron"})
    if scope == "global":
        if p.role not in _NAT_ADMIN_ROLES:
            raise HTTPException(403, detail={"error": "forbidden",
                                              "message": "Only national_admin or system_admin can create global tags."})
    elif scope == "wing":
        if p.role not in _WING_WRITE_ROLES:
            raise HTTPException(403, detail={"error": "forbidden",
                                              "message": "Only wing_admin or above can create wing tags."})
        effective_wing = wing_id or p.wing_id
        if p.role == "wing_admin" and effective_wing != p.wing_id:
            raise HTTPException(403, detail={"error": "out_of_scope",
                                              "message": "Wing admin can only create tags for their own wing."})
    else:  # squadron
        if p.role == "sqn_admin":
            if squadron_id and squadron_id != p.squadron_id:
                raise HTTPException(403, detail={"error": "out_of_scope",
                                                  "message": "Squadron admin can only create tags for their own squadron."})
        elif p.role == "wing_admin":
            if not p.acting_squadron_id:
                raise HTTPException(403, detail={"error": "proxy_required",
                                                  "message": "Wing Admin must enter Proxy Mode to create a squadron-scope tag."})
            if squadron_id and squadron_id != p.acting_squadron_id:
                raise HTTPException(403, detail={"error": "out_of_scope",
                                                  "message": "Can only create a squadron-scope tag for the squadron currently in Proxy Mode."})
        elif p.role in _NAT_ADMIN_ROLES:
            if not p.acting_squadron_id:
                raise HTTPException(403, detail={"error": "intervention_required",
                                                  "message": "National Admin must enter Delegated Intervention Mode to create a squadron-scope tag."})
            if squadron_id and squadron_id != p.acting_squadron_id:
                raise HTTPException(403, detail={"error": "out_of_scope",
                                                  "message": "Can only create a squadron-scope tag for the squadron currently in Delegated Intervention Mode."})
        else:
            raise HTTPException(403, detail={"error": "forbidden"})


def _visible_tag_scope(p: Principal) -> tuple[str | None, str | None]:
    """Resolve (wing_id, squadron_id) for tag-visibility filtering -- same
    acting-scope-first fallback _visible_phases uses, so a wing_admin/
    national_admin's own list view sees exactly what they're allowed to
    create at that scope."""
    wing_id = p.acting_wing_id or p.wing_id
    sq_id = p.acting_squadron_id or p.squadron_id
    return wing_id, sq_id


def _tag_out(t: SubjectAreaTag) -> dict:
    return {
        "tag_id": t.id,
        "display_name": t.display_name,
        "normalised_name": t.normalised_name,
        "scope": t.scope,
        "squadron_id": t.squadron_id,
        "wing_id": t.wing_id,
        "is_active": t.is_active,
        "created_by": t.created_by,
        "created_at": t.created_at.isoformat() if t.created_at else None,
    }


class SubjectAreaTagIn(BaseModel):
    display_name: str
    scope: str = "squadron"
    wing_id: str | None = None
    squadron_id: str | None = None


@router.get("/subject-area-tags")
def list_subject_area_tags(
    include_archived: bool = False,
    p: Principal = Depends(get_principal),
    db: DBSession = Depends(get_db),
):
    """Return active (or all, when include_archived=true) tags visible to the caller."""
    wing_id, sq_id = _visible_tag_scope(p)
    require_can_view_squadron(p, sq_id or "", wing_id)
    conditions = [SubjectAreaTag.scope == "global"]
    if wing_id:
        conditions.append((SubjectAreaTag.scope == "wing") & (SubjectAreaTag.wing_id == wing_id))
    elif p.role in _NAT_ADMIN_ROLES:
        conditions.append(SubjectAreaTag.scope == "wing")
    if sq_id:
        conditions.append((SubjectAreaTag.scope == "squadron") & (SubjectAreaTag.squadron_id == sq_id))
    elif p.role in _NAT_ADMIN_ROLES:
        conditions.append(SubjectAreaTag.scope == "squadron")
    from sqlalchemy import or_ as _or_tags
    q = db.query(SubjectAreaTag).filter(_or_tags(*conditions))
    if not include_archived:
        q = q.filter(SubjectAreaTag.is_active == True)  # noqa: E712
    rows = q.order_by(SubjectAreaTag.display_name).all()
    return [_tag_out(t) for t in rows]


@router.post("/subject-area-tags", status_code=201)
def create_subject_area_tag(
    body: SubjectAreaTagIn,
    p: Principal = Depends(get_principal),
    db: DBSession = Depends(get_db),
):
    """Create a new subject-area tag at the requested scope (sqn_admin: squadron
    only; wing_admin/national_admin/system_admin: their own wing/global, or a
    squadron scope while Proxy/Delegated Intervention is active)."""
    scope = body.scope if body.scope in ("global", "wing", "squadron") else "squadron"
    _can_create_tag(p, scope, body.wing_id, body.squadron_id)

    display = body.display_name.strip()
    if not display:
        raise HTTPException(400, detail={"error": "tag_name_blank"})
    if len(display) > 80:
        raise HTTPException(400, detail={"error": "tag_name_too_long", "max": 80})

    norm = _normalise_tag(display)

    wing_id = body.wing_id or ((p.acting_wing_id or p.wing_id) if scope in ("wing", "squadron") else None)
    squadron_id = body.squadron_id or ((p.acting_squadron_id or p.squadron_id) if scope == "squadron" else None)

    existing = (
        db.query(SubjectAreaTag)
        .filter(
            SubjectAreaTag.normalised_name == norm,
            SubjectAreaTag.is_active == True,  # noqa: E712
            (
                (SubjectAreaTag.squadron_id == squadron_id) |
                (SubjectAreaTag.wing_id == wing_id) |
                (SubjectAreaTag.scope == "global")
            ),
        )
        .first()
    )
    if existing:
        raise HTTPException(409, detail={"error": "tag_already_exists", "existing_id": existing.id})

    tag = SubjectAreaTag(
        squadron_id=squadron_id,
        wing_id=wing_id,
        scope=scope,
        display_name=display,
        normalised_name=norm,
        is_active=True,
        created_by=p.user_id,
    )
    db.add(tag)
    db.commit()
    db.refresh(tag)
    audit(db, p, object_type="SubjectAreaTag", object_id=tag.id, action="create",
          new={"display_name": tag.display_name, "scope": tag.scope})
    return _tag_out(tag)


@router.delete("/subject-area-tags/{tag_id}", status_code=200)
def archive_subject_area_tag(
    tag_id: str,
    p: Principal = Depends(get_principal),
    db: DBSession = Depends(get_db),
):
    """Archive a subject-area tag (does not remove existing facilitator references).
    Scope-checked the same way as creation -- previously any sqn_admin could
    archive ANY tag regardless of owning scope (a role-only check, no
    ownership verification at all)."""
    tag = db.get(SubjectAreaTag, tag_id)
    if not tag:
        raise HTTPException(404, detail={"error": "tag_not_found"})
    _can_create_tag(p, tag.scope, tag.wing_id, tag.squadron_id)
    tag.is_active = False
    db.commit()
    audit(db, p, object_type="SubjectAreaTag", object_id=tag_id, action="archive")
    return {"ok": True}


@router.post("/subject-area-tags/{tag_id}/restore", status_code=200)
def restore_subject_area_tag(
    tag_id: str,
    p: Principal = Depends(get_principal),
    db: DBSession = Depends(get_db),
):
    """REM-23: restore an archived subject-area tag."""
    tag = db.get(SubjectAreaTag, tag_id)
    if not tag:
        raise HTTPException(404, detail={"error": "tag_not_found"})
    _can_create_tag(p, tag.scope, tag.wing_id, tag.squadron_id)
    if tag.is_active:
        raise HTTPException(409, detail={"error": "tag_already_active"})
    tag.is_active = True
    db.commit()
    audit(db, p, object_type="SubjectAreaTag", object_id=tag_id, action="restore")
    return {"ok": True}


# ── Facilitator Type reference data — mirrors subject-area-tags exactly
# (remediation program Section 6, Stage 3). Facilitator.type stays free-text;
# this is advisory/governed reference data, not a hard foreign key. ──

def _fac_type_out(t: FacilitatorTypeTag) -> dict:
    return {
        "tag_id": t.id,
        "display_name": t.display_name,
        "normalised_name": t.normalised_name,
        "scope": t.scope,
        "squadron_id": t.squadron_id,
        "wing_id": t.wing_id,
        "is_active": t.is_active,
        "created_by": t.created_by,
        "created_at": t.created_at.isoformat() if t.created_at else None,
    }


class FacilitatorTypeTagIn(BaseModel):
    display_name: str
    scope: str = "squadron"
    wing_id: str | None = None
    squadron_id: str | None = None


@router.get("/facilitator-type-tags")
def list_facilitator_type_tags(
    include_archived: bool = False,
    p: Principal = Depends(get_principal),
    db: DBSession = Depends(get_db),
):
    """Return active (or all, when include_archived=true) facilitator-type tags."""
    wing_id, sq_id = _visible_tag_scope(p)
    require_can_view_squadron(p, sq_id or "", wing_id)
    conditions = [FacilitatorTypeTag.scope == "global"]
    if wing_id:
        conditions.append((FacilitatorTypeTag.scope == "wing") & (FacilitatorTypeTag.wing_id == wing_id))
    elif p.role in _NAT_ADMIN_ROLES:
        conditions.append(FacilitatorTypeTag.scope == "wing")
    if sq_id:
        conditions.append((FacilitatorTypeTag.scope == "squadron") & (FacilitatorTypeTag.squadron_id == sq_id))
    elif p.role in _NAT_ADMIN_ROLES:
        conditions.append(FacilitatorTypeTag.scope == "squadron")
    from sqlalchemy import or_ as _or_factype
    q = db.query(FacilitatorTypeTag).filter(_or_factype(*conditions))
    if not include_archived:
        q = q.filter(FacilitatorTypeTag.is_active == True)  # noqa: E712
    rows = q.order_by(FacilitatorTypeTag.display_name).all()
    return [_fac_type_out(t) for t in rows]


@router.post("/facilitator-type-tags", status_code=201)
def create_facilitator_type_tag(
    body: FacilitatorTypeTagIn,
    p: Principal = Depends(get_principal),
    db: DBSession = Depends(get_db),
):
    """Create a new facilitator-type tag at the requested scope (sqn_admin: squadron
    only; wing_admin/national_admin/system_admin: their own wing/global, or a
    squadron scope while Proxy/Delegated Intervention is active)."""
    scope = body.scope if body.scope in ("global", "wing", "squadron") else "squadron"
    _can_create_tag(p, scope, body.wing_id, body.squadron_id)

    display = body.display_name.strip()
    if not display:
        raise HTTPException(400, detail={"error": "tag_name_blank"})
    if len(display) > 80:
        raise HTTPException(400, detail={"error": "tag_name_too_long", "max": 80})

    norm = _normalise_tag(display)

    wing_id = body.wing_id or ((p.acting_wing_id or p.wing_id) if scope in ("wing", "squadron") else None)
    squadron_id = body.squadron_id or ((p.acting_squadron_id or p.squadron_id) if scope == "squadron" else None)

    existing = (
        db.query(FacilitatorTypeTag)
        .filter(
            FacilitatorTypeTag.normalised_name == norm,
            FacilitatorTypeTag.is_active == True,  # noqa: E712
            (
                (FacilitatorTypeTag.squadron_id == squadron_id) |
                (FacilitatorTypeTag.wing_id == wing_id) |
                (FacilitatorTypeTag.scope == "global")
            ),
        )
        .first()
    )
    if existing:
        raise HTTPException(409, detail={"error": "tag_already_exists", "existing_id": existing.id})

    tag = FacilitatorTypeTag(
        squadron_id=squadron_id,
        wing_id=wing_id,
        scope=scope,
        display_name=display,
        normalised_name=norm,
        is_active=True,
        created_by=p.user_id,
    )
    db.add(tag)
    db.commit()
    db.refresh(tag)
    audit(db, p, object_type="FacilitatorTypeTag", object_id=tag.id, action="create",
          new={"display_name": tag.display_name, "scope": tag.scope})
    return _fac_type_out(tag)


@router.delete("/facilitator-type-tags/{tag_id}", status_code=200)
def archive_facilitator_type_tag(
    tag_id: str,
    p: Principal = Depends(get_principal),
    db: DBSession = Depends(get_db),
):
    """Archive a facilitator-type tag (does not remove existing facilitator references).
    Scope-checked the same way as creation -- previously any sqn_admin could
    archive ANY tag regardless of owning scope (a role-only check, no
    ownership verification at all)."""
    tag = db.get(FacilitatorTypeTag, tag_id)
    if not tag:
        raise HTTPException(404, detail={"error": "tag_not_found"})
    _can_create_tag(p, tag.scope, tag.wing_id, tag.squadron_id)
    tag.is_active = False
    db.commit()
    audit(db, p, object_type="FacilitatorTypeTag", object_id=tag_id, action="archive")
    return {"ok": True}


@router.post("/facilitator-type-tags/{tag_id}/restore", status_code=200)
def restore_facilitator_type_tag(
    tag_id: str,
    p: Principal = Depends(get_principal),
    db: DBSession = Depends(get_db),
):
    """REM-23: restore an archived facilitator-type tag."""
    tag = db.get(FacilitatorTypeTag, tag_id)
    if not tag:
        raise HTTPException(404, detail={"error": "tag_not_found"})
    _can_create_tag(p, tag.scope, tag.wing_id, tag.squadron_id)
    if tag.is_active:
        raise HTTPException(409, detail={"error": "tag_already_active"})
    tag.is_active = True
    db.commit()
    audit(db, p, object_type="FacilitatorTypeTag", object_id=tag_id, action="restore")
    return {"ok": True}


# ── Session Status Reason reference data — mirrors facilitator-type-tags
# exactly (REM-23 continuation). The #or-reason dropdown's reason text
# stays free-text on SessionStatusHistory; this is advisory/governed
# reference data, not a hard foreign key. ──

def _reason_out(t: SessionStatusReasonTag) -> dict:
    return {
        "tag_id": t.id,
        "display_name": t.display_name,
        "normalised_name": t.normalised_name,
        "scope": t.scope,
        "squadron_id": t.squadron_id,
        "wing_id": t.wing_id,
        "is_active": t.is_active,
        "created_by": t.created_by,
        "created_at": t.created_at.isoformat() if t.created_at else None,
    }


class SessionStatusReasonTagIn(BaseModel):
    display_name: str
    scope: str = "squadron"
    wing_id: str | None = None
    squadron_id: str | None = None


@router.get("/session-status-reason-tags")
def list_session_status_reason_tags(
    include_archived: bool = False,
    p: Principal = Depends(get_principal),
    db: DBSession = Depends(get_db),
):
    """Return active (or all, when include_archived=true) session-status reason tags."""
    wing_id, sq_id = _visible_tag_scope(p)
    require_can_view_squadron(p, sq_id or "", wing_id)
    conditions = [SessionStatusReasonTag.scope == "global"]
    if wing_id:
        conditions.append((SessionStatusReasonTag.scope == "wing") & (SessionStatusReasonTag.wing_id == wing_id))
    elif p.role in _NAT_ADMIN_ROLES:
        conditions.append(SessionStatusReasonTag.scope == "wing")
    if sq_id:
        conditions.append((SessionStatusReasonTag.scope == "squadron") & (SessionStatusReasonTag.squadron_id == sq_id))
    elif p.role in _NAT_ADMIN_ROLES:
        conditions.append(SessionStatusReasonTag.scope == "squadron")
    from sqlalchemy import or_ as _or_reason
    q = db.query(SessionStatusReasonTag).filter(_or_reason(*conditions))
    if not include_archived:
        q = q.filter(SessionStatusReasonTag.is_active == True)  # noqa: E712
    rows = q.order_by(SessionStatusReasonTag.display_name).all()
    return [_reason_out(t) for t in rows]


@router.post("/session-status-reason-tags", status_code=201)
def create_session_status_reason_tag(
    body: SessionStatusReasonTagIn,
    p: Principal = Depends(get_principal),
    db: DBSession = Depends(get_db),
):
    """Create a new session-status reason tag at the requested scope (sqn_admin: squadron
    only; wing_admin/national_admin/system_admin: their own wing/global, or a
    squadron scope while Proxy/Delegated Intervention is active)."""
    scope = body.scope if body.scope in ("global", "wing", "squadron") else "squadron"
    _can_create_tag(p, scope, body.wing_id, body.squadron_id)

    display = body.display_name.strip()
    if not display:
        raise HTTPException(400, detail={"error": "tag_name_blank"})
    if len(display) > 80:
        raise HTTPException(400, detail={"error": "tag_name_too_long", "max": 80})

    norm = _normalise_tag(display)

    wing_id = body.wing_id or ((p.acting_wing_id or p.wing_id) if scope in ("wing", "squadron") else None)
    squadron_id = body.squadron_id or ((p.acting_squadron_id or p.squadron_id) if scope == "squadron" else None)

    existing = (
        db.query(SessionStatusReasonTag)
        .filter(
            SessionStatusReasonTag.normalised_name == norm,
            SessionStatusReasonTag.is_active == True,  # noqa: E712
            (
                (SessionStatusReasonTag.squadron_id == squadron_id) |
                (SessionStatusReasonTag.wing_id == wing_id) |
                (SessionStatusReasonTag.scope == "global")
            ),
        )
        .first()
    )
    if existing:
        raise HTTPException(409, detail={"error": "tag_already_exists", "existing_id": existing.id})

    tag = SessionStatusReasonTag(
        squadron_id=squadron_id,
        wing_id=wing_id,
        scope=scope,
        display_name=display,
        normalised_name=norm,
        is_active=True,
        created_by=p.user_id,
    )
    db.add(tag)
    db.commit()
    db.refresh(tag)
    audit(db, p, object_type="SessionStatusReasonTag", object_id=tag.id, action="create",
          new={"display_name": tag.display_name, "scope": tag.scope})
    return _reason_out(tag)


@router.delete("/session-status-reason-tags/{tag_id}", status_code=200)
def archive_session_status_reason_tag(
    tag_id: str,
    p: Principal = Depends(get_principal),
    db: DBSession = Depends(get_db),
):
    """Archive a session-status reason tag (does not remove existing history rows).
    Scope-checked the same way as creation."""
    tag = db.get(SessionStatusReasonTag, tag_id)
    if not tag:
        raise HTTPException(404, detail={"error": "tag_not_found"})
    _can_create_tag(p, tag.scope, tag.wing_id, tag.squadron_id)
    tag.is_active = False
    db.commit()
    audit(db, p, object_type="SessionStatusReasonTag", object_id=tag_id, action="archive")
    return {"ok": True}


@router.post("/session-status-reason-tags/{tag_id}/restore", status_code=200)
def restore_session_status_reason_tag(
    tag_id: str,
    p: Principal = Depends(get_principal),
    db: DBSession = Depends(get_db),
):
    """REM-23: restore an archived session-status reason tag."""
    tag = db.get(SessionStatusReasonTag, tag_id)
    if not tag:
        raise HTTPException(404, detail={"error": "tag_not_found"})
    _can_create_tag(p, tag.scope, tag.wing_id, tag.squadron_id)
    if tag.is_active:
        raise HTTPException(409, detail={"error": "tag_already_active"})
    tag.is_active = True
    db.commit()
    audit(db, p, object_type="SessionStatusReasonTag", object_id=tag_id, action="restore")
    return {"ok": True}


# ── Parade Night Templates (WORK-17) ──────────────────────────────────────────


class TemplateIn(BaseModel):
    name: str = Field(..., max_length=100)
    description: str | None = Field(None, max_length=500)


class TemplateSessionOut(BaseModel):
    id: str
    period_number: int
    curriculum_item_id: str | None
    custom_title: str | None
    facilitator_id: str | None
    training_area_id: str | None
    cadet_group: str | None
    notes: str | None


class TemplateOut(BaseModel):
    id: str
    name: str
    description: str | None
    squadron_id: str
    session_count: int
    created_at: str
    sessions: list[TemplateSessionOut] | None = None


def _template_out(t: ParadeNightTemplate, include_sessions: bool = False) -> dict:
    return {
        "id": t.id,
        "name": t.name,
        "description": t.description,
        "squadron_id": t.squadron_id,
        "session_count": len(t.sessions) if t.sessions is not None else 0,
        "created_at": t.created_at.isoformat() if t.created_at else None,
        "sessions": [
            {
                "id": s.id,
                "period_number": s.period_number,
                "curriculum_item_id": s.curriculum_item_id,
                "custom_title": s.custom_title,
                "facilitator_id": s.facilitator_id,
                "training_area_id": s.training_area_id,
                "cadet_group": s.cadet_group,
                "notes": s.notes,
            }
            for s in t.sessions
        ] if include_sessions else None,
    }


@router.post("/parade-nights/{pnid}/save-as-template")
def save_parade_night_as_template(
    pnid: str,
    body: TemplateIn,
    db: DBSession = Depends(get_db),
    p: Principal = Depends(get_principal),
):
    """WORK-17: save a parade night's session structure as a reusable named template."""
    pn = db.get(ParadeNight, pnid)
    if not pn or pn.is_archived:
        raise HTTPException(404, detail={"error": "source_not_found"})
    require_can_write_squadron(p, pn.squadron_id, pn.wing_id)

    sessions = (
        db.query(Session)
        .filter(Session.parade_night_id == pnid, Session.is_archived.is_(False))
        .order_by(Session.period_number)
        .all()
    )
    if not sessions:
        raise HTTPException(400, detail={"error": "no_sessions", "msg": "This parade night has no sessions to save as a template."})

    import uuid as _uuid
    tmpl = ParadeNightTemplate(
        id=str(_uuid.uuid4()),
        squadron_id=pn.squadron_id,
        name=body.name.strip(),
        description=body.description.strip() if body.description else None,
        created_by=p.user_id,
    )
    db.add(tmpl)
    db.flush()

    for s in sessions:
        db.add(ParadeNightTemplateSession(
            id=str(_uuid.uuid4()),
            template_id=tmpl.id,
            period_number=s.period_number,
            curriculum_item_id=s.curriculum_item_id,
            custom_title=s.custom_title,
            facilitator_id=s.facilitator_id,
            training_area_id=s.training_area_id,
            cadet_group=s.cadet_group,
            notes=None,
        ))

    db.commit()
    db.refresh(tmpl)
    audit(db, p, object_type="parade_night_template", object_id=tmpl.id,
          action="create", new={"name": tmpl.name, "from_parade_night_id": pnid, "session_count": len(sessions)})
    return _template_out(tmpl, include_sessions=True)


@router.get("/parade-night-templates")
def list_parade_night_templates(
    db: DBSession = Depends(get_db),
    p: Principal = Depends(get_principal),
):
    """WORK-17: list all non-archived templates for the current squadron."""
    sqn_id = p.squadron_id
    if not sqn_id:
        raise HTTPException(400, detail={"error": "squadron_required"})
    templates = (
        db.query(ParadeNightTemplate)
        .filter(
            ParadeNightTemplate.squadron_id == sqn_id,
            ParadeNightTemplate.is_archived.is_(False),
        )
        .order_by(ParadeNightTemplate.created_at.desc())
        .all()
    )
    return [_template_out(t) for t in templates]


@router.get("/parade-night-templates/{tid}")
def get_parade_night_template(
    tid: str,
    db: DBSession = Depends(get_db),
    p: Principal = Depends(get_principal),
):
    """WORK-17: get a single template with its sessions."""
    t = db.get(ParadeNightTemplate, tid)
    if not t or t.is_archived:
        raise HTTPException(404, detail={"error": "template_not_found"})
    sqn = db.get(Squadron, t.squadron_id)
    require_can_view_squadron(p, t.squadron_id, sqn.wing_id if sqn else None)
    return _template_out(t, include_sessions=True)


@router.post("/parade-night-templates/{tid}/apply/{pnid}")
def apply_parade_night_template(
    tid: str,
    pnid: str,
    db: DBSession = Depends(get_db),
    p: Principal = Depends(get_principal),
):
    """WORK-17: apply a template's session structure to a parade night.

    Periods already occupied in the target are skipped, not overwritten.
    Returns {ok, applied, skipped}.
    """
    import uuid as _uuid
    t = db.get(ParadeNightTemplate, tid)
    if not t or t.is_archived:
        raise HTTPException(404, detail={"error": "template_not_found"})
    pn = db.get(ParadeNight, pnid)
    if not pn or pn.is_archived:
        raise HTTPException(404, detail={"error": "target_not_found"})
    require_can_write_squadron(p, pn.squadron_id, pn.wing_id)
    if t.squadron_id != pn.squadron_id:
        raise HTTPException(400, detail={"error": "cross_squadron_copy_not_permitted"})

    existing_periods = {
        s.period_number
        for s in db.query(Session).filter(
            Session.parade_night_id == pnid, Session.is_archived.is_(False)
        )
    }

    applied, skipped = 0, 0
    for ts in sorted(t.sessions, key=lambda s: s.period_number):
        if ts.period_number in existing_periods:
            skipped += 1
            continue
        ci = db.get(CurriculumItem, ts.curriculum_item_id) if ts.curriculum_item_id else None
        fac = db.get(Facilitator, ts.facilitator_id) if ts.facilitator_id else None
        ta = db.get(TrainingArea, ts.training_area_id) if ts.training_area_id else None
        new_sess = Session(
            id=str(_uuid.uuid4()),
            parade_night_id=pnid,
            squadron_id=pn.squadron_id,
            period_number=ts.period_number,
            curriculum_item_id=ts.curriculum_item_id,
            curriculum_code_at_time=ci.code if ci else None,
            curriculum_title_at_time=ci.title if ci else None,
            custom_title=ts.custom_title,
            facilitator_id=ts.facilitator_id,
            facilitator_rank_at_time=fac.current_rank if fac else None,
            facilitator_display_name_at_time=(
                f"{fac.current_rank or ''} {fac.last_name}".strip() if fac else None
            ),
            training_area_id=ts.training_area_id,
            training_area_name_at_time=ta.name if ta else None,
            cadet_group=ts.cadet_group,
            status="planned",
        )
        db.add(new_sess)
        existing_periods.add(ts.period_number)
        applied += 1

    db.commit()
    audit(db, p, object_type="parade_night", object_id=pnid,
          action="apply_template", new={"template_id": tid, "applied": applied})
    return {"ok": True, "applied": applied, "skipped": skipped}


@router.delete("/parade-night-templates/{tid}")
def archive_parade_night_template(
    tid: str,
    db: DBSession = Depends(get_db),
    p: Principal = Depends(get_principal),
):
    """WORK-17: archive (soft-delete) a parade night template."""
    t = db.get(ParadeNightTemplate, tid)
    if not t:
        raise HTTPException(404, detail={"error": "template_not_found"})
    sqn = db.get(Squadron, t.squadron_id)
    require_can_write_squadron(p, t.squadron_id, sqn.wing_id if sqn else None)
    if t.is_archived:
        raise HTTPException(409, detail={"error": "already_archived"})
    t.is_archived = True
    db.commit()
    audit(db, p, object_type="parade_night_template", object_id=tid, action="archive")
    return {"ok": True}


class _BulkApplyTemplateIn(BaseModel):
    template_id: str
    parade_night_ids: List[str]
    conflict_resolution: Literal["skip_night", "alongside", "replace_draft"] = "skip_night"
    dry_run: bool = False


@router.post("/parade-nights/bulk-apply-template")
def bulk_apply_template(
    body: _BulkApplyTemplateIn,
    db: DBSession = Depends(get_db),
    p: Principal = Depends(get_principal),
):
    """BULK-01: apply a saved template to multiple selected parade nights.

    conflict_resolution:
      skip_night    — skip the entire night if it has any existing sessions
      alongside     — add template sessions even if the period is already occupied
      replace_draft — remove planned/draft sessions first, then apply template

    dry_run=true returns counts without writing; does NOT copy delivered outcomes.
    """
    import uuid as _uuid

    t = db.get(ParadeNightTemplate, body.template_id)
    if not t or t.is_archived:
        raise HTTPException(404, detail={"error": "template_not_found"})

    _DRAFT_STATUSES = {"planned", "rescheduled"}
    batch_id = str(_uuid.uuid4())

    nights_processed, nights_skipped = 0, 0
    sessions_added, sessions_skipped = 0, 0
    conflict_nights = 0

    for pnid in body.parade_night_ids:
        pn = db.get(ParadeNight, pnid)
        if not pn or pn.is_archived:
            nights_skipped += 1
            continue
        require_can_write_squadron(p, pn.squadron_id, pn.wing_id)
        if t.squadron_id != pn.squadron_id:
            nights_skipped += 1
            continue

        existing_sessions = db.query(Session).filter(
            Session.parade_night_id == pnid,
            Session.is_archived.is_(False),
        ).all()
        existing_periods = {s.period_number for s in existing_sessions}

        has_conflict = bool(existing_sessions)
        if has_conflict:
            conflict_nights += 1

        if body.conflict_resolution == "skip_night" and has_conflict:
            nights_skipped += 1
            sessions_skipped += len(t.sessions)
            continue

        if body.conflict_resolution == "replace_draft" and not body.dry_run:
            for s in existing_sessions:
                if s.status in _DRAFT_STATUSES:
                    s.is_archived = True
            existing_periods = {
                s.period_number for s in existing_sessions
                if s.status not in _DRAFT_STATUSES
            }

        nights_processed += 1
        for ts in sorted(t.sessions, key=lambda s: s.period_number):
            if body.conflict_resolution == "skip_night":
                # already handled above
                pass
            if body.conflict_resolution != "alongside" and ts.period_number in existing_periods:
                sessions_skipped += 1
                continue
            ci = db.get(CurriculumItem, ts.curriculum_item_id) if ts.curriculum_item_id else None
            fac = db.get(Facilitator, ts.facilitator_id) if ts.facilitator_id else None
            ta = db.get(TrainingArea, ts.training_area_id) if ts.training_area_id else None
            if not body.dry_run:
                new_sess = Session(
                    id=str(_uuid.uuid4()),
                    parade_night_id=pnid,
                    squadron_id=pn.squadron_id,
                    period_number=ts.period_number,
                    curriculum_item_id=ts.curriculum_item_id,
                    curriculum_code_at_time=ci.code if ci else None,
                    curriculum_title_at_time=ci.title if ci else None,
                    custom_title=ts.custom_title,
                    facilitator_id=ts.facilitator_id,
                    facilitator_rank_at_time=fac.current_rank if fac else None,
                    facilitator_display_name_at_time=(
                        f"{fac.current_rank or ''} {fac.last_name}".strip() if fac else None
                    ),
                    training_area_id=ts.training_area_id,
                    training_area_name_at_time=ta.name if ta else None,
                    cadet_group=ts.cadet_group,
                    status="planned",
                )
                db.add(new_sess)
            sessions_added += 1

        if not body.dry_run:
            audit(db, p, object_type="parade_night", object_id=pnid,
                  action="bulk_apply_template",
                  new={"template_id": body.template_id, "batch_id": batch_id,
                       "conflict_resolution": body.conflict_resolution})

    if not body.dry_run:
        db.commit()

    return {
        "ok": True,
        "dry_run": body.dry_run,
        "batch_id": batch_id,
        "nights_processed": nights_processed,
        "nights_skipped": nights_skipped,
        "conflict_nights": conflict_nights,
        "sessions_added": sessions_added,
        "sessions_skipped": sessions_skipped,
    }
