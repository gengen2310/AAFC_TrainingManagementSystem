#!/usr/bin/env python3
"""Import Wing HQ Calendar from 7WG Yearly Schedule of Events XLSX.

Usage:
    # Dry-run first — always check counts before committing
    railway run python scripts/import_wing_hq_calendar.py \\
        --file "7WG Yearly Schedule of Events - MASTER.xlsx" \\
        --wing-code 7WG --year 2026 --dry-run

    # Commit after dry-run looks correct
    railway run python scripts/import_wing_hq_calendar.py \\
        --file "7WG Yearly Schedule of Events - MASTER.xlsx" \\
        --wing-code 7WG --year 2026

Expected XLSX format (7WG Master Schedule):
    The importer auto-detects columns by keyword matching. Supported column
    name variants (case-insensitive):

    - Date column:  "Date", "Dates", "Date Range", "Start", "When"
    - Title column: "Title", "Event", "Activity", "Name", "Description"
    - Type column:  "Type", "Category", "Event Type"
    - Audience:     "Audience", "Who", "For"
    - Importance:   "Importance", "Priority", "Category"
    - Notes:        "Notes", "Details", "Source", "Remarks"
    - Location:     "Location", "Venue", "Where"

    Date formats accepted:
        "21 Feb 26"                  → 2026-02-21
        "21 - 22 Feb 26"             → 2026-02-21 to 2026-02-22
        "14 Jan 26 - 22 Mar 26"      → 2026-01-14 to 2026-03-22
        "14/01/2026"                 → 2026-01-14
        "2026-01-14"                 → 2026-01-14

    Rows skipped automatically:
        - Blank title
        - Title matches: "Insert new rows", "N/A", month headings
        - Date cell is a month name only (e.g. "January")
        - Row is entirely blank

    Event type mapping (title keyword → event_type):
        "parade night"   → home_parade
        "public holiday" → public_holiday
        "school holiday" → school_holiday
        "conference"     → meeting
        "biv", "bivouac" → cadet_training
        "camp", "course" → cadet_training
        "competition"    → competition
        "ceremony"       → ceremony
        "ball"           → ceremony
        "staff"          → staff_activity
        (default)        → wing_event

    Importance mapping (title keyword → planning_importance):
        "must attend" OR type in (competition, ceremony) → must_attend
        "key event"   → key_event
        "biv"         → must_attend
        "camp"        → must_attend
        "conference"  → key_event
        (default)     → recommended

Idempotent upsert:
    Source reference = "{source_system}:{sheet_name}:row{row_number}"
    On re-run, existing records with the same source_reference are
    updated (not duplicated) if the title or dates have changed.
    Records NOT in the new import are NOT deleted (safe to re-run).

Safety:
    - Never prints DATABASE_URL or credentials
    - Transaction-safe: all rows in one transaction; any failure rolls back
    - Dry-run shows exactly what would be created/updated without writing
"""
import argparse
import os
import re
import sys
import uuid
from datetime import date, datetime, timedelta
from pathlib import Path

# ── Bootstrap: add backend to PYTHONPATH ────────────────────────────────────
_SCRIPT_DIR = Path(__file__).resolve().parent
_BACKEND_DIR = _SCRIPT_DIR.parent
sys.path.insert(0, str(_BACKEND_DIR))

from sqlalchemy import create_engine
from sqlalchemy.orm import sessionmaker

DATABASE_URL = os.environ.get("DATABASE_URL", "")
if not DATABASE_URL:
    print("ERROR: DATABASE_URL environment variable not set.", file=sys.stderr)
    sys.exit(1)

# SQLAlchemy 2.0: asyncpg URIs must use postgresql+psycopg2 for sync scripts
if DATABASE_URL.startswith("postgres://"):
    DATABASE_URL = DATABASE_URL.replace("postgres://", "postgresql://", 1)

engine = create_engine(DATABASE_URL, pool_pre_ping=True)
SessionLocal = sessionmaker(bind=engine)

from app.models import Wing
from app.models.wing_calendar import WingHQEvent
from app.database import Base

# ── Date parsing helpers ─────────────────────────────────────────────────────

_MONTHS = {
    "jan": 1, "feb": 2, "mar": 3, "apr": 4, "may": 5, "jun": 6,
    "jul": 7, "aug": 8, "sep": 9, "oct": 10, "nov": 11, "dec": 12,
}
_MONTH_NAMES_FULL = {
    "january": 1, "february": 2, "march": 3, "april": 4, "may": 5, "june": 6,
    "july": 7, "august": 8, "september": 9, "october": 10, "november": 11, "december": 12,
}


def _two_digit_year(yy: int, base_year: int) -> int:
    """Convert 2-digit year to 4-digit using base_year as reference."""
    century = (base_year // 100) * 100
    full = century + yy
    if full > base_year + 50:
        full -= 100
    return full


def parse_date_range(raw: str, base_year: int) -> tuple[str | None, str | None]:
    """Parse a date or date range string into (start_date, end_date) ISO strings.

    Handles:
      "21 Feb 26"              → 2026-02-21
      "21 - 22 Feb 26"         → 2026-02-21, 2026-02-22
      "28 Feb 26 - 1 Mar 26"   → 2026-02-28, 2026-03-01
      "14/01/2026"             → 2026-01-14
      "2026-01-14"             → 2026-01-14
    Returns (None, None) if unparseable.
    """
    if not raw:
        return None, None
    raw = str(raw).strip()

    # Already ISO
    if re.match(r"^\d{4}-\d{2}-\d{2}$", raw):
        return raw, None

    # Slash format: 14/01/2026
    m = re.match(r"^(\d{1,2})/(\d{1,2})/(\d{4})$", raw)
    if m:
        d, mo, y = int(m.group(1)), int(m.group(2)), int(m.group(3))
        try:
            return date(y, mo, d).isoformat(), None
        except ValueError:
            return None, None

    # Normalise dash/en-dash/em-dash separators → " TO "
    raw_norm = re.sub(r"\s*[-–—]\s*", " TO ", raw)
    parts = raw_norm.split(" TO ")

    def _parse_one(s: str, fallback_month: int | None = None, fallback_year: int | None = None):
        s = s.strip()
        tokens = s.split()
        if not tokens:
            return None
        day_t = month_t = year_t = None
        for tok in tokens:
            t = tok.strip(".,")
            if t.lower() in _MONTHS:
                month_t = _MONTHS[t.lower()]
            elif t.lower() in _MONTH_NAMES_FULL:
                month_t = _MONTH_NAMES_FULL[t.lower()]
            elif re.match(r"^\d{4}$", t):
                year_t = int(t)
            elif re.match(r"^\d{1,2}$", t):
                val = int(t)
                if day_t is None:
                    # First small number → day
                    day_t = val
                elif year_t is None:
                    # Second small number: if we already know the month it's a year ("D M YY")
                    if month_t is not None:
                        year_t = _two_digit_year(val, base_year) if val < 100 else val
                    elif val > 31:
                        year_t = _two_digit_year(val, base_year)
                    # else: ignore (ambiguous extra number)
        # Fill from fallbacks
        if month_t is None:
            month_t = fallback_month
        if year_t is None:
            year_t = fallback_year if fallback_year is not None else base_year
        if year_t and year_t < 100:
            year_t = _two_digit_year(year_t, base_year)
        if day_t and month_t and year_t:
            try:
                return date(year_t, month_t, day_t).isoformat()
            except ValueError:
                return None
        return None

    if len(parts) == 1:
        return _parse_one(parts[0]), None

    if len(parts) == 2:
        # Parse end first to get month/year fallbacks for a bare-day start ("21 TO 22 Feb 26")
        ed = _parse_one(parts[1])
        fallback_m = int(ed[5:7]) if ed else None
        fallback_y = int(ed[:4]) if ed else None
        sd = _parse_one(parts[0], fallback_m, fallback_y)
        return sd, ed

    return None, None


# ── Event type / importance inference ────────────────────────────────────────

def _infer_event_type(title: str, raw_type: str = "") -> str:
    t = title.lower()
    rt = raw_type.lower()
    if "parade night" in t or "home parade" in t:
        return "home_parade"
    if "public holiday" in t or rt == "public_holiday":
        return "public_holiday"
    if "school holiday" in t or rt == "school_holiday":
        return "school_holiday"
    if "conference" in t or "meeting" in rt:
        return "meeting"
    if any(x in t for x in ("biv", "bivouac", "camp", "cadet training", "cadet course")):
        return "cadet_training"
    if any(x in t for x in ("adult training", "adult course", "staff training")):
        return "adult_training"
    if "competition" in t or rt == "competition":
        return "competition"
    if any(x in t for x in ("ceremony", "ball", "freedom of entry", "dining in", "dining-in")):
        return "ceremony"
    if any(x in t for x in ("staff", "oic conference", "trgo")):
        return "staff_activity"
    if "course" in t:
        return "course"
    return "wing_event"


def _infer_importance(title: str, event_type: str, raw_importance: str = "") -> str:
    ri = raw_importance.lower()
    if ri in ("must_attend", "must attend", "mandatory", "compulsory"):
        return "must_attend"
    if ri in ("key_event", "key event", "key"):
        return "key_event"
    if ri in ("optional", "noting"):
        return ri
    t = title.lower()
    if event_type in ("competition", "ceremony"):
        return "must_attend"
    if any(x in t for x in ("biv", "bivouac", "wing biv", "training weekend", "gold cadet")):
        return "must_attend"
    if any(x in t for x in ("conference", "staff ball", "freedom of entry", "drill competition")):
        return "key_event"
    if event_type in ("public_holiday", "school_holiday", "home_parade"):
        return "home_parade" if event_type == "home_parade" else "noting"
    return "recommended"


def _infer_audience(title: str, raw_audience: str = "") -> list[str]:
    t = title.lower()
    ra = raw_audience.lower()
    aud = []
    if any(x in t or x in ra for x in ("staff", "oic", "trgo", "adult training", "adult course")):
        aud.append("staff")
    if any(x in t or x in ra for x in ("gold", "senior", "sclp", "gclp")):
        aud.append("seniors")
    if any(x in t or x in ra for x in ("all cadets", "all cadet", "everyone")):
        return ["all_cadets"]
    if any(x in t or x in ra for x in ("all personnel", "all staff")):
        return ["all_personnel"]
    if any(x in t or x in ra for x in ("cadet", "biv", "bivouac", "camp", "training weekend")):
        if "staff" not in aud:
            aud.append("all_cadets")
    return aud or ["all_personnel"]


# ── Row skip logic ────────────────────────────────────────────────────────────

_SKIP_PATTERNS = (
    "insert new rows",
    "n/a",
    "timeline",
    "activity",
    "event",
    "date",   # header rows
)


def _should_skip(title: str, date_str: str) -> bool:
    if not title or not title.strip():
        return True
    t = title.strip().lower()
    if t in _MONTH_NAMES_FULL:  # e.g. "January" as a heading
        return True
    for pat in _SKIP_PATTERNS:
        if t == pat:
            return True
    if t.startswith("insert") or t.startswith("//") or t.startswith("#"):
        return True
    # Date cell is just a month name
    if date_str and date_str.strip().lower() in _MONTH_NAMES_FULL:
        return True
    return False


# ── Column detection ──────────────────────────────────────────────────────────

_COL_KEYWORDS = {
    "date":       ("date", "dates", "date range", "start", "when"),
    "title":      ("title", "event", "activity", "name", "description"),
    "type":       ("type", "category", "event type"),
    "audience":   ("audience", "who", "for"),
    "importance": ("importance", "priority"),
    "notes":      ("notes", "details", "source", "remarks"),
    "location":   ("location", "venue", "where"),
}


def _detect_columns(header_row: list) -> dict[str, int | None]:
    """Return column index mapping for a header row."""
    mapping: dict[str, int | None] = {k: None for k in _COL_KEYWORDS}
    for i, cell in enumerate(header_row):
        if cell is None:
            continue
        h = str(cell).strip().lower()
        for col_name, variants in _COL_KEYWORDS.items():
            if mapping[col_name] is None:
                for v in variants:
                    if v in h:
                        mapping[col_name] = i
                        break
    return mapping


# ── Main import logic ─────────────────────────────────────────────────────────

def _cell(row: tuple, idx: int | None) -> str:
    if idx is None or idx >= len(row):
        return ""
    v = row[idx]
    if v is None:
        return ""
    return str(v).strip()


def run_import(
    filepath: str,
    wing_code: str,
    year: int,
    dry_run: bool,
    sheet_name: str | None = None,
) -> None:
    try:
        import openpyxl
    except ImportError:
        print("ERROR: openpyxl not installed. Run: pip install openpyxl")
        sys.exit(1)

    wb = openpyxl.load_workbook(filepath, read_only=True, data_only=True)
    sheet = wb[sheet_name] if sheet_name and sheet_name in wb.sheetnames else wb.active
    print(f"Reading sheet: '{sheet.title}' from {filepath}")
    print(f"Target: wing_code={wing_code}, year={year}, dry_run={dry_run}")

    rows_iter = sheet.iter_rows(values_only=True)

    # Find header row — scan first 5 rows for date+title columns
    header_row = None
    col_map = None
    preamble = []
    for _ in range(5):
        row = next(rows_iter, None)
        if row is None:
            break
        preamble.append(row)
        mapping = _detect_columns(list(row))
        if mapping["date"] is not None or mapping["title"] is not None:
            header_row = row
            col_map = mapping
            break

    if col_map is None:
        # Fallback: treat first row as headers, columns A=date, B=title, C=type, D=notes
        col_map = {"date": 0, "title": 1, "type": 2, "audience": None,
                   "importance": None, "notes": 3, "location": 4}
        print("  WARNING: Could not auto-detect header row. Using default column positions (A=date, B=title, C=type, D=notes).")

    print(f"  Column mapping: {col_map}")

    db = SessionLocal()
    try:
        wing = db.query(Wing).filter(Wing.code == wing_code).first()
        if not wing:
            print(f"ERROR: Wing with code '{wing_code}' not found in database.")
            sys.exit(1)
        print(f"  Wing found: {wing.name} (id={wing.id})")

        source_system = f"{wing_code}_MASTER_{year}"
        created = updated = skipped = failed = 0
        parse_errors: list[str] = []

        for row_num, row in enumerate(rows_iter, start=2 + len(preamble)):
            if all(v is None for v in row):
                continue

            date_raw = _cell(row, col_map["date"])
            title_raw = _cell(row, col_map["title"])
            type_raw = _cell(row, col_map.get("type"))
            aud_raw = _cell(row, col_map.get("audience"))
            imp_raw = _cell(row, col_map.get("importance"))
            notes_raw = _cell(row, col_map.get("notes"))
            loc_raw = _cell(row, col_map.get("location"))

            if _should_skip(title_raw, date_raw):
                skipped += 1
                continue

            start_date, end_date = parse_date_range(date_raw, year)
            if not start_date:
                if title_raw:
                    parse_errors.append(f"Row {row_num}: could not parse date '{date_raw}' for '{title_raw[:40]}'")
                    failed += 1
                else:
                    skipped += 1
                continue

            # Only import events in the target year (allow Dec prev year / Jan next year)
            event_year = int(start_date[:4])
            if event_year not in (year - 1, year, year + 1):
                skipped += 1
                continue

            event_type = _infer_event_type(title_raw, type_raw)
            importance = _infer_importance(title_raw, event_type, imp_raw)
            audience = _infer_audience(title_raw, aud_raw)

            source_ref = f"{source_system}:{sheet.title}:row{row_num}"

            # Check for existing record (idempotent upsert)
            existing = db.query(WingHQEvent).filter(
                WingHQEvent.source_reference == source_ref,
                WingHQEvent.wing_id == wing.id,
            ).first()

            row_data = {
                "title": title_raw[:300],
                "event_type": event_type,
                "start_date": start_date,
                "end_date": end_date,
                "planning_importance": importance,
                "audience": audience,
                "location": loc_raw[:200] if loc_raw else None,
                "notes": notes_raw or None,
                "source_event_code": type_raw[:80] if type_raw else None,
            }

            if dry_run:
                action = "UPDATE" if existing else "CREATE"
                print(f"  [{action}] {start_date}"
                      f"{' → '+end_date if end_date else ''} | "
                      f"{importance:12} | {event_type:15} | {title_raw[:60]}")
                if existing:
                    updated += 1
                else:
                    created += 1
                continue

            if existing:
                changed = False
                for k, v in row_data.items():
                    if getattr(existing, k, None) != v:
                        setattr(existing, k, v)
                        changed = True
                if changed:
                    existing.year = event_year
                    existing.updated_at = datetime.utcnow()
                    updated += 1
                else:
                    skipped += 1
            else:
                ev = WingHQEvent(
                    id=str(uuid.uuid4()),
                    wing_id=wing.id,
                    year=event_year,
                    source_system=source_system,
                    source_reference=source_ref,
                    created_at=datetime.utcnow(),
                    updated_at=datetime.utcnow(),
                    is_archived=False,
                    **row_data,
                )
                db.add(ev)
                created += 1

        if not dry_run:
            db.commit()
            print(f"\nCommitted to database.")

        print(f"\n{'=== DRY RUN ===' if dry_run else '=== IMPORT COMPLETE ==='}")
        print(f"  Created:  {created}")
        print(f"  Updated:  {updated}")
        print(f"  Skipped:  {skipped}")
        print(f"  Failed:   {failed}")
        if parse_errors:
            print(f"\nParse errors ({len(parse_errors)}):")
            for e in parse_errors[:20]:
                print(f"  {e}")
            if len(parse_errors) > 20:
                print(f"  ... and {len(parse_errors)-20} more")

    except Exception as exc:
        db.rollback()
        print(f"\nERROR: {exc}")
        raise
    finally:
        db.close()


def main():
    parser = argparse.ArgumentParser(description="Import Wing HQ Calendar from XLSX")
    parser.add_argument("--file", required=True, help="Path to XLSX file")
    parser.add_argument("--wing-code", required=True, help="Wing code (e.g. 7WG)")
    parser.add_argument("--year", type=int, default=2026, help="Target year (default: 2026)")
    parser.add_argument("--sheet", default=None, help="Sheet name (default: active sheet)")
    parser.add_argument("--dry-run", action="store_true", help="Show what would be imported without writing")
    args = parser.parse_args()

    if not Path(args.file).exists():
        print(f"ERROR: File not found: {args.file}")
        sys.exit(1)

    run_import(
        filepath=args.file,
        wing_code=args.wing_code,
        year=args.year,
        dry_run=args.dry_run,
        sheet_name=args.sheet,
    )


if __name__ == "__main__":
    main()
